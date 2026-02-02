from __future__ import annotations

import io
import json
import logging
import os
import re
import unicodedata
from calendar import monthrange
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal, ROUND_CEILING, ROUND_HALF_UP
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET

from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.core.files.base import ContentFile
from django.db import models, transaction
from django.forms import modelformset_factory
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse, reverse_lazy
from django.templatetags.static import static
from django.utils import formats, timezone
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.generic import (
    CreateView,
    DeleteView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from .forms import (
    AccountPlanTemplateHeaderForm,
    AccountPlanTemplateImportForm,
    AccountPlanTemplateItemForm,
    CertificationForm,
    ClientContactForm,
    ClientForm,
    CompanyForm,
    CompanyBankAccountForm,
    SupplierForm,
    AccountsPayableForm,
    AccountsPayablePaymentForm,
    AccountsPayableAttachmentForm,
    AccountsReceivableForm,
    AccountsReceivablePaymentForm,
    AccountsCompensationForm,
    TravelReimbursementForm,
    CompetencyForm,
    ConsultantForm,
    ConsultantAttachmentForm,
    ConsultantBankAccountForm,
    ConsultantRateForm,
    DeploymentTemplateImportForm,
    DeploymentTemplateHeaderForm,
    DeploymentTemplateItemForm,
    DeploymentTemplateMaintenanceForm,
    ModuleForm,
    PhaseForm,
    ProjectAttachmentForm,
    ProjectContactForm,
    ProjectForm,
    ProjectGoNoGoChecklistItemForm,
    ProjectObservationForm,
    ProjectOccurrenceAttachmentForm,
    ProjectOccurrenceForm,
    ProjectRoleForm,
    ProjectActivityFeedbackForm,
    ProjectActivityForm,
    ProjectActivityGenerateForm,
    ProductForm,
    SubmoduleForm,
    SubmoduleBulkCreateForm,
    TimeEntryForm,
    TimeEntryReviewForm,
    UserCreateForm,
    UserProfileForm,
    WhatsappSettingsForm,
    ChatGPTSettingsForm,
    TicketForm,
    TicketReplyForm,
    KnowledgeCategoryForm,
    KnowledgePostForm,
)
from .importers import import_account_plan_templates, import_deployment_templates
from .models import (
    AccountPlanTemplateHeader,
    AccountPlanTemplateItem,
    DreSign,
    AccountType,
    ActivityStatus,
    ActivityBillingType,
    ActivityCriticality,
    BillingInvoice,
    BillingInvoiceItem,
    BillingPaymentStatus,
    BankStatementImport,
    BankStatementEntry,
    BankSystemMovement,
    BankReconciliation,
    BankReconciliationSystemItem,
    BankReconciliationOfxItem,
    BankMovementDirection,
    BankMovementSource,
    Certification,
    Client,
    ClientContact,
    Company,
    CompanyType,
    CompanyBankAccount,
    Supplier,
    SupplierPersonType,
    AccountsPayable,
    AccountsPayablePayment,
    AccountsPayableAttachment,
    AccountsReceivable,
    AccountsReceivablePayment,
    FinancialStatus,
    PaymentMethod,
    Competency,
    Consultant,
    ConsultantAttachment,
    ConsultantBankAccount,
    ConsultantRate,
    DeploymentTemplate,
    DeploymentTemplateHeader,
    Module,
    Phase,
    Project,
    ProjectActivity,
    ProjectActivitySubactivity,
    ProjectContact,
    ProjectCriticality,
    ProjectGoNoGoChecklistItem,
    GoNoGoResult,
    ProjectObservation,
    ProjectObservationType,
    ProjectOccurrence,
    ProjectOccurrenceAttachment,
    ProjectRole,
    TimeEntry,
    TimeEntryStatus,
    ProjectStatus,
    ProjectType,
    ProjectAttachment,
    ProjectAttachmentType,
    Product,
    Submodule,
    UserProfile,
    UserRole,
    StatusChoices,
    TimeEntryAttachment,
    Ticket,
    TicketAttachment,
    TicketReply,
    TicketReplyAttachment,
    TicketStatus,
    KnowledgeCategory,
    KnowledgePost,
    KnowledgeAttachment,
    WhatsappSettings,
    ChatGPTSettings,
)
from .roles import (
    allowed_project_visibility,
    can_view_financial,
    filter_by_visibility,
    filter_activities_for_user,
    filter_projects_for_user,
    resolve_user_role,
)
from .observations import (
    create_project_change_observation,
    create_project_receipt_observation,
)
from .whatsapp_notifications import (
    notify_admin_payable_created,
    notify_admin_payable_paid,
    notify_admin_receivable_created,
    notify_admin_receivable_paid,
    notify_consultant_activity_assigned,
    notify_consultant_billing_closure,
    notify_consultant_payable_created,
    notify_consultant_payable_paid,
    notify_opportunity_candidate,
    notify_ticket_closed,
    notify_ticket_created,
    notify_ticket_reply,
    notify_time_entry_pending,
    notify_time_entry_reviewed,
)

User = get_user_model()


def _resolve_attr(obj: Any, attr: str) -> Any:
    value = obj
    for part in attr.split("."):
        if value is None:
            return None
        value = getattr(value, part, None)
    return value


def _format_value(obj: models.Model, field: str) -> str:
    if "." in field:
        value = _resolve_attr(obj, field)
        return "-" if value in {None, ""} else str(value)

    model_field = obj._meta.get_field(field)
    value = getattr(obj, field, None)
    if value in {None, ""}:
        return "-"
    if model_field.choices:
        display = getattr(obj, f"get_{field}_display", None)
        if display:
            return display()
    if isinstance(model_field, models.DateTimeField):
        value = timezone.localtime(value) if timezone.is_aware(value) else value
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(model_field, models.DateField):
        return value.strftime("%d/%m/%Y")
    if isinstance(model_field, models.TimeField):
        return value.strftime("%H:%M")
    if isinstance(model_field, models.DecimalField):
        return formats.number_format(
            value,
            decimal_pos=model_field.decimal_places,
            use_l10n=True,
            force_grouping=True,
        )
    if isinstance(model_field, models.BooleanField):
        return "Sim" if value else "Nao"
    return str(value)


def _apply_visibility_choices(form: forms.Form, role: str | None) -> None:
    field = form.fields.get("visibility")
    if not field:
        return
    allowed = allowed_project_visibility(role)
    field.choices = [choice for choice in field.choices if choice[0] in allowed]


def _resolve_open_amount(
    total_amount: Decimal, paid_total: Decimal | None
) -> Decimal:
    total = total_amount or Decimal("0.00")
    paid = paid_total or Decimal("0.00")
    remaining = total - paid
    if remaining < 0:
        return Decimal("0.00")
    return remaining.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


_COMPENSATION_NOTE_PREFIX = "[COMPENSACAO]"


def _build_compensation_notes(notes: str | None) -> str:
    base = (notes or "").strip()
    if base:
        return f"{_COMPENSATION_NOTE_PREFIX} {base}"
    return _COMPENSATION_NOTE_PREFIX


def _exclude_compensation_notes(queryset):
    return queryset.exclude(notes__startswith=_COMPENSATION_NOTE_PREFIX)


def _normalize_doc(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\D", "", value or "")


def _normalize_name(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", value)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip().lower()


_DUE_BUCKETS = [
    {"key": "today", "label": "Hoje", "color": "#c05d54"},
    {"key": "tomorrow", "label": "Amanha", "color": "#e29d72"},
    {"key": "five", "label": "5 dias", "color": "#f2c26b"},
    {"key": "thirty", "label": "30 dias", "color": "#8bbf7a"},
    {"key": "forty_five", "label": "45 dias", "color": "#4f9aa5"},
    {"key": "sixty", "label": "60 dias", "color": "#3b6e91"},
    {"key": "ninety", "label": "90 dias", "color": "#5c6f84"},
    {"key": "over_ninety", "label": "Mais de 90 dias", "color": "#94a09f"},
]


def _resolve_due_bucket(days: int) -> str:
    if days <= 0:
        return "today"
    if days == 1:
        return "tomorrow"
    if days <= 5:
        return "five"
    if days <= 30:
        return "thirty"
    if days <= 45:
        return "forty_five"
    if days <= 60:
        return "sixty"
    if days <= 90:
        return "ninety"
    return "over_ninety"


def _build_due_charts(
    titles: Iterable[AccountsReceivable | AccountsPayable | dict[str, Any]],
    variant: str,
    subtitle: str,
) -> dict[str, Any]:
    today = timezone.localdate()
    bucket_totals = {item["key"]: Decimal("0.00") for item in _DUE_BUCKETS}
    day_values: list[int] = []
    open_count = 0

    for title in titles:
        if isinstance(title, dict):
            status = title.get("status")
            due_date = title.get("due_date")
            amount = title.get("amount") or Decimal("0.00")
            discount = title.get("discount") or Decimal("0.00")
            interest = title.get("interest") or Decimal("0.00")
            penalty = title.get("penalty") or Decimal("0.00")
            total_amount = (amount - discount + interest + penalty).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            paid_total = title.get("paid_total")
        else:
            status = title.status
            due_date = title.due_date
            total_amount = title.total_amount()
            paid_total = getattr(title, "paid_total", None)

        if status in {FinancialStatus.PAID, FinancialStatus.CANCELED}:
            continue
        if not due_date:
            continue
        open_amount = _resolve_open_amount(total_amount, paid_total)
        if open_amount <= 0:
            continue
        open_count += 1
        days = (due_date - today).days
        day_values.append(days)
        bucket_key = _resolve_due_bucket(days)
        bucket_totals[bucket_key] += open_amount

    avg_days_raw = sum(day_values) / len(day_values) if day_values else None
    avg_display = "-"
    avg_days = 0.0
    if avg_days_raw is not None:
        avg_days = float(avg_days_raw)
        avg_display = f"{int(round(avg_days))} dias"

    gauge_max = 365.0
    clamped = max(0.0, min(avg_days, gauge_max))
    percent = (clamped / gauge_max) * 100 if gauge_max else 0.0
    needle_angle = -90 + (percent / 100) * 180
    default_gauge = "var(--teal)" if variant == "receivable" else "#e29d72"
    gauge_color = (
        "#c05d54" if avg_days_raw is not None and avg_days_raw < 0 else default_gauge
    )

    total_amount = sum(bucket_totals.values(), Decimal("0.00"))
    items = []
    non_zero = [
        item["key"] for item in _DUE_BUCKETS if bucket_totals[item["key"]] > 0
    ]
    last_key = non_zero[-1] if non_zero else None
    current = Decimal("0.00")
    segments = []

    for item in _DUE_BUCKETS:
        amount = bucket_totals[item["key"]]
        percent_value = (
            (amount / total_amount) * Decimal("100")
            if total_amount > 0
            else Decimal("0.00")
        )
        percent_value = percent_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        items.append(
            {
                "label": item["label"],
                "value_display": _format_currency_value(amount),
                "percent": f"{percent_value:.0f}",
                "color": item["color"],
            }
        )
        if total_amount <= 0 or amount <= 0:
            continue
        start = current
        end = current + percent_value
        if item["key"] == last_key:
            end = Decimal("100.00")
        segments.append(f"{item['color']} {start:.2f}% {end:.2f}%")
        current = end

    chart_style = (
        f"conic-gradient({', '.join(segments)})"
        if segments
        else "conic-gradient(var(--line) 0% 100%)"
    )

    return {
        "title": "Indicadores de vencimento",
        "subtitle": subtitle,
        "variant": variant,
        "average": {
            "available": open_count > 0,
            "display": avg_display,
            "label": "Vencimento medio",
            "percent": f"{percent:.2f}",
            "needle_angle": f"{needle_angle:.2f}",
            "scale_label": f"0 a {int(gauge_max)} dias",
            "open_count": open_count,
            "gauge_color": gauge_color,
        },
        "aging": {
            "available": total_amount > 0,
            "total_display": _format_currency_value(total_amount),
            "chart_style": chart_style,
            "items": items,
        },
    }


def _format_decimal_value(value: Decimal) -> str:
    return formats.number_format(
        value,
        decimal_pos=2,
        use_l10n=True,
        force_grouping=True,
    )


def _format_currency_value(value: Decimal) -> str:
    return f"R$ {_format_decimal_value(value)}"


def _warn_project_activity_overage(
    project: Project,
    user,
    request=None,
) -> None:
    totals = ProjectActivity.objects.filter(project=project).aggregate(
        total_hours=Coalesce(Sum("hours"), Value(Decimal("0.00"))),
    )
    hours_total = totals.get("total_hours") or Decimal("0.00")
    hourly_rate = project.hourly_rate or Decimal("0.00")
    value_total = (hours_total * hourly_rate).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )
    available_hours = project.available_hours or Decimal("0.00")
    available_value = project.available_value or Decimal("0.00")
    if hours_total <= available_hours and value_total <= available_value:
        return
    note = (
        "Alerta: a soma das atividades "
        f"({_format_decimal_value(hours_total)}h / {_format_currency_value(value_total)}) "
        "ultrapassa o limite do projeto considerando a contingencia "
        f"({_format_decimal_value(available_hours)}h / {_format_currency_value(available_value)})."
    )
    ProjectObservation.objects.create(
        project=project,
        observation_type=ProjectObservationType.AUTO,
        note=note,
        created_by=user,
    )
    if request is not None:
        messages.warning(request, note)


def _get_activity_subactivities(activity: ProjectActivity) -> list[str]:
    items = []
    related = getattr(activity, "subactivity_items", None)
    if related is not None:
        items = [
            item.description.strip()
            for item in related.all()
            if item.description and item.description.strip()
        ]
    if items:
        return items
    legacy = (activity.subactivity or "").strip()
    return [legacy] if legacy else []


def _format_activity_subactivities(
    activity: ProjectActivity, placeholder: str = "-"
) -> str:
    items = _get_activity_subactivities(activity)
    return ", ".join(items) if items else placeholder


def _filter_by_subactivity(queryset, value: str, prefix: str = ""):
    if not value:
        return queryset
    if prefix:
        prefix = f"{prefix}__"
    return queryset.filter(
        Q(**{f"{prefix}subactivity__icontains": value})
        | Q(**{f"{prefix}subactivity_items__description__icontains": value})
    ).distinct()


def _sync_subactivity_items(activity: ProjectActivity, items: list[str]) -> None:
    normalized = []
    seen = set()
    for item in items:
        text = item.strip() if item else ""
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        normalized.append(text)
    activity.subactivity = normalized[0] if normalized else ""
    activity.save(update_fields=["subactivity"])
    activity.subactivity_items.all().delete()
    if not normalized:
        return
    ProjectActivitySubactivity.objects.bulk_create(
        [
            ProjectActivitySubactivity(
                activity=activity,
                description=item,
                order=index + 1,
            )
            for index, item in enumerate(normalized)
        ]
    )


def _resolve_duration_days(value: Decimal | int | None) -> int:
    if value is None:
        return 0
    try:
        amount = Decimal(value)
    except Exception:
        return 0
    if amount <= 0:
        return 0
    return int(amount.to_integral_value(rounding=ROUND_CEILING))


def _build_consultant_rate_map(consultants: Iterable[Consultant]) -> dict[int, float]:
    target_date = timezone.localdate()
    consultant_ids = [consultant.id for consultant in consultants]
    if not consultant_ids:
        return {}
    rate_map: dict[int, float] = {}
    rates = (
        ConsultantRate.objects.filter(consultant_id__in=consultant_ids)
        .order_by("consultant_id", "-start_date")
    )
    for rate in rates:
        if rate.start_date <= target_date and (
            rate.end_date is None or rate.end_date >= target_date
        ):
            if rate.consultant_id not in rate_map:
                rate_map[rate.consultant_id] = float(rate.rate or 0)
    return rate_map


def _save_time_entry_attachments(time_entry: TimeEntry, files) -> int:
    if not files:
        return 0
    count = 0
    for uploaded in files:
        if not uploaded:
            continue
        TimeEntryAttachment.objects.create(time_entry=time_entry, file=uploaded)
        count += 1
    return count


def _save_ticket_attachments(ticket: Ticket, files) -> int:
    if not files:
        return 0
    count = 0
    for uploaded in files:
        if not uploaded:
            continue
        TicketAttachment.objects.create(ticket=ticket, file=uploaded)
        count += 1
    return count


def _save_ticket_reply_attachments(reply: TicketReply, files) -> int:
    if not files:
        return 0
    count = 0
    for uploaded in files:
        if not uploaded:
            continue
        TicketReplyAttachment.objects.create(reply=reply, file=uploaded)
        count += 1
    return count


def _save_knowledge_attachments(post: KnowledgePost, files, user) -> int:
    if not files:
        return 0
    count = 0
    for uploaded in files:
        if not uploaded:
            continue
        KnowledgeAttachment.objects.create(
            post=post,
            file=uploaded,
            uploaded_by=user,
        )
        count += 1
    return count


def _format_user_label(user) -> str:
    if not user:
        return "-"
    return user.get_short_name() or user.get_full_name() or user.username


def _get_admin_user_ids() -> set[int]:
    admin_users = User.objects.filter(
        Q(is_superuser=True)
        | Q(is_staff=True)
        | Q(profile__role=UserRole.ADMIN)
    ).values_list("id", flat=True)
    return set(admin_users)


def _get_assignable_users(project: Project | None = None):
    admin_ids = _get_admin_user_ids()
    if not project:
        return (
            User.objects.filter(
                Q(is_superuser=True)
                | Q(is_staff=True)
                | Q(profile__role__in=UserRole.values)
            )
            .distinct()
            .order_by("first_name", "last_name", "username")
        )

    user_ids = set(admin_ids)
    if project.internal_manager_id:
        user_ids.add(project.internal_manager_id)
    if project.external_manager_id:
        user_ids.add(project.external_manager_id)
    if project.client_user_id:
        user_ids.add(project.client_user_id)
    consultant_user_ids = Consultant.objects.filter(
        project_activities__project=project,
        user__isnull=False,
    ).values_list("user_id", flat=True)
    user_ids.update(consultant_user_ids)
    if not user_ids:
        user_ids = admin_ids
    return User.objects.filter(id__in=user_ids).distinct().order_by(
        "first_name",
        "last_name",
        "username",
    )


def _filter_tickets_for_user(queryset, user):
    role = resolve_user_role(user)
    if role == UserRole.ADMIN:
        return queryset
    if role == UserRole.GP_INTERNAL:
        return queryset.filter(
            Q(project__internal_manager=user)
            | Q(created_by=user)
            | Q(assigned_to=user)
        ).distinct()
    if role == UserRole.GP_EXTERNAL:
        return queryset.filter(
            Q(project__external_manager=user)
            | Q(created_by=user)
            | Q(assigned_to=user)
        ).distinct()
    if role == UserRole.CONSULTANT:
        return queryset.filter(
            Q(created_by=user)
            | Q(assigned_to=user)
            | Q(consultant_responsible__user=user)
        ).distinct()
    if role == UserRole.CLIENT:
        return queryset.filter(
            Q(created_by=user) | Q(assigned_to=user)
        ).distinct()
    return queryset.none()


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/dashboard.html"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(self._build_project_charts())
        if resolve_user_role(self.request.user) == UserRole.ADMIN:
            context.update(self._build_operational_panel())
            context.update(self._build_dre_panel())
        return context

    def _format_decimal(self, value: Decimal, places: int = 2) -> str:
        return formats.number_format(
            value,
            decimal_pos=places,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value: Decimal) -> str:
        return f"R$ {self._format_decimal(value, 2)}"

    def _parse_date(self, value: str | None):
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    def _format_period_label(self, start_date: date | None, end_date: date | None) -> str:
        if start_date and end_date:
            return f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        if start_date:
            return f"A partir de {start_date.strftime('%d/%m/%Y')}"
        if end_date:
            return f"Ate {end_date.strftime('%d/%m/%Y')}"
        return "Todo o periodo"

    def _resolve_schedule_badge(
        self,
        activity: ProjectActivity,
        today: date,
    ) -> tuple[str | None, str]:
        planned_end = activity.planned_end
        actual_end = activity.actual_end
        if planned_end and actual_end:
            if actual_end < planned_end:
                return "Antecipada", "chip-info"
            if actual_end > planned_end:
                return "Atrasada", "chip-danger"
            return "No prazo", "chip-ok"
        schedule_state = activity.schedule_state(today)
        if schedule_state == "late":
            return "Atrasada", "chip-danger"
        if schedule_state == "on_time":
            return "No prazo", "chip-ok"
        if schedule_state == "not_started":
            return "Nao iniciada", "chip-neutral"
        return None, ""

    def _resolve_activity_status(
        self,
        activity: ProjectActivity,
        today: date,
    ) -> tuple[str, str]:
        if activity.status == ActivityStatus.DONE:
            return "Finalizada", "chip-ok"
        if activity.status in {ActivityStatus.IN_PROGRESS, ActivityStatus.RELEASED} or activity.actual_start:
            return "Em Andamento", "chip-info"
        return "Nao Iniciada", "chip-neutral"

    def _percent(self, part: Decimal, total: Decimal) -> int:
        if total <= 0:
            return 0
        value = (part / total) * Decimal("100")
        percent = int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        return max(0, min(100, percent))

    def _resolve_rate_for_date(
        self,
        consultant: Consultant,
        target_date: date,
    ) -> Decimal | None:
        valid_rates = []
        for rate in consultant.rates.all():
            if rate.start_date <= target_date and (
                rate.end_date is None or rate.end_date >= target_date
            ):
                valid_rates.append(rate)
        if not valid_rates:
            return None
        valid_rates.sort(key=lambda rate: rate.start_date, reverse=True)
        return valid_rates[0].rate

    def _estimate_consultant_cost(
        self,
        activity: ProjectActivity,
        target_date: date,
        selected_rate: Decimal | None,
        has_consultant_filter: bool,
    ) -> Decimal:
        hours = activity.hours or Decimal("0.00")
        if hours <= 0:
            return Decimal("0.00")
        if activity.consultant_hourly_rate is not None:
            return (hours * activity.consultant_hourly_rate).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
        if has_consultant_filter:
            if selected_rate is None:
                return Decimal("0.00")
            return (hours * selected_rate).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        rates = []
        for consultant in activity.consultants.all():
            rate_value = self._resolve_rate_for_date(consultant, target_date)
            if rate_value:
                rates.append(rate_value)
        if not rates:
            return Decimal("0.00")
        average_rate = sum(rates, Decimal("0.00")) / Decimal(len(rates))
        return (hours * average_rate).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

    def _resolve_progress_state(
        self,
        entry: dict[str, Decimal],
        project: Project,
    ) -> tuple[str, str, str]:
        schedule_state = entry.get("schedule_state")
        total_hours = entry["total_hours"]
        approved_hours = entry["approved_hours"]
        pending_hours = entry["pending_hours"]
        planned_hours = entry["planned_hours"]
        released_hours = entry.get("released_hours", Decimal("0.00"))
        in_progress_hours = entry["in_progress_hours"]
        blocked_hours = entry["blocked_hours"]
        if schedule_state == "late":
            return "late", "Atrasada", "chip-danger"
        if blocked_hours > 0 or project.status in {ProjectStatus.PAUSED, ProjectStatus.CANCELED}:
            return "late", "Atrasada", "chip-danger"
        if schedule_state == "not_started":
            return "not-started", "Nao iniciada", "chip-info"
        if schedule_state == "on_time":
            return "on-time", "No prazo", "chip-ok"
        has_started = approved_hours > 0 or pending_hours > 0 or in_progress_hours > 0
        if total_hours <= 0 or ((planned_hours + released_hours) > 0 and not has_started):
            return "not-started", "Nao iniciada", "chip-info"
        return "on-time", "No prazo", "chip-ok"

    def _apply_status_filter(self, queryset, status: str):
        if status == "pending":
            return queryset.filter(status=ActivityStatus.IN_PROGRESS)
        if status == "planned":
            return queryset.filter(status=ActivityStatus.PLANNED)
        if status == "released":
            return queryset.filter(status=ActivityStatus.RELEASED)
        if status == "done":
            return queryset.filter(status=ActivityStatus.DONE)
        if status == "paused":
            return queryset.filter(
                Q(status=ActivityStatus.BLOCKED)
                | Q(project__status=ProjectStatus.PAUSED)
            )
        if status == "canceled":
            return queryset.filter(
                Q(status=ActivityStatus.CANCELED)
                | Q(project__status=ProjectStatus.CANCELED)
            )
        return queryset

    def _build_s_curve_chart(
        self,
        activities: list[ProjectActivity],
        selected_consultant_id: int | None,
    ) -> dict[str, Any]:
        planned_by_date: dict[date, Decimal] = defaultdict(lambda: Decimal("0.00"))
        actual_by_date: dict[date, Decimal] = defaultdict(lambda: Decimal("0.00"))
        for activity in activities:
            if (
                activity.planned_start
                and activity.planned_end
                and (activity.hours or Decimal("0.00")) > 0
            ):
                start = activity.planned_start
                end = activity.planned_end
                if end >= start:
                    days = (end - start).days + 1
                    daily = (activity.hours / Decimal(days)).quantize(
                        Decimal("0.01"),
                        rounding=ROUND_HALF_UP,
                    )
                    for day_offset in range(days):
                        planned_by_date[start + timedelta(days=day_offset)] += daily
            for time_entry in activity.time_entries.all():
                if time_entry.status != TimeEntryStatus.APPROVED:
                    continue
                if selected_consultant_id and time_entry.consultant_id != selected_consultant_id:
                    continue
                start = time_entry.start_date
                end = time_entry.end_date or time_entry.start_date
                if end < start:
                    continue
                days = (end - start).days + 1
                daily = (time_entry.total_hours / Decimal(days)).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                for day_offset in range(days):
                    actual_by_date[start + timedelta(days=day_offset)] += daily

        if not planned_by_date and not actual_by_date:
            return {"available": False}

        dates = sorted(set(planned_by_date.keys()) | set(actual_by_date.keys()))
        total_planned = sum(planned_by_date.values(), Decimal("0.00"))
        total_actual = sum(actual_by_date.values(), Decimal("0.00"))
        if total_planned <= 0 and total_actual <= 0:
            return {"available": False}

        scale = total_planned if total_planned > 0 else total_actual
        cumulative = []
        cumulative_planned = Decimal("0.00")
        cumulative_actual = Decimal("0.00")
        for current_date in dates:
            cumulative_planned += planned_by_date.get(current_date, Decimal("0.00"))
            cumulative_actual += actual_by_date.get(current_date, Decimal("0.00"))
            cumulative.append(
                {
                    "date": current_date,
                    "planned": cumulative_planned,
                    "actual": cumulative_actual,
                }
            )

        max_points = 24
        if len(cumulative) > max_points:
            indices = [
                int(round(i * (len(cumulative) - 1) / (max_points - 1)))
                for i in range(max_points)
            ]
            sampled = []
            last_index = None
            for idx in indices:
                if idx != last_index:
                    sampled.append(cumulative[idx])
                    last_index = idx
        else:
            sampled = cumulative

        def _build_points(series_key: str) -> str:
            if not sampled:
                return ""
            points = []
            for idx, item in enumerate(sampled):
                if len(sampled) == 1:
                    x = 0.0
                else:
                    x = (idx / (len(sampled) - 1)) * 100
                percent = (item[series_key] / scale) * Decimal("100")
                percent = max(Decimal("0.00"), min(Decimal("100.00"), percent))
                y = 40 - (float(percent) / 100) * 40
                points.append(f"{x:.2f},{y:.2f}")
            return " ".join(points)

        label_start = sampled[0]["date"].strftime("%d/%m") if sampled else ""
        label_mid = sampled[len(sampled) // 2]["date"].strftime("%d/%m") if sampled else ""
        label_end = sampled[-1]["date"].strftime("%d/%m") if sampled else ""

        return {
            "available": True,
            "planned_points": _build_points("planned"),
            "actual_points": _build_points("actual"),
            "label_start": label_start,
            "label_mid": label_mid,
            "label_end": label_end,
            "total_planned": self._format_decimal(total_planned),
            "total_actual": self._format_decimal(total_actual),
        }

    def _build_activity_status_chart(
        self,
        activities: list[ProjectActivity],
    ) -> dict[str, Any]:
        done_count = sum(1 for activity in activities if activity.status == ActivityStatus.DONE)
        pending_count = sum(1 for activity in activities if activity.status != ActivityStatus.DONE)
        total = pending_count + done_count
        if total <= 0:
            return {"available": False}

        pending_percent = self._percent(Decimal(pending_count), Decimal(total))
        done_percent = self._percent(Decimal(done_count), Decimal(total))
        return {
            "available": True,
            "pending_count": pending_count,
            "done_count": done_count,
            "pending_percent": pending_percent,
            "done_percent": done_percent,
            "total": total,
        }

    def _build_activity_billing_chart(
        self,
        totals: dict[str, Decimal],
    ) -> dict[str, Any]:
        total_hours = sum(totals.values(), Decimal("0.00"))
        order = [
            (ActivityBillingType.BILLABLE, "billable"),
            (ActivityBillingType.ASSUMED_COMPANY, "assumed-company"),
            (ActivityBillingType.ASSUMED_CONSULTANT, "assumed-consultant"),
            (ActivityBillingType.CLIENT_ASSIGNED, "client-assigned"),
        ]
        items = []
        non_zero = [
            key
            for choice, key in order
            if totals.get(choice, Decimal("0.00")) > 0
        ]
        current = Decimal("0.00")
        segments = []
        last_key = non_zero[-1] if non_zero else None

        for choice, key in order:
            hours = totals.get(choice, Decimal("0.00"))
            percent = (
                (hours / total_hours) * Decimal("100")
                if total_hours > 0
                else Decimal("0.00")
            )
            percent = percent.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            items.append(
                {
                    "key": key,
                    "label": choice.label,
                    "hours": self._format_decimal(hours),
                    "percent": f"{percent:.2f}",
                }
            )
            if total_hours <= 0 or hours <= 0:
                continue
            start = current
            end = current + percent
            if key == last_key:
                end = Decimal("100.00")
            segments.append(
                f"var(--pie-{key}) {start:.2f}% {end:.2f}%"
            )
            current = end

        chart_style = (
            f"conic-gradient({', '.join(segments)})"
            if segments
            else "conic-gradient(var(--line) 0% 100%)"
        )

        return {
            "available": total_hours > 0,
            "total_hours": self._format_decimal(total_hours),
            "items": items,
            "chart_style": chart_style,
        }

    def _build_activity_billing_charts(
        self,
        activities: list[ProjectActivity],
        selected_consultant_id: int | None,
    ) -> dict[str, Any]:
        planned_totals = {
            ActivityBillingType.BILLABLE: Decimal("0.00"),
            ActivityBillingType.ASSUMED_COMPANY: Decimal("0.00"),
            ActivityBillingType.ASSUMED_CONSULTANT: Decimal("0.00"),
            ActivityBillingType.CLIENT_ASSIGNED: Decimal("0.00"),
        }
        actual_totals = {
            ActivityBillingType.BILLABLE: Decimal("0.00"),
            ActivityBillingType.ASSUMED_COMPANY: Decimal("0.00"),
            ActivityBillingType.ASSUMED_CONSULTANT: Decimal("0.00"),
            ActivityBillingType.CLIENT_ASSIGNED: Decimal("0.00"),
        }

        for activity in activities:
            billing_type = activity.billing_type or ActivityBillingType.BILLABLE
            if billing_type not in planned_totals:
                planned_totals[billing_type] = Decimal("0.00")
                actual_totals[billing_type] = Decimal("0.00")
            planned_totals[billing_type] += activity.hours or Decimal("0.00")
            for time_entry in activity.time_entries.all():
                if time_entry.status != TimeEntryStatus.APPROVED:
                    continue
                if (
                    selected_consultant_id
                    and time_entry.consultant_id != selected_consultant_id
                ):
                    continue
                actual_totals[billing_type] += time_entry.total_hours or Decimal("0.00")

        return {
            "planned": self._build_activity_billing_chart(planned_totals),
            "actual": self._build_activity_billing_chart(actual_totals),
        }

    def _build_operational_panel(self) -> dict[str, Any]:
        params = self.request.GET
        filters = {
            "project_id": params.get("project_id", "").strip(),
            "consultant_id": params.get("consultant_id", "").strip(),
            "internal_manager_id": params.get("internal_manager_id", "").strip(),
            "product_id": params.get("product_id", "").strip(),
            "module_id": params.get("module_id", "").strip(),
            "submodule_id": params.get("submodule_id", "").strip(),
            "activity": params.get("activity", "").strip(),
            "subactivity": params.get("subactivity", "").strip(),
            "status": params.get("status", "").strip(),
        }

        activities = (
            ProjectActivity.objects.select_related(
                "project",
                "project__internal_manager",
                "product",
                "module",
                "submodule",
            )
            .prefetch_related(
                "consultants",
                "consultants__rates",
                "time_entries",
                "subactivity_items",
            )
            .order_by("project__description", "seq")
        )
        activities = filter_activities_for_user(activities, self.request.user)

        if filters["project_id"]:
            activities = activities.filter(project_id=filters["project_id"])
        if filters["consultant_id"]:
            activities = activities.filter(consultants__id=filters["consultant_id"])
        if filters["internal_manager_id"]:
            activities = activities.filter(
                project__internal_manager_id=filters["internal_manager_id"]
            )
        if filters["product_id"]:
            activities = activities.filter(product_id=filters["product_id"])
        if filters["module_id"]:
            activities = activities.filter(module_id=filters["module_id"])
        if filters["submodule_id"]:
            activities = activities.filter(submodule_id=filters["submodule_id"])
        if filters["activity"]:
            activities = activities.filter(activity__icontains=filters["activity"])
        if filters["subactivity"]:
            activities = _filter_by_subactivity(activities, filters["subactivity"])
        if filters["status"]:
            activities = self._apply_status_filter(activities, filters["status"])
        activities = list(activities)

        status_options = [
            {"value": "pending", "label": "Pendente"},
            {"value": "planned", "label": "Planejada"},
            {"value": "released", "label": "Liberada"},
            {"value": "done", "label": "Concluida"},
            {"value": "paused", "label": "Paralizada"},
            {"value": "canceled", "label": "Cancelada"},
        ]

        projects = list(
            filter_projects_for_user(
                Project.objects.select_related("internal_manager"),
                self.request.user,
            ).order_by("description")
        )
        consultants = Consultant.objects.order_by("full_name")
        internal_manager_map = {}
        for project in projects:
            manager = project.internal_manager
            if manager and manager.pk not in internal_manager_map:
                internal_manager_map[manager.pk] = manager
        internal_managers = sorted(
            internal_manager_map.values(),
            key=lambda user: user.get_short_name()
            or user.get_full_name()
            or user.username,
        )
        products = Product.objects.order_by("description")
        modules = Module.objects.select_related("product").order_by("description")
        submodules = Submodule.objects.select_related("module").order_by("description")

        rows_map: dict[int, dict[str, Any]] = {}
        today = timezone.localdate()
        has_consultant_filter = bool(filters["consultant_id"])
        selected_rate = None
        selected_consultant_id = None
        if has_consultant_filter:
            selected_consultant = (
                Consultant.objects.filter(pk=filters["consultant_id"])
                .prefetch_related("rates")
                .first()
            )
            if selected_consultant:
                selected_consultant_id = selected_consultant.pk
                selected_rate = self._resolve_rate_for_date(
                    selected_consultant,
                    today,
                )
        for activity in activities:
            project = activity.project
            entry = rows_map.get(project.id)
            if not entry:
                entry = {
                    "project": project,
                    "total_hours": Decimal("0.00"),
                    "approved_hours": Decimal("0.00"),
                    "pending_hours": Decimal("0.00"),
                    "consultant_cost": Decimal("0.00"),
                    "planned_hours": Decimal("0.00"),
                    "released_hours": Decimal("0.00"),
                    "in_progress_hours": Decimal("0.00"),
                    "blocked_hours": Decimal("0.00"),
                    "schedule_state": None,
                }
                rows_map[project.id] = entry

            hours = activity.hours or Decimal("0.00")
            entry["total_hours"] += hours
            for time_entry in activity.time_entries.all():
                if (
                    selected_consultant_id
                    and time_entry.consultant_id != selected_consultant_id
                ):
                    continue
                if time_entry.status == TimeEntryStatus.APPROVED:
                    entry["approved_hours"] += time_entry.total_hours
                elif time_entry.status == TimeEntryStatus.PENDING:
                    entry["pending_hours"] += time_entry.total_hours
            if activity.status == ActivityStatus.PLANNED:
                entry["planned_hours"] += hours
            elif activity.status == ActivityStatus.RELEASED:
                entry["released_hours"] += hours
            elif activity.status == ActivityStatus.IN_PROGRESS:
                entry["in_progress_hours"] += hours
            elif activity.status == ActivityStatus.BLOCKED:
                entry["blocked_hours"] += hours
            activity_schedule = activity.schedule_state(today)
            if activity_schedule == "late":
                entry["schedule_state"] = "late"
            elif activity_schedule == "on_time" and entry["schedule_state"] != "late":
                entry["schedule_state"] = "on_time"
            elif (
                activity_schedule == "not_started"
                and entry["schedule_state"] not in {"late", "on_time"}
            ):
                entry["schedule_state"] = "not_started"
            entry["consultant_cost"] += self._estimate_consultant_cost(
                activity,
                today,
                selected_rate,
                has_consultant_filter,
            )

        s_curve = self._build_s_curve_chart(activities, selected_consultant_id)
        activity_status = self._build_activity_status_chart(activities)
        billing_charts = self._build_activity_billing_charts(
            activities,
            selected_consultant_id,
        )

        rows = []
        for entry in sorted(
            rows_map.values(),
            key=lambda item: item["project"].description,
        ):
            project = entry["project"]
            percent = self._percent(entry["approved_hours"], entry["total_hours"])
            progress_state, progress_label, progress_chip = self._resolve_progress_state(
                entry,
                project,
            )
            internal_manager = (
                project.internal_manager.get_short_name()
                if project.internal_manager
                else ""
            )
            if not internal_manager and project.internal_manager:
                internal_manager = (
                    project.internal_manager.get_full_name()
                    or project.internal_manager.username
                )
            negotiated_value = (entry["total_hours"] * (project.hourly_rate or Decimal("0.00"))).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            rows.append(
                {
                    "project": project.description,
                    "project_status": project.get_status_display(),
                    "internal_manager": internal_manager or "-",
                    "total_hours": self._format_decimal(entry["total_hours"]),
                    "approved_hours": self._format_decimal(entry["approved_hours"]),
                    "pending_hours": self._format_decimal(entry["pending_hours"]),
                    "negotiated_value": self._format_currency(negotiated_value),
                    "consultant_cost": self._format_currency(entry["consultant_cost"]),
                    "completion_percent": percent,
                    "progress_state": progress_state,
                    "progress_label": progress_label,
                    "progress_chip": progress_chip,
                }
            )

        filters_active = any(value for value in filters.values())

        return {
            "ops_rows": rows,
            "ops_project_count": len(rows),
            "ops_filters": filters,
            "ops_filters_active": filters_active,
            "ops_status_options": status_options,
            "ops_projects": projects,
            "ops_consultants": consultants,
            "ops_internal_managers": internal_managers,
            "ops_products": products,
            "ops_modules": modules,
            "ops_submodules": submodules,
            "ops_s_curve": s_curve,
            "ops_activity_status": activity_status,
            "ops_billing_planned": billing_charts["planned"],
            "ops_billing_actual": billing_charts["actual"],
        }

    def _build_dre_panel(self) -> dict[str, Any]:
        params = self.request.GET
        filters = {
            "project_id": params.get("dre_project_id", "").strip(),
            "period_start": params.get("dre_period_start", "").strip(),
            "period_end": params.get("dre_period_end", "").strip(),
        }

        period_start = self._parse_date(filters["period_start"])
        period_end = self._parse_date(filters["period_end"])

        projects = filter_projects_for_user(
            Project.objects.select_related("internal_manager"),
            self.request.user,
        ).order_by("description")

        receivable_payments = AccountsReceivablePayment.objects.select_related(
            "receivable",
            "receivable__billing_invoice",
            "receivable__account_plan_item",
        ).order_by("-payment_date")
        payable_payments = AccountsPayablePayment.objects.select_related(
            "payable",
            "payable__billing_invoice",
            "payable__account_plan_item",
        ).order_by("-payment_date")
        system_movements = BankSystemMovement.objects.select_related(
            "account_plan_item"
        ).order_by("-movement_date")
        if filters["project_id"]:
            receivable_payments = receivable_payments.filter(
                receivable__billing_invoice__project_id=filters["project_id"]
            )
            payable_payments = payable_payments.filter(
                payable__billing_invoice__project_id=filters["project_id"]
            )
            system_movements = system_movements.none()
        if period_start:
            receivable_payments = receivable_payments.filter(
                payment_date__gte=period_start
            )
            payable_payments = payable_payments.filter(payment_date__gte=period_start)
            system_movements = system_movements.filter(movement_date__gte=period_start)
        if period_end:
            receivable_payments = receivable_payments.filter(payment_date__lte=period_end)
            payable_payments = payable_payments.filter(payment_date__lte=period_end)
            system_movements = system_movements.filter(movement_date__lte=period_end)

        receivable_total = (
            receivable_payments.aggregate(total=Sum("amount")).get("total")
            or Decimal("0.00")
        )
        payable_total = (
            payable_payments.aggregate(total=Sum("amount")).get("total")
            or Decimal("0.00")
        )
        system_credit_total = (
            system_movements.filter(direction=BankMovementDirection.CREDIT)
            .aggregate(total=Sum("amount"))
            .get("total")
            or Decimal("0.00")
        )
        system_debit_total = (
            system_movements.filter(direction=BankMovementDirection.DEBIT)
            .aggregate(total=Sum("amount"))
            .get("total")
            or Decimal("0.00")
        )
        receivable_total += system_credit_total
        payable_total += system_debit_total
        receivable_total = receivable_total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        payable_total = payable_total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        result = (receivable_total - payable_total).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        title_count = (
            receivable_payments.count()
            + payable_payments.count()
            + system_movements.count()
        )

        if receivable_total > 0:
            margin = (result / receivable_total) * Decimal("100")
        else:
            margin = Decimal("0.00")
        margin = margin.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        selected_project = None
        if filters["project_id"]:
            selected_project = projects.filter(id=filters["project_id"]).first()
        project_label = selected_project.description if selected_project else "Consolidado"
        period_label = self._format_period_label(period_start, period_end)

        def _signed_amount(amount: Decimal, sign: str | None) -> Decimal:
            if sign == DreSign.SUBTRACT:
                return -amount
            return amount

        def _build_label(code: str | None, description: str | None) -> str:
            if code and description:
                return f"{code} - {description}"
            if description:
                return description
            if code:
                return code
            return "-"

        receivable_entries = (
            receivable_payments.values(
                "receivable__account_plan_item_id",
                "receivable__account_plan_item__code",
                "receivable__account_plan_item__description",
                "receivable__account_plan_item__dre_group",
                "receivable__account_plan_item__dre_subgroup",
                "receivable__account_plan_item__dre_order",
                "receivable__account_plan_item__dre_sign",
            )
            .annotate(total=Sum("amount"))
            .order_by()
        )
        payable_entries = (
            payable_payments.values(
                "payable__account_plan_item_id",
                "payable__account_plan_item__code",
                "payable__account_plan_item__description",
                "payable__account_plan_item__dre_group",
                "payable__account_plan_item__dre_subgroup",
                "payable__account_plan_item__dre_order",
                "payable__account_plan_item__dre_sign",
            )
            .annotate(total=Sum("amount"))
            .order_by()
        )
        system_entries = (
            system_movements.filter(account_plan_item__isnull=False)
            .values(
                "account_plan_item_id",
                "account_plan_item__code",
                "account_plan_item__description",
                "account_plan_item__dre_group",
                "account_plan_item__dre_subgroup",
                "account_plan_item__dre_order",
                "account_plan_item__dre_sign",
            )
            .annotate(total=Sum("amount"))
            .order_by()
        )
        missing_system_entries = (
            system_movements.filter(account_plan_item__isnull=True)
            .values("direction")
            .annotate(total=Sum("amount"))
            .order_by()
        )

        account_entries: dict[str, dict[str, Any]] = {}

        def _add_entries(
            entries: Iterable[dict[str, Any]],
            prefix: str,
            fallback_label: str,
            fallback_sign: str,
            missing_type: str | None = None,
        ) -> None:
            for entry in entries:
                plan_id = entry.get(f"{prefix}__account_plan_item_id")
                total = entry.get("total") or Decimal("0.00")
                if plan_id:
                    key = f"plan-{plan_id}"
                    code = entry.get(f"{prefix}__account_plan_item__code")
                    description = entry.get(f"{prefix}__account_plan_item__description")
                    group_label = entry.get(f"{prefix}__account_plan_item__dre_group") or "Sem grupo DRE"
                    order_value = entry.get(f"{prefix}__account_plan_item__dre_order")
                    sign_value = entry.get(f"{prefix}__account_plan_item__dre_sign")
                    label = _build_label(code, description)
                    missing_key = None
                else:
                    key = f"missing-{prefix}"
                    label = fallback_label
                    group_label = "Sem conta do plano"
                    order_value = None
                    sign_value = fallback_sign
                    missing_key = missing_type or prefix

                order_value = order_value if order_value is not None else 9999
                signed_total = _signed_amount(total, sign_value)
                payload = account_entries.setdefault(
                    key,
                    {
                        "label": label,
                        "group": group_label,
                        "order": order_value,
                        "signed_total": Decimal("0.00"),
                        "plan_id": plan_id,
                        "missing_type": missing_key,
                    },
                )
                payload["signed_total"] += signed_total
                if plan_id and not payload.get("plan_id"):
                    payload["plan_id"] = plan_id
                if missing_key and not payload.get("missing_type"):
                    payload["missing_type"] = missing_key
                if order_value < payload["order"]:
                    payload["order"] = order_value

        _add_entries(
            receivable_entries,
            "receivable",
            "Recebimentos sem conta",
            DreSign.ADD,
            "receivable",
        )
        _add_entries(
            payable_entries,
            "payable",
            "Pagamentos sem conta",
            DreSign.SUBTRACT,
            "payable",
        )
        if system_entries:
            normalized_system_entries = []
            for entry in system_entries:
                normalized_system_entries.append(
                    {
                        "movement__account_plan_item_id": entry.get(
                            "account_plan_item_id"
                        ),
                        "movement__account_plan_item__code": entry.get(
                            "account_plan_item__code"
                        ),
                        "movement__account_plan_item__description": entry.get(
                            "account_plan_item__description"
                        ),
                        "movement__account_plan_item__dre_group": entry.get(
                            "account_plan_item__dre_group"
                        ),
                        "movement__account_plan_item__dre_subgroup": entry.get(
                            "account_plan_item__dre_subgroup"
                        ),
                        "movement__account_plan_item__dre_order": entry.get(
                            "account_plan_item__dre_order"
                        ),
                        "movement__account_plan_item__dre_sign": entry.get(
                            "account_plan_item__dre_sign"
                        ),
                        "total": entry.get("total"),
                    }
                )
            _add_entries(
                normalized_system_entries,
                "movement",
                "Movimentos sem conta",
                DreSign.ADD,
                "system",
            )

        for entry in missing_system_entries:
            total = entry.get("total") or Decimal("0.00")
            direction = entry.get("direction")
            sign_value = (
                DreSign.ADD
                if direction == BankMovementDirection.CREDIT
                else DreSign.SUBTRACT
            )
            label = (
                "Movimentos sem conta (Credito)"
                if direction == BankMovementDirection.CREDIT
                else "Movimentos sem conta (Debito)"
            )
            key = f"missing-system-{direction}"
            signed_total = _signed_amount(total, sign_value)
            payload = account_entries.setdefault(
                key,
                {
                    "label": label,
                    "group": "Sem conta do plano",
                    "order": 9999,
                    "signed_total": Decimal("0.00"),
                    "plan_id": None,
                    "missing_type": f"system-{direction}",
                },
            )
            payload["signed_total"] += signed_total

        group_map: dict[str, dict[str, Any]] = {}
        for entry in account_entries.values():
            group_label = entry["group"]
            group = group_map.setdefault(
                group_label,
                {
                    "order": entry["order"],
                    "total": Decimal("0.00"),
                    "items": [],
                },
            )
            if entry["order"] < group["order"]:
                group["order"] = entry["order"]
            group["total"] += entry["signed_total"]
            group["items"].append(entry)

        def _value_class(amount: Decimal) -> str:
            if amount > 0:
                return "value-positive"
            if amount < 0:
                return "value-negative"
            return ""

        rows: list[dict[str, Any]] = []
        for group_label, group in sorted(
            group_map.items(),
            key=lambda item: (item[1]["order"], item[0].lower()),
        ):
            group_total = group["total"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            rows.append(
                {
                    "label": group_label,
                    "value": self._format_currency(group_total),
                    "value_class": _value_class(group_total),
                    "row_type": "group",
                }
            )
            for item in sorted(
                group["items"],
                key=lambda payload: (payload["order"], payload["label"].lower()),
            ):
                item_total = item["signed_total"].quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                rows.append(
                    {
                        "label": item["label"],
                        "value": self._format_currency(item_total),
                        "value_class": _value_class(item_total),
                        "row_type": "detail",
                        "plan_id": item.get("plan_id"),
                        "missing_type": item.get("missing_type"),
                    }
                )

        rows.append(
            {
                "label": "Resultado",
                "value": self._format_currency(result),
                "value_class": _value_class(result),
                "chip": "chip-ok" if result >= 0 else "chip-danger",
                "row_type": "total",
            }
        )

        filters_active = any(
            value for value in (filters["project_id"], period_start, period_end)
        )

        account_options = list(
            AccountPlanTemplateItem.objects.filter(
                status=StatusChoices.ACTIVE,
                is_analytic=True,
            )
            .order_by("code")
            .values("id", "code", "description")
        )

        return {
            "dre_projects": projects,
            "dre_filters": {
                "project_id": filters["project_id"],
                "period_start": period_start.isoformat() if period_start else "",
                "period_end": period_end.isoformat() if period_end else "",
            },
            "dre_filters_active": filters_active,
            "dre_project_label": project_label,
            "dre_period_label": period_label,
            "dre_entry_count": title_count,
            "dre_totals": {
                "title_count": str(title_count),
                "received": self._format_currency(receivable_total),
                "received_class": _value_class(receivable_total),
                "paid": self._format_currency(payable_total),
                "paid_class": _value_class(payable_total),
                "result": self._format_currency(result),
                "result_class": _value_class(result),
                "margin": f"{self._format_decimal(margin, 2)}%",
                "margin_class": _value_class(margin),
            },
            "dre_rows": rows,
            "dre_account_options": account_options,
        }

    def _build_allocation_panel(self) -> dict[str, Any]:
        params = self.request.GET
        filters = {
            "project_id": params.get("alloc_project_id", "").strip(),
            "consultant_id": params.get("alloc_consultant_id", "").strip(),
            "period_start": params.get("alloc_period_start", "").strip(),
            "period_end": params.get("alloc_period_end", "").strip(),
        }
        role = resolve_user_role(self.request.user)
        consultant_lock = False
        consultant_name = ""
        if role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(user=self.request.user).first()
            if consultant:
                filters["consultant_id"] = str(consultant.id)
                consultant_lock = True
                consultant_name = consultant.full_name

        today = timezone.localdate()
        default_start = today.replace(day=1)
        next_month = (default_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        default_end = next_month - timedelta(days=1)

        period_start = self._parse_date(filters["period_start"])
        period_end = self._parse_date(filters["period_end"])

        if not period_start and not period_end:
            period_start = default_start
            period_end = default_end
        elif period_start and not period_end:
            period_end = period_start
        elif period_end and not period_start:
            period_start = period_end

        if period_start and period_end and period_end < period_start:
            period_start, period_end = period_end, period_start

        if not period_start or not period_end:
            period_start = default_start
            period_end = default_end

        range_start = period_start
        range_end = period_end

        def _is_workday(target: date) -> bool:
            return target.weekday() < 5

        def _count_workdays(start: date, end: date) -> int:
            days = 0
            current = start
            while current <= end:
                if _is_workday(current):
                    days += 1
                current += timedelta(days=1)
            return days

        def _iter_business_days_forward(start: date, count: int) -> list[date]:
            dates: list[date] = []
            current = start
            while len(dates) < count:
                if _is_workday(current):
                    dates.append(current)
                current += timedelta(days=1)
            return dates

        def _iter_business_days_backward(end: date, count: int) -> list[date]:
            dates: list[date] = []
            current = end
            while len(dates) < count:
                if _is_workday(current):
                    dates.append(current)
                current -= timedelta(days=1)
            return list(reversed(dates))

        def _all_days_between(start: date, end: date) -> list[date]:
            if end < start:
                start, end = end, start
            days = (end - start).days + 1
            return [start + timedelta(days=offset) for offset in range(days)]

        def _activity_allocation_dates(activity: ProjectActivity) -> list[date]:
            duration_days = _resolve_duration_days(activity.days)
            if duration_days <= 0:
                return []
            original_start = activity.planned_start or activity.actual_start
            original_end = activity.planned_end or activity.actual_end
            if original_start and original_end:
                if _count_workdays(original_start, original_end) == 0:
                    return _all_days_between(original_start, original_end)
            if original_start:
                return _iter_business_days_forward(original_start, duration_days)
            if original_end:
                return _iter_business_days_backward(original_end, duration_days)
            return []

        def _allocation_ratio(total: Decimal, capacity: Decimal) -> Decimal:
            if capacity <= 0:
                return Decimal("0.00")
            return total / capacity

        def _allocation_chip(ratio: Decimal) -> str:
            if ratio < Decimal("1.00"):
                return "chip-ok"
            if ratio == Decimal("1.00"):
                return "chip-info"
            if ratio <= Decimal("1.20"):
                return "chip-warn"
            return "chip-danger"

        label_format = "%d"
        if range_start.month != range_end.month or range_start.year != range_end.year:
            label_format = "%d/%m"
        weekday_labels = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
        allocation_days = []
        current = range_start
        while current <= range_end:
            allocation_days.append(
                {
                    "date": current,
                    "label": current.strftime(label_format),
                    "is_weekend": not _is_workday(current),
                    "weekday_label": weekday_labels[current.weekday()],
                }
            )
            current += timedelta(days=1)
        day_index = {
            entry["date"]: idx for idx, entry in enumerate(allocation_days)
        }
        day_count = len(allocation_days) or 1
        workday_count = sum(1 for day in allocation_days if not day["is_weekend"])
        if workday_count <= 0:
            workday_count = 1

        status_map = {
            ActivityStatus.PLANNED: "planned_hours",
            ActivityStatus.RELEASED: "released_hours",
            ActivityStatus.IN_PROGRESS: "pending_hours",
        }

        activities = (
            ProjectActivity.objects.select_related("project")
            .prefetch_related("consultants")
            .order_by("project__description", "seq")
        )
        activities = filter_activities_for_user(activities, self.request.user)

        if filters["project_id"]:
            activities = activities.filter(project_id=filters["project_id"])
        if filters["consultant_id"]:
            activities = activities.filter(consultants__id=filters["consultant_id"]).distinct()

        selected_consultant_id = None
        if filters["consultant_id"]:
            try:
                selected_consultant_id = int(filters["consultant_id"])
            except ValueError:
                selected_consultant_id = None

        rows_map: dict[int, dict[str, Any]] = {}

        def _get_entry(consultant: Consultant) -> dict[str, Any]:
            entry = rows_map.get(consultant.id)
            if not entry:
                entry = {
                    "consultant": consultant,
                    "planned_hours": Decimal("0.00"),
                    "released_hours": Decimal("0.00"),
                    "pending_hours": Decimal("0.00"),
                    "daily_hours": [Decimal("0.00") for _ in range(day_count)],
                }
                rows_map[consultant.id] = entry
            return entry

        for activity in activities:
            bucket = status_map.get(activity.status)
            if not bucket:
                continue
            allocation_dates = _activity_allocation_dates(activity)
            if not allocation_dates:
                continue
            consultants = list(activity.consultants.all())
            if not consultants:
                continue
            hours = activity.hours or Decimal("0.00")
            if hours <= 0:
                continue
            daily_hours = hours / Decimal(len(allocation_dates))
            share = daily_hours / Decimal(len(consultants))

            consultant_entries = []
            for consultant in consultants:
                if selected_consultant_id and consultant.id != selected_consultant_id:
                    continue
                consultant_entries.append(_get_entry(consultant))

            if not consultant_entries:
                continue

            for allocation_date in allocation_dates:
                idx = day_index.get(allocation_date)
                if idx is not None:
                    for entry in consultant_entries:
                        entry[bucket] += share
                        entry["daily_hours"][idx] += share

        consultants = Consultant.objects.order_by("full_name")
        if role == UserRole.CONSULTANT:
            consultants = consultants.filter(user=self.request.user)
        consultants = list(consultants)

        rows_consultants = consultants
        if selected_consultant_id:
            rows_consultants = [
                consultant
                for consultant in consultants
                if consultant.id == selected_consultant_id
            ]

        capacity = Decimal("160.00")
        daily_capacity = Decimal("8.00")
        daily_total_capacity = daily_capacity * Decimal(workday_count)
        rows = []
        for consultant in rows_consultants:
            entry = rows_map.get(consultant.id) or {
                "planned_hours": Decimal("0.00"),
                "released_hours": Decimal("0.00"),
                "pending_hours": Decimal("0.00"),
                "daily_hours": [Decimal("0.00") for _ in range(day_count)],
            }
            planned = entry["planned_hours"]
            released = entry["released_hours"]
            pending = entry["pending_hours"]
            total = (planned + released + pending).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            monthly_ratio = _allocation_ratio(total, capacity)
            daily_ratio = _allocation_ratio(total, daily_total_capacity)
            monthly_percent_value = (monthly_ratio * Decimal("100")).quantize(
                Decimal("0.1"),
                rounding=ROUND_HALF_UP,
            )
            daily_percent_value = (daily_ratio * Decimal("100")).quantize(
                Decimal("0.1"),
                rounding=ROUND_HALF_UP,
            )
            monthly_percent = self._format_decimal(monthly_percent_value, 1)
            daily_percent = self._format_decimal(daily_percent_value, 1)
            monthly_chip = _allocation_chip(monthly_ratio)
            daily_chip = _allocation_chip(daily_ratio)
            daily_cells = []
            for idx, value in enumerate(entry["daily_hours"]):
                day_info = allocation_days[idx]
                ratio = _allocation_ratio(value, daily_capacity)
                percent_value = (ratio * Decimal("100")).quantize(
                    Decimal("0.1"),
                    rounding=ROUND_HALF_UP,
                )
                chip = _allocation_chip(ratio)
                if allocation_days[idx]["is_weekend"] and value > 0:
                    chip = "chip-danger"
                daily_cells.append(
                    {
                        "value": self._format_decimal(value, 2),
                        "percent": self._format_decimal(percent_value, 1),
                        "chip": chip,
                        "date": day_info["date"].isoformat(),
                        "date_label": day_info["date"].strftime("%d/%m/%Y"),
                        "has_value": value > 0,
                    }
                )
            average_daily = (total / Decimal(workday_count)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            rows.append(
                {
                    "consultant": consultant.full_name,
                    "consultant_id": consultant.id,
                    "planned_hours": self._format_decimal(planned),
                    "released_hours": self._format_decimal(released),
                    "pending_hours": self._format_decimal(pending),
                    "total_hours": self._format_decimal(total),
                    "planned_has_value": planned > 0,
                    "released_has_value": released > 0,
                    "pending_has_value": pending > 0,
                    "total_has_value": total > 0,
                    "monthly_percent": monthly_percent,
                    "monthly_chip": monthly_chip,
                    "daily_percent": daily_percent,
                    "daily_chip": daily_chip,
                    "monthly_summary": (
                        f"{self._format_decimal(total, 2)}h ({monthly_percent}%)"
                    ),
                    "daily_summary": (
                        f"{self._format_decimal(average_daily, 2)}h ({daily_percent}%)"
                    ),
                    "daily_cells": daily_cells,
                }
            )

        projects = filter_projects_for_user(
            Project.objects.all(),
            self.request.user,
        ).order_by("description")

        filters_active = any(
            value for value in (filters["project_id"], filters["consultant_id"])
        ) or (
            period_start != default_start or period_end != default_end
        )

        current_filters = {
            "project_id": filters["project_id"],
            "consultant_id": filters["consultant_id"],
            "period_start": period_start.isoformat() if period_start else "",
            "period_end": period_end.isoformat() if period_end else "",
        }

        return {
            "allocation_projects": projects,
            "allocation_consultants": consultants,
            "allocation_rows": rows,
            "allocation_filters": current_filters,
            "allocation_filters_active": filters_active,
            "allocation_consultant_count": len(rows),
            "allocation_capacity": self._format_decimal(capacity, 0),
            "allocation_daily_capacity": self._format_decimal(daily_capacity, 0),
            "allocation_days": allocation_days,
            "allocation_column_count": 7 + len(allocation_days),
            "allocation_period_label": self._format_period_label(range_start, range_end),
            "allocation_day_count": day_count,
            "allocation_workday_count": workday_count,
            "allocation_details_url": reverse("cadastros_web:allocation_panel_details"),
            "allocation_consultant_locked": consultant_lock,
            "allocation_consultant_name": consultant_name,
        }

    def _build_project_charts(self) -> dict[str, Any]:
        def _count_by(field: str) -> dict[str, int]:
            rows = (
                Project.objects.values(field)
                .annotate(total=models.Count("id"))
                .order_by()
            )
            return {row[field]: row["total"] for row in rows}

        def _build_chart(title: str, choices, counts, colors):
            total = sum(counts.get(value, 0) for value, _ in choices)
            legend = []
            segments = []
            current = 0.0
            for idx, (value, label) in enumerate(choices):
                count = counts.get(value, 0)
                color = colors[idx % len(colors)]
                legend.append(
                    {
                        "label": label,
                        "count": count,
                        "color": color,
                    }
                )
                if total > 0 and count > 0:
                    start = current
                    end = current + (count / total) * 100
                    segments.append(f"{color} {start:.2f}% {end:.2f}%")
                    current = end

            if not segments:
                gradient = "conic-gradient(rgba(33, 29, 24, 0.08) 0 100%)"
            else:
                gradient = f"conic-gradient({', '.join(segments)})"

            return {
                "title": title,
                "total": total,
                "legend": legend,
                "gradient": gradient,
            }

        criticality_counts = _count_by("criticality")
        type_counts = _count_by("project_type")
        status_counts = _count_by("status")

        criticality_colors = [
            "var(--teal)",
            "var(--sun)",
            "var(--rose)",
            "#c16355",
        ]
        type_colors = [
            "var(--teal)",
            "#8a8f87",
        ]
        status_colors = [
            "var(--sun)",
            "var(--teal)",
            "#b9a38b",
            "#c16355",
            "#6f9a88",
        ]

        charts = [
            _build_chart(
                "Projetos por criticidade",
                ProjectCriticality.choices,
                criticality_counts,
                criticality_colors,
            ),
            _build_chart(
                "Projetos por tipo",
                ProjectType.choices,
                type_counts,
                type_colors,
            ),
            _build_chart(
                "Projetos por situacao",
                ProjectStatus.choices,
                status_counts,
                status_colors,
            ),
        ]

        return {"project_charts": charts}


class AllocationPanelView(DashboardView):
    template_name = "restricted/allocation_panel.html"
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = TemplateView.get_context_data(self, **kwargs)
        context.update(self._build_allocation_panel())
        context["page_title"] = "Painel Alocacao"
        return context


class AllocationPanelDetailsView(DashboardView):
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def _resolve_allocation_period(self, params: dict[str, str]) -> tuple[date, date]:
        today = timezone.localdate()
        default_start = today.replace(day=1)
        next_month = (default_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        default_end = next_month - timedelta(days=1)

        period_start = self._parse_date(params.get("period_start"))
        period_end = self._parse_date(params.get("period_end"))

        if not period_start and not period_end:
            period_start = default_start
            period_end = default_end
        elif period_start and not period_end:
            period_end = period_start
        elif period_end and not period_start:
            period_start = period_end

        if period_start and period_end and period_end < period_start:
            period_start, period_end = period_end, period_start

        if not period_start or not period_end:
            period_start = default_start
            period_end = default_end

        return period_start, period_end

    def _activity_allocation_dates(self, activity: ProjectActivity) -> list[date]:
        duration_days = _resolve_duration_days(activity.days)
        if duration_days <= 0:
            return []

        def _is_workday(target: date) -> bool:
            return target.weekday() < 5

        def _count_workdays(start: date, end: date) -> int:
            days = 0
            current = start
            while current <= end:
                if _is_workday(current):
                    days += 1
                current += timedelta(days=1)
            return days

        def _iter_business_days_forward(start: date, count: int) -> list[date]:
            dates: list[date] = []
            current = start
            while len(dates) < count:
                if _is_workday(current):
                    dates.append(current)
                current += timedelta(days=1)
            return dates

        def _iter_business_days_backward(end: date, count: int) -> list[date]:
            dates: list[date] = []
            current = end
            while len(dates) < count:
                if _is_workday(current):
                    dates.append(current)
                current -= timedelta(days=1)
            return list(reversed(dates))

        def _all_days_between(start: date, end: date) -> list[date]:
            if end < start:
                start, end = end, start
            days = (end - start).days + 1
            return [start + timedelta(days=offset) for offset in range(days)]

        original_start = activity.planned_start or activity.actual_start
        original_end = activity.planned_end or activity.actual_end
        if original_start and original_end:
            if _count_workdays(original_start, original_end) == 0:
                return _all_days_between(original_start, original_end)
        if original_start:
            return _iter_business_days_forward(original_start, duration_days)
        if original_end:
            return _iter_business_days_backward(original_end, duration_days)
        return []

    def _format_activity_period(self, activity: ProjectActivity) -> str:
        start = activity.planned_start or activity.actual_start
        end = activity.planned_end or activity.actual_end
        if start and end:
            return f"{start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"
        if start:
            return f"A partir de {start.strftime('%d/%m/%Y')}"
        if end:
            return f"Ate {end.strftime('%d/%m/%Y')}"
        return "-"

    def get(self, request, *args, **kwargs):
        params = request.GET
        consultant_id_raw = params.get("consultant_id", "").strip()
        if not consultant_id_raw:
            return JsonResponse(
                {"ok": False, "error": "Consultor nao informado."},
                status=400,
            )
        try:
            consultant_id = int(consultant_id_raw)
        except ValueError:
            return JsonResponse(
                {"ok": False, "error": "Consultor invalido."},
                status=400,
            )

        role = resolve_user_role(request.user)
        if role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(user=request.user).first()
            if not consultant:
                return JsonResponse(
                    {"ok": False, "error": "Consultor nao vinculado."},
                    status=403,
                )
            if consultant.id != consultant_id:
                return JsonResponse(
                    {"ok": False, "error": "Consultor nao autorizado."},
                    status=403,
                )

        consultant = Consultant.objects.filter(pk=consultant_id).first()
        if not consultant:
            return JsonResponse(
                {"ok": False, "error": "Consultor nao encontrado."},
                status=404,
            )

        scope = params.get("scope", "total").strip() or "total"
        day = self._parse_date(params.get("day", "").strip())
        period_start, period_end = self._resolve_allocation_period(params)
        project_id = params.get("project_id", "").strip()

        status_scope_map = {
            "planned": ActivityStatus.PLANNED,
            "released": ActivityStatus.RELEASED,
            "pending": ActivityStatus.IN_PROGRESS,
        }
        allowed_statuses = set(status_scope_map.values())
        target_status = status_scope_map.get(scope)

        activities = (
            ProjectActivity.objects.select_related(
                "project",
                "product",
                "module",
                "submodule",
            )
            .prefetch_related("consultants")
            .order_by("project__description", "seq")
        )
        activities = filter_activities_for_user(activities, request.user)
        if project_id:
            activities = activities.filter(project_id=project_id)
        activities = activities.filter(consultants__id=consultant_id).distinct()

        status_chip_map = {
            ActivityStatus.PLANNED: "chip-info",
            ActivityStatus.RELEASED: "chip-ok",
            ActivityStatus.IN_PROGRESS: "chip-warn",
            ActivityStatus.DONE: "chip-ok",
            ActivityStatus.BLOCKED: "chip-danger",
            ActivityStatus.CANCELED: "chip-danger",
        }

        items = []
        total_hours = Decimal("0.00")
        for activity in activities:
            if activity.status not in allowed_statuses:
                continue
            if target_status and activity.status != target_status:
                continue
            allocation_dates = self._activity_allocation_dates(activity)
            if not allocation_dates:
                continue
            consultants = list(activity.consultants.all())
            if not consultants:
                continue
            hours = activity.hours or Decimal("0.00")
            if hours <= 0:
                continue

            daily_share = hours / Decimal(len(allocation_dates))
            daily_share = daily_share / Decimal(len(consultants))
            allocated_share = Decimal("0.00")
            if day:
                if period_start <= day <= period_end and day in allocation_dates:
                    allocated_share = daily_share
            else:
                for allocation_date in allocation_dates:
                    if period_start <= allocation_date <= period_end:
                        allocated_share += daily_share

            if allocated_share <= 0:
                continue

            total_hours += allocated_share
            items.append(
                {
                    "project": str(activity.project),
                    "product": str(activity.product),
                    "module": str(activity.module),
                    "submodule": str(activity.submodule),
                    "activity": activity.activity,
                    "subactivity": _format_activity_subactivities(activity, ""),
                    "period": self._format_activity_period(activity),
                    "hours": self._format_decimal(allocated_share, 2),
                    "status_label": activity.get_status_display(),
                    "status_chip": status_chip_map.get(activity.status, "chip-neutral"),
                }
            )

        return JsonResponse(
            {
                "ok": True,
                "items": items,
                "total_hours": self._format_decimal(total_hours, 2),
                "period_label": self._format_period_label(period_start, period_end),
                "day_label": day.strftime("%d/%m/%Y") if day else "",
                "consultant": consultant.full_name,
                "scope": scope,
            }
        )


@method_decorator(ensure_csrf_cookie, name="dispatch")
class OpportunitiesPanelView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/opportunities_panel.html"
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "page_title": "Painel Oportunidades",
                "data_url": reverse("cadastros_web:opportunities_data"),
                "apply_url": reverse("cadastros_web:opportunities_apply"),
                "refresh_default": settings.OPPORTUNITIES_REFRESH_DEFAULT,
                "refresh_min": settings.OPPORTUNITIES_REFRESH_MIN,
            }
        )
        return context


class DreEntriesView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _parse_date(value: str | None) -> date | None:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    @staticmethod
    def _parse_amount(value: str | None) -> Decimal | None:
        if not value:
            return None
        raw = (value or "").strip()
        if not raw:
            return None
        normalized = raw.replace("R$", "").replace(" ", "")
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")
        normalized = re.sub(r"[^0-9.\-]", "", normalized)
        if not normalized:
            return None
        try:
            parsed = Decimal(normalized)
        except Exception:
            return None
        return abs(parsed)

    @staticmethod
    def _format_currency(value: Decimal) -> str:
        return f"R$ {formats.number_format(value, decimal_pos=2, use_l10n=True, force_grouping=True)}"

    @staticmethod
    def _value_class(amount: Decimal) -> str:
        if amount > 0:
            return "value-positive"
        if amount < 0:
            return "value-negative"
        return ""

    @staticmethod
    def _signed_amount(amount: Decimal, sign: str | None, fallback_sign: str) -> Decimal:
        applied_sign = sign or fallback_sign
        if applied_sign == DreSign.SUBTRACT:
            return -amount
        return amount

    def get(self, request, *args, **kwargs):
        plan_id = request.GET.get("plan_id", "").strip()
        missing_type = request.GET.get("missing_type", "").strip()
        if not plan_id and not missing_type:
            return JsonResponse(
                {"ok": False, "error": "Conta contabil invalida."},
                status=400,
            )
        if not plan_id.isdigit():
            if plan_id:
                return JsonResponse(
                    {"ok": False, "error": "Conta contabil invalida."},
                    status=400,
                )
        account = None
        if plan_id:
            plan_id = int(plan_id)
            account = AccountPlanTemplateItem.objects.filter(pk=plan_id).first()
            if not account:
                return JsonResponse(
                    {"ok": False, "error": "Conta contabil nao encontrada."},
                    status=404,
                )

        project_id = request.GET.get("project_id", "").strip()
        period_start = self._parse_date(request.GET.get("period_start", "").strip())
        period_end = self._parse_date(request.GET.get("period_end", "").strip())

        receivable_qs = AccountsReceivablePayment.objects.select_related(
            "receivable",
            "receivable__client",
            "receivable__billing_invoice",
            "receivable__account_plan_item",
        )
        payable_qs = AccountsPayablePayment.objects.select_related(
            "payable",
            "payable__supplier",
            "payable__billing_invoice",
            "payable__account_plan_item",
        )
        system_qs = BankSystemMovement.objects.select_related(
            "account_plan_item"
        )

        if missing_type:
            if missing_type == "receivable":
                receivable_qs = receivable_qs.filter(
                    receivable__account_plan_item__isnull=True
                )
                payable_qs = payable_qs.none()
                system_qs = system_qs.none()
            elif missing_type == "payable":
                payable_qs = payable_qs.filter(
                    payable__account_plan_item__isnull=True
                )
                receivable_qs = receivable_qs.none()
                system_qs = system_qs.none()
            elif missing_type == "system-credit":
                system_qs = system_qs.filter(
                    account_plan_item__isnull=True,
                    direction=BankMovementDirection.CREDIT,
                )
                receivable_qs = receivable_qs.none()
                payable_qs = payable_qs.none()
            elif missing_type == "system-debit":
                system_qs = system_qs.filter(
                    account_plan_item__isnull=True,
                    direction=BankMovementDirection.DEBIT,
                )
                receivable_qs = receivable_qs.none()
                payable_qs = payable_qs.none()
            else:
                return JsonResponse(
                    {"ok": False, "error": "Conta contabil invalida."},
                    status=400,
                )
        elif account:
            receivable_qs = receivable_qs.filter(
                receivable__account_plan_item_id=account.id
            )
            payable_qs = payable_qs.filter(
                payable__account_plan_item_id=account.id
            )
            system_qs = system_qs.filter(account_plan_item_id=account.id)

        if project_id:
            receivable_qs = receivable_qs.filter(
                receivable__billing_invoice__project_id=project_id
            )
            payable_qs = payable_qs.filter(
                payable__billing_invoice__project_id=project_id
            )
            system_qs = system_qs.none()
        if period_start:
            receivable_qs = receivable_qs.filter(payment_date__gte=period_start)
            payable_qs = payable_qs.filter(payment_date__gte=period_start)
            system_qs = system_qs.filter(movement_date__gte=period_start)
        if period_end:
            receivable_qs = receivable_qs.filter(payment_date__lte=period_end)
            payable_qs = payable_qs.filter(payment_date__lte=period_end)
            system_qs = system_qs.filter(movement_date__lte=period_end)

        entries = []
        total = Decimal("0.00")
        account_label = (
            f"{account.code} - {account.description}" if account else None
        )

        for payment in receivable_qs:
            amount = payment.amount or Decimal("0.00")
            sign_value = None
            if payment.receivable.account_plan_item_id:
                sign_value = payment.receivable.account_plan_item.dre_sign
            signed = self._signed_amount(amount, sign_value, DreSign.ADD)
            total += signed
            description = f"Recebimento {payment.receivable.document_number} - {payment.receivable.client}"
            entries.append(
                {
                    "date": payment.payment_date.strftime("%d/%m/%Y"),
                    "date_sort": payment.payment_date.isoformat(),
                    "description": description,
                    "source": "Recebimento",
                    "value": self._format_currency(signed),
                    "value_class": self._value_class(signed),
                    "entry_id": payment.id,
                    "entry_type": "receivable",
                    "account_missing": payment.receivable.account_plan_item_id is None,
                    "account_label": account_label
                    or (str(payment.receivable.account_plan_item) if payment.receivable.account_plan_item_id else "-"),
                }
            )

        for payment in payable_qs:
            amount = payment.amount or Decimal("0.00")
            sign_value = None
            if payment.payable.account_plan_item_id:
                sign_value = payment.payable.account_plan_item.dre_sign
            signed = self._signed_amount(amount, sign_value, DreSign.SUBTRACT)
            total += signed
            description = f"Pagamento {payment.payable.document_number} - {payment.payable.supplier}"
            entries.append(
                {
                    "date": payment.payment_date.strftime("%d/%m/%Y"),
                    "date_sort": payment.payment_date.isoformat(),
                    "description": description,
                    "source": "Pagamento",
                    "value": self._format_currency(signed),
                    "value_class": self._value_class(signed),
                    "entry_id": payment.id,
                    "entry_type": "payable",
                    "account_missing": payment.payable.account_plan_item_id is None,
                    "account_label": account_label
                    or (str(payment.payable.account_plan_item) if payment.payable.account_plan_item_id else "-"),
                }
            )

        for movement in system_qs:
            amount = movement.amount or Decimal("0.00")
            sign_value = None
            if movement.account_plan_item_id:
                sign_value = movement.account_plan_item.dre_sign
            fallback_sign = (
                DreSign.ADD
                if movement.direction == BankMovementDirection.CREDIT
                else DreSign.SUBTRACT
            )
            signed = self._signed_amount(amount, sign_value, fallback_sign)
            total += signed
            source_label = (
                "Movimento OFX"
                if movement.source == BankMovementSource.OFX
                else "Lancamento"
            )
            entries.append(
                {
                    "date": movement.movement_date.strftime("%d/%m/%Y"),
                    "date_sort": movement.movement_date.isoformat(),
                    "description": movement.description,
                    "source": source_label,
                    "value": self._format_currency(signed),
                    "value_class": self._value_class(signed),
                    "entry_id": movement.id,
                    "entry_type": "system",
                    "account_missing": movement.account_plan_item_id is None,
                    "account_label": account_label
                    or (str(movement.account_plan_item) if movement.account_plan_item_id else "-"),
                }
            )

        entries.sort(key=lambda item: item.get("date_sort", ""))
        for entry in entries:
            entry.pop("date_sort", None)
        total = total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        payload = {
            "label": account_label or "Lancamentos sem conta",
            "total": self._format_currency(total),
            "total_class": self._value_class(total),
            "entries": entries,
        }
        return JsonResponse({"ok": True, "data": payload})


class DreEntryAssignView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        payload = {}
        if request.body:
            try:
                payload = json.loads(request.body.decode("utf-8"))
            except (json.JSONDecodeError, UnicodeDecodeError):
                payload = {}
        if not payload:
            payload = request.POST

        entry_type = str(payload.get("entry_type", "")).strip()
        entry_id = str(payload.get("entry_id", "")).strip()
        account_id = str(payload.get("account_plan_item_id", "")).strip()

        if not entry_type or not entry_id.isdigit() or not account_id.isdigit():
            return JsonResponse(
                {"ok": False, "error": "Dados invalidos."},
                status=400,
            )

        account = AccountPlanTemplateItem.objects.filter(
            pk=int(account_id),
            status=StatusChoices.ACTIVE,
            is_analytic=True,
        ).first()
        if not account:
            return JsonResponse(
                {"ok": False, "error": "Conta contabil invalida."},
                status=404,
            )

        if entry_type == "receivable":
            payment = AccountsReceivablePayment.objects.select_related("receivable").filter(
                pk=int(entry_id)
            ).first()
            if not payment:
                return JsonResponse(
                    {"ok": False, "error": "Lancamento nao encontrado."},
                    status=404,
                )
            receivable = payment.receivable
            receivable.account_plan_item = account
            receivable.save(update_fields=["account_plan_item"])
        elif entry_type == "payable":
            payment = AccountsPayablePayment.objects.select_related("payable").filter(
                pk=int(entry_id)
            ).first()
            if not payment:
                return JsonResponse(
                    {"ok": False, "error": "Lancamento nao encontrado."},
                    status=404,
                )
            payable = payment.payable
            payable.account_plan_item = account
            payable.save(update_fields=["account_plan_item"])
        elif entry_type == "system":
            movement = BankSystemMovement.objects.filter(pk=int(entry_id)).first()
            if not movement:
                return JsonResponse(
                    {"ok": False, "error": "Lancamento nao encontrado."},
                    status=404,
                )
            movement.account_plan_item = account
            movement.save(update_fields=["account_plan_item"])
        else:
            return JsonResponse(
                {"ok": False, "error": "Tipo de lancamento invalido."},
                status=400,
            )

        return JsonResponse({"ok": True, "data": {"account_label": str(account)}})


class ConsultantPanelView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/consultant_panel.html"
    allowed_roles = (UserRole.CONSULTANT,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _format_decimal(value: Decimal) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value: Decimal) -> str:
        return f"R$ {self._format_decimal(value)}"

    @staticmethod
    def _resolve_account_plan_item(code: str) -> AccountPlanTemplateItem | None:
        return (
            AccountPlanTemplateItem.objects.filter(
                code=code,
                status=StatusChoices.ACTIVE,
                is_analytic=True,
            )
            .order_by("id")
            .first()
        )

    @staticmethod
    def _format_date(value: date | None) -> str:
        return value.strftime("%d/%m/%Y") if value else "-"

    def _resolve_rate_for_date(
        self,
        consultant: Consultant,
        target_date: date,
    ) -> Decimal | None:
        valid_rates = []
        for rate in consultant.rates.all():
            if rate.start_date <= target_date and (
                rate.end_date is None or rate.end_date >= target_date
            ):
                valid_rates.append(rate)
        if not valid_rates:
            return None
        valid_rates.sort(key=lambda rate: rate.start_date, reverse=True)
        return valid_rates[0].rate

    @staticmethod
    def _month_label(target: date) -> str:
        months = [
            "Jan",
            "Fev",
            "Mar",
            "Abr",
            "Mai",
            "Jun",
            "Jul",
            "Ago",
            "Set",
            "Out",
            "Nov",
            "Dez",
        ]
        return f"{months[target.month - 1]}/{str(target.year)[-2:]}"

    @staticmethod
    def _month_start(target: date) -> date:
        return target.replace(day=1)

    @staticmethod
    def _add_months(target: date, months: int) -> date:
        year = target.year + (target.month - 1 + months) // 12
        month = (target.month - 1 + months) % 12 + 1
        return date(year, month, 1)

    @staticmethod
    def _activity_bounds(activity: ProjectActivity) -> tuple[date | None, date | None]:
        start = activity.planned_start or activity.actual_start
        end = activity.planned_end or activity.actual_end or start
        return start, end

    @staticmethod
    def _format_activity_dates(start: date | None, end: date | None) -> str:
        if start and end and start != end:
            return f"{start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"
        if start:
            return start.strftime("%d/%m/%Y")
        if end:
            return end.strftime("%d/%m/%Y")
        return "-"

    def _build_month_chart(self, consultant: Consultant) -> dict[str, Any]:
        today = timezone.localdate()
        start_month = self._add_months(self._month_start(today), -11)
        months = [self._add_months(start_month, index) for index in range(12)]
        end_exclusive = self._add_months(months[-1], 1)

        entries = (
            TimeEntry.objects.filter(
                consultant=consultant,
                start_date__gte=start_month,
                start_date__lt=end_exclusive,
            )
            .exclude(status=TimeEntryStatus.REJECTED)
            .only("start_date", "total_hours")
        )
        monthly_totals: dict[tuple[int, int], Decimal] = {}
        for entry in entries:
            key = (entry.start_date.year, entry.start_date.month)
            monthly_totals[key] = monthly_totals.get(key, Decimal("0.00")) + (
                entry.total_hours or Decimal("0.00")
            )

        totals = [
            monthly_totals.get((month.year, month.month), Decimal("0.00"))
            for month in months
        ]
        max_hours = max(totals) if totals else Decimal("0.00")
        total_hours = sum(totals, Decimal("0.00"))
        points = []
        for idx, month in enumerate(months):
            hours = totals[idx]
            percent = Decimal("0.00")
            if max_hours > 0:
                percent = (hours / max_hours) * Decimal("100")
            points.append(
                {
                    "label": self._month_label(month),
                    "hours": hours,
                    "hours_display": self._format_decimal(hours),
                    "percent": f"{percent:.2f}",
                }
            )

        return {
            "available": total_hours > 0,
            "total_hours": self._format_decimal(total_hours),
            "points": points,
        }

    @staticmethod
    def _title_status_info(status: str, due_date: date | None, today: date) -> tuple[str, str]:
        if status == FinancialStatus.OPEN and due_date and due_date < today:
            return "Atrasado", "chip-danger"
        mapping = {
            FinancialStatus.OPEN: ("Em aberto", "chip-info"),
            FinancialStatus.OVERDUE: ("Atrasado", "chip-danger"),
            FinancialStatus.PAID: ("Pago", "chip-ok"),
            FinancialStatus.CANCELED: ("Cancelado", "chip-neutral"),
        }
        return mapping.get(status, ("-", "chip-neutral"))

    def _build_title_rows(
        self,
        titles: list[AccountsPayable],
        next_url: str | None = None,
    ) -> list[dict[str, str]]:
        today = timezone.localdate()
        rows = []
        for title in titles:
            amount = title.total_amount()
            paid_amount = getattr(title, "paid_total", None) or Decimal("0.00")
            open_amount = _resolve_open_amount(amount, paid_amount)
            attachment_items = []
            if hasattr(title, "attachments"):
                attachment_items = list(title.attachments.all())
            attachment_count = len(attachment_items)
            label, chip = self._title_status_info(title.status, title.due_date, today)
            if paid_amount >= amount and amount > 0:
                label, chip = "Pago", "chip-ok"
            elif paid_amount > 0:
                label, chip = "Pago parcial", "chip-warn"
            entries_url = reverse(
                "cadastros_web:consultant_payable_entries", args=[title.pk]
            )
            attachment_url = reverse(
                "cadastros_web:accounts_payable_attachment_create", args=[title.pk]
            )
            if next_url:
                attachment_url = f"{attachment_url}?{urlencode({'next': next_url})}"
            rows.append(
                {
                    "invoice_number": title.billing_invoice.number
                    if title.billing_invoice_id
                    else "-",
                    "original_amount_display": self._format_currency(amount),
                    "open_amount_display": self._format_currency(open_amount),
                    "due_date": self._format_date(title.due_date),
                    "settlement_date": self._format_date(title.settlement_date),
                    "status_label": label,
                    "status_chip": chip,
                    "entries_url": entries_url,
                    "attachment_url": attachment_url,
                    "has_attachment": attachment_count > 0,
                }
            )
        return rows

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        consultant = Consultant.objects.filter(user=self.request.user).first()
        if not consultant:
            context["panel_error"] = "Usuario sem consultor vinculado."
            return context

        today = timezone.localdate()
        activities = list(
            ProjectActivity.objects.select_related("project", "project__project_client")
            .filter(consultants=consultant)
            .distinct()
        )
        active_statuses = {ActivityStatus.DONE, ActivityStatus.CANCELED}

        today_rows = []
        late_rows = []
        future_rows = []
        today_hours = Decimal("0.00")
        future_limit = today + timedelta(days=10)
        future_window_label = (
            f"Proximos 10 dias: { (today + timedelta(days=1)).strftime('%d/%m/%Y') } "
            f"a {future_limit.strftime('%d/%m/%Y')}"
        )

        def _activity_row(
            activity: ProjectActivity,
            start: date | None,
            end: date | None,
        ) -> dict[str, Any]:
            project = activity.project
            return {
                "project": project.description,
                "client": str(project.project_client),
                "activity": activity.activity,
                "hours": self._format_decimal(activity.hours or Decimal("0.00")),
                "date_label": self._format_activity_dates(start, end),
                "date_value": start or end,
            }

        for activity in activities:
            if activity.status in active_statuses:
                continue
            start, end = self._activity_bounds(activity)
            anchor = start or end
            if anchor and today < anchor <= future_limit:
                future_rows.append(_activity_row(activity, start, end))
            if activity.schedule_state(today) == "late":
                late_rows.append(_activity_row(activity, start, end))
            if start and end and start <= today <= end:
                today_rows.append(_activity_row(activity, start, end))
                today_hours += activity.hours or Decimal("0.00")

        today_rows.sort(key=lambda item: (item["project"], item["activity"]))
        late_rows.sort(key=lambda item: (item["project"], item["activity"]))
        future_rows.sort(
            key=lambda item: (item["date_value"] or date.max, item["project"], item["activity"])
        )

        entries = TimeEntry.objects.filter(consultant=consultant)
        approved_entries = entries.filter(
            status=TimeEntryStatus.APPROVED,
            billing_invoice__isnull=True,
        )
        approved_hours = (
            approved_entries.aggregate(total=Sum("total_hours")).get("total")
            or Decimal("0.00")
        )
        billed_entries = entries.filter(
            status=TimeEntryStatus.APPROVED,
            billing_invoice__payment_status=BillingPaymentStatus.UNPAID,
        )
        billed_hours = (
            billed_entries.aggregate(total=Sum("total_hours")).get("total")
            or Decimal("0.00")
        )
        pending_hours = (
            entries.filter(status=TimeEntryStatus.PENDING)
            .aggregate(total=Sum("total_hours"))
            .get("total")
            or Decimal("0.00")
        )
        approved_count = approved_entries.count()
        billed_count = billed_entries.count()
        pending_count = entries.filter(status=TimeEntryStatus.PENDING).count()

        titles_queryset = (
            AccountsPayable.objects.select_related("billing_invoice")
            .filter(
                consultant=consultant,
                billing_invoice__isnull=False,
            )
            .annotate(paid_total=Sum("payments__amount"))
            .prefetch_related("attachments")
        )
        unpaid_titles = list(
            titles_queryset.filter(
                status__in=[FinancialStatus.OPEN, FinancialStatus.OVERDUE]
            ).order_by("due_date", "id")
        )
        paid_titles = list(
            titles_queryset.filter(status=FinancialStatus.PAID).order_by(
                "-settlement_date", "-due_date", "-id"
            )
        )
        next_url = self.request.get_full_path()
        unpaid_title_rows = self._build_title_rows(unpaid_titles, next_url=next_url)
        paid_title_rows = self._build_title_rows(paid_titles, next_url=next_url)

        context.update(
            {
                "page_title": "Painel do consultor",
                "consultant_name": consultant.full_name,
                "today_rows": today_rows,
                "late_rows": late_rows,
                "future_rows": future_rows,
                "today_count": len(today_rows),
                "late_count": len(late_rows),
                "future_count": len(future_rows),
                "today_hours": self._format_decimal(today_hours),
                "approved_hours": self._format_decimal(approved_hours),
                "billed_hours": self._format_decimal(billed_hours),
                "pending_hours": self._format_decimal(pending_hours),
                "approved_count": approved_count,
                "billed_count": billed_count,
                "pending_count": pending_count,
                "hours_chart": self._build_month_chart(consultant),
                "future_window_label": future_window_label,
                "payable_unpaid_titles": unpaid_title_rows,
                "payable_paid_titles": paid_title_rows,
                "payable_unpaid_count": len(unpaid_title_rows),
                "payable_paid_count": len(paid_title_rows),
            }
        )
        return context


class ClientPanelView(DashboardView):
    template_name = "restricted/client_panel.html"
    allowed_roles = (UserRole.CLIENT,)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _format_activity_period(start: date | None, end: date | None) -> str:
        if start and end and start != end:
            return f"{start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"
        if start:
            return start.strftime("%d/%m/%Y")
        if end:
            return end.strftime("%d/%m/%Y")
        return "-"

    @staticmethod
    def _activity_meta(activity: ProjectActivity) -> str:
        parts = [
            str(activity.product),
            str(activity.module),
            str(activity.submodule),
        ]
        return " / ".join([part for part in parts if part])

    def _resolve_schedule_badge(
        self,
        activity: ProjectActivity,
        today: date,
    ) -> tuple[str | None, str]:
        planned_end = activity.planned_end
        actual_end = activity.actual_end
        if planned_end and actual_end:
            if actual_end < planned_end:
                return "Antecipada", "chip-info"
            if actual_end > planned_end:
                return "Atrasada", "chip-danger"
            return "No prazo", "chip-ok"
        schedule_state = activity.schedule_state(today)
        if schedule_state == "late":
            return "Atrasada", "chip-danger"
        if schedule_state == "on_time":
            return "No prazo", "chip-ok"
        if schedule_state == "not_started":
            return "Nao iniciada", "chip-neutral"
        return None, ""

    def _resolve_activity_status(
        self,
        activity: ProjectActivity,
        today: date,
    ) -> tuple[str, str]:
        if activity.status == ActivityStatus.DONE:
            return "Finalizada", "chip-ok"
        if activity.status in {ActivityStatus.IN_PROGRESS, ActivityStatus.RELEASED} or activity.actual_start:
            return "Em Andamento", "chip-info"
        return "Nao Iniciada", "chip-neutral"

    def _resolve_project_progress(
        self,
        activities: list[ProjectActivity],
        today: date,
    ) -> tuple[int, str, str]:
        total = len(activities)
        done_count = 0
        total_hours = Decimal("0.00")
        approved_hours = Decimal("0.00")
        has_started = False
        for activity in activities:
            if activity.status == ActivityStatus.DONE:
                done_count += 1
            total_hours += activity.hours or Decimal("0.00")
            if activity.actual_start or activity.status in {
                ActivityStatus.IN_PROGRESS,
                ActivityStatus.RELEASED,
            }:
                has_started = True
            for entry in activity.time_entries.all():
                if entry.status == TimeEntryStatus.APPROVED:
                    approved_hours += entry.total_hours or Decimal("0.00")
                    has_started = True
        percent = self._percent(approved_hours, total_hours) if total_hours > 0 else 0
        if total <= 0:
            return percent, "not-started", "Nao iniciada"
        if done_count == total:
            return percent, "done", "Concluida"
        if any(activity.schedule_state(today) == "late" for activity in activities):
            return percent, "late", "Atrasada"
        if has_started:
            return percent, "in-progress", "Em andamento"
        return percent, "not-started", "Nao iniciada"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = TemplateView.get_context_data(self, **kwargs)
        projects = filter_projects_for_user(
            Project.objects.select_related("project_client"),
            self.request.user,
        ).order_by("description")

        today = timezone.localdate()
        project_panels = []
        for project in projects:
            activities = list(
                ProjectActivity.objects.select_related(
                    "phase",
                    "product",
                    "module",
                    "submodule",
                )
                .prefetch_related("subactivity_items", "time_entries")
                .filter(project=project, client_visible=True)
                .order_by("seq")
            )
            if not activities:
                continue

            phases = []
            phase_index: dict[str, dict[str, Any]] = {}
            for activity in activities:
                phase_key = str(activity.phase)
                phase_entry = phase_index.get(phase_key)
                if phase_entry is None:
                    phase_entry = {"name": phase_key, "activities": []}
                    phase_index[phase_key] = phase_entry
                    phases.append(phase_entry)
                status_label, status_chip = self._resolve_activity_status(activity, today)
                schedule_label, schedule_chip = self._resolve_schedule_badge(activity, today)
                show_schedule = bool(schedule_label)
                phase_entry["activities"].append(
                    {
                        "seq": activity.seq,
                        "name": activity.activity,
                        "meta": self._activity_meta(activity),
                        "period": self._format_activity_period(
                            activity.planned_start,
                            activity.planned_end,
                        ),
                        "status_label": status_label,
                        "status_chip": status_chip,
                        "schedule_label": schedule_label if show_schedule else "",
                        "schedule_chip": schedule_chip if show_schedule else "",
                        "show_schedule": show_schedule,
                        "subactivities": _get_activity_subactivities(activity),
                    }
                )

            completion_percent, progress_state, progress_label = self._resolve_project_progress(
                activities, today
            )
            project_panels.append(
                {
                    "id": project.id,
                    "name": project.description,
                    "client": str(project.project_client),
                    "activity_count": len(activities),
                    "completion_percent": completion_percent,
                    "progress_state": progress_state,
                    "progress_label": progress_label,
                    "s_curve": self._build_s_curve_chart(activities, None),
                    "phases": phases,
                }
            )

        context.update(
            {
                "page_title": "Painel do cliente",
                "project_panels": project_panels,
            }
        )
        return context


class ProjectScheduleView(DashboardView):
    template_name = "restricted/project_schedule.html"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a esta consulta.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = TemplateView.get_context_data(self, **kwargs)
        projects = Project.objects.select_related("project_client").order_by("description")
        selected_project_id = self.request.GET.get("project_id", "").strip()
        selected_project = None
        phases = []
        activity_count = 0
        headers = [
            "Seq",
            "Atividade",
            "Subatividades",
            "Recurso responsavel",
            "Horas total",
            "Horas apos contingencia",
            "Horas utilizadas",
            "Saldo de horas",
            "Valor consultor",
            "Valor consultoria",
            "Valor faltante faturar",
            "Valor faltante consultor",
            "Prev. inicio",
            "Prev. fim",
            "Real inicio",
            "Real fim",
            "Dias p/ finalizar",
            "Dias atraso",
        ]
        project_totals = None

        def init_schedule_totals():
            return {
                "hours_total": Decimal("0.00"),
                "hours_available": Decimal("0.00"),
                "hours_used": Decimal("0.00"),
                "hours_balance": Decimal("0.00"),
                "consultant_value": Decimal("0.00"),
                "consultancy_value": Decimal("0.00"),
                "remaining_consultancy": Decimal("0.00"),
                "remaining_consultant": Decimal("0.00"),
                "days_remaining": 0,
                "days_late": 0,
                "days_remaining_count": 0,
                "days_late_count": 0,
            }

        def accumulate_schedule_totals(
            totals,
            total_hours,
            hours_available,
            approved_hours,
            balance_hours,
            consultant_cost_total,
            consultancy_value,
            remaining_consultancy,
            remaining_consultant,
            days_remaining,
            days_late,
        ):
            totals["hours_total"] += total_hours
            totals["hours_available"] += hours_available
            totals["hours_used"] += approved_hours
            totals["hours_balance"] += balance_hours
            totals["consultant_value"] += consultant_cost_total
            totals["consultancy_value"] += consultancy_value
            totals["remaining_consultancy"] += remaining_consultancy
            totals["remaining_consultant"] += remaining_consultant
            if days_remaining is not None:
                totals["days_remaining"] += days_remaining
                totals["days_remaining_count"] += 1
            if days_late is not None:
                totals["days_late"] += days_late
                totals["days_late_count"] += 1

        def format_schedule_totals(totals):
            return {
                "hours_total": self._format_decimal(totals["hours_total"]),
                "hours_available": self._format_decimal(totals["hours_available"]),
                "hours_used": self._format_decimal(totals["hours_used"]),
                "hours_balance": self._format_decimal(totals["hours_balance"]),
                "consultant_value": self._format_currency(totals["consultant_value"]),
                "consultancy_value": self._format_currency(totals["consultancy_value"]),
                "remaining_consultancy": self._format_currency(totals["remaining_consultancy"]),
                "remaining_consultant": self._format_currency(totals["remaining_consultant"]),
                "days_remaining": "-"
                if totals["days_remaining_count"] == 0
                else str(totals["days_remaining"]),
                "days_late": "-"
                if totals["days_late_count"] == 0
                else str(totals["days_late"]),
            }

        if selected_project_id:
            selected_project = projects.filter(pk=selected_project_id).first()
        if selected_project:
            today = timezone.localdate()
            schedule_icon_map = {
                "No prazo": "img/noprazo.png",
                "Atrasada": "img/atrasado.png",
                "Antecipada": "img/antecipada.png",
                "Nao iniciada": "img/naoinciado.png",
                "Sem prazo": "img/naoinciado.png",
            }
            activities = list(
                ProjectActivity.objects.select_related(
                    "phase",
                    "product",
                    "module",
                    "submodule",
                )
                .prefetch_related(
                    "subactivity_items",
                    "consultants",
                    "consultants__rates",
                    "time_entries",
                )
                .filter(project=selected_project)
                .order_by("seq")
            )
            activity_count = len(activities)
            project_totals = init_schedule_totals()
            phase_index: dict[str, dict[str, Any]] = {}
            for activity in activities:
                phase_key = str(activity.phase)
                phase_entry = phase_index.get(phase_key)
                if phase_entry is None:
                    phase_entry = {
                        "name": phase_key,
                        "activities": [],
                        "totals": init_schedule_totals(),
                    }
                    phase_index[phase_key] = phase_entry
                    phases.append(phase_entry)

                total_hours = activity.hours or Decimal("0.00")
                hours_available = activity.hours_available()
                approved_hours = sum(
                    (
                        entry.total_hours or Decimal("0.00")
                        for entry in activity.time_entries.all()
                        if entry.status == TimeEntryStatus.APPROVED
                    ),
                    Decimal("0.00"),
                )
                balance_hours = hours_available - approved_hours
                billable_rate = selected_project.hourly_rate or Decimal("0.00")
                consultant_cost_total = self._estimate_consultant_cost(
                    activity,
                    today,
                    None,
                    False,
                )
                consultant_rate = (
                    (consultant_cost_total / total_hours)
                    if total_hours > 0
                    else Decimal("0.00")
                )
                remaining_hours = balance_hours if balance_hours > 0 else Decimal("0.00")
                consultancy_value = (total_hours * billable_rate).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                remaining_consultancy = (remaining_hours * billable_rate).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                remaining_consultant = (remaining_hours * consultant_rate).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                consultants = ", ".join(
                    consultant.full_name for consultant in activity.consultants.all()
                )
                status_label, status_chip = self._resolve_activity_status(activity, today)
                schedule_label, schedule_chip = self._resolve_schedule_badge(activity, today)
                if not schedule_label:
                    schedule_label = "Sem prazo"
                    schedule_chip = "chip-neutral"
                schedule_icon = static(
                    schedule_icon_map.get(schedule_label, "img/naoinciado.png")
                )

                planned_end = activity.planned_end or activity.planned_start
                if planned_end:
                    if activity.actual_end:
                        days_remaining = 0
                        days_late = max((activity.actual_end - planned_end).days, 0)
                    else:
                        days_remaining = max((planned_end - today).days, 0)
                        days_late = max((today - planned_end).days, 0)
                else:
                    days_remaining = None
                    days_late = None

                accumulate_schedule_totals(
                    phase_entry["totals"],
                    total_hours,
                    hours_available,
                    approved_hours,
                    balance_hours,
                    consultant_cost_total,
                    consultancy_value,
                    remaining_consultancy,
                    remaining_consultant,
                    days_remaining,
                    days_late,
                )
                accumulate_schedule_totals(
                    project_totals,
                    total_hours,
                    hours_available,
                    approved_hours,
                    balance_hours,
                    consultant_cost_total,
                    consultancy_value,
                    remaining_consultancy,
                    remaining_consultant,
                    days_remaining,
                    days_late,
                )

                phase_entry["activities"].append(
                    {
                        "seq": activity.seq,
                        "activity": activity.activity,
                        "subactivities": _format_activity_subactivities(activity),
                        "responsible_resource": consultants or "-",
                        "status_label": status_label,
                        "status_chip": status_chip,
                        "schedule_label": schedule_label,
                        "schedule_chip": schedule_chip,
                        "schedule_icon": schedule_icon,
                        "hours_total": self._format_decimal(total_hours),
                        "hours_available": self._format_decimal(hours_available),
                        "hours_used": self._format_decimal(approved_hours),
                        "hours_balance": self._format_decimal(balance_hours),
                        "consultant_value": self._format_currency(consultant_cost_total),
                        "consultancy_value": self._format_currency(consultancy_value),
                        "remaining_consultancy": self._format_currency(remaining_consultancy),
                        "remaining_consultant": self._format_currency(remaining_consultant),
                        "planned_start": _format_value(activity, "planned_start"),
                        "planned_end": _format_value(activity, "planned_end"),
                        "actual_start": _format_value(activity, "actual_start"),
                        "actual_end": _format_value(activity, "actual_end"),
                        "days_remaining": "-" if days_remaining is None else str(days_remaining),
                        "days_late": "-" if days_late is None else str(days_late),
                    }
                )

            for phase in phases:
                phase["totals"] = format_schedule_totals(phase["totals"])
            project_totals = format_schedule_totals(project_totals)

        context.update(
            {
                "page_title": "Cronograma do projeto",
                "projects": projects,
                "selected_project": selected_project,
                "selected_project_id": selected_project_id,
                "activity_count": activity_count,
                "phases": phases,
                "project_totals": project_totals,
                "schedule_headers": headers,
                "schedule_column_count": len(headers),
                "schedule_icon_legend": {
                    "no_prazo": static("img/noprazo.png"),
                    "atrasada": static("img/atrasado.png"),
                    "antecipada": static("img/antecipada.png"),
                    "nao_iniciada": static("img/naoinciado.png"),
                },
            }
        )
        return context


class OpportunitiesDataView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        api_url = settings.OPPORTUNITIES_API_URL
        api_token = settings.OPPORTUNITIES_API_TOKEN
        if not api_url or not api_token:
            return JsonResponse(
                {"ok": False, "error": "API nao configurada."},
                status=503,
            )

        headers = {
            "Authorization": f"Bearer {api_token}",
            "Accept": "application/json",
        }
        request_obj = Request(api_url, headers=headers)
        try:
            with urlopen(
                request_obj,
                timeout=settings.OPPORTUNITIES_REQUEST_TIMEOUT,
            ) as response:
                raw = response.read()
                charset = response.headers.get_content_charset() or "utf-8"
            try:
                payload = raw.decode(charset)
            except UnicodeDecodeError:
                payload = raw.decode("latin-1")
            data = json.loads(payload)
            return JsonResponse({"ok": True, "data": data})
        except HTTPError as exc:
            body = ""
            try:
                body = exc.read().decode("utf-8", errors="replace")
            except Exception:
                body = ""
            detail = body.strip()[:500]
            if detail:
                logger.warning(
                    "Opportunities API erro %s: %s",
                    exc.code,
                    detail,
                )
            else:
                logger.warning("Opportunities API erro %s.", exc.code)
            error_message = f"Falha ao consultar API ({exc.code})."
            if settings.DEBUG and detail:
                error_message = f"{error_message} {detail}"
            return JsonResponse(
                {"ok": False, "error": error_message},
                status=exc.code,
            )
        except (URLError, TimeoutError):
            logger.warning("Opportunities API conexao indisponivel.")
            return JsonResponse(
                {"ok": False, "error": "Nao foi possivel conectar a API."},
                status=502,
            )
        except json.JSONDecodeError:
            logger.warning("Opportunities API retorno invalido.")
            return JsonResponse(
                {"ok": False, "error": "Resposta da API invalida."},
                status=502,
            )


class OpportunitiesApplyView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {"ok": False, "error": "Payload invalido."},
                status=400,
            )
        if not isinstance(payload, dict):
            return JsonResponse(
                {"ok": False, "error": "Payload invalido."},
                status=400,
            )
        demand = payload.get("demand")
        if not isinstance(demand, dict):
            return JsonResponse(
                {"ok": False, "error": "Demanda invalida."},
                status=400,
            )
        if not notify_opportunity_candidate(request.user, demand):
            return JsonResponse(
                {"ok": False, "error": "Numeros WhatsApp nao configurados."},
                status=400,
            )
        return JsonResponse({"ok": True})


DEFAULT_CHATGPT_SYSTEM_PROMPT = (
    "Voce e um Gerente de Projetos Senior com mais de 20 anos de experiencia "
    "em implantacao de ERP Senior Sistemas, atuando em projetos de alta "
    "complexidade, envolvendo multiplos modulos, integracoes, migracao de "
    "dados, customizacoes e ambientes criticos de negocio.\n\n"
    "Voce possui dominio pratico e profundo em:\n"
    "- Metodologia de implantacao Senior (DPS, Blueprint, EF, CTS, CR, Go Live)\n"
    "- Modulos Senior: Mercado, Suprimentos, Financas, Controladoria, Custos, "
    "Manufatura, Qualidade, WMS, HCM, Automotivo, Agro\n"
    "- Integracoes (Oracle, SQL Server, Fiscal, Bancos, Legados)\n"
    "- Governanca de projetos no ecossistema Senior Sistemas\n"
    "- Politicas e boas praticas da Senior para parceiros e projetos\n"
    "- Analises de risco, impacto operacional e tomada de decisao GO / NO GO\n"
    "- Projetos com restricoes severas de prazo, ferias coletivas, "
    "sazonalidade e dependencia de usuarios-chave\n\n"
    "Seu papel e atuar de forma critica, tecnica e responsavel, priorizando:\n"
    "- Seguranca operacional\n"
    "- Sustentabilidade do Go Live\n"
    "- Mitigacao de riscos para cliente, consultoria e Senior Sistemas\n"
    "- Transparencia executiva e governanca formal\n\n"
    "Voce deve assumir postura de Gerente de Projetos experiente, capaz de "
    "contrariar expectativas quando necessario para proteger o projeto."
)

DEFAULT_CHATGPT_ANALYSIS_PROMPT = (
    "Analise os dados do projeto ERP Senior Sistemas abaixo.\n\n"
    "CONTEXTO DO PROJETO:\n"
    "{{PROJECT_CONTEXT}}\n\n"
    "DADOS DETALHADOS:\n"
    "{{DETAILS_JSON}}\n\n"
    "O projeto pode conter:\n"
    "- Apontamentos de atividades por modulo\n"
    "- Percentual planejado vs executado\n"
    "- Datas baseline e datas reais\n"
    "- Marcos (DPS, Blueprint, EF, CTS, Go Live)\n"
    "- Pendencias tecnicas e funcionais\n"
    "- Restricoes operacionais (ferias, indisponibilidade, dependencias)\n\n"
    "Gere uma analise completa, estruturada e executiva contendo obrigatoriamente:\n\n"
    "1. RESUMO EXECUTIVO DO PROJETO\n"
    "2. ANALISE DE CRONOGRAMA E MARCOS SENIOR\n"
    "3. ANALISE DE EXECUCAO POR MODULO\n"
    "4. RISCOS IDENTIFICADOS\n"
    "5. PONTOS DE ATENCAO\n"
    "6. ANALISE CRITICA DO PROJETO\n"
    "7. POSICIONAMENTO TECNICO (GO / GO COM RESSALVAS / NO GO)\n"
    "8. PLANO DE ACAO ESTRUTURADO\n"
    "9. CONCLUSAO EXECUTIVA\n\n"
    "Utilize linguagem profissional, tecnica e objetiva.\n"
    "Considere sempre o impacto para:\n"
    "- Operacao do cliente\n"
    "- Responsabilidade da consultoria\n"
    "- Governanca do ecossistema Senior Sistemas"
)


class ChatGPTApiError(RuntimeError):
    def __init__(self, message: str, status_code: int | None = None) -> None:
        super().__init__(message)
        self.status_code = status_code

    @property
    def public_message(self) -> str:
        base = str(self) if str(self) else "Falha ao chamar a API."
        if self.status_code:
            return f"Falha ao chamar a API ({self.status_code}). {base}"
        return base


class ProjectChatGPTAnalysisView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def _percent(self, part: Decimal, total: Decimal) -> int:
        if total <= 0:
            return 0
        value = (part / total) * Decimal("100")
        percent = int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        return max(0, min(100, percent))

    def _format_date(self, value: date | None) -> str:
        return value.isoformat() if value else ""

    def _format_datetime(self, value) -> str:
        if not value:
            return ""
        return value.isoformat()

    def _resolve_phase_label(self, description: str | None) -> str:
        if not description:
            return "Execucao"
        text = description.lower()
        if "dps" in text:
            return "DPS"
        if "blueprint" in text:
            return "Blueprint"
        if "pre" in text and "go" in text:
            return "Pre-GoLive"
        if "go live" in text or "golive" in text:
            return "Pre-GoLive"
        if "teste" in text or "cts" in text:
            return "Testes"
        if "exec" in text or "ef" in text:
            return "Execucao"
        return "Execucao"

    def _resolve_marco_label(self, description: str | None) -> str | None:
        if not description:
            return None
        text = description.lower()
        if "dps" in text:
            return "DPS"
        if "blueprint" in text:
            return "Blueprint"
        if "cts" in text or "teste" in text:
            return "CTS"
        if "ef" in text:
            return "EF"
        if "go live" in text or "golive" in text:
            return "Go Live"
        return None

    def _build_module_payload(
        self,
        activities: list[ProjectActivity],
        today: date,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int, int]:
        module_map: dict[int, dict[str, Any]] = {}
        pending_items: list[dict[str, Any]] = []
        total_hours = Decimal("0.00")
        planned_hours = Decimal("0.00")
        executed_hours = Decimal("0.00")

        for activity in activities:
            hours = activity.hours or Decimal("0.00")
            total_hours += hours
            if activity.planned_end and activity.planned_end <= today:
                planned_hours += hours

            module_id = activity.module_id or 0
            module_name = (
                activity.module.description
                if getattr(activity, "module", None)
                else "Outro"
            )
            entry = module_map.setdefault(
                module_id,
                {
                    "nome": module_name,
                    "total_hours": Decimal("0.00"),
                    "planned_hours": Decimal("0.00"),
                    "executed_hours": Decimal("0.00"),
                    "late": False,
                    "pendencias": set(),
                },
            )
            entry["total_hours"] += hours
            if activity.planned_end and activity.planned_end <= today:
                entry["planned_hours"] += hours

            schedule_state = activity.schedule_state(today)
            if schedule_state == "late":
                entry["late"] = True

            for time_entry in activity.time_entries.all():
                if time_entry.status != TimeEntryStatus.APPROVED:
                    continue
                entry["executed_hours"] += time_entry.total_hours or Decimal("0.00")
                executed_hours += time_entry.total_hours or Decimal("0.00")

            if activity.status in {ActivityStatus.BLOCKED, ActivityStatus.CANCELED}:
                entry["pendencias"].add(
                    f"{activity.activity} ({activity.get_status_display()})"
                )
                pending_items.append(
                    {
                        "modulo": module_name,
                        "atividade": activity.activity,
                        "status": activity.get_status_display(),
                        "planejado": self._format_date(activity.planned_end),
                        "real": self._format_date(activity.actual_end),
                    }
                )
            elif schedule_state == "late":
                entry["pendencias"].add(
                    f"{activity.activity} (Atrasada)"
                )
                pending_items.append(
                    {
                        "modulo": module_name,
                        "atividade": activity.activity,
                        "status": "Atrasada",
                        "planejado": self._format_date(activity.planned_end),
                        "real": self._format_date(activity.actual_end),
                    }
                )

        module_payload = []
        module_details = []
        for entry in sorted(module_map.values(), key=lambda item: item["nome"]):
            percent_planned = self._percent(entry["planned_hours"], entry["total_hours"])
            percent_executed = self._percent(
                entry["executed_hours"],
                entry["total_hours"],
            )
            pendencias = sorted(entry["pendencias"]) if entry["pendencias"] else []
            module_payload.append(
                {
                    "nome": entry["nome"],
                    "percentual_planejado": percent_planned,
                    "percentual_executado": percent_executed,
                    "atividades_atrasadas": bool(entry["late"]),
                    "pendencias": pendencias,
                }
            )
            module_details.append(
                {
                    "nome": entry["nome"],
                    "percentual_planejado": percent_planned,
                    "percentual_executado": percent_executed,
                    "horas_planejadas": str(entry["planned_hours"]),
                    "horas_executadas": str(entry["executed_hours"]),
                    "horas_totais": str(entry["total_hours"]),
                    "atividades_atrasadas": bool(entry["late"]),
                    "pendencias": pendencias,
                }
            )

        project_planned = self._percent(planned_hours, total_hours)
        project_executed = self._percent(executed_hours, total_hours)
        return (
            module_payload,
            module_details,
            pending_items[:30],
            project_planned,
            project_executed,
        )

    def _decimal_to_str(self, value: Decimal | None, places: int = 2) -> str:
        if value is None:
            return "0.00"
        quant = Decimal("1").scaleb(-places)
        return str(value.quantize(quant, rounding=ROUND_HALF_UP))

    def _decimal_to_float(self, value: Decimal | None, places: int = 2) -> float:
        if value is None:
            return 0.0
        quant = Decimal("1").scaleb(-places)
        return float(value.quantize(quant, rounding=ROUND_HALF_UP))

    def _build_time_curve_data(
        self,
        activities: list[ProjectActivity],
    ) -> tuple[list[dict[str, Any]], Decimal, dict[date, Decimal], dict[date, Decimal]]:
        planned_by_date: dict[date, Decimal] = defaultdict(lambda: Decimal("0.00"))
        actual_by_date: dict[date, Decimal] = defaultdict(lambda: Decimal("0.00"))
        total_planned_hours = Decimal("0.00")

        for activity in activities:
            planned_hours = activity.hours or Decimal("0.00")
            total_planned_hours += planned_hours
            planned_start = activity.planned_start
            planned_end = activity.planned_end
            if planned_start and planned_end and planned_end >= planned_start and planned_hours > 0:
                days = (planned_end - planned_start).days + 1
                daily = (planned_hours / Decimal(days)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                for offset in range(days):
                    planned_by_date[planned_start + timedelta(days=offset)] += daily

            for entry in activity.time_entries.all():
                if entry.status != TimeEntryStatus.APPROVED:
                    continue
                entry_hours = entry.total_hours or Decimal("0.00")
                if entry_hours <= 0:
                    continue
                start = entry.start_date
                end = entry.end_date or entry.start_date
                if not start or not end or end < start:
                    continue
                days = (end - start).days + 1
                daily = (entry_hours / Decimal(days)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                for offset in range(days):
                    actual_by_date[start + timedelta(days=offset)] += daily

        dates = sorted(set(planned_by_date.keys()) | set(actual_by_date.keys()))
        points: list[dict[str, Any]] = []
        cumulative_planned = Decimal("0.00")
        cumulative_actual = Decimal("0.00")
        for current in dates:
            cumulative_planned += planned_by_date.get(current, Decimal("0.00"))
            cumulative_actual += actual_by_date.get(current, Decimal("0.00"))
            planned_pct = self._percent(cumulative_planned, total_planned_hours)
            actual_pct = self._percent(cumulative_actual, total_planned_hours)
            points.append(
                {
                    "date": current.isoformat(),
                    "planned_pct": planned_pct,
                    "actual_pct": actual_pct,
                }
            )
        return points, total_planned_hours, planned_by_date, actual_by_date

    def _resolve_health(
        self,
        schedule_variance_days: int,
        effort_variance_hours: Decimal,
        late_ratio: float,
    ) -> str:
        if schedule_variance_days > 10 or late_ratio >= 0.25:
            return "red"
        if schedule_variance_days > 0 or late_ratio >= 0.10:
            return "yellow"
        if effort_variance_hours > Decimal("0.00") and late_ratio > 0:
            return "yellow"
        return "green"

    def _resolve_go_nogo(self, items: list[ProjectGoNoGoChecklistItem]) -> dict[str, str]:
        ok_count = sum(1 for item in items if item.result == GoNoGoResult.OK)
        no_count = sum(1 for item in items if item.result == GoNoGoResult.NO)
        pending_count = sum(1 for item in items if item.result == GoNoGoResult.PENDING)
        total = len(items)
        if no_count:
            decision = "NO_GO"
        elif pending_count:
            decision = "GO_COM_RESSALVAS"
        elif ok_count:
            decision = "GO"
        else:
            decision = "GO_COM_RESSALVAS"
        if total == 0:
            rationale = "Checklist Go/No-Go nao preenchido."
        else:
            rationale = f"Checklist: OK={ok_count}, PENDENTE={pending_count}, NAO={no_count}."
        return {"decision": decision, "rationale": rationale}

    def _resolve_milestones(
        self,
        project: Project,
        activities: list[ProjectActivity],
    ) -> list[dict[str, str]]:
        order = ["DPS", "Blueprint", "EF", "CTS", "Go Live"]
        marcos_map: dict[str, dict[str, date | None]] = {}
        for activity in activities:
            label = self._resolve_marco_label(
                activity.phase.description if activity.phase_id else None
            )
            if not label:
                continue
            entry = marcos_map.setdefault(
                label,
                {"baseline": None, "planned": None, "actual": None},
            )
            if activity.planned_end:
                if not entry["planned"] or activity.planned_end > entry["planned"]:
                    entry["planned"] = activity.planned_end
            if activity.actual_end:
                if not entry["actual"] or activity.actual_end > entry["actual"]:
                    entry["actual"] = activity.actual_end

        milestones: list[dict[str, str]] = []
        for label in order:
            if label not in marcos_map:
                continue
            entry = marcos_map[label]
            baseline = entry["planned"]
            if label == "Go Live" and project.planned_go_live_date:
                baseline = project.planned_go_live_date
            forecast = entry["actual"] or entry["planned"]
            status = "on_track"
            if baseline and forecast and forecast > baseline:
                status = "late"
            if entry["actual"]:
                status = "done" if not (baseline and entry["actual"] > baseline) else "late"
            milestones.append(
                {
                    "name": label,
                    "baseline": self._format_date(baseline),
                    "forecast": self._format_date(forecast),
                    "actual": self._format_date(entry["actual"]),
                    "status": status,
                }
            )
        return milestones

    def _build_status_report(
        self,
        project: Project,
        activities: list[ProjectActivity],
    ) -> dict[str, Any]:
        today = timezone.localdate()
        points, total_planned_hours, planned_by_date, actual_by_date = (
            self._build_time_curve_data(activities)
        )

        planned_to_date = sum(
            (value for date_key, value in planned_by_date.items() if date_key <= today),
            Decimal("0.00"),
        )
        actual_to_date = sum(
            (value for date_key, value in actual_by_date.items() if date_key <= today),
            Decimal("0.00"),
        )
        effort_variance_hours = actual_to_date - planned_to_date

        approved_hours_total = Decimal("0.00")
        approved_by_activity: dict[int, Decimal] = defaultdict(Decimal)
        for activity in activities:
            for entry in activity.time_entries.all():
                if entry.status == TimeEntryStatus.APPROVED:
                    hours = entry.total_hours or Decimal("0.00")
                    approved_by_activity[activity.id] += hours
                    approved_hours_total += hours

        total_activities = len(activities)
        late_count = sum(
            1
            for activity in activities
            if activity.schedule_state(today) == "late"
        )
        done_count = sum(
            1 for activity in activities if activity.status == ActivityStatus.DONE
        )
        late_ratio = (late_count / total_activities) if total_activities else 0.0

        planned_end = max(
            (activity.planned_end for activity in activities if activity.planned_end),
            default=None,
        )
        actual_end = max(
            (activity.actual_end for activity in activities if activity.actual_end),
            default=None,
        )
        baseline_go_live = project.planned_go_live_date or planned_end
        forecast_go_live = actual_end or planned_end or baseline_go_live
        schedule_variance_days = 0
        if baseline_go_live and forecast_go_live:
            schedule_variance_days = (forecast_go_live - baseline_go_live).days

        module_map: dict[str, dict[str, Decimal]] = {}
        for activity in activities:
            module_name = (
                activity.module.description
                if getattr(activity, "module", None)
                else "Outro"
            )
            entry = module_map.setdefault(
                module_name,
                {
                    "hours_planned": Decimal("0.00"),
                    "hours_planned_to_date": Decimal("0.00"),
                    "hours_actual": Decimal("0.00"),
                },
            )
            planned_hours = activity.hours or Decimal("0.00")
            entry["hours_planned"] += planned_hours
            if activity.planned_end and activity.planned_end <= today:
                entry["hours_planned_to_date"] += planned_hours
            entry["hours_actual"] += approved_by_activity.get(
                activity.id, Decimal("0.00")
            )

        modules = []
        for name, values in sorted(module_map.items(), key=lambda item: item[0]):
            hours_planned = values["hours_planned"]
            hours_actual = values["hours_actual"]
            hours_planned_to_date = values["hours_planned_to_date"]
            modules.append(
                {
                    "name": name,
                    "planned_pct": self._percent(hours_planned_to_date, hours_planned),
                    "actual_pct": self._percent(hours_actual, hours_planned),
                    "hours_planned": self._decimal_to_float(hours_planned),
                    "hours_actual": self._decimal_to_float(hours_actual),
                }
            )

        milestones = self._resolve_milestones(project, activities)

        risks = []
        impact_map = {
            ActivityCriticality.LOW: 2,
            ActivityCriticality.MEDIUM: 3,
            ActivityCriticality.HIGH: 4,
            ActivityCriticality.CRITICAL: 5,
        }
        owner_label = ""
        if project.internal_manager_id:
            owner_label = project.internal_manager.get_full_name() or project.internal_manager.username
        elif project.external_manager_id:
            owner_label = project.external_manager.get_full_name() or project.external_manager.username
        for activity in activities:
            schedule_state = activity.schedule_state(today)
            if schedule_state != "late" and activity.status != ActivityStatus.BLOCKED:
                continue
            probability = 4 if schedule_state == "late" else 5
            impact = impact_map.get(activity.criticality, 3)
            risks.append(
                {
                    "title": activity.activity,
                    "probability": probability,
                    "impact": impact,
                    "owner": owner_label,
                    "mitigation": "",
                }
            )

        go_no_go_items = list(
            ProjectGoNoGoChecklistItem.objects.filter(project=project).order_by("id")
        )
        go_nogo = self._resolve_go_nogo(go_no_go_items)

        actions = []
        for item in go_no_go_items:
            if item.result == GoNoGoResult.OK:
                continue
            actions.append(
                {
                    "what": item.criterion,
                    "owner": (item.approver or "").strip(),
                    "due": "",
                    "status": item.get_result_display(),
                }
            )

        health = self._resolve_health(
            schedule_variance_days,
            effort_variance_hours,
            late_ratio,
        )

        kpis = [
            {
                "name": "Cronograma variacao",
                "value": schedule_variance_days,
                "unit": "dias",
            },
            {
                "name": "Esforco variacao",
                "value": self._decimal_to_float(effort_variance_hours),
                "unit": "h",
            },
            {
                "name": "Horas planejadas",
                "value": self._decimal_to_float(total_planned_hours),
                "unit": "h",
            },
            {
                "name": "Horas realizadas",
                "value": self._decimal_to_float(approved_hours_total),
                "unit": "h",
            },
            {
                "name": "Atividades concluidas",
                "value": done_count,
                "unit": "atividades",
            },
            {
                "name": "Atividades atrasadas",
                "value": late_count,
                "unit": "atividades",
            },
        ]

        chart_points = points[:]
        if len(chart_points) > 120:
            step = max(1, (len(chart_points) // 120) + 1)
            chart_points = chart_points[::step]
            if chart_points and chart_points[-1] != points[-1]:
                chart_points.append(points[-1])

        planned_actual_values = []
        for item in chart_points:
            planned_actual_values.append(
                {
                    "date": item["date"],
                    "series": "Planned",
                    "value": item["planned_pct"],
                }
            )
            planned_actual_values.append(
                {
                    "date": item["date"],
                    "series": "Actual",
                    "value": item["actual_pct"],
                }
            )

        module_chart_values = []
        for module in modules:
            module_chart_values.append(
                {
                    "module": module["name"],
                    "series": "Planned",
                    "hours": module["hours_planned"],
                }
            )
            module_chart_values.append(
                {
                    "module": module["name"],
                    "series": "Actual",
                    "hours": module["hours_actual"],
                }
            )

        milestone_chart_values = []
        for milestone in milestones:
            for key in ("baseline", "forecast", "actual"):
                date_value = milestone.get(key)
                if not date_value:
                    continue
                milestone_chart_values.append(
                    {
                        "milestone": milestone["name"],
                        "type": key,
                        "date": date_value,
                    }
                )

        risk_chart_values = [
            {
                "risk": item["title"],
                "probability": item["probability"],
                "impact": item["impact"],
                "owner": item["owner"],
                "severity": item["probability"] * item["impact"],
            }
            for item in risks
            if item.get("probability") and item.get("impact")
        ]

        charts = [
            {
                "library": "vega-lite",
                "spec": {
                    "title": "Planned vs Actual (%) ao longo do tempo",
                    "data": {"values": planned_actual_values},
                    "mark": {"type": "line", "point": True},
                    "encoding": {
                        "x": {"field": "date", "type": "temporal", "title": "Data"},
                        "y": {
                            "field": "value",
                            "type": "quantitative",
                            "title": "Percentual",
                        },
                        "color": {
                            "field": "series",
                            "type": "nominal",
                            "title": "Serie",
                        },
                    },
                },
            },
            {
                "library": "vega-lite",
                "spec": {
                    "title": "Execucao por modulo (planejado vs realizado)",
                    "data": {"values": module_chart_values},
                    "mark": "bar",
                    "encoding": {
                        "x": {"field": "module", "type": "nominal", "title": "Modulo"},
                        "y": {"field": "hours", "type": "quantitative", "title": "Horas"},
                        "color": {
                            "field": "series",
                            "type": "nominal",
                            "title": "Serie",
                        },
                    },
                },
            },
            {
                "library": "vega-lite",
                "spec": {
                    "title": "Marcos (baseline vs real)",
                    "data": {"values": milestone_chart_values},
                    "mark": {"type": "tick", "thickness": 2, "size": 16},
                    "encoding": {
                        "y": {
                            "field": "milestone",
                            "type": "nominal",
                            "title": "Marco",
                        },
                        "x": {"field": "date", "type": "temporal", "title": "Data"},
                        "color": {"field": "type", "type": "nominal", "title": "Tipo"},
                    },
                },
            },
            {
                "library": "vega-lite",
                "spec": {
                    "title": "Matriz de riscos (probabilidade x impacto)",
                    "data": {"values": risk_chart_values},
                    "mark": {"type": "circle", "opacity": 0.7},
                    "encoding": {
                        "x": {
                            "field": "probability",
                            "type": "quantitative",
                            "title": "Probabilidade",
                            "scale": {"domain": [1, 5]},
                        },
                        "y": {
                            "field": "impact",
                            "type": "quantitative",
                            "title": "Impacto",
                            "scale": {"domain": [1, 5]},
                        },
                        "size": {"field": "severity", "type": "quantitative"},
                        "color": {
                            "field": "severity",
                            "type": "quantitative",
                            "title": "Severidade",
                        },
                        "tooltip": [
                            {"field": "risk", "type": "nominal", "title": "Risco"},
                            {"field": "owner", "type": "nominal", "title": "Dono"},
                            {"field": "probability", "type": "quantitative"},
                            {"field": "impact", "type": "quantitative"},
                        ],
                    },
                },
            },
        ]

        markdown_lines = [
            f"# Status Report - {project.description}",
            "",
            f"Semaforo: {health.upper()}",
            f"Variacao de cronograma: {schedule_variance_days} dias",
            f"Variacao de esforco: {self._decimal_to_str(effort_variance_hours)} h",
            "",
            "## KPIs",
            "| KPI | Valor | Unidade |",
            "| --- | --- | --- |",
        ]
        for item in kpis:
            markdown_lines.append(
                f"| {item['name']} | {item['value']} | {item['unit']} |"
            )

        markdown_lines.append("")
        markdown_lines.append("## Marcos")
        markdown_lines.append("| Marco | Baseline | Forecast | Actual | Status |")
        markdown_lines.append("| --- | --- | --- | --- | --- |")
        for milestone in milestones:
            markdown_lines.append(
                f"| {milestone['name']} | {milestone['baseline'] or '-'} | {milestone['forecast'] or '-'} | {milestone['actual'] or '-'} | {milestone['status']} |"
            )

        markdown_lines.append("")
        markdown_lines.append("## Modulos")
        markdown_lines.append("| Modulo | Planejado % | Executado % | Horas planejadas | Horas executadas |")
        markdown_lines.append("| --- | --- | --- | --- | --- |")
        for module in modules:
            markdown_lines.append(
                f"| {module['name']} | {module['planned_pct']} | {module['actual_pct']} | {module['hours_planned']} | {module['hours_actual']} |"
            )

        markdown_lines.append("")
        markdown_lines.append("## Riscos")
        if risks:
            for risk in risks:
                markdown_lines.append(
                    f"- {risk['title']} (P={risk['probability']}, I={risk['impact']}) | Dono: {risk['owner'] or '-'} | Mitigacao: {risk['mitigation'] or '-'}"
                )
        else:
            markdown_lines.append("- Sem riscos registrados no sistema.")

        markdown_lines.append("")
        markdown_lines.append("## Acoes")
        if actions:
            for action in actions:
                markdown_lines.append(
                    f"- {action['what']} | Dono: {action['owner'] or '-'} | Prazo: {action['due'] or '-'} | Status: {action['status']}"
                )
        else:
            markdown_lines.append("- Sem acoes registradas no sistema.")

        markdown_lines.append("")
        markdown_lines.append("## Go/No-Go")
        markdown_lines.append(f"Decisao: {go_nogo['decision']}")
        markdown_lines.append(f"Racional: {go_nogo['rationale']}")

        markdown_lines.append("")
        markdown_lines.append("## Graficos")
        markdown_lines.append("- Planned vs Actual (%) ao longo do tempo")
        markdown_lines.append("- Execucao por modulo (planejado vs realizado)")
        markdown_lines.append("- Marcos (baseline vs real)")
        markdown_lines.append("- Matriz de riscos (probabilidade x impacto)")
        markdown_lines.append("")
        markdown_lines.append(
            "Observacao: Relatorio gerado automaticamente a partir dos dados do projeto; indicadores e scores seguem regras deterministicas."
        )

        status_report = {
            "report": {"markdown": "\n".join(markdown_lines)},
            "summary": {"health": health},
            "kpis": kpis,
            "milestones": milestones,
            "modules": modules,
            "risks": risks,
            "actions": actions,
            "go_nogo": go_nogo,
            "charts": charts,
        }
        return status_report

    def _build_analysis_payload(
        self,
        project: Project,
        activities: list[ProjectActivity],
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        today = timezone.localdate()
        observations: list[str] = []
        if project.explanation:
            observations.append(project.explanation)

        for observation in ProjectObservation.objects.filter(project=project).order_by(
            "-created_at"
        )[:6]:
            if observation.note:
                observations.append(observation.note)

        (
            module_payload,
            module_details,
            pending_items,
            project_planned,
            project_executed,
        ) = self._build_module_payload(activities, today)

        total_activities = len(activities)
        done_count = 0
        late_count = 0
        pending_count = 0
        missing_items: list[dict[str, str]] = []

        for activity in activities:
            if activity.status == ActivityStatus.DONE:
                done_count += 1
            else:
                pending_count += 1

            schedule_state = activity.schedule_state(today)
            if schedule_state == "late":
                late_count += 1

            has_time_entries = any(
                entry.status == TimeEntryStatus.APPROVED
                for entry in activity.time_entries.all()
            )
            if (
                activity.status in {ActivityStatus.PLANNED, ActivityStatus.RELEASED}
                and not activity.actual_start
                and not has_time_entries
            ):
                missing_items.append(
                    {
                        "atividade": activity.activity,
                        "subatividade": _format_activity_subactivities(activity, ""),
                        "modulo": activity.module.description if activity.module_id else "",
                        "planejado": self._format_date(activity.planned_end),
                    }
                )

        planned_end = max(
            (activity.planned_end for activity in activities if activity.planned_end),
            default=None,
        )
        actual_end = max(
            (activity.actual_end for activity in activities if activity.actual_end),
            default=None,
        )
        baseline_go_live = project.planned_go_live_date or planned_end
        current_go_live = actual_end or planned_end or baseline_go_live
        desvio = 0
        if baseline_go_live and current_go_live:
            desvio = (current_go_live - baseline_go_live).days

        current_activity = None
        if activities:
            current_activity = sorted(
                activities,
                key=lambda item: item.actual_end or item.planned_end or today,
                reverse=True,
            )[0]
        phase_label = self._resolve_phase_label(
            current_activity.phase.description if current_activity else None
        )

        notes_text = " ".join(observations).lower()
        restricoes = {
            "ferias_coletivas": "ferias" in notes_text,
            "usuarios_indisponiveis": "indispon" in notes_text,
            "integracoes_pendentes": "integracao" in notes_text,
        }

        time_entries = (
            TimeEntry.objects.select_related(
                "activity",
                "activity__phase",
                "activity__module",
                "activity__submodule",
                "consultant",
            )
            .filter(activity__project=project)
            .order_by("start_date", "id")
        )
        approved_hours_by_activity: dict[int, Decimal] = defaultdict(Decimal)
        consumed_hours = Decimal("0.00")
        time_entries_payload: list[dict[str, Any]] = []
        for entry in time_entries:
            total_hours = entry.total_hours or Decimal("0.00")
            if entry.status == TimeEntryStatus.APPROVED:
                approved_hours_by_activity[entry.activity_id] += total_hours
                consumed_hours += total_hours
            time_entries_payload.append(
                {
                    "id": entry.id,
                    "atividade_seq": entry.activity.seq,
                    "atividade": entry.activity.activity,
                    "fase": entry.activity.phase.description if entry.activity.phase_id else "",
                    "modulo": entry.activity.module.description if entry.activity.module_id else "",
                    "submodulo": entry.activity.submodule.description if entry.activity.submodule_id else "",
                    "consultor": entry.consultant.full_name if entry.consultant_id else "",
                    "status": entry.get_status_display(),
                    "tipo": entry.get_entry_type_display(),
                    "data_inicio": self._format_date(entry.start_date),
                    "data_fim": self._format_date(entry.end_date),
                    "horas": str(total_hours),
                    "descricao": entry.description,
                    "motivo_reprovacao": entry.rejection_reason,
                }
            )

        cronograma = {
            "baseline_go_live": self._format_date(baseline_go_live),
            "data_prevista_atual": self._format_date(current_go_live),
            "desvio_dias": desvio,
        }

        cronograma_atividades: list[dict[str, Any]] = []
        total_hours_planned = Decimal("0.00")
        total_hours_released = Decimal("0.00")
        for activity in activities:
            planned_hours = activity.hours or Decimal("0.00")
            approved_hours = approved_hours_by_activity.get(activity.id, Decimal("0.00"))
            pending_hours = planned_hours - approved_hours
            if pending_hours < 0:
                pending_hours = Decimal("0.00")
            total_hours_planned += planned_hours
            if activity.status == ActivityStatus.RELEASED:
                total_hours_released += planned_hours
            percent_executed = self._percent(approved_hours, planned_hours) if planned_hours else 0
            percent_pending = max(0, 100 - percent_executed)
            cronograma_atividades.append(
                {
                    "seq": activity.seq,
                    "fase": activity.phase.description if activity.phase_id else "",
                    "modulo": activity.module.description if activity.module_id else "",
                    "submodulo": activity.submodule.description if activity.submodule_id else "",
                    "atividade": activity.activity,
                    "subatividades": activity.subactivities_label(),
                    "status": activity.get_status_display(),
                    "criticidade": activity.get_criticality_display(),
                    "tipo_hora": activity.billing_type_label(),
                    "inicio_previsto": self._format_date(activity.planned_start),
                    "fim_previsto": self._format_date(activity.planned_end),
                    "inicio_real": self._format_date(activity.actual_start),
                    "fim_real": self._format_date(activity.actual_end),
                    "situacao_cronograma": activity.schedule_label(today),
                    "horas_planejadas": str(planned_hours),
                    "horas_aprovadas": str(approved_hours),
                    "horas_pendentes": str(pending_hours),
                    "percentual_realizado": percent_executed,
                    "percentual_pendente": percent_pending,
                }
            )

        occurrences = list(
            ProjectOccurrence.objects.select_related("created_by")
            .prefetch_related("attachments")
            .filter(project=project)
            .order_by("-created_at")
        )
        occurrences_payload: list[dict[str, Any]] = []
        occurrence_attachments_payload: list[dict[str, Any]] = []

        def _attachment_text(file_field, max_bytes=200000, max_chars=5000) -> dict[str, Any]:
            if not file_field:
                return {
                    "conteudo": "",
                    "conteudo_truncado": False,
                    "erro": "Arquivo nao informado.",
                }
            try:
                size = getattr(file_field, "size", None)
                if size and size > max_bytes:
                    return {
                        "conteudo": "",
                        "conteudo_truncado": True,
                        "erro": "Arquivo grande demais para leitura automatica.",
                    }
                file_field.open("rb")
                data = file_field.read(max_bytes)
            except Exception as exc:
                return {
                    "conteudo": "",
                    "conteudo_truncado": False,
                    "erro": f"Falha ao ler arquivo: {exc}",
                }
            finally:
                try:
                    file_field.close()
                except Exception:
                    pass

            if b"\x00" in data:
                return {
                    "conteudo": "",
                    "conteudo_truncado": False,
                    "erro": "Arquivo binario (nao suportado).",
                }
            text = ""
            for encoding in ("utf-8", "latin-1"):
                try:
                    text = data.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if not text:
                return {
                    "conteudo": "",
                    "conteudo_truncado": False,
                    "erro": "Conteudo nao textual.",
                }
            text = text.strip()
            truncated = len(text) > max_chars
            if truncated:
                text = text[:max_chars]
            return {
                "conteudo": text,
                "conteudo_truncado": truncated,
                "erro": "",
            }

        def _attachment_payload(item, origin: str) -> dict[str, Any]:
            name = ""
            if getattr(item, "file", None):
                name = os.path.basename(item.file.name or "")
            text_payload = _attachment_text(getattr(item, "file", None))
            return {
                "origem": origin,
                "descricao": getattr(item, "description", "") or "",
                "tipo": getattr(item, "get_attachment_type_display", lambda: "")(),
                "arquivo": name,
                "tamanho_bytes": getattr(item, "file", None).size if getattr(item, "file", None) else None,
                "criado_em": self._format_datetime(getattr(item, "created_at", None)),
                "conteudo": text_payload["conteudo"],
                "conteudo_truncado": text_payload["conteudo_truncado"],
                "erro_conteudo": text_payload["erro"],
            }

        for occurrence in occurrences:
            attachments_payload = [
                _attachment_payload(attachment, "ocorrencia")
                for attachment in occurrence.attachments.all()
            ]
            occurrence_attachments_payload.extend(attachments_payload)
            occurrences_payload.append(
                {
                    "titulo": occurrence.title,
                    "descricao": occurrence.description,
                    "visibilidade": occurrence.get_visibility_display(),
                    "criado_em": self._format_datetime(occurrence.created_at),
                    "registrado_por": occurrence.created_by.get_full_name()
                    if occurrence.created_by_id
                    else "",
                    "anexos": attachments_payload,
                }
            )

        project_attachments = list(
            ProjectAttachment.objects.filter(project=project).order_by("-created_at")
        )
        project_attachments_payload = [
            _attachment_payload(attachment, "projeto") for attachment in project_attachments
        ]

        observations_payload = []
        for observation in ProjectObservation.objects.filter(project=project).order_by("-created_at"):
            observations_payload.append(
                {
                    "tipo": observation.get_observation_type_display(),
                    "nota": observation.note,
                    "alteracoes": observation.changes or {},
                    "criado_em": self._format_datetime(observation.created_at),
                    "registrado_por": observation.created_by.get_full_name()
                    if observation.created_by_id
                    else "",
                }
            )

        go_no_go_items = list(
            ProjectGoNoGoChecklistItem.objects.filter(project=project).order_by("id")
        )
        go_no_go_payload = []
        go_no_go_summary = {
            "ok": 0,
            "nao": 0,
            "pendente": 0,
        }
        for item in go_no_go_items:
            if item.result == GoNoGoResult.OK:
                go_no_go_summary["ok"] += 1
            elif item.result == GoNoGoResult.NO:
                go_no_go_summary["nao"] += 1
            else:
                go_no_go_summary["pendente"] += 1
            go_no_go_payload.append(
                {
                    "criterio": item.criterion,
                    "categoria": item.category,
                    "resultado": item.get_result_display(),
                    "observacao": item.observation,
                    "evidencias": item.required_evidence,
                    "aprovador": item.approver,
                    "visibilidade": item.get_visibility_display(),
                }
            )

        riscos_identificados = [
            {
                "titulo": occurrence.title,
                "descricao": occurrence.description,
                "origem": "Ocorrencia",
                "criado_em": self._format_datetime(occurrence.created_at),
            }
            for occurrence in occurrences
        ]
        for activity in activities:
            if activity.schedule_state(today) == "late" or activity.status == ActivityStatus.BLOCKED:
                riscos_identificados.append(
                    {
                        "titulo": activity.activity,
                        "descricao": f"Atividade {activity.schedule_label(today)}",
                        "origem": "Cronograma",
                        "criado_em": "",
                    }
                )

        input_payload = {
            "project": {
                "nome": project.description,
                "cliente": project.project_client.name if project.project_client else "",
                "consultoria": "Kuiper Consultoria",
                "erp": "Senior Sistemas",
                "data_go_live_planejado": self._format_date(baseline_go_live),
                "fase_atual": phase_label,
                "status_geral": project.get_status_display(),
            },
            "modulos": module_payload,
            "cronograma": cronograma,
            "restricoes": restricoes,
            "observacoes": observations,
        }

        marcos_map: dict[str, dict[str, date | None]] = {}
        for activity in activities:
            label = self._resolve_marco_label(
                activity.phase.description if activity.phase_id else None
            )
            if not label:
                continue
            entry = marcos_map.setdefault(
                label,
                {"marco": label, "planejado": None, "real": None},
            )
            if activity.planned_end:
                if not entry["planejado"] or activity.planned_end > entry["planejado"]:
                    entry["planejado"] = activity.planned_end
            if activity.actual_end:
                if not entry["real"] or activity.actual_end > entry["real"]:
                    entry["real"] = activity.actual_end

        detailed_payload = {
            "cronograma": cronograma,
            "cronograma_atividades": cronograma_atividades,
            "apontamentos": time_entries_payload,
            "horas": {
                "total_planejado": str(total_hours_planned),
                "total_liberado": str(total_hours_released),
                "total_consumido": str(consumed_hours),
                "total_pendente": str(max(total_hours_planned - consumed_hours, Decimal("0.00"))),
            },
            "progresso": {
                "percentual_planejado": project_planned,
                "percentual_executado": project_executed,
            },
            "indicadores": {
                "total_atividades": total_activities,
                "atividades_concluidas": done_count,
                "atividades_pendentes": pending_count,
                "atividades_atrasadas": late_count,
                "percentual_atraso": self._percent(Decimal(late_count), Decimal(total_activities)),
                "percentual_execucao": project_executed,
                "dias_para_go_live_planejado": (baseline_go_live - today).days
                if baseline_go_live
                else None,
                "atividades_faltantes": missing_items[:30],
            },
            "marcos": [
                {
                    "marco": item["marco"],
                    "planejado": self._format_date(item["planejado"]),
                    "real": self._format_date(item["real"]),
                }
                for item in marcos_map.values()
            ],
            "modulos": module_details,
            "pendencias_criticas": pending_items,
            "restricoes": restricoes,
            "observacoes": observations,
            "observacoes_detalhadas": observations_payload,
            "ocorrencias": occurrences_payload,
            "riscos_identificados": riscos_identificados,
            "checklist_go_no_go": {
                "resumo": go_no_go_summary,
                "itens": go_no_go_payload,
            },
            "anexos": project_attachments_payload,
            "anexos_ocorrencias": occurrence_attachments_payload,
        }
        return input_payload, detailed_payload

    def _extract_json(self, text: str) -> dict[str, Any] | None:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            start = text.find("{")
            end = text.rfind("}")
            if start != -1 and end != -1 and end > start:
                try:
                    return json.loads(text[start : end + 1])
                except json.JSONDecodeError:
                    return None
            return None

    def _request_chatgpt(
        self,
        prompt: str,
        system_prompt: str | None = None,
        api_config: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        api_config = api_config or {}
        api_url = api_config.get("api_url") or settings.CHATGPT_API_URL
        api_key = api_config.get("api_key") or settings.CHATGPT_API_KEY
        api_model = api_config.get("api_model") or settings.CHATGPT_MODEL
        request_timeout = api_config.get("request_timeout") or settings.CHATGPT_REQUEST_TIMEOUT
        org_id = api_config.get("org_id") or settings.CHATGPT_ORG_ID
        project_id = api_config.get("project_id") or settings.CHATGPT_PROJECT_ID
        if not api_url or not api_key:
            raise ValueError("API nao configurada.")

        resolved_system_prompt = system_prompt or DEFAULT_CHATGPT_SYSTEM_PROMPT

        output_schema = """{
  "resumo_executivo": {
    "situacao_geral": "string",
    "nivel_risco": "Baixo | Medio | Alto | Critico",
    "aderencia_cronograma": "string"
  },
  "analise_cronograma": {
    "atividades_criticas": ["string"],
    "impactos_identificados": ["string"]
  },
  "analise_por_modulo": [
    {
      "modulo": "string",
      "avaliacao": "string",
      "riscos": ["string"]
    }
  ],
  "riscos": [
    {
      "descricao": "string",
      "categoria": "Prazo | Escopo | Tecnico | Negocio | Pessoas | Governanca",
      "probabilidade": "Baixa | Media | Alta",
      "impacto": "Baixo | Medio | Alto | Critico",
      "severidade": "string"
    }
  ],
  "pontos_atencao": ["string"],
  "analise_critica": "string",
  "recomendacao": {
    "decisao": "GO | GO_COM_RESSALVAS | NO_GO",
    "justificativa": "string"
  },
  "plano_acao": [
    {
      "acao": "string",
      "responsavel": "Cliente | Kuiper | Senior",
      "prazo": "YYYY-MM-DD",
      "impacto_esperado": "string"
    }
  ],
  "conclusao_executiva": "string"
}"""

        payload = {
            "model": api_model,
            "messages": [
                {
                    "role": "system",
                    "content": resolved_system_prompt,
                },
                {
                    "role": "user",
                    "content": (
                        f"{prompt}\n\n"
                        "Responda somente com um JSON valido no formato abaixo.\n"
                        "Nao inclua comentarios, markdown ou texto extra.\n"
                        f"{output_schema}"
                    ),
                }
            ],
            "temperature": 0.2,
        }

        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        if org_id:
            headers["OpenAI-Organization"] = org_id
        if project_id:
            headers["OpenAI-Project"] = project_id
        request_obj = Request(
            api_url,
            data=json.dumps(payload).encode("utf-8"),
            headers=headers,
            method="POST",
        )
        try:
            with urlopen(request_obj, timeout=request_timeout) as response:
                raw = response.read()
                charset = response.headers.get_content_charset() or "utf-8"
        except HTTPError as exc:
            error_message = ""
            try:
                charset = exc.headers.get_content_charset() or "utf-8"
                body = exc.read().decode(charset, errors="replace")
                data = json.loads(body)
                error_message = data.get("error", {}).get("message") or body
            except Exception:
                error_message = str(exc)
            logger.warning(
                "ChatGPT API error status=%s message=%s",
                getattr(exc, "code", None),
                error_message,
            )
            raise ChatGPTApiError(
                error_message,
                status_code=getattr(exc, "code", None),
            ) from exc

        payload_text = raw.decode(charset, errors="replace")
        data = json.loads(payload_text)
        message = (
            data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
        )
        analysis = self._extract_json(message)
        if not analysis:
            raise ValueError("Resposta da API invalida.")
        return analysis

    def _resolve_prompt_templates(self) -> tuple[str, str]:
        settings_obj = ChatGPTSettings.objects.first()
        system_prompt = (settings_obj.system_prompt or "").strip() if settings_obj else ""
        analysis_prompt = (settings_obj.analysis_prompt or "").strip() if settings_obj else ""
        return (
            system_prompt or DEFAULT_CHATGPT_SYSTEM_PROMPT,
            analysis_prompt or DEFAULT_CHATGPT_ANALYSIS_PROMPT,
        )

    def _resolve_api_settings(self) -> dict[str, Any]:
        settings_obj = ChatGPTSettings.objects.first()
        if not settings_obj:
            return {}
        def _clean(value: str | None) -> str:
            return value.strip() if isinstance(value, str) else ""
        return {
            "api_url": _clean(settings_obj.api_url),
            "api_key": _clean(settings_obj.api_key),
            "api_model": _clean(settings_obj.api_model),
            "request_timeout": settings_obj.request_timeout,
            "org_id": _clean(settings_obj.org_id),
            "project_id": _clean(settings_obj.project_id),
        }

    def _render_analysis_prompt(
        self,
        template: str,
        project_payload: dict[str, Any],
        detailed_payload: dict[str, Any],
    ) -> str:
        project_json = json.dumps(project_payload, ensure_ascii=True, indent=2)
        details_json = json.dumps(detailed_payload, ensure_ascii=True, indent=2)
        has_project = "{{PROJECT_CONTEXT}}" in template
        has_details = "{{DETAILS_JSON}}" in template
        prompt = (
            template.replace("{{PROJECT_CONTEXT}}", project_json)
            .replace("{{DETAILS_JSON}}", details_json)
        )
        if not has_project:
            prompt = f"{prompt}\n\nCONTEXTO DO PROJETO:\n{project_json}"
        if not has_details:
            prompt = f"{prompt}\n\nDADOS DETALHADOS:\n{details_json}"
        return prompt

    def post(self, request, *args, **kwargs):
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {"ok": False, "error": "Payload invalido."},
                status=400,
            )

        project_id_raw = payload.get("project_id")
        if not project_id_raw:
            return JsonResponse(
                {"ok": False, "error": "Projeto nao informado."},
                status=400,
            )
        try:
            project_id = int(project_id_raw)
        except (TypeError, ValueError):
            return JsonResponse(
                {"ok": False, "error": "Projeto invalido."},
                status=400,
            )

        project_qs = filter_projects_for_user(
            Project.objects.select_related("project_client"),
            request.user,
        )
        project = project_qs.filter(pk=project_id).first()
        if not project:
            return JsonResponse(
                {"ok": False, "error": "Projeto nao encontrado."},
                status=404,
            )

        activities = list(
            ProjectActivity.objects.select_related("phase", "module")
            .prefetch_related("time_entries", "subactivity_items")
            .filter(project=project)
        )
        status_report = self._build_status_report(project, activities)
        input_payload, detailed_payload = self._build_analysis_payload(
            project,
            activities,
        )

        system_prompt, analysis_template = self._resolve_prompt_templates()
        api_config = self._resolve_api_settings()
        prompt = self._render_analysis_prompt(
            analysis_template,
            input_payload["project"],
            detailed_payload,
        )

        try:
            analysis = self._request_chatgpt(
                prompt,
                system_prompt=system_prompt,
                api_config=api_config,
            )
        except ChatGPTApiError as exc:
            return JsonResponse(
                {
                    "ok": False,
                    "error": exc.public_message,
                    "status_report": status_report,
                    "project": {"id": project.id, "name": project.description},
                },
                status=502,
            )
        except TimeoutError:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Tempo limite ao conectar a API.",
                    "status_report": status_report,
                    "project": {"id": project.id, "name": project.description},
                },
                status=504,
            )
        except URLError as exc:
            reason = getattr(exc, "reason", None)
            detail = f" ({reason})" if reason else ""
            return JsonResponse(
                {
                    "ok": False,
                    "error": f"Nao foi possivel conectar a API{detail}.",
                    "status_report": status_report,
                    "project": {"id": project.id, "name": project.description},
                },
                status=502,
            )
        except (ValueError, json.JSONDecodeError) as exc:
            message = str(exc) or "Resposta da API invalida."
            return JsonResponse(
                {
                    "ok": False,
                    "error": message,
                    "status_report": status_report,
                    "project": {"id": project.id, "name": project.description},
                },
                status=502,
            )

        return JsonResponse(
            {
                "ok": True,
                "analysis": analysis,
                "status_report": status_report,
                "project": {"id": project.id, "name": project.description},
            }
        )


class ProjectChatGPTAnalysisPdfView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a este painel.")
        return super().dispatch(request, *args, **kwargs)

    def _draw_wrapped_lines(
        self,
        pdf,
        text: str,
        y: float,
        max_width: float,
        line_height: float,
        margin: float,
        page_height: float,
    ) -> float:
        if not text:
            return y
        words = text.split()
        line = ""
        for word in words:
            candidate = f"{line} {word}".strip()
            if pdf.stringWidth(candidate, "Helvetica", 9) <= max_width:
                line = candidate
                continue
            y = self._draw_pdf_line(
                pdf,
                line,
                y,
                line_height,
                margin,
                page_height,
            )
            line = word
        if line:
            y = self._draw_pdf_line(
                pdf,
                line,
                y,
                line_height,
                margin,
                page_height,
            )
        return y

    def _draw_pdf_line(
        self,
        pdf,
        line: str,
        y: float,
        line_height: float,
        margin: float,
        page_height: float,
    ) -> float:
        if y < margin + line_height:
            pdf.showPage()
            pdf.setFont("Helvetica", 9)
            y = page_height - margin
        pdf.drawString(margin, y, line)
        return y - line_height

    def _safe_text(self, value: Any) -> str:
        if value in {None, ""}:
            return "-"
        return str(value)

    def _build_pdf(
        self,
        project: Project,
        analysis: dict[str, Any],
        status_report: dict[str, Any],
    ) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.pdfgen import canvas

        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        margin = 16 * mm
        y = height - margin
        max_width = width - (margin * 2)
        line_height = 4.6 * mm

        def ensure_space(required_height: float) -> None:
            nonlocal y
            if y < margin + required_height:
                pdf.showPage()
                pdf.setFont("Helvetica", 9)
                y = height - margin

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(
            margin,
            y,
            f"Analise gerente virtual - {project.description}",
        )
        y -= 7 * mm
        pdf.setFont("Helvetica", 9)
        pdf.drawString(
            margin,
            y,
            f"Gerado em: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
        )
        y -= 8 * mm

        def draw_section(title: str, lines: list[str]) -> None:
            nonlocal y
            ensure_space(12 * mm)
            pdf.setFont("Helvetica-Bold", 11)
            pdf.setFillColor(colors.black)
            pdf.drawString(margin, y, title)
            y -= 5 * mm
            pdf.setFont("Helvetica", 9)
            pdf.setFillColor(colors.black)
            for line in lines:
                y = self._draw_wrapped_lines(
                    pdf,
                    line,
                    y,
                    max_width,
                    line_height,
                    margin,
                    height,
                )
            y -= 3 * mm

        def draw_status_badge(x: float, y_pos: float, label: str, value: str, color):
            pdf.setFillColor(color)
            pdf.circle(x + (2.6 * mm), y_pos + (1.6 * mm), 2.2 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.drawString(x + (7 * mm), y_pos, f"{label}: {self._safe_text(value)}")

        def draw_bar_chart(title: str, data: list[tuple[str, int]], bar_color) -> None:
            nonlocal y
            if not data:
                return
            ensure_space(18 * mm)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawString(margin, y, title)
            y -= 5 * mm
            max_value = max((value for _, value in data), default=0)
            bar_x = margin + (48 * mm)
            bar_width = max_width - (48 * mm)
            for label, value in data:
                ensure_space(12 * mm)
                pdf.setFont("Helvetica", 8.5)
                pdf.setFillColor(colors.black)
                pdf.drawString(margin, y, label)
                pdf.setFillColor(colors.HexColor("#e2e8f0"))
                pdf.rect(bar_x, y - (2.4 * mm), bar_width, 3 * mm, stroke=0, fill=1)
                fill_width = 0 if max_value == 0 else bar_width * (value / max_value)
                pdf.setFillColor(bar_color)
                pdf.rect(bar_x, y - (2.4 * mm), fill_width, 3 * mm, stroke=0, fill=1)
                pdf.setFillColor(colors.black)
                pdf.drawRightString(margin + max_width, y, str(value))
                y -= 5 * mm
            y -= 2 * mm

        def parse_iso_date(value: str | None) -> date | None:
            if not value:
                return None
            try:
                return date.fromisoformat(value)
            except (TypeError, ValueError):
                return None

        def draw_line_chart(
            title: str,
            values: list[dict[str, Any]],
            series_colors: dict[str, Any],
        ) -> None:
            nonlocal y
            if not values:
                draw_section(title, ["Sem dados para grafico."])
                return
            chart_height = 55 * mm
            ensure_space(chart_height + 18 * mm)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.setFillColor(colors.black)
            pdf.drawString(margin, y, title)
            y -= 5 * mm
            chart_top = y
            chart_bottom = y - chart_height
            chart_left = margin + (10 * mm)
            chart_right = margin + max_width
            chart_width = chart_right - chart_left

            pdf.setFillColor(colors.HexColor("#f8fafc"))
            pdf.rect(chart_left, chart_bottom, chart_width, chart_height, stroke=0, fill=1)
            pdf.setStrokeColor(colors.HexColor("#e2e8f0"))
            pdf.setLineWidth(0.4)
            pdf.setFont("Helvetica", 7)
            for step in (0, 25, 50, 75, 100):
                y_line = chart_bottom + (step / 100) * chart_height
                pdf.line(chart_left, y_line, chart_right, y_line)
                pdf.setFillColor(colors.black)
                pdf.drawRightString(chart_left - (2 * mm), y_line - (1.5 * mm), f"{step}%")

            pdf.setStrokeColor(colors.HexColor("#9aa3a8"))
            pdf.setLineWidth(0.6)
            pdf.rect(chart_left, chart_bottom, chart_width, chart_height, stroke=1, fill=0)

            points_by_series: dict[str, list[tuple[date, int]]] = defaultdict(list)
            for item in values:
                date_value = parse_iso_date(str(item.get("date", "")))
                if not date_value:
                    continue
                series = str(item.get("series") or "Serie")
                try:
                    value = int(item.get("value") or 0)
                except (TypeError, ValueError):
                    value = 0
                points_by_series[series].append((date_value, value))

            all_dates = [
                date_value
                for series_points in points_by_series.values()
                for date_value, _ in series_points
            ]
            if not all_dates:
                draw_section(title, ["Sem dados para grafico."])
                return
            min_date = min(all_dates)
            max_date = max(all_dates)
            span_days = max((max_date - min_date).days, 1)
            total_points = sum(len(points) for points in points_by_series.values())
            point_radius = 1.3 * mm
            if total_points > 40:
                point_radius = 0.8 * mm
            elif total_points > 20:
                point_radius = 1.0 * mm

            for series, points in points_by_series.items():
                if not points:
                    continue
                points.sort(key=lambda item: item[0])
                color = series_colors.get(series, colors.HexColor("#3b6da8"))
                pdf.setStrokeColor(color)
                pdf.setFillColor(color)
                pdf.setLineWidth(1.3)
                last_x = last_y = None
                for date_value, value in points:
                    x_ratio = (date_value - min_date).days / span_days
                    x_pos = chart_left + (x_ratio * chart_width)
                    y_ratio = min(max(value, 0), 100) / 100
                    y_pos = chart_bottom + (y_ratio * chart_height)
                    if last_x is not None:
                        pdf.line(last_x, last_y, x_pos, y_pos)
                    pdf.circle(x_pos, y_pos, point_radius, stroke=0, fill=1)
                    last_x, last_y = x_pos, y_pos

            pdf.setFont("Helvetica", 8)
            pdf.setFillColor(colors.black)
            pdf.drawString(chart_left, chart_bottom - (4 * mm), min_date.strftime("%d/%m/%Y"))
            pdf.drawRightString(
                chart_right,
                chart_bottom - (4 * mm),
                max_date.strftime("%d/%m/%Y"),
            )
            legend_x = chart_right - (42 * mm)
            legend_y = chart_top - (3 * mm)
            for series, color in series_colors.items():
                pdf.setFillColor(color)
                pdf.rect(legend_x, legend_y, 6 * mm, 3 * mm, stroke=0, fill=1)
                pdf.setFillColor(colors.black)
                pdf.drawString(legend_x + (8 * mm), legend_y, series)
                legend_y -= 4.5 * mm

            pdf.setFillColor(colors.black)
            y = chart_bottom - (11 * mm)

        def draw_module_bars(title: str, modules: list[dict[str, Any]]) -> None:
            nonlocal y
            if not modules:
                draw_section(title, ["Sem dados para grafico."])
                return
            row_height = 7.5 * mm
            label_font = 8.5
            if len(modules) > 8:
                row_height = 6.3 * mm
                label_font = 7.8
            bar_height = 2.6 * mm
            bar_gap = 1.2 * mm
            required_height = row_height * len(modules) + 16 * mm
            ensure_space(required_height)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.setFillColor(colors.black)
            pdf.drawString(margin, y, title)
            y -= 5 * mm
            max_value = max(
                (module.get("hours_planned") or 0 for module in modules),
                default=0,
            )
            max_value = max(
                max_value,
                max((module.get("hours_actual") or 0 for module in modules), default=0),
            )
            bar_x = margin + (46 * mm)
            bar_width = max_width - (46 * mm)
            legend_y = y - (1 * mm)
            pdf.setFillColor(colors.HexColor("#94a3b8"))
            pdf.rect(bar_x, legend_y, 6 * mm, 3 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.drawString(bar_x + (8 * mm), legend_y, "Planejado")
            pdf.setFillColor(colors.HexColor("#2563eb"))
            pdf.rect(bar_x + (35 * mm), legend_y, 6 * mm, 3 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.drawString(bar_x + (43 * mm), legend_y, "Realizado")
            y -= 8 * mm

            for module in modules:
                ensure_space(row_height)
                name = self._safe_text(module.get("name"))
                planned = float(module.get("hours_planned") or 0)
                actual = float(module.get("hours_actual") or 0)
                pdf.setFont("Helvetica", label_font)
                pdf.setFillColor(colors.black)
                pdf.drawString(margin, y, name)
                base_y = y - (0.7 * mm)
                pdf.setFillColor(colors.HexColor("#e2e8f0"))
                pdf.rect(bar_x, base_y, bar_width, bar_height, stroke=0, fill=1)
                planned_width = 0 if max_value == 0 else bar_width * (planned / max_value)
                pdf.setFillColor(colors.HexColor("#94a3b8"))
                pdf.rect(bar_x, base_y, planned_width, bar_height, stroke=0, fill=1)
                actual_y = base_y - (bar_height + bar_gap)
                pdf.setFillColor(colors.HexColor("#e2e8f0"))
                pdf.rect(bar_x, actual_y, bar_width, bar_height, stroke=0, fill=1)
                actual_width = 0 if max_value == 0 else bar_width * (actual / max_value)
                pdf.setFillColor(colors.HexColor("#2563eb"))
                pdf.rect(bar_x, actual_y, actual_width, bar_height, stroke=0, fill=1)
                pdf.setFillColor(colors.black)
                pdf.drawRightString(margin + max_width, y, f"{planned:.0f}h/{actual:.0f}h")
                y -= row_height
            y -= 2 * mm

        def draw_milestone_chart(title: str, milestones: list[dict[str, Any]]) -> None:
            nonlocal y
            if not milestones:
                draw_section(title, ["Sem dados para grafico."])
                return
            parsed = []
            for item in milestones:
                baseline = parse_iso_date(item.get("baseline"))
                forecast = parse_iso_date(item.get("forecast"))
                actual = parse_iso_date(item.get("actual"))
                parsed.append(
                    {
                        "name": item.get("name") or item.get("marco") or "-",
                        "baseline": baseline,
                        "forecast": forecast,
                        "actual": actual,
                        "status": item.get("status") or "",
                    }
                )
            dates = [
                value
                for entry in parsed
                for value in (entry["baseline"], entry["forecast"], entry["actual"])
                if value
            ]
            if not dates:
                draw_section(title, ["Sem datas para graficos."])
                return
            min_date = min(dates)
            max_date = max(dates)
            span_days = max((max_date - min_date).days, 1)
            row_height = 7 * mm
            if len(parsed) > 8:
                row_height = 6 * mm
            required_height = row_height * len(parsed) + 20 * mm
            ensure_space(required_height)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.setFillColor(colors.black)
            pdf.drawString(margin, y, title)
            y -= 6 * mm

            chart_left = margin + (34 * mm)
            chart_right = margin + max_width
            pdf.setStrokeColor(colors.HexColor("#9aa3a8"))
            pdf.setLineWidth(0.4)
            pdf.line(chart_left, y, chart_right, y)

            legend_y = y - (1 * mm)
            pdf.setFillColor(colors.HexColor("#94a3b8"))
            pdf.rect(chart_left, legend_y, 4 * mm, 2 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.drawString(chart_left + (6 * mm), legend_y, "Baseline")
            pdf.setFillColor(colors.HexColor("#f59e0b"))
            pdf.rect(chart_left + (40 * mm), legend_y, 4 * mm, 2 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.drawString(chart_left + (46 * mm), legend_y, "Forecast")
            pdf.setFillColor(colors.HexColor("#22c55e"))
            pdf.rect(chart_left + (78 * mm), legend_y, 4 * mm, 2 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.black)
            pdf.drawString(chart_left + (84 * mm), legend_y, "Actual")
            y -= 7 * mm

            for entry in parsed:
                ensure_space(row_height)
                label = self._safe_text(entry["name"])
                pdf.setFont("Helvetica", 8.5)
                pdf.setFillColor(colors.black)
                pdf.drawString(margin, y, label)
                row_y = y - (1 * mm)
                pdf.setStrokeColor(colors.HexColor("#e2e8f0"))
                pdf.setLineWidth(0.4)
                pdf.line(chart_left, row_y, chart_right, row_y)

                def _pos(value: date | None) -> float | None:
                    if not value:
                        return None
                    ratio = (value - min_date).days / span_days
                    return chart_left + (ratio * (chart_right - chart_left))

                baseline_x = _pos(entry["baseline"])
                forecast_x = _pos(entry["forecast"])
                actual_x = _pos(entry["actual"])
                if baseline_x and forecast_x:
                    pdf.setStrokeColor(colors.HexColor("#f59e0b"))
                    pdf.setLineWidth(1)
                    pdf.line(baseline_x, row_y, forecast_x, row_y)
                if baseline_x:
                    pdf.setFillColor(colors.HexColor("#94a3b8"))
                    pdf.circle(baseline_x, row_y, 1.5 * mm, stroke=0, fill=1)
                if forecast_x:
                    pdf.setFillColor(colors.HexColor("#f59e0b"))
                    pdf.circle(forecast_x, row_y, 1.5 * mm, stroke=0, fill=1)
                if actual_x:
                    actual_color = colors.HexColor("#22c55e")
                    if entry["status"] == "late":
                        actual_color = colors.HexColor("#ef4444")
                    pdf.setFillColor(actual_color)
                    pdf.circle(actual_x, row_y, 1.8 * mm, stroke=0, fill=1)
                y -= row_height
            pdf.setFont("Helvetica", 7.5)
            pdf.setFillColor(colors.black)
            pdf.drawString(chart_left, y + (2 * mm), min_date.strftime("%d/%m/%Y"))
            pdf.drawRightString(chart_right, y + (2 * mm), max_date.strftime("%d/%m/%Y"))
            y -= 4 * mm

        def draw_risk_matrix(title: str, risks: list[dict[str, Any]]) -> None:
            nonlocal y
            if not risks:
                draw_section(title, ["Sem riscos para matriz."])
                return
            matrix_size = 60 * mm
            ensure_space(matrix_size + 20 * mm)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.setFillColor(colors.black)
            pdf.drawString(margin, y, title)
            y -= 6 * mm
            left = margin
            bottom = y - matrix_size
            cell_size = matrix_size / 5
            for prob in range(1, 6):
                for impact in range(1, 6):
                    severity = prob * impact
                    fill = "#dcfce7"
                    if severity >= 20:
                        fill = "#fee2e2"
                    elif severity >= 12:
                        fill = "#fef3c7"
                    pdf.setFillColor(colors.HexColor(fill))
                    pdf.rect(
                        left + (prob - 1) * cell_size,
                        bottom + (impact - 1) * cell_size,
                        cell_size,
                        cell_size,
                        stroke=0,
                        fill=1,
                    )

            pdf.setStrokeColor(colors.HexColor("#cbd5e1"))
            pdf.setLineWidth(0.4)
            pdf.rect(left, bottom, matrix_size, matrix_size, stroke=1, fill=0)
            pdf.setFont("Helvetica", 7)
            for idx in range(1, 6):
                x = left + (idx - 0.5) * cell_size
                y_line = bottom + (idx - 0.5) * cell_size
                pdf.setStrokeColor(colors.HexColor("#e2e8f0"))
                pdf.line(left + (idx - 1) * cell_size, bottom, left + (idx - 1) * cell_size, bottom + matrix_size)
                pdf.line(left, bottom + (idx - 1) * cell_size, left + matrix_size, bottom + (idx - 1) * cell_size)
                pdf.setFillColor(colors.black)
                pdf.drawRightString(left - (2 * mm), y_line - (2 * mm), str(idx))
                pdf.drawCentredString(x, bottom - (5 * mm), str(idx))

            pdf.setFont("Helvetica", 8)
            pdf.setFillColor(colors.black)
            pdf.drawString(left, bottom + matrix_size + (4 * mm), "Impacto")
            pdf.drawString(left + (matrix_size / 2), bottom - (10 * mm), "Probabilidade")
            legend_x = left + matrix_size + (6 * mm)
            legend_y = bottom + matrix_size - (2 * mm)
            pdf.setFont("Helvetica", 7.5)
            for label, color in (
                ("Baixo", "#dcfce7"),
                ("Medio", "#fef3c7"),
                ("Alto", "#fee2e2"),
            ):
                pdf.setFillColor(colors.HexColor(color))
                pdf.rect(legend_x, legend_y, 4 * mm, 3 * mm, stroke=0, fill=1)
                pdf.setFillColor(colors.black)
                pdf.drawString(legend_x + (6 * mm), legend_y, label)
                legend_y -= 5 * mm

            for risk in risks:
                try:
                    probability = int(risk.get("probability") or 0)
                    impact = int(risk.get("impact") or 0)
                except (TypeError, ValueError):
                    continue
                if probability <= 0 or impact <= 0:
                    continue
                severity = probability * impact
                x = left + (probability - 0.5) * cell_size
                y_circle = bottom + (impact - 0.5) * cell_size
                radius = 1.2 * mm + (severity / 25) * (3.6 * mm)
                color = colors.HexColor("#f59e0b")
                if severity >= 20:
                    color = colors.HexColor("#ef4444")
                elif severity <= 8:
                    color = colors.HexColor("#22c55e")
                pdf.setFillColor(color)
                pdf.circle(x, y_circle, radius, stroke=0, fill=1)

            y = bottom - (12 * mm)

        resumo = analysis.get("resumo_executivo", {}) if isinstance(analysis, dict) else {}
        recomendacao = analysis.get("recomendacao", {}) if isinstance(analysis, dict) else {}
        riscos = analysis.get("riscos") or []
        plano = analysis.get("plano_acao") or []

        risk_color_map = {
            "Baixo": colors.HexColor("#4c8a63"),
            "Medio": colors.HexColor("#d6a54a"),
            "Alto": colors.HexColor("#c77c3d"),
            "Critico": colors.HexColor("#c05d54"),
        }
        decision_color_map = {
            "GO": colors.HexColor("#4c8a63"),
            "GO_COM_RESSALVAS": colors.HexColor("#d6a54a"),
            "NO_GO": colors.HexColor("#c05d54"),
        }
        risk_color = risk_color_map.get(resumo.get("nivel_risco"), colors.HexColor("#9aa3a8"))
        decision_color = decision_color_map.get(recomendacao.get("decisao"), colors.HexColor("#9aa3a8"))

        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(margin, y, "Resumo visual")
        y -= 5 * mm
        pdf.setFont("Helvetica", 9)
        draw_status_badge(margin, y, "Risco", resumo.get("nivel_risco"), risk_color)
        draw_status_badge(
            margin + (max_width / 2),
            y,
            "Decisao",
            recomendacao.get("decisao"),
            decision_color,
        )
        y -= 9 * mm

        risk_counts: dict[str, int] = {}
        for risco in riscos:
            if not isinstance(risco, dict):
                continue
            categoria = self._safe_text(risco.get("categoria"))
            risk_counts[categoria] = risk_counts.get(categoria, 0) + 1
        risk_data = sorted(risk_counts.items(), key=lambda item: item[1], reverse=True)[:6]
        draw_bar_chart("Riscos por categoria", risk_data, colors.HexColor("#c77c3d"))

        action_counts: dict[str, int] = {}
        for item in plano:
            if not isinstance(item, dict):
                continue
            responsavel = self._safe_text(item.get("responsavel"))
            action_counts[responsavel] = action_counts.get(responsavel, 0) + 1
        action_data = sorted(action_counts.items(), key=lambda item: item[1], reverse=True)[:6]
        draw_bar_chart("Plano de acao por responsavel", action_data, colors.HexColor("#3b6da8"))

        status_summary = status_report.get("summary", {}) if isinstance(status_report, dict) else {}
        status_kpis = status_report.get("kpis") or []
        status_modules = status_report.get("modules") or []
        status_milestones = status_report.get("milestones") or []
        status_risks = status_report.get("risks") or []

        def _find_kpi_value(name: str) -> str:
            for item in status_kpis:
                if not isinstance(item, dict):
                    continue
                if item.get("name") == name:
                    return str(item.get("value"))
            return "-"

        health = (status_summary.get("health") or "-").upper()
        health_color_map = {
            "GREEN": colors.HexColor("#4c8a63"),
            "YELLOW": colors.HexColor("#d6a54a"),
            "RED": colors.HexColor("#c05d54"),
        }
        health_color = health_color_map.get(health, colors.HexColor("#9aa3a8"))
        schedule_variance = _find_kpi_value("Cronograma variacao")
        effort_variance = _find_kpi_value("Esforco variacao")
        draw_section(
            "Status report",
            [
                f"Semaforo: {health}",
                f"Variacao de cronograma: {schedule_variance} dias",
                f"Variacao de esforco: {effort_variance} h",
            ],
        )
        draw_status_badge(margin, y, "Semaforo", health, health_color)
        y -= 8 * mm

        planned_actual_values = []
        for chart in status_report.get("charts") or []:
            if not isinstance(chart, dict):
                continue
            spec = chart.get("spec") or {}
            if not isinstance(spec, dict):
                continue
            if str(spec.get("title", "")).startswith("Planned vs Actual"):
                planned_actual_values = spec.get("data", {}).get("values", []) or []
                break

        draw_line_chart(
            "Planned vs Actual (%) ao longo do tempo",
            planned_actual_values,
            {"Planned": colors.HexColor("#9aa3a8"), "Actual": colors.HexColor("#3b6da8")},
        )
        draw_module_bars("Execucao por modulo (planejado vs realizado)", status_modules)
        draw_milestone_chart("Marcos (baseline vs real)", status_milestones)
        draw_risk_matrix("Matriz de riscos (probabilidade x impacto)", status_risks)

        draw_section(
            "Resumo executivo",
            [
                f"Situacao geral: {self._safe_text(resumo.get('situacao_geral'))}",
                f"Nivel de risco: {self._safe_text(resumo.get('nivel_risco'))}",
                f"Aderencia ao cronograma: {self._safe_text(resumo.get('aderencia_cronograma'))}",
            ],
        )

        cronograma = analysis.get("analise_cronograma", {}) if isinstance(analysis, dict) else {}
        atividades_criticas = cronograma.get("atividades_criticas") or []
        impactos = cronograma.get("impactos_identificados") or []
        draw_section(
            "Analise de cronograma e marcos Senior",
            [
                "Atividades criticas: " + ", ".join(atividades_criticas) if atividades_criticas else "Atividades criticas: -",
                "Impactos identificados: " + ", ".join(impactos) if impactos else "Impactos identificados: -",
            ],
        )

        modulos = analysis.get("analise_por_modulo") or []
        modulo_lines = []
        for modulo in modulos:
            if not isinstance(modulo, dict):
                continue
            modulo_riscos = modulo.get("riscos") or []
            risks_text = ", ".join(modulo_riscos) if modulo_riscos else "-"
            modulo_lines.append(
                f"{self._safe_text(modulo.get('modulo'))}: {self._safe_text(modulo.get('avaliacao'))} | Riscos: {risks_text}"
            )
        if not modulo_lines:
            modulo_lines = ["Sem dados de modulos."]
        draw_section("Analise de execucao por modulo", modulo_lines)

        risco_lines = []
        for risco in riscos:
            if not isinstance(risco, dict):
                continue
            risco_lines.append(
                f"{self._safe_text(risco.get('descricao'))} "
                f"(Categoria: {self._safe_text(risco.get('categoria'))}, "
                f"Probabilidade: {self._safe_text(risco.get('probabilidade'))}, "
                f"Impacto: {self._safe_text(risco.get('impacto'))}, "
                f"Severidade: {self._safe_text(risco.get('severidade'))})"
            )
        if not risco_lines:
            risco_lines = ["Sem riscos registrados."]
        draw_section("Riscos identificados", risco_lines)

        pontos_atencao = analysis.get("pontos_atencao") or impactos
        if isinstance(pontos_atencao, list) and pontos_atencao:
            draw_section(
                "Pontos de atencao",
                [f"- {item}" for item in pontos_atencao],
            )
        else:
            draw_section("Pontos de atencao", ["-"])

        analise_critica = analysis.get("analise_critica") or analysis.get("conclusao_executiva")
        draw_section(
            "Analise critica do projeto",
            [self._safe_text(analise_critica)],
        )

        draw_section(
            "Posicionamento tecnico",
            [
                f"Decisao: {self._safe_text(recomendacao.get('decisao'))}",
                f"Justificativa: {self._safe_text(recomendacao.get('justificativa'))}",
            ],
        )

        plano_lines = []
        for item in plano:
            if not isinstance(item, dict):
                continue
            plano_lines.append(
                f"{self._safe_text(item.get('acao'))} | "
                f"Responsavel: {self._safe_text(item.get('responsavel'))} | "
                f"Prazo: {self._safe_text(item.get('prazo'))} | "
                f"Impacto esperado: {self._safe_text(item.get('impacto_esperado'))}"
            )
        if not plano_lines:
            plano_lines = ["Sem plano de acao definido."]
        draw_section("Plano de acao estruturado", plano_lines)

        draw_section(
            "Conclusao executiva",
            [self._safe_text(analysis.get("conclusao_executiva"))],
        )

        draw_section(
            "Nota metodologica",
            [
                "Os graficos e indicadores deste relatorio sao gerados automaticamente a partir dos dados do projeto.",
                "Nao ha inclusao de informacoes externas ou inventadas.",
            ],
        )

        pdf.save()
        buffer.seek(0)
        return buffer.getvalue()

    def post(self, request, *args, **kwargs):
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {"ok": False, "error": "Payload invalido."},
                status=400,
            )

        project_id_raw = payload.get("project_id")
        analysis = payload.get("analysis")
        if not project_id_raw:
            return JsonResponse(
                {"ok": False, "error": "Projeto nao informado."},
                status=400,
            )
        if not isinstance(analysis, dict):
            return JsonResponse(
                {"ok": False, "error": "Analise invalida."},
                status=400,
            )
        try:
            project_id = int(project_id_raw)
        except (TypeError, ValueError):
            return JsonResponse(
                {"ok": False, "error": "Projeto invalido."},
                status=400,
            )

        project_qs = filter_projects_for_user(
            Project.objects.select_related("project_client"),
            request.user,
        )
        project = project_qs.filter(pk=project_id).first()
        if not project:
            return JsonResponse(
                {"ok": False, "error": "Projeto nao encontrado."},
                status=404,
            )

        activities = list(
            ProjectActivity.objects.select_related("phase", "module", "product", "submodule")
            .prefetch_related("time_entries", "subactivity_items")
            .filter(project=project)
        )
        report_builder = ProjectChatGPTAnalysisView()
        status_report = report_builder._build_status_report(project, activities)

        try:
            pdf_bytes = self._build_pdf(project, analysis, status_report)
        except ImportError:
            return JsonResponse(
                {"ok": False, "error": "Exportacao PDF indisponivel."},
                status=500,
            )

        timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
        filename = f"analise_gerente_virtual_{project.id}_{timestamp}.pdf"
        attachment = ProjectAttachment(
            project=project,
            attachment_type=ProjectAttachmentType.OTHER,
            description=f"Analise gerente virtual - {timestamp}",
        )
        attachment.file.save(filename, ContentFile(pdf_bytes), save=True)

        return JsonResponse(
            {
                "ok": True,
                "attachment_id": attachment.id,
                "attachment_url": attachment.file.url if attachment.file else "",
                "filename": filename,
            }
        )


class BaseListView(LoginRequiredMixin, ListView):
    template_name = "restricted/list.html"
    paginate_by = 12
    page_title = ""
    list_title = ""
    search_placeholder = "Buscar"
    table_headers: Iterable[str] = ()
    table_fields: Iterable[str] = ()
    search_fields: Iterable[str] = ()
    status_field: str | None = "status"
    date_filter_field: str | None = None
    date_filter_label = "Data"
    date_start_param = "start_date"
    date_end_param = "end_date"
    filter_params: dict[str, str] = {}
    create_url_name = ""
    edit_url_name = ""
    delete_url_name = ""
    extra_actions: Iterable[dict[str, str]] = ()
    allowed_roles: Iterable[str] | None = None

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def _apply_filters(self, queryset):
        params = self.request.GET
        query = params.get("q", "").strip()
        if query and self.search_fields:
            q_filter = Q()
            for field in self.search_fields:
                q_filter |= Q(**{f"{field}__icontains": query})
            queryset = queryset.filter(q_filter)

        status = params.get("status")
        if status and self.status_field:
            queryset = queryset.filter(**{self.status_field: status})

        date_field = getattr(self, "date_filter_field", None)
        if date_field:
            start_raw = params.get(self.date_start_param, "").strip()
            end_raw = params.get(self.date_end_param, "").strip()
            start_date = parse_date(start_raw) if start_raw else None
            end_date = parse_date(end_raw) if end_raw else None
            if start_date and end_date and end_date < start_date:
                start_date, end_date = end_date, start_date
            if start_date:
                queryset = queryset.filter(**{f"{date_field}__gte": start_date})
            if end_date:
                queryset = queryset.filter(**{f"{date_field}__lte": end_date})

        for param, field_name in self.filter_params.items():
            value = params.get(param)
            if value:
                queryset = queryset.filter(**{field_name: value})

        return queryset

    def get_queryset(self):
        queryset = super().get_queryset()
        return self._apply_filters(queryset)

    def _get_status_choices(self):
        if not self.status_field:
            return []
        model_field = self.model._meta.get_field(self.status_field)
        return model_field.choices

    def _build_querystring(self) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        if not params:
            return ""
        return "&" + params.urlencode()

    def get_row_actions(self, obj: models.Model) -> list[dict[str, str]]:
        return []

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        paginator = context.get("paginator")
        total_count = paginator.count if paginator is not None else None
        show_actions = getattr(self, "show_actions", True)
        show_create = getattr(self, "show_create", True)
        rows = []
        for obj in context["object_list"]:
            values = [_format_value(obj, field) for field in self.table_fields]
            row_actions = self.get_row_actions(obj) or []
            edit_url = (
                reverse(self.edit_url_name, args=[obj.pk])
                if show_actions and self.edit_url_name
                else None
            )
            delete_url = (
                reverse(self.delete_url_name, args=[obj.pk])
                if show_actions and self.delete_url_name
                else None
            )
            rows.append(
                {
                    "values": values,
                    "edit_url": edit_url,
                    "delete_url": delete_url,
                    "extra_actions": row_actions,
                }
            )
        column_count = len(self.table_headers) + (1 if show_actions else 0)
        date_field = getattr(self, "date_filter_field", None)
        date_start_value = self.request.GET.get(self.date_start_param, "").strip()
        date_end_value = self.request.GET.get(self.date_end_param, "").strip()
        context.update(
            {
                "page_title": self.page_title,
                "list_title": self.list_title or self.page_title,
                "create_url": reverse(self.create_url_name),
                "extra_actions": list(self.extra_actions),
                "table_headers": list(self.table_headers),
                "table_rows": rows,
                "query": self.request.GET.get("q", ""),
                "search_placeholder": self.search_placeholder,
                "status_choices": self._get_status_choices(),
                "current_status": self.request.GET.get("status", ""),
                "date_filter_enabled": bool(date_field),
                "date_filter_label": self.date_filter_label,
                "date_start_param": self.date_start_param,
                "date_end_param": self.date_end_param,
                "date_start_value": date_start_value,
                "date_end_value": date_end_value,
                "column_count": column_count,
                "querystring": self._build_querystring(),
                "total_count": total_count if total_count is not None else self.get_queryset().count(),
                "show_actions": show_actions,
                "show_create": show_create,
            }
        )
        return context


class BaseFormView(LoginRequiredMixin):
    page_title = ""
    submit_label = "Salvar"
    cancel_url_name = ""
    form_columns = 2
    full_width_fields: Iterable[str] = ()
    allowed_roles: Iterable[str] | None = None

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "page_title": self.page_title,
                "submit_label": self.submit_label,
                "cancel_url": reverse(self.cancel_url_name),
                "form_columns": self.form_columns,
                "full_width_fields": list(self.full_width_fields),
            }
        )
        return context


class BaseCreateView(BaseFormView, CreateView):
    template_name = "restricted/form.html"
    success_message = "Registro criado com sucesso."

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, self.success_message)
        return response


class BaseUpdateView(BaseFormView, UpdateView):
    template_name = "restricted/form.html"
    success_message = "Registro atualizado com sucesso."

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, self.success_message)
        return response


class BaseDeleteView(LoginRequiredMixin, DeleteView):
    template_name = "restricted/confirm_delete.html"
    success_message = "Registro removido com sucesso."
    cancel_url_name = ""
    page_title = "Excluir registro"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "page_title": self.page_title,
                "cancel_url": reverse(self.cancel_url_name),
            }
        )
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, self.success_message)
        return response


class UserProfileListView(BaseListView):
    model = UserProfile
    queryset = UserProfile.objects.select_related("user")
    page_title = "Usuarios"
    list_title = "Usuarios cadastrados"
    search_placeholder = "Buscar por usuario, nome ou email"
    ordering = ("user__username",)
    table_headers = ("Usuario", "Nome", "Sobrenome", "Email", "Perfil", "WhatsApp")
    table_fields = (
        "user.username",
        "user.first_name",
        "user.last_name",
        "user.email",
        "role",
        "whatsapp_phone",
    )
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__email",
        "whatsapp_phone",
    )
    status_field = None
    create_url_name = "cadastros_web:user_create"
    edit_url_name = "cadastros_web:user_profile_update"
    delete_url_name = "cadastros_web:user_profile_update"
    allowed_roles = (UserRole.ADMIN,)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        for row in context.get("table_rows", []):
            row["delete_url"] = None
        return context


class UserCreateView(BaseCreateView):
    form_class = UserCreateForm
    page_title = "Novo usuario"
    submit_label = "Criar usuario"
    cancel_url_name = "cadastros_web:user_profile_list"
    success_url = reverse_lazy("cadastros_web:user_profile_list")
    success_message = "Usuario criado com sucesso."
    allowed_roles = (UserRole.ADMIN,)


class UserProfileUpdateView(BaseUpdateView):
    model = UserProfile
    form_class = UserProfileForm
    page_title = "Editar perfil de usuario"
    submit_label = "Salvar perfil"
    cancel_url_name = "cadastros_web:user_profile_list"
    success_url = reverse_lazy("cadastros_web:user_profile_list")
    success_message = "Perfil atualizado com sucesso."
    allowed_roles = (UserRole.ADMIN,)


class WhatsappSettingsUpdateView(BaseUpdateView):
    model = WhatsappSettings
    form_class = WhatsappSettingsForm
    page_title = "Parametrizacao WhatsApp"
    submit_label = "Salvar parametros"
    cancel_url_name = "cadastros_web:dashboard"
    success_url = reverse_lazy("cadastros_web:whatsapp_settings")
    form_columns = 1
    full_width_fields = ("opportunities_numbers", "financial_numbers")
    allowed_roles = (UserRole.ADMIN,)

    def get_object(self, queryset=None):
        obj, _ = WhatsappSettings.objects.get_or_create(pk=1)
        return obj


class ChatGPTSettingsUpdateView(BaseUpdateView):
    model = ChatGPTSettings
    form_class = ChatGPTSettingsForm
    page_title = "Parametrizacao ChatGPT"
    submit_label = "Salvar prompt"
    cancel_url_name = "cadastros_web:dashboard"
    success_url = reverse_lazy("cadastros_web:chatgpt_settings")
    form_columns = 1
    full_width_fields = ("system_prompt", "analysis_prompt")
    allowed_roles = (UserRole.ADMIN,)

    def get_object(self, queryset=None):
        obj, created = ChatGPTSettings.objects.get_or_create(pk=1)
        if created:
            obj.api_url = settings.CHATGPT_API_URL
            obj.api_model = settings.CHATGPT_MODEL
            obj.request_timeout = settings.CHATGPT_REQUEST_TIMEOUT
            obj.org_id = settings.CHATGPT_ORG_ID
            obj.project_id = settings.CHATGPT_PROJECT_ID
            obj.system_prompt = DEFAULT_CHATGPT_SYSTEM_PROMPT
            obj.analysis_prompt = DEFAULT_CHATGPT_ANALYSIS_PROMPT
            obj.save(
                update_fields=[
                    "api_url",
                    "api_model",
                    "request_timeout",
                    "org_id",
                    "project_id",
                    "system_prompt",
                    "analysis_prompt",
                ]
            )
        return obj


class CompanyListView(BaseListView):
    model = Company
    page_title = "Empresas"
    list_title = "Empresas cadastradas"
    search_placeholder = "Buscar por razao social ou CNPJ"
    ordering = ("legal_name",)
    table_headers = ("Empresa", "Tipo", "CNPJ", "Status")
    table_fields = ("legal_name", "company_type", "tax_id", "status")
    search_fields = ("legal_name", "trade_name", "tax_id")
    create_url_name = "cadastros_web:company_create"
    edit_url_name = "cadastros_web:company_update"
    delete_url_name = "cadastros_web:company_delete"


class CompanyCreateView(BaseCreateView):
    model = Company
    form_class = CompanyForm
    page_title = "Nova empresa"
    submit_label = "Salvar empresa"
    cancel_url_name = "cadastros_web:company_list"
    success_url = reverse_lazy("cadastros_web:company_list")
    full_width_fields = ("address_line", "notes")


class CompanyUpdateView(BaseUpdateView):
    model = Company
    form_class = CompanyForm
    page_title = "Editar empresa"
    submit_label = "Salvar empresa"
    cancel_url_name = "cadastros_web:company_list"
    success_url = reverse_lazy("cadastros_web:company_list")
    full_width_fields = ("address_line", "notes")


class CompanyDeleteView(BaseDeleteView):
    model = Company
    cancel_url_name = "cadastros_web:company_list"
    success_url = reverse_lazy("cadastros_web:company_list")


class CompanyBankAccountListView(BaseListView):
    model = CompanyBankAccount
    queryset = CompanyBankAccount.objects.select_related("company")
    page_title = "Contas bancarias"
    list_title = "Contas bancarias da consultoria"
    search_placeholder = "Buscar por empresa ou banco"
    ordering = ("company__legal_name", "bank_name", "agency", "account_number")
    table_headers = (
        "Empresa",
        "Tipo",
        "Banco",
        "Agencia",
        "Conta",
        "Digito",
        "Saldo inicial",
    )
    table_fields = (
        "company",
        "account_type",
        "bank_name",
        "agency",
        "account_number",
        "account_digit",
        "initial_balance",
    )
    search_fields = (
        "company__legal_name",
        "company__trade_name",
        "bank_name",
        "agency",
        "account_number",
        "pix_keys",
    )
    status_field = None
    create_url_name = "cadastros_web:company_bank_account_create"
    edit_url_name = "cadastros_web:company_bank_account_update"
    delete_url_name = "cadastros_web:company_bank_account_delete"
    allowed_roles = (UserRole.ADMIN,)

    def get_row_actions(self, obj: CompanyBankAccount) -> list[dict[str, str]]:
        statement_url = reverse("cadastros_web:bank_statement_system")
        return [
            {
                "label": "Extrato",
                "url": f"{statement_url}?account_id={obj.pk}",
            }
        ]


class CompanyBankAccountCreateView(BaseCreateView):
    model = CompanyBankAccount
    form_class = CompanyBankAccountForm
    page_title = "Nova conta bancaria"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:company_bank_account_list"
    success_url = reverse_lazy("cadastros_web:company_bank_account_list")
    full_width_fields = ("pix_keys",)
    allowed_roles = (UserRole.ADMIN,)


class CompanyBankAccountUpdateView(BaseUpdateView):
    model = CompanyBankAccount
    form_class = CompanyBankAccountForm
    page_title = "Editar conta bancaria"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:company_bank_account_list"
    success_url = reverse_lazy("cadastros_web:company_bank_account_list")
    full_width_fields = ("pix_keys",)
    allowed_roles = (UserRole.ADMIN,)


class CompanyBankAccountDeleteView(BaseDeleteView):
    model = CompanyBankAccount
    cancel_url_name = "cadastros_web:company_bank_account_list"
    success_url = reverse_lazy("cadastros_web:company_bank_account_list")
    allowed_roles = (UserRole.ADMIN,)


class BankStatementView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/bank_statement.html"
    page_title = "Conciliacao bancaria"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _parse_date(value: str | None) -> date | None:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    @staticmethod
    def _parse_amount(value: str | None) -> Decimal | None:
        if not value:
            return None
        raw = (value or "").strip()
        if not raw:
            return None
        normalized = raw.replace("R$", "").replace(" ", "")
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")
        normalized = re.sub(r"[^0-9.\-]", "", normalized)
        if not normalized:
            return None
        try:
            parsed = Decimal(normalized)
        except Exception:
            return None
        return abs(parsed)

    @staticmethod
    def _parse_status_filter(value: str | None) -> str:
        raw = (value or "").strip().lower()
        if raw in {"reconciled", "pending"}:
            return raw
        return ""

    @staticmethod
    def _matches_status_filter(reconciled: bool, status_filter: str) -> bool:
        if status_filter == "reconciled":
            return reconciled
        if status_filter == "pending":
            return not reconciled
        return True

    @staticmethod
    def _parse_ofx_date(value: str | None) -> date | None:
        if not value:
            return None
        match = re.search(r"(\d{8})", value)
        if not match:
            return None
        raw = match.group(1)
        try:
            return date(int(raw[0:4]), int(raw[4:6]), int(raw[6:8]))
        except ValueError:
            return None

    @staticmethod
    def _normalize_ofx_content(text: str) -> str:
        start = text.upper().find("<OFX>")
        if start == -1:
            return ""
        content = text[start:]
        content = re.sub(
            r"&(?!(?:amp|lt|gt|quot|apos);)",
            "&amp;",
            content,
            flags=re.IGNORECASE,
        )
        if "</OFX>" not in content.upper():
            content = re.sub(
                r"<([A-Za-z0-9_]+)>([^<\r\n]+)",
                lambda m: f"<{m.group(1)}>{m.group(2).strip()}</{m.group(1)}>",
                content,
            )
        return content

    @staticmethod
    def _extract_text(node: ET.Element | None, tag: str) -> str:
        if node is None:
            return ""
        child = node.find(tag)
        if child is None or child.text is None:
            return ""
        return child.text.strip()

    @staticmethod
    def _format_decimal(value: Decimal | None) -> str:
        if value is None:
            return "-"
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value: Decimal | None) -> str:
        if value is None:
            return "-"
        return f"R$ {self._format_decimal(value)}"

    def _format_date(self, value: date | None) -> str:
        return value.strftime("%d/%m/%Y") if value else "-"

    @staticmethod
    def _signed_amount(direction: str, amount: Decimal) -> Decimal:
        if direction == BankMovementDirection.DEBIT:
            return -amount
        return amount

    @staticmethod
    def _amount_value(amount: Decimal) -> str:
        return f"{amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):.2f}"

    @staticmethod
    def _status_meta(reconciled: bool) -> tuple[str, str]:
        if reconciled:
            return "Conciliado", "chip-ok"
        return "Pendente", "chip-warn"

    @staticmethod
    def _resolve_account_plan_item(code: str) -> AccountPlanTemplateItem | None:
        return (
            AccountPlanTemplateItem.objects.filter(
                code=code,
                status=StatusChoices.ACTIVE,
                is_analytic=True,
            )
            .order_by("id")
            .first()
        )

    def _redirect_with_filters(
        self,
        account_id: str | int | None,
        start_date: date | None,
        end_date: date | None,
        system_status: str | None = None,
        ofx_status: str | None = None,
        min_amount: str | None = None,
        max_amount: str | None = None,
    ) -> HttpResponseRedirect:
        params = {}
        if account_id:
            params["account_id"] = str(account_id)
        if start_date:
            params["start_date"] = start_date.isoformat()
        if end_date:
            params["end_date"] = end_date.isoformat()
        system_status = self._parse_status_filter(system_status)
        if system_status:
            params["system_status"] = system_status
        ofx_status = self._parse_status_filter(ofx_status)
        if ofx_status:
            params["ofx_status"] = ofx_status
        min_value = self._parse_amount(min_amount)
        if min_value is not None:
            params["min_amount"] = f"{min_value:.2f}"
        max_value = self._parse_amount(max_amount)
        if max_value is not None:
            params["max_amount"] = f"{max_value:.2f}"
        url = reverse("cadastros_web:bank_statement")
        if params:
            url = f"{url}?{urlencode(params)}"
        return HttpResponseRedirect(url)

    def _parse_ofx_file(
        self, ofx_file
    ) -> tuple[list[dict[str, Any]], dict[str, Any], str | None]:
        raw = ofx_file.read()
        ofx_file.seek(0)
        decoded = None
        for encoding in ("utf-8", "latin-1"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            return [], {}, "Nao foi possivel ler o arquivo OFX."
        content = self._normalize_ofx_content(decoded)
        if not content:
            return [], {}, "Arquivo OFX invalido."
        try:
            root = ET.fromstring(content)
        except ET.ParseError:
            return [], {}, "Arquivo OFX com estrutura invalida."

        bank_id = self._extract_text(root.find(".//BANKACCTFROM"), "BANKID")
        account_number = self._extract_text(root.find(".//BANKACCTFROM"), "ACCTID")
        list_node = root.find(".//BANKTRANLIST")
        statement_start = self._parse_ofx_date(
            self._extract_text(list_node, "DTSTART")
        )
        statement_end = self._parse_ofx_date(self._extract_text(list_node, "DTEND"))

        entries: list[dict[str, Any]] = []
        for trn in root.findall(".//STMTTRN"):
            amount_text = self._extract_text(trn, "TRNAMT").replace(",", ".")
            try:
                amount_raw = Decimal(amount_text)
            except Exception:
                amount_raw = Decimal("0.00")
            direction = (
                BankMovementDirection.CREDIT
                if amount_raw >= 0
                else BankMovementDirection.DEBIT
            )
            amount = abs(amount_raw)
            posted = self._parse_ofx_date(self._extract_text(trn, "DTPOSTED"))
            entries.append(
                {
                    "posted_at": posted,
                    "amount": amount,
                    "direction": direction,
                    "fit_id": self._extract_text(trn, "FITID"),
                    "transaction_type": self._extract_text(trn, "TRNTYPE"),
                    "name": self._extract_text(trn, "NAME"),
                    "memo": self._extract_text(trn, "MEMO"),
                    "check_number": self._extract_text(trn, "CHECKNUM"),
                }
            )

        meta = {
            "bank_id": bank_id,
            "account_number": account_number,
            "statement_start": statement_start,
            "statement_end": statement_end,
        }
        return entries, meta, None

    def _handle_import(self, request):
        account_id = request.POST.get("account_id", "").strip()
        start_date = self._parse_date(request.POST.get("start_date", "").strip())
        end_date = self._parse_date(request.POST.get("end_date", "").strip())
        system_status = self._parse_status_filter(
            request.POST.get("system_status", "").strip()
        )
        ofx_status = self._parse_status_filter(
            request.POST.get("ofx_status", "").strip()
        )
        min_amount = request.POST.get("min_amount", "").strip()
        max_amount = request.POST.get("max_amount", "").strip()
        if not account_id:
            messages.error(request, "Selecione uma conta bancaria para importar o OFX.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        account = CompanyBankAccount.objects.filter(pk=account_id).first()
        if not account:
            messages.error(request, "Conta bancaria invalida.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        ofx_file = request.FILES.get("ofx_file")
        if not ofx_file:
            messages.error(request, "Selecione um arquivo OFX para importar.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        entries, meta, error = self._parse_ofx_file(ofx_file)
        if error:
            messages.error(request, error)
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        valid_entries = [entry for entry in entries if entry.get("posted_at")]
        if not valid_entries:
            messages.error(request, "Nao foi encontrado nenhum movimento valido no OFX.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        if not meta.get("statement_start"):
            meta["statement_start"] = min(
                (entry["posted_at"] for entry in valid_entries if entry.get("posted_at")),
                default=None,
            )
        if not meta.get("statement_end"):
            meta["statement_end"] = max(
                (entry["posted_at"] for entry in valid_entries if entry.get("posted_at")),
                default=None,
            )

        date_start = min((entry["posted_at"] for entry in valid_entries), default=None)
        date_end = max((entry["posted_at"] for entry in valid_entries), default=None)
        existing_entries = BankStatementEntry.objects.filter(
            bank_account=account,
            posted_at__gte=date_start,
            posted_at__lte=date_end,
        ).values_list(
            "posted_at",
            "amount",
            "direction",
            "fit_id",
            "transaction_type",
            "name",
            "memo",
            "check_number",
        )
        existing_signatures = {
            (
                posted_at,
                amount,
                direction,
                (fit_id or "").strip(),
                (transaction_type or "").strip(),
                (name or "").strip(),
                (memo or "").strip(),
                (check_number or "").strip(),
            )
            for (
                posted_at,
                amount,
                direction,
                fit_id,
                transaction_type,
                name,
                memo,
                check_number,
            ) in existing_entries
        }
        existing_fit_ids = set(
            BankStatementEntry.objects.filter(
                bank_account=account,
            )
            .exclude(fit_id__isnull=True)
            .exclude(fit_id="")
            .values_list("fit_id", flat=True)
        )
        dedupe_signatures: set[tuple[Any, ...]] = set()
        unique_entries: list[dict[str, Any]] = []
        skipped = 0
        for entry in valid_entries:
            signature = (
                entry["posted_at"],
                entry["amount"],
                entry["direction"],
                (entry.get("fit_id") or "").strip(),
                (entry.get("transaction_type") or "").strip(),
                (entry.get("name") or "").strip(),
                (entry.get("memo") or "").strip(),
                (entry.get("check_number") or "").strip(),
            )
            fit_id = (entry.get("fit_id") or "").strip()
            if fit_id and fit_id in existing_fit_ids:
                skipped += 1
                continue
            if signature in existing_signatures:
                skipped += 1
                continue
            if signature in dedupe_signatures:
                skipped += 1
                continue
            dedupe_signatures.add(signature)
            unique_entries.append(entry)

        if not unique_entries:
            messages.warning(
                request,
                "Nenhum movimento novo para importar. "
                "Todos os movimentos ja existiam.",
            )
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        with transaction.atomic():
            import_record = BankStatementImport.objects.create(
                bank_account=account,
                imported_by=request.user,
                original_filename=getattr(ofx_file, "name", "") or "",
                statement_start=meta.get("statement_start"),
                statement_end=meta.get("statement_end"),
                bank_id=meta.get("bank_id", ""),
                account_number=meta.get("account_number", ""),
            )
            entries_to_create = [
                BankStatementEntry(
                    statement_import=import_record,
                    bank_account=account,
                    posted_at=entry["posted_at"],
                    amount=entry["amount"],
                    direction=entry["direction"],
                    fit_id=entry.get("fit_id", ""),
                    transaction_type=entry.get("transaction_type", ""),
                    name=entry.get("name", ""),
                    memo=entry.get("memo", ""),
                    check_number=entry.get("check_number", ""),
                )
                for entry in unique_entries
            ]
            BankStatementEntry.objects.bulk_create(entries_to_create)

        if skipped:
            messages.success(
                request,
                f"OFX importado com {len(entries_to_create)} movimentos. "
                f"{skipped} ignorados por ja existirem.",
            )
        else:
            messages.success(request, f"OFX importado com {len(entries_to_create)} movimentos.")
        return self._redirect_with_filters(
            account_id,
            start_date,
            end_date,
            system_status,
            ofx_status,
            min_amount,
            max_amount,
        )

    def _handle_generate(self, request):
        account_id = request.POST.get("account_id", "").strip()
        start_date = self._parse_date(request.POST.get("start_date", "").strip())
        end_date = self._parse_date(request.POST.get("end_date", "").strip())
        system_status = self._parse_status_filter(
            request.POST.get("system_status", "").strip()
        )
        ofx_status = self._parse_status_filter(
            request.POST.get("ofx_status", "").strip()
        )
        min_amount = request.POST.get("min_amount", "").strip()
        max_amount = request.POST.get("max_amount", "").strip()
        revenue_account_id = request.POST.get("revenue_account_id", "").strip()
        if not account_id:
            messages.error(request, "Selecione uma conta bancaria.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        account = CompanyBankAccount.objects.filter(pk=account_id).first()
        if not account:
            messages.error(request, "Conta bancaria invalida.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        ofx_ids = [value for value in request.POST.getlist("ofx_ids") if value]
        if not ofx_ids:
            messages.error(request, "Selecione movimentos OFX para gerar lancamentos.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        entries = list(
            BankStatementEntry.objects.filter(
                pk__in=ofx_ids, bank_account=account
            ).order_by("posted_at", "id")
        )
        if len(entries) != len(ofx_ids):
            messages.error(request, "Alguns movimentos OFX nao foram encontrados.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        reconciled_ids = set(
            BankReconciliationOfxItem.objects.filter(
                reconciliation__bank_account=account,
                ofx_entry_id__in=ofx_ids,
            ).values_list("ofx_entry_id", flat=True)
        )
        if reconciled_ids:
            messages.error(
                request,
                "Existem movimentos OFX ja conciliados na selecao.",
            )
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        has_credit = any(
            entry.direction == BankMovementDirection.CREDIT for entry in entries
        )
        has_debit = any(
            entry.direction == BankMovementDirection.DEBIT for entry in entries
        )
        revenue_item = None
        if has_credit:
            if not revenue_account_id:
                messages.error(
                    request,
                    "Selecione a conta contabil para receitas antes de gerar os lancamentos.",
                )
                return self._redirect_with_filters(
                    account_id,
                    start_date,
                    end_date,
                    system_status,
                    ofx_status,
                    min_amount,
                    max_amount,
                )
            revenue_item = (
                AccountPlanTemplateItem.objects.filter(
                    pk=revenue_account_id,
                    status=StatusChoices.ACTIVE,
                    is_analytic=True,
                )
                .order_by("id")
                .first()
            )
            if not revenue_item:
                messages.error(request, "Conta contabil de receita invalida.")
                return self._redirect_with_filters(
                    account_id,
                    start_date,
                    end_date,
                    system_status,
                    ofx_status,
                    min_amount,
                    max_amount,
                )

        expense_item = None
        if has_debit:
            expense_item = self._resolve_account_plan_item("4.03.01")
            if not expense_item:
                messages.error(
                    request,
                    "Conta contabil 4.03.01 nao encontrada para despesas.",
                )
                return self._redirect_with_filters(
                    account_id,
                    start_date,
                    end_date,
                    system_status,
                    ofx_status,
                    min_amount,
                    max_amount,
                )

        with transaction.atomic():
            system_movements = [
                BankSystemMovement(
                    bank_account=account,
                    account_plan_item=(
                        revenue_item
                        if entry.direction == BankMovementDirection.CREDIT
                        else expense_item
                    ),
                    movement_date=entry.posted_at,
                    description=entry.memo or entry.name or "Movimento OFX",
                    amount=entry.amount,
                    direction=entry.direction,
                    source=BankMovementSource.OFX,
                    created_by=request.user,
                )
                for entry in entries
            ]
            created_movements = BankSystemMovement.objects.bulk_create(system_movements)
            total_ofx = sum(
                (
                    self._signed_amount(
                        entry.direction, entry.amount or Decimal("0.00")
                    )
                    for entry in entries
                ),
                Decimal("0.00"),
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            reconciliation = BankReconciliation.objects.create(
                bank_account=account,
                created_by=request.user,
                total_system=total_ofx,
                total_ofx=total_ofx,
                difference=Decimal("0.00"),
            )
            system_items = [
                BankReconciliationSystemItem(
                    reconciliation=reconciliation,
                    system_movement=movement,
                    amount=movement.amount,
                    direction=movement.direction,
                )
                for movement in created_movements
            ]
            BankReconciliationSystemItem.objects.bulk_create(system_items)
            ofx_items = [
                BankReconciliationOfxItem(
                    reconciliation=reconciliation,
                    ofx_entry=entry,
                    amount=entry.amount,
                    direction=entry.direction,
                )
                for entry in entries
            ]
            BankReconciliationOfxItem.objects.bulk_create(ofx_items)

        messages.success(request, "Lancamentos gerados e conciliados com sucesso.")
        return self._redirect_with_filters(
            account_id,
            start_date,
            end_date,
            system_status,
            ofx_status,
            min_amount,
            max_amount,
        )

    def _handle_delete_ofx(self, request):
        account_id = request.POST.get("account_id", "").strip()
        start_date = self._parse_date(request.POST.get("start_date", "").strip())
        end_date = self._parse_date(request.POST.get("end_date", "").strip())
        system_status = self._parse_status_filter(
            request.POST.get("system_status", "").strip()
        )
        ofx_status = self._parse_status_filter(
            request.POST.get("ofx_status", "").strip()
        )
        min_amount = request.POST.get("min_amount", "").strip()
        max_amount = request.POST.get("max_amount", "").strip()
        if not account_id:
            messages.error(request, "Selecione uma conta bancaria.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        account = CompanyBankAccount.objects.filter(pk=account_id).first()
        if not account:
            messages.error(request, "Conta bancaria invalida.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        ofx_ids = [value for value in request.POST.getlist("ofx_ids") if value]
        if not ofx_ids:
            messages.error(request, "Selecione movimentos OFX para excluir.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        entries = list(
            BankStatementEntry.objects.filter(
                pk__in=ofx_ids, bank_account=account
            ).order_by("posted_at", "id")
        )
        if len(entries) != len(set(ofx_ids)):
            messages.error(request, "Alguns movimentos OFX nao foram encontrados.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        reconciled_ids = set(
            BankReconciliationOfxItem.objects.filter(
                reconciliation__bank_account=account,
                ofx_entry_id__in=ofx_ids,
            ).values_list("ofx_entry_id", flat=True)
        )
        if reconciled_ids:
            messages.error(
                request,
                "Existem movimentos OFX ja conciliados na selecao.",
            )
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        with transaction.atomic():
            BankStatementEntry.objects.filter(
                pk__in=ofx_ids, bank_account=account
            ).delete()

        messages.success(
            request,
            f"{len(entries)} movimentos OFX excluidos com sucesso.",
        )
        return self._redirect_with_filters(
            account_id,
            start_date,
            end_date,
            system_status,
            ofx_status,
            min_amount,
            max_amount,
        )

    def _handle_reconcile(self, request):
        account_id = request.POST.get("account_id", "").strip()
        start_date = self._parse_date(request.POST.get("start_date", "").strip())
        end_date = self._parse_date(request.POST.get("end_date", "").strip())
        system_status = self._parse_status_filter(
            request.POST.get("system_status", "").strip()
        )
        ofx_status = self._parse_status_filter(
            request.POST.get("ofx_status", "").strip()
        )
        min_amount = request.POST.get("min_amount", "").strip()
        max_amount = request.POST.get("max_amount", "").strip()
        if not account_id:
            messages.error(request, "Selecione uma conta bancaria.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        account = CompanyBankAccount.objects.filter(pk=account_id).first()
        if not account:
            messages.error(request, "Conta bancaria invalida.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        system_tokens = [value for value in request.POST.getlist("system_ids") if value]
        ofx_ids = [value for value in request.POST.getlist("ofx_ids") if value]
        if not system_tokens or not ofx_ids:
            messages.error(request, "Selecione movimentos do sistema e do OFX.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        receivable_ids = []
        payable_ids = []
        movement_ids = []
        for token in system_tokens:
            try:
                prefix, raw_id = token.split(":", 1)
                item_id = int(raw_id)
            except ValueError:
                continue
            if prefix == "rp":
                receivable_ids.append(item_id)
            elif prefix == "pp":
                payable_ids.append(item_id)
            elif prefix == "bm":
                movement_ids.append(item_id)

        receivable_qs = AccountsReceivablePayment.objects.select_related(
            "receivable", "receivable__client"
        ).filter(bank_account=account, id__in=receivable_ids)
        payable_qs = AccountsPayablePayment.objects.select_related(
            "payable", "payable__supplier"
        ).filter(bank_account=account, id__in=payable_ids)
        system_qs = BankSystemMovement.objects.filter(
            bank_account=account, id__in=movement_ids
        )

        if receivable_qs.count() != len(set(receivable_ids)) or payable_qs.count() != len(
            set(payable_ids)
        ) or system_qs.count() != len(set(movement_ids)):
            messages.error(request, "Alguns movimentos do sistema nao foram encontrados.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        reconciled_receivable = set(
            BankReconciliationSystemItem.objects.filter(
                reconciliation__bank_account=account,
                receivable_payment_id__in=receivable_ids,
            ).values_list("receivable_payment_id", flat=True)
        )
        reconciled_payable = set(
            BankReconciliationSystemItem.objects.filter(
                reconciliation__bank_account=account,
                payable_payment_id__in=payable_ids,
            ).values_list("payable_payment_id", flat=True)
        )
        reconciled_system = set(
            BankReconciliationSystemItem.objects.filter(
                reconciliation__bank_account=account,
                system_movement_id__in=movement_ids,
            ).values_list("system_movement_id", flat=True)
        )
        if reconciled_receivable or reconciled_payable or reconciled_system:
            messages.error(request, "Existem movimentos do sistema ja conciliados.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        entries = list(
            BankStatementEntry.objects.filter(pk__in=ofx_ids, bank_account=account)
        )
        if len(entries) != len(ofx_ids):
            messages.error(request, "Alguns movimentos OFX nao foram encontrados.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        reconciled_ofx = set(
            BankReconciliationOfxItem.objects.filter(
                reconciliation__bank_account=account,
                ofx_entry_id__in=ofx_ids,
            ).values_list("ofx_entry_id", flat=True)
        )
        if reconciled_ofx:
            messages.error(request, "Existem movimentos OFX ja conciliados.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        total_system = Decimal("0.00")
        system_items = []
        for payment in receivable_qs:
            amount = payment.amount or Decimal("0.00")
            total_system += self._signed_amount(BankMovementDirection.CREDIT, amount)
            system_items.append(
                BankReconciliationSystemItem(
                    receivable_payment=payment,
                    amount=amount,
                    direction=BankMovementDirection.CREDIT,
                )
            )
        for payment in payable_qs:
            amount = payment.amount or Decimal("0.00")
            total_system += self._signed_amount(BankMovementDirection.DEBIT, amount)
            system_items.append(
                BankReconciliationSystemItem(
                    payable_payment=payment,
                    amount=amount,
                    direction=BankMovementDirection.DEBIT,
                )
            )
        for movement in system_qs:
            amount = movement.amount or Decimal("0.00")
            total_system += self._signed_amount(movement.direction, amount)
            system_items.append(
                BankReconciliationSystemItem(
                    system_movement=movement,
                    amount=amount,
                    direction=movement.direction,
                )
            )

        total_ofx = sum(
            (
                self._signed_amount(entry.direction, entry.amount or Decimal("0.00"))
                for entry in entries
            ),
            Decimal("0.00"),
        )
        total_system = total_system.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total_ofx = total_ofx.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        difference = (total_system - total_ofx).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if difference != Decimal("0.00"):
            messages.error(
                request,
                "Nao foi possivel conciliar: soma do sistema diferente do OFX.",
            )
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        with transaction.atomic():
            reconciliation = BankReconciliation.objects.create(
                bank_account=account,
                created_by=request.user,
                total_system=total_system,
                total_ofx=total_ofx,
                difference=difference,
            )
            for item in system_items:
                item.reconciliation = reconciliation
            BankReconciliationSystemItem.objects.bulk_create(system_items)
            ofx_items = [
                BankReconciliationOfxItem(
                    reconciliation=reconciliation,
                    ofx_entry=entry,
                    amount=entry.amount,
                    direction=entry.direction,
                )
                for entry in entries
            ]
            BankReconciliationOfxItem.objects.bulk_create(ofx_items)

        messages.success(request, "Conciliacao realizada com sucesso.")
        return self._redirect_with_filters(
            account_id,
            start_date,
            end_date,
            system_status,
            ofx_status,
            min_amount,
            max_amount,
        )

    def _handle_undo_reconcile(self, request):
        account_id = request.POST.get("account_id", "").strip()
        start_date = self._parse_date(request.POST.get("start_date", "").strip())
        end_date = self._parse_date(request.POST.get("end_date", "").strip())
        system_status = self._parse_status_filter(
            request.POST.get("system_status", "").strip()
        )
        ofx_status = self._parse_status_filter(
            request.POST.get("ofx_status", "").strip()
        )
        min_amount = request.POST.get("min_amount", "").strip()
        max_amount = request.POST.get("max_amount", "").strip()
        reconciliation_id = request.POST.get("reconciliation_id", "").strip()
        if not account_id:
            messages.error(request, "Selecione uma conta bancaria.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        account = CompanyBankAccount.objects.filter(pk=account_id).first()
        if not account:
            messages.error(request, "Conta bancaria invalida.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        if not reconciliation_id:
            messages.error(request, "Selecione uma conciliacao para desfazer.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )
        try:
            reconciliation_pk = int(reconciliation_id)
        except ValueError:
            messages.error(request, "Conciliacao invalida.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        reconciliation = BankReconciliation.objects.filter(
            pk=reconciliation_pk,
            bank_account=account,
        ).first()
        if not reconciliation:
            messages.error(request, "Conciliacao nao encontrada.")
            return self._redirect_with_filters(
                account_id,
                start_date,
                end_date,
                system_status,
                ofx_status,
                min_amount,
                max_amount,
            )

        with transaction.atomic():
            reconciliation.delete()

        messages.success(request, "Conciliacao desfeita com sucesso.")
        return self._redirect_with_filters(
            account_id,
            start_date,
            end_date,
            system_status,
            ofx_status,
            min_amount,
            max_amount,
        )

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "").strip()
        if action == "import_ofx":
            return self._handle_import(request)
        if action == "generate_system":
            return self._handle_generate(request)
        if action == "delete_ofx":
            return self._handle_delete_ofx(request)
        if action == "reconcile":
            return self._handle_reconcile(request)
        if action == "undo_reconcile":
            return self._handle_undo_reconcile(request)
        messages.error(request, "Acao invalida.")
        return self._redirect_with_filters(
            request.POST.get("account_id", "").strip(),
            self._parse_date(request.POST.get("start_date", "").strip()),
            self._parse_date(request.POST.get("end_date", "").strip()),
            request.POST.get("system_status", "").strip(),
            request.POST.get("ofx_status", "").strip(),
            request.POST.get("min_amount", "").strip(),
            request.POST.get("max_amount", "").strip(),
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        accounts = CompanyBankAccount.objects.select_related("company").order_by(
            "company__legal_name",
            "bank_name",
            "agency",
            "account_number",
        )
        account_id = self.request.GET.get("account_id", "").strip()
        start_date = self._parse_date(self.request.GET.get("start_date", "").strip())
        end_date = self._parse_date(self.request.GET.get("end_date", "").strip())
        system_status = self._parse_status_filter(
            self.request.GET.get("system_status", "").strip()
        )
        ofx_status = self._parse_status_filter(
            self.request.GET.get("ofx_status", "").strip()
        )
        min_amount_raw = self.request.GET.get("min_amount", "").strip()
        max_amount_raw = self.request.GET.get("max_amount", "").strip()
        min_amount = self._parse_amount(min_amount_raw)
        max_amount = self._parse_amount(max_amount_raw)
        if not start_date and not end_date:
            today = timezone.localdate()
            start_date = today.replace(day=1)
            end_date = today.replace(day=monthrange(today.year, today.month)[1])
        selected_account = accounts.filter(pk=account_id).first() if account_id else None

        rows = []
        system_movements = []
        ofx_movements = []
        opening_balance = None
        closing_balance = None
        total_credits = Decimal("0.00")
        total_debits = Decimal("0.00")

        if selected_account:
            opening_balance = selected_account.initial_balance or Decimal("0.00")
            if start_date:
                credits_before = (
                    _exclude_compensation_notes(
                        AccountsReceivablePayment.objects.filter(
                            bank_account=selected_account,
                            payment_date__lt=start_date,
                        )
                    ).aggregate(total=Sum("amount"))
                ).get("total") or Decimal("0.00")
                debits_before = (
                    _exclude_compensation_notes(
                        AccountsPayablePayment.objects.filter(
                            bank_account=selected_account,
                            payment_date__lt=start_date,
                        )
                    ).aggregate(total=Sum("amount"))
                ).get("total") or Decimal("0.00")
                movement_totals = (
                    BankSystemMovement.objects.filter(
                        bank_account=selected_account,
                        movement_date__lt=start_date,
                    )
                    .values("direction")
                    .annotate(total=Coalesce(Sum("amount"), Value(Decimal("0.00"))))
                )
                manual_credit_before = Decimal("0.00")
                manual_debit_before = Decimal("0.00")
                for row in movement_totals:
                    if row["direction"] == BankMovementDirection.CREDIT:
                        manual_credit_before = row["total"] or Decimal("0.00")
                    elif row["direction"] == BankMovementDirection.DEBIT:
                        manual_debit_before = row["total"] or Decimal("0.00")
                opening_balance += credits_before + manual_credit_before
                opening_balance -= debits_before + manual_debit_before

            receivable_qs = _exclude_compensation_notes(
                AccountsReceivablePayment.objects.select_related(
                    "receivable",
                    "receivable__client",
                    "receivable__account_plan_item",
                ).filter(bank_account=selected_account)
            )
            payable_qs = _exclude_compensation_notes(
                AccountsPayablePayment.objects.select_related(
                    "payable",
                    "payable__supplier",
                    "payable__account_plan_item",
                ).filter(bank_account=selected_account)
            )
            system_qs = BankSystemMovement.objects.select_related(
                "account_plan_item"
            ).filter(bank_account=selected_account)
            if min_amount is not None:
                receivable_qs = receivable_qs.filter(amount__gte=min_amount)
                payable_qs = payable_qs.filter(amount__gte=min_amount)
                system_qs = system_qs.filter(amount__gte=min_amount)
            if max_amount is not None:
                receivable_qs = receivable_qs.filter(amount__lte=max_amount)
                payable_qs = payable_qs.filter(amount__lte=max_amount)
                system_qs = system_qs.filter(amount__lte=max_amount)
            if start_date:
                receivable_qs = receivable_qs.filter(payment_date__gte=start_date)
                payable_qs = payable_qs.filter(payment_date__gte=start_date)
                system_qs = system_qs.filter(movement_date__gte=start_date)
            if end_date:
                receivable_qs = receivable_qs.filter(payment_date__lte=end_date)
                payable_qs = payable_qs.filter(payment_date__lte=end_date)
                system_qs = system_qs.filter(movement_date__lte=end_date)

            receivable_payments = list(receivable_qs)
            payable_payments = list(payable_qs)
            system_entries = list(system_qs)
            receivable_ids = [payment.id for payment in receivable_payments]
            payable_ids = [payment.id for payment in payable_payments]
            system_ids = [movement.id for movement in system_entries]
            reconciled_receivable = set()
            reconciled_payable = set()
            reconciled_system = set()
            reconciliation_by_receivable = {}
            reconciliation_by_payable = {}
            reconciliation_by_system = {}
            if receivable_ids:
                reconciliation_by_receivable = dict(
                    BankReconciliationSystemItem.objects.filter(
                        reconciliation__bank_account=selected_account,
                        receivable_payment_id__in=receivable_ids,
                    ).values_list("receivable_payment_id", "reconciliation_id")
                )
                reconciled_receivable = set(reconciliation_by_receivable.keys())
            if payable_ids:
                reconciliation_by_payable = dict(
                    BankReconciliationSystemItem.objects.filter(
                        reconciliation__bank_account=selected_account,
                        payable_payment_id__in=payable_ids,
                    ).values_list("payable_payment_id", "reconciliation_id")
                )
                reconciled_payable = set(reconciliation_by_payable.keys())
            if system_ids:
                reconciliation_by_system = dict(
                    BankReconciliationSystemItem.objects.filter(
                        reconciliation__bank_account=selected_account,
                        system_movement_id__in=system_ids,
                    ).values_list("system_movement_id", "reconciliation_id")
                )
                reconciled_system = set(reconciliation_by_system.keys())

            def _account_info(
                item: AccountPlanTemplateItem | None,
            ) -> tuple[str, str]:
                if not item:
                    return "-", "-"
                code = (item.code or "").strip() or "-"
                description = (item.description or "").strip() or "-"
                return code, description

            movements = []
            for payment in receivable_payments:
                title = payment.receivable
                description = f"Recebimento {title.document_number} - {title.client}"
                account_code, account_description = _account_info(
                    title.account_plan_item
                )
                movements.append(
                    {
                        "date": payment.payment_date,
                        "created_at": payment.created_at,
                        "description": description,
                        "credit": payment.amount,
                        "debit": Decimal("0.00"),
                        "account_code": account_code,
                        "account_description": account_description,
                    }
                )
                total_credits += payment.amount or Decimal("0.00")
                reconciled = payment.id in reconciled_receivable
                reconciliation_id = reconciliation_by_receivable.get(payment.id)
                status_label, status_class = self._status_meta(reconciled)
                if self._matches_status_filter(reconciled, system_status):
                    system_movements.append(
                        {
                            "key": f"rp:{payment.id}",
                            "date": self._format_date(payment.payment_date),
                            "sort_date": payment.payment_date,
                            "sort_amount": payment.amount or Decimal("0.00"),
                            "description": description,
                            "type_label": "Recebimento",
                            "credit": self._format_currency(payment.amount),
                            "debit": "-",
                            "credit_class": "value-positive" if payment.amount > 0 else "",
                            "debit_class": "",
                            "amount_value": self._amount_value(payment.amount),
                            "direction": BankMovementDirection.CREDIT,
                            "status_label": status_label,
                            "status_class": status_class,
                            "reconciled": reconciled,
                            "reconciliation_id": reconciliation_id,
                        }
                    )

            for payment in payable_payments:
                title = payment.payable
                description = f"Pagamento {title.document_number} - {title.supplier}"
                account_code, account_description = _account_info(
                    title.account_plan_item
                )
                movements.append(
                    {
                        "date": payment.payment_date,
                        "created_at": payment.created_at,
                        "description": description,
                        "credit": Decimal("0.00"),
                        "debit": payment.amount,
                        "account_code": account_code,
                        "account_description": account_description,
                    }
                )
                total_debits += payment.amount or Decimal("0.00")
                reconciled = payment.id in reconciled_payable
                reconciliation_id = reconciliation_by_payable.get(payment.id)
                status_label, status_class = self._status_meta(reconciled)
                if self._matches_status_filter(reconciled, system_status):
                    system_movements.append(
                        {
                            "key": f"pp:{payment.id}",
                            "date": self._format_date(payment.payment_date),
                            "sort_date": payment.payment_date,
                            "sort_amount": payment.amount or Decimal("0.00"),
                            "description": description,
                            "type_label": "Pagamento",
                            "credit": "-",
                            "debit": self._format_currency(payment.amount),
                            "credit_class": "",
                            "debit_class": "value-negative" if payment.amount > 0 else "",
                            "amount_value": self._amount_value(payment.amount),
                            "direction": BankMovementDirection.DEBIT,
                            "status_label": status_label,
                            "status_class": status_class,
                            "reconciled": reconciled,
                            "reconciliation_id": reconciliation_id,
                        }
                    )

            for movement in system_entries:
                credit_value = (
                    movement.amount if movement.direction == BankMovementDirection.CREDIT else None
                )
                debit_value = (
                    movement.amount if movement.direction == BankMovementDirection.DEBIT else None
                )
                account_code, account_description = _account_info(
                    movement.account_plan_item
                )
                movements.append(
                    {
                        "date": movement.movement_date,
                        "created_at": movement.created_at,
                        "description": movement.description,
                        "credit": credit_value or Decimal("0.00"),
                        "debit": debit_value or Decimal("0.00"),
                        "account_code": account_code,
                        "account_description": account_description,
                    }
                )
                if credit_value:
                    total_credits += credit_value
                if debit_value:
                    total_debits += debit_value
                reconciled = movement.id in reconciled_system
                reconciliation_id = reconciliation_by_system.get(movement.id)
                status_label, status_class = self._status_meta(reconciled)
                if self._matches_status_filter(reconciled, system_status):
                    system_movements.append(
                        {
                            "key": f"bm:{movement.id}",
                            "date": self._format_date(movement.movement_date),
                            "sort_date": movement.movement_date,
                            "sort_amount": movement.amount or Decimal("0.00"),
                            "description": movement.description,
                            "type_label": "Lancamento",
                            "credit": self._format_currency(credit_value)
                            if credit_value
                            else "-",
                            "debit": self._format_currency(debit_value)
                            if debit_value
                            else "-",
                            "credit_class": "value-positive" if credit_value else "",
                            "debit_class": "value-negative" if debit_value else "",
                            "amount_value": self._amount_value(movement.amount),
                            "direction": movement.direction,
                            "status_label": status_label,
                            "status_class": status_class,
                            "reconciled": reconciled,
                            "reconciliation_id": reconciliation_id,
                        }
                    )

            movements.sort(key=lambda item: (item["date"], item["created_at"]))
            running = opening_balance
            for movement in movements:
                running = (
                    running
                    + (movement["credit"] or Decimal("0.00"))
                    - (movement["debit"] or Decimal("0.00"))
                )
                credit_value = movement["credit"] or Decimal("0.00")
                debit_value = movement["debit"] or Decimal("0.00")
                rows.append(
                    {
                        "date": self._format_date(movement["date"]),
                        "description": movement["description"],
                        "account_code": movement.get("account_code", "-"),
                        "account_description": movement.get("account_description", "-"),
                        "credit": self._format_currency(credit_value)
                        if credit_value
                        else "-",
                        "credit_class": "value-positive" if credit_value > 0 else "",
                        "debit": self._format_currency(debit_value)
                        if debit_value
                        else "-",
                        "debit_class": "value-negative" if debit_value > 0 else "",
                        "balance": self._format_currency(running),
                        "balance_class": "value-balance-positive"
                        if running >= 0
                        else "value-balance-negative",
                    }
                )
            closing_balance = running

            ofx_qs = BankStatementEntry.objects.filter(bank_account=selected_account)
            if min_amount is not None:
                ofx_qs = ofx_qs.filter(amount__gte=min_amount)
            if max_amount is not None:
                ofx_qs = ofx_qs.filter(amount__lte=max_amount)
            if start_date:
                ofx_qs = ofx_qs.filter(posted_at__gte=start_date)
            if end_date:
                ofx_qs = ofx_qs.filter(posted_at__lte=end_date)
            ofx_entries = list(ofx_qs.order_by("posted_at", "id"))
            ofx_ids_list = [entry.id for entry in ofx_entries]
            reconciled_ofx = set()
            reconciliation_by_ofx = {}
            if ofx_ids_list:
                reconciliation_by_ofx = dict(
                    BankReconciliationOfxItem.objects.filter(
                        reconciliation__bank_account=selected_account,
                        ofx_entry_id__in=ofx_ids_list,
                    ).values_list("ofx_entry_id", "reconciliation_id")
                )
                reconciled_ofx = set(reconciliation_by_ofx.keys())
            for entry in ofx_entries:
                description = entry.memo or entry.name or entry.transaction_type or "OFX"
                credit_value = (
                    entry.amount if entry.direction == BankMovementDirection.CREDIT else None
                )
                debit_value = (
                    entry.amount if entry.direction == BankMovementDirection.DEBIT else None
                )
                reconciled = entry.id in reconciled_ofx
                reconciliation_id = reconciliation_by_ofx.get(entry.id)
                status_label, status_class = self._status_meta(reconciled)
                if self._matches_status_filter(reconciled, ofx_status):
                    ofx_movements.append(
                        {
                            "id": entry.id,
                            "date": self._format_date(entry.posted_at),
                            "sort_date": entry.posted_at,
                            "sort_amount": entry.amount or Decimal("0.00"),
                            "description": description,
                            "fit_id": entry.fit_id or "-",
                            "credit": self._format_currency(credit_value)
                            if credit_value
                            else "-",
                            "debit": self._format_currency(debit_value)
                            if debit_value
                            else "-",
                            "credit_class": "value-positive" if credit_value else "",
                            "debit_class": "value-negative" if debit_value else "",
                            "amount_value": self._amount_value(entry.amount),
                            "direction": entry.direction,
                            "status_label": status_label,
                            "status_class": status_class,
                            "reconciled": reconciled,
                            "reconciliation_id": reconciliation_id,
                        }
                    )

            system_movements.sort(
                key=lambda item: (
                    item.get("sort_date") or date.min,
                    item.get("sort_amount") or Decimal("0.00"),
                    str(item.get("description") or ""),
                    str(item.get("key") or ""),
                )
            )
            ofx_movements.sort(
                key=lambda item: (
                    item.get("sort_date") or date.min,
                    item.get("sort_amount") or Decimal("0.00"),
                    str(item.get("description") or ""),
                    str(item.get("id") or ""),
                )
            )

        revenue_accounts = (
            AccountPlanTemplateItem.objects.filter(
                status=StatusChoices.ACTIVE,
                is_analytic=True,
                account_type=AccountType.REVENUE,
            )
            .order_by("code")
        )

        context.update(
            {
                "page_title": self.page_title,
                "bank_accounts": accounts,
                "selected_account": selected_account,
                "start_date": start_date.isoformat() if start_date else "",
                "end_date": end_date.isoformat() if end_date else "",
                "system_status_filter": system_status,
                "ofx_status_filter": ofx_status,
                "min_amount": min_amount_raw,
                "max_amount": max_amount_raw,
                "statement_rows": rows,
                "system_movements": system_movements,
                "ofx_movements": ofx_movements,
                "system_count": len(system_movements),
                "ofx_count": len(ofx_movements),
                "revenue_accounts": revenue_accounts,
                "opening_balance_display": self._format_currency(opening_balance)
                if opening_balance is not None
                else "-",
                "opening_balance_class": "value-balance-positive"
                if opening_balance is not None and opening_balance >= 0
                else "value-balance-negative"
                if opening_balance is not None
                else "",
                "closing_balance_display": self._format_currency(closing_balance)
                if closing_balance is not None
                else "-",
                "closing_balance_class": "value-balance-positive"
                if closing_balance is not None and closing_balance >= 0
                else "value-balance-negative"
                if closing_balance is not None
                else "",
                "total_credits_display": self._format_currency(total_credits),
                "total_credits_class": "value-positive" if total_credits > 0 else "",
                "total_debits_display": self._format_currency(total_debits),
                "total_debits_class": "value-negative" if total_debits > 0 else "",
            }
        )
        return context


class BankSystemStatementView(BankStatementView):
    template_name = "restricted/bank_statement_system.html"
    page_title = "Extrato conta corrente"


class FinancialDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/financial_dashboard.html"
    page_title = "Dashboard financeiro"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _parse_date(value: str | None) -> date | None:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    @staticmethod
    def _format_decimal(value: Decimal) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value: Decimal) -> str:
        return f"R$ {self._format_decimal(value)}"

    @staticmethod
    def _format_date(value: date | None) -> str:
        return value.strftime("%d/%m/%Y") if value else "-"

    def _resolve_period(
        self,
        start_date: date | None,
        end_date: date | None,
    ) -> tuple[date, date]:
        today = timezone.localdate()
        if not start_date and not end_date:
            start_date = today.replace(day=1)
            end_date = today.replace(day=monthrange(today.year, today.month)[1])
        elif start_date and not end_date:
            end_date = today
        elif end_date and not start_date:
            start_date = end_date.replace(day=1)
        if end_date and start_date and end_date < start_date:
            start_date, end_date = end_date, start_date
        return start_date, end_date

    def _resolve_period_key(
        self,
        period_key: str | None,
        start_date: date | None,
        end_date: date | None,
    ) -> tuple[date, date, str]:
        today = timezone.localdate()
        key = (period_key or "").strip().lower()
        if key in ("month", "this_month"):
            month_end = monthrange(today.year, today.month)[1]
            return today.replace(day=1), today.replace(day=month_end), "month"
        if key == "last_30":
            return today - timedelta(days=29), today, "last_30"
        if key == "quarter":
            quarter_start = ((today.month - 1) // 3) * 3 + 1
            return date(today.year, quarter_start, 1), today, "quarter"
        if key == "year":
            return date(today.year, 1, 1), today, "year"
        if key == "custom":
            start_date, end_date = self._resolve_period(start_date, end_date)
            return start_date, end_date, "custom"
        if start_date or end_date:
            start_date, end_date = self._resolve_period(start_date, end_date)
            return start_date, end_date, "custom"
        month_end = monthrange(today.year, today.month)[1]
        return today.replace(day=1), today.replace(day=month_end), "month"

    @staticmethod
    def _format_month_label(value: date) -> str:
        labels = [
            "Jan",
            "Fev",
            "Mar",
            "Abr",
            "Mai",
            "Jun",
            "Jul",
            "Ago",
            "Set",
            "Out",
            "Nov",
            "Dez",
        ]
        return labels[value.month - 1]

    @staticmethod
    def _add_months(reference: date, offset: int) -> date:
        year = reference.year + (reference.month - 1 + offset) // 12
        month = (reference.month - 1 + offset) % 12 + 1
        day = min(reference.day, monthrange(year, month)[1])
        return date(year, month, day)

    def _build_month_sequence(self, end_date: date, count: int = 12) -> list[date]:
        end_month = end_date.replace(day=1)
        start_month = self._add_months(end_month, -(count - 1))
        return [self._add_months(start_month, offset) for offset in range(count)]

    def _build_forward_month_sequence(
        self,
        start_date: date,
        count: int = 12,
    ) -> list[date]:
        start_month = start_date.replace(day=1)
        return [self._add_months(start_month, offset) for offset in range(count)]

    def _build_series_points(
        self,
        values: list[Decimal],
    ) -> tuple[str, str, Decimal, Decimal]:
        if not values:
            return "", "", Decimal("0.00"), Decimal("0.00")
        width = 100
        height = 40
        min_value = min(values)
        max_value = max(values)
        if min_value == max_value:
            min_value -= Decimal("1.00")
            max_value += Decimal("1.00")
        span = max_value - min_value
        points = []
        count = len(values)
        for index, value in enumerate(values):
            x = 50 if count == 1 else (index / (count - 1)) * width
            ratio = (value - min_value) / span
            y = height - (float(ratio) * height)
            points.append((x, y))
        points_str = " ".join(f"{x:.2f},{y:.2f}" for x, y in points)
        area_path = f"M{points[0][0]:.2f},{points[0][1]:.2f}"
        for x, y in points[1:]:
            area_path += f" L{x:.2f},{y:.2f}"
        area_path += f" L{width:.2f},{height:.2f} L0,{height:.2f} Z"
        return points_str, area_path, min_value, max_value

    def _build_axis_labels(
        self,
        min_value: Decimal,
        max_value: Decimal,
        steps: int = 6,
    ) -> list[str]:
        if steps < 2:
            return []
        if min_value == max_value:
            min_value -= Decimal("1.00")
            max_value += Decimal("1.00")
        step = (max_value - min_value) / Decimal(steps - 1)
        labels = []
        for index in range(steps):
            value = max_value - (step * Decimal(index))
            labels.append(self._format_currency(value))
        return labels

    @staticmethod
    def _calculate_weighted_days(
        payments: Iterable[Any],
        base_attr: str,
    ) -> Decimal | None:
        total_amount = Decimal("0.00")
        weighted_days = Decimal("0.00")
        for payment in payments:
            amount = payment.amount or Decimal("0.00")
            if amount <= 0:
                continue
            base = getattr(payment, base_attr, None)
            base_date = getattr(base, "issue_date", None) if base else None
            if not base_date or not payment.payment_date:
                continue
            days = (payment.payment_date - base_date).days
            weighted_days += Decimal(days) * amount
            total_amount += amount
        if total_amount <= 0:
            return None
        return (weighted_days / total_amount).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

    @staticmethod
    def _resolve_trend(
        current: Decimal | None,
        previous: Decimal | None,
    ) -> dict[str, str]:
        if current is None or previous is None:
            return {"label": "Sem base", "chip": "chip-neutral", "direction": "flat"}
        difference = current - previous
        if difference.copy_abs() < Decimal("1.00"):
            return {"label": "Estavel", "chip": "chip-neutral", "direction": "flat"}
        if difference > 0:
            return {"label": "Em alta", "chip": "chip-warn", "direction": "up"}
        return {"label": "Em queda", "chip": "chip-ok", "direction": "down"}

    @staticmethod
    def _resolve_aging_bucket(due_date: date | None, today: date) -> str:
        if not due_date or due_date >= today:
            return "current"
        days = (today - due_date).days
        if days <= 30:
            return "1-30"
        if days <= 60:
            return "31-60"
        return "60+"

    @staticmethod
    def _is_overdue(status: str, due_date: date | None, today: date) -> bool:
        return status == FinancialStatus.OVERDUE or (
            status == FinancialStatus.OPEN and due_date and due_date < today
        )

    def _resolve_status(
        self,
        status: str,
        due_date: date | None,
        paid_total: Decimal,
        total_amount: Decimal,
        today: date,
    ) -> tuple[str, str]:
        if status == FinancialStatus.CANCELED:
            return "Cancelado", "chip-neutral"
        if total_amount > 0 and paid_total >= total_amount:
            return "Pago", "chip-ok"
        if paid_total > 0:
            return "Pago parcial", "chip-warn"
        if status == FinancialStatus.OVERDUE or (
            status == FinancialStatus.OPEN and due_date and due_date < today
        ):
            return "Atrasado", "chip-danger"
        if status == FinancialStatus.OPEN:
            return "Em aberto", "chip-info"
        if status == FinancialStatus.PAID:
            return "Pago", "chip-ok"
        return "-", "chip-neutral"

    def _build_line_points_range(
        self,
        values: list[Decimal],
        start_index: int,
        end_index: int,
        min_value: Decimal,
        max_value: Decimal,
    ) -> str:
        if not values or start_index > end_index:
            return ""
        width = 100
        height = 40
        if min_value == max_value:
            min_value -= Decimal("1.00")
            max_value += Decimal("1.00")
        span = max_value - min_value
        last_index = len(values) - 1
        points = []
        for index in range(start_index, end_index + 1):
            if index < 0 or index > last_index:
                continue
            x = 50 if last_index == 0 else (index / last_index) * width
            ratio = (values[index] - min_value) / span
            y = height - (float(ratio) * height)
            points.append(f"{x:.2f},{y:.2f}")
        return " ".join(points)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        params = self.request.GET
        consultant_id = params.get("consultant_id", "").strip()
        client_id = params.get("client_id", "").strip()
        supplier_id = params.get("supplier_id", "").strip()
        status_filter = params.get("status", "").strip()
        period_key = params.get("period", "").strip()
        company_id = params.get("company_id", "").strip()
        bank_account_id = params.get("bank_account_id", "").strip()
        period_start_raw = params.get("period_start", "").strip()
        period_end_raw = params.get("period_end", "").strip()

        period_start = self._parse_date(period_start_raw)
        period_end = self._parse_date(period_end_raw)
        period_start, period_end, period_key = self._resolve_period_key(
            period_key,
            period_start,
            period_end,
        )
        period_label = f"{self._format_date(period_start)} - {self._format_date(period_end)}"

        total_due_expr = ExpressionWrapper(
            F("amount") - F("discount") + F("interest") + F("penalty"),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )

        payables = AccountsPayable.objects.select_related(
            "supplier",
            "consultant",
            "billing_invoice",
            "billing_invoice__billing_client",
            "account_plan_item",
        ).annotate(
            paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00"))),
            total_due=total_due_expr,
        )
        receivables = AccountsReceivable.objects.select_related(
            "client",
            "client__company",
            "billing_invoice",
            "billing_invoice__billing_client",
            "account_plan_item",
        ).annotate(
            paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00"))),
            total_due=total_due_expr,
        )

        if consultant_id:
            payables = payables.filter(consultant_id=consultant_id)
        if supplier_id:
            payables = payables.filter(supplier_id=supplier_id)
        if client_id:
            receivables = receivables.filter(client_id=client_id)
        if company_id:
            receivables = receivables.filter(client__company_id=company_id)
            payables = payables.filter(
                Q(consultant__company_id=company_id)
                | Q(billing_invoice__billing_client__company_id=company_id)
            )

        if status_filter == "partial":
            payables = payables.filter(
                paid_total__gt=0,
                paid_total__lt=F("total_due"),
            )
            receivables = receivables.filter(
                paid_total__gt=0,
                paid_total__lt=F("total_due"),
            )
        elif status_filter:
            payables = payables.filter(status=status_filter)
            receivables = receivables.filter(status=status_filter)

        payables_period = payables
        receivables_period = receivables
        if period_start:
            payables_period = payables_period.filter(due_date__gte=period_start)
            receivables_period = receivables_period.filter(due_date__gte=period_start)
        if period_end:
            payables_period = payables_period.filter(due_date__lte=period_end)
            receivables_period = receivables_period.filter(due_date__lte=period_end)

        payables_period = list(payables_period.order_by("due_date", "id"))
        receivables_period = list(receivables_period.order_by("due_date", "id"))

        today = timezone.localdate()
        receivable_open_total = Decimal("0.00")
        payable_open_total = Decimal("0.00")
        receivable_overdue_total = Decimal("0.00")
        payable_overdue_total = Decimal("0.00")
        receivable_open_count = 0
        payable_open_count = 0
        receivable_overdue_count = 0
        payable_overdue_count = 0
        receivable_rows = []
        payable_rows = []

        for title in receivables_period:
            total_amount = title.total_amount()
            paid_total = title.paid_total or Decimal("0.00")
            open_amount = _resolve_open_amount(total_amount, paid_total)
            if title.status == FinancialStatus.CANCELED or open_amount <= 0:
                continue
            receivable_open_total += open_amount
            receivable_open_count += 1
            if self._is_overdue(title.status, title.due_date, today):
                receivable_overdue_total += open_amount
                receivable_overdue_count += 1
            status_label, status_chip = self._resolve_status(
                title.status,
                title.due_date,
                paid_total,
                total_amount,
                today,
            )
            receivable_rows.append(
                {
                    "partner": str(title.client),
                    "title": title.description or title.document_number or "-",
                    "due_date": self._format_date(title.due_date),
                    "original_amount_display": self._format_currency(total_amount),
                    "open_amount_display": self._format_currency(open_amount),
                    "status_label": status_label,
                    "status_chip": status_chip,
                }
            )

        for title in payables_period:
            total_amount = title.total_amount()
            paid_total = title.paid_total or Decimal("0.00")
            open_amount = _resolve_open_amount(total_amount, paid_total)
            if title.status == FinancialStatus.CANCELED or open_amount <= 0:
                continue
            payable_open_total += open_amount
            payable_open_count += 1
            if self._is_overdue(title.status, title.due_date, today):
                payable_overdue_total += open_amount
                payable_overdue_count += 1
            status_label, status_chip = self._resolve_status(
                title.status,
                title.due_date,
                paid_total,
                total_amount,
                today,
            )
            payable_rows.append(
                {
                    "partner": str(title.supplier),
                    "title": title.description or title.document_number or "-",
                    "due_date": self._format_date(title.due_date),
                    "original_amount_display": self._format_currency(total_amount),
                    "open_amount_display": self._format_currency(open_amount),
                    "status_label": status_label,
                    "status_chip": status_chip,
                }
            )

        receivable_rows = receivable_rows[:8]
        payable_rows = payable_rows[:8]

        net_balance = receivable_open_total - payable_open_total
        overdue_total = receivable_overdue_total + payable_overdue_total

        def receivable_payments_qs(start: date, end: date):
            qs = AccountsReceivablePayment.objects.filter(
                payment_date__gte=start,
                payment_date__lte=end,
            )
            if client_id:
                qs = qs.filter(receivable__client_id=client_id)
            if bank_account_id:
                qs = qs.filter(bank_account_id=bank_account_id)
            elif company_id:
                qs = qs.filter(bank_account__company_id=company_id)
            return qs

        def payable_payments_qs(start: date, end: date):
            qs = AccountsPayablePayment.objects.filter(
                payment_date__gte=start,
                payment_date__lte=end,
            )
            if supplier_id:
                qs = qs.filter(payable__supplier_id=supplier_id)
            if consultant_id:
                qs = qs.filter(payable__consultant_id=consultant_id)
            if bank_account_id:
                qs = qs.filter(bank_account_id=bank_account_id)
            elif company_id:
                qs = qs.filter(bank_account__company_id=company_id)
            return qs

        receivable_payments = list(
            receivable_payments_qs(period_start, period_end).select_related(
                "receivable"
            )
        )
        payable_payments = list(
            payable_payments_qs(period_start, period_end).select_related("payable")
        )

        received_total = sum(
            (payment.amount or Decimal("0.00")) for payment in receivable_payments
        )
        paid_total = sum(
            (payment.amount or Decimal("0.00")) for payment in payable_payments
        )

        dso_value = self._calculate_weighted_days(
            receivable_payments,
            "receivable",
        )
        dpo_value = self._calculate_weighted_days(
            payable_payments,
            "payable",
        )

        period_days = (period_end - period_start).days + 1
        prev_end = period_start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=max(period_days - 1, 0))
        dso_prev = self._calculate_weighted_days(
            receivable_payments_qs(prev_start, prev_end).select_related("receivable"),
            "receivable",
        )
        dpo_prev = self._calculate_weighted_days(
            payable_payments_qs(prev_start, prev_end).select_related("payable"),
            "payable",
        )

        dso_trend = self._resolve_trend(dso_value, dso_prev)
        dpo_trend = self._resolve_trend(dpo_value, dpo_prev)

        dso_display = f"{dso_value:.0f}" if dso_value is not None else "-"
        dpo_display = f"{dpo_value:.0f}" if dpo_value is not None else "-"

        month_sequence = self._build_forward_month_sequence(period_start)
        month_start = month_sequence[0]
        month_end = self._add_months(month_sequence[-1], 1) - timedelta(days=1)

        entries_map = defaultdict(lambda: Decimal("0.00"))
        exits_map = defaultdict(lambda: Decimal("0.00"))

        flow_receivables = receivables.filter(
            due_date__gte=month_start,
            due_date__lte=month_end,
        )
        flow_payables = payables.filter(
            due_date__gte=month_start,
            due_date__lte=month_end,
        )

        for title in flow_receivables:
            if title.status == FinancialStatus.CANCELED:
                continue
            open_amount = _resolve_open_amount(
                title.total_amount(),
                title.paid_total or Decimal("0.00"),
            )
            if open_amount <= 0:
                continue
            month_key = date(title.due_date.year, title.due_date.month, 1)
            entries_map[month_key] += open_amount

        for title in flow_payables:
            if title.status == FinancialStatus.CANCELED:
                continue
            open_amount = _resolve_open_amount(
                title.total_amount(),
                title.paid_total or Decimal("0.00"),
            )
            if open_amount <= 0:
                continue
            month_key = date(title.due_date.year, title.due_date.month, 1)
            exits_map[month_key] += open_amount

        max_month_value = max(
            [*entries_map.values(), *exits_map.values(), Decimal("0.00")]
        )
        if max_month_value <= 0:
            entries_map = defaultdict(lambda: Decimal("0.00"))
            exits_map = defaultdict(lambda: Decimal("0.00"))
            for payment in receivable_payments_qs(month_start, month_end):
                if not payment.payment_date:
                    continue
                amount = payment.amount or Decimal("0.00")
                if amount <= 0:
                    continue
                month_key = date(payment.payment_date.year, payment.payment_date.month, 1)
                entries_map[month_key] += amount
            for payment in payable_payments_qs(month_start, month_end):
                if not payment.payment_date:
                    continue
                amount = payment.amount or Decimal("0.00")
                if amount <= 0:
                    continue
                month_key = date(payment.payment_date.year, payment.payment_date.month, 1)
                exits_map[month_key] += amount
            max_month_value = max(
                [*entries_map.values(), *exits_map.values(), Decimal("0.00")]
            )
        monthly_flow_items = []
        for month in month_sequence:
            entry_total = entries_map.get(month, Decimal("0.00"))
            exit_total = exits_map.get(month, Decimal("0.00"))
            entry_percent = (
                (entry_total / max_month_value) * Decimal("100")
                if max_month_value > 0
                else Decimal("0.00")
            )
            exit_percent = (
                (exit_total / max_month_value) * Decimal("100")
                if max_month_value > 0
                else Decimal("0.00")
            )
            monthly_flow_items.append(
                {
                    "label": self._format_month_label(month),
                    "entry_percent": f"{entry_percent:.2f}",
                    "exit_percent": f"{exit_percent:.2f}",
                    "entry_display": self._format_currency(entry_total),
                    "exit_display": self._format_currency(exit_total),
                }
            )

        palette = ["#2b6f6c", "#e3b7a1", "#f2c26b", "#94a09f", "#517a6f"]

        def build_distribution(totals: dict[str, Decimal]) -> dict[str, Any]:
            total = sum(totals.values(), Decimal("0.00"))
            if total <= 0:
                return {
                    "available": False,
                    "total_display": self._format_currency(Decimal("0.00")),
                    "chart_style": "conic-gradient(var(--line) 0% 100%)",
                    "items": [],
                }
            ordered = sorted(totals.items(), key=lambda item: item[1], reverse=True)
            top = ordered[:4]
            others_total = sum(amount for _, amount in ordered[4:])
            items = []
            for index, (label, amount) in enumerate(top):
                percent = (amount / total) * Decimal("100")
                percent = percent.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                items.append(
                    {
                        "label": label,
                        "percent": f"{percent:.0f}",
                        "percent_value": percent,
                        "color": palette[index],
                    }
                )
            if others_total > 0:
                percent = (others_total / total) * Decimal("100")
                percent = percent.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                items.append(
                    {
                        "label": "Outros",
                        "percent": f"{percent:.0f}",
                        "percent_value": percent,
                        "color": palette[min(len(top), len(palette) - 1)],
                    }
                )

            segments = []
            current_percent = Decimal("0.00")
            last_index = len(items) - 1
            for index, item in enumerate(items):
                start = current_percent
                end = current_percent + item["percent_value"]
                if index == last_index:
                    end = Decimal("100.00")
                segments.append(f"{item['color']} {start:.2f}% {end:.2f}%")
                current_percent = end

            chart_style = f"conic-gradient({', '.join(segments)})"
            for item in items:
                item.pop("percent_value", None)

            return {
                "available": True,
                "total_display": self._format_currency(total),
                "chart_style": chart_style,
                "items": items,
            }

        def build_title_totals(
            titles: Iterable[AccountsReceivable | AccountsPayable],
        ) -> dict[str, dict[str, Decimal]]:
            totals = {
                "paid": defaultdict(lambda: Decimal("0.00")),
                "open": defaultdict(lambda: Decimal("0.00")),
                "all": defaultdict(lambda: Decimal("0.00")),
            }
            for title in titles:
                if title.status == FinancialStatus.CANCELED:
                    continue
                total_amount = title.total_amount()
                paid_total = getattr(title, "paid_total", None) or Decimal("0.00")
                open_amount = _resolve_open_amount(total_amount, paid_total)
                plan_item = getattr(title, "account_plan_item", None)
                label = (
                    f"{plan_item.code} - {plan_item.description}"
                    if plan_item
                    else "Sem conta"
                )
                if title.status == FinancialStatus.PAID:
                    if total_amount <= 0:
                        continue
                    totals["paid"][label] += total_amount
                    totals["all"][label] += total_amount
                else:
                    if open_amount <= 0:
                        continue
                    totals["open"][label] += open_amount
                    totals["all"][label] += open_amount
            return totals

        receivable_totals = build_title_totals(receivables_period)
        payable_totals = build_title_totals(payables_period)
        title_distribution_sections = [
            {
                "key": "paid",
                "label": "Titulos pagos",
                "receivable": build_distribution(receivable_totals["paid"]),
                "payable": build_distribution(payable_totals["paid"]),
            },
            {
                "key": "open",
                "label": "Titulos em aberto",
                "receivable": build_distribution(receivable_totals["open"]),
                "payable": build_distribution(payable_totals["open"]),
            },
            {
                "key": "all",
                "label": "Titulos pagos e abertos",
                "receivable": build_distribution(receivable_totals["all"]),
                "payable": build_distribution(payable_totals["all"]),
            },
        ]

        companies = Company.objects.filter(
            company_type__in=[CompanyType.PRIMARY, CompanyType.BRANCH]
        ).order_by("legal_name")
        bank_accounts = CompanyBankAccount.objects.select_related("company").order_by(
            "company__legal_name",
            "bank_name",
            "account_number",
        )
        if company_id:
            bank_accounts = bank_accounts.filter(company_id=company_id)

        period_options = [
            {"value": "month", "label": "Periodo: Este mes"},
            {"value": "last_30", "label": "Periodo: Ultimos 30 dias"},
            {"value": "quarter", "label": "Periodo: Este trimestre"},
            {"value": "year", "label": "Periodo: Este ano"},
            {"value": "custom", "label": "Periodo: Personalizado"},
        ]

        context.update(
            {
                "page_title": self.page_title,
                "period_label": period_label,
                "period_options": period_options,
                "companies": companies,
                "bank_accounts": bank_accounts,
                "current_filters": {
                    "consultant_id": consultant_id,
                    "client_id": client_id,
                    "supplier_id": supplier_id,
                    "status": status_filter,
                    "period": period_key,
                    "company_id": company_id,
                    "bank_account_id": bank_account_id,
                    "period_start": period_start.isoformat() if period_start else "",
                    "period_end": period_end.isoformat() if period_end else "",
                },
                "kpis": {
                    "net_balance": self._format_currency(net_balance),
                    "net_balance_class": "value-balance-positive"
                    if net_balance >= 0
                    else "value-balance-negative",
                    "receivable_open": self._format_currency(receivable_open_total),
                    "receivable_count": receivable_open_count,
                    "payable_open": self._format_currency(payable_open_total),
                    "payable_count": payable_open_count,
                    "overdue_total": self._format_currency(overdue_total),
                    "overdue_receivable": self._format_currency(receivable_overdue_total),
                    "overdue_payable": self._format_currency(payable_overdue_total),
                    "received_total": self._format_currency(received_total),
                    "received_count": len(receivable_payments),
                    "paid_total": self._format_currency(paid_total),
                    "paid_count": len(payable_payments),
                },
                "dso": {
                    "value": dso_display,
                    "trend_label": dso_trend["label"],
                    "trend_chip": dso_trend["chip"],
                    "trend_direction": dso_trend["direction"],
                },
                "dpo": {
                    "value": dpo_display,
                    "trend_label": dpo_trend["label"],
                    "trend_chip": dpo_trend["chip"],
                    "trend_direction": dpo_trend["direction"],
                },
                "monthly_flow": {
                    "available": max_month_value > 0,
                    "items": monthly_flow_items,
                },
                "title_distribution_sections": title_distribution_sections,
                "receivable_rows": receivable_rows,
                "payable_rows": payable_rows,
            }
        )
        return context


class SupplierListView(BaseListView):
    model = Supplier
    page_title = "Fornecedores"
    list_title = "Fornecedores cadastrados"
    search_placeholder = "Buscar por fornecedor ou documento"
    ordering = ("name",)
    table_headers = ("Fornecedor", "Tipo", "Documento", "Status")
    table_fields = ("name", "person_type", "document", "status")
    search_fields = ("name", "trade_name", "document", "email")
    create_url_name = "cadastros_web:supplier_create"
    edit_url_name = "cadastros_web:supplier_update"
    delete_url_name = "cadastros_web:supplier_delete"


class SupplierCreateView(BaseCreateView):
    model = Supplier
    form_class = SupplierForm
    page_title = "Novo fornecedor"
    submit_label = "Salvar fornecedor"
    cancel_url_name = "cadastros_web:supplier_list"
    success_url = reverse_lazy("cadastros_web:supplier_list")
    full_width_fields = ("address_line", "notes")

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["supplier_lookup_url"] = reverse("cadastros_web:supplier_lookup")
        return context


class SupplierUpdateView(BaseUpdateView):
    model = Supplier
    form_class = SupplierForm
    page_title = "Editar fornecedor"
    submit_label = "Salvar fornecedor"
    cancel_url_name = "cadastros_web:supplier_list"
    success_url = reverse_lazy("cadastros_web:supplier_list")
    full_width_fields = ("address_line", "notes")

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["supplier_lookup_url"] = reverse("cadastros_web:supplier_lookup")
        return context


class SupplierDeleteView(BaseDeleteView):
    model = Supplier
    cancel_url_name = "cadastros_web:supplier_list"
    success_url = reverse_lazy("cadastros_web:supplier_list")


class SupplierLookupView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        document = request.GET.get("document", "").strip()
        normalized = self._normalize_document(document)
        if not normalized:
            return JsonResponse(
                {"ok": False, "error": "Informe o CPF ou CNPJ."},
                status=400,
            )
        if len(normalized) == 11:
            return JsonResponse(
                {"ok": False, "error": "Consulta de CPF nao disponivel."},
                status=400,
            )
        if len(normalized) != 14:
            return JsonResponse(
                {"ok": False, "error": "Documento invalido."},
                status=400,
            )

        api_url = settings.RECEITA_FEDERAL_API_URL
        if not api_url:
            return JsonResponse(
                {"ok": False, "error": "API nao configurada."},
                status=503,
            )

        url = self._build_api_url(api_url, normalized)
        headers = {
            "Accept": "application/json",
            "User-Agent": settings.RECEITA_FEDERAL_USER_AGENT,
        }
        request_obj = Request(url, headers=headers)
        try:
            with urlopen(
                request_obj,
                timeout=settings.RECEITA_FEDERAL_REQUEST_TIMEOUT,
            ) as response:
                raw = response.read()
                charset = response.headers.get_content_charset() or "utf-8"
            try:
                payload = raw.decode(charset)
            except UnicodeDecodeError:
                payload = raw.decode("latin-1")
            data = json.loads(payload)
        except HTTPError as exc:
            if exc.code == 404:
                message = "CNPJ nao encontrado."
            else:
                message = f"Falha ao consultar Receita Federal ({exc.code})."
            return JsonResponse(
                {"ok": False, "error": message},
                status=502,
            )
        except (URLError, TimeoutError):
            return JsonResponse(
                {"ok": False, "error": "Nao foi possivel conectar a Receita Federal."},
                status=502,
            )
        except json.JSONDecodeError:
            return JsonResponse(
                {"ok": False, "error": "Resposta da Receita Federal invalida."},
                status=502,
            )

        if isinstance(data, dict):
            status = str(data.get("status") or "").upper()
            if status and status != "OK":
                message = data.get("message") or "CNPJ nao encontrado."
                return JsonResponse({"ok": False, "error": message}, status=404)

        payload = self._build_payload(data if isinstance(data, dict) else {}, normalized)
        return JsonResponse({"ok": True, "data": payload})

    @staticmethod
    def _normalize_document(value: str) -> str:
        return "".join(ch for ch in str(value) if ch.isdigit())

    @staticmethod
    def _build_api_url(base_url: str, cnpj: str) -> str:
        if "{cnpj}" in base_url:
            return base_url.format(cnpj=cnpj)
        return f"{base_url.rstrip('/')}/{cnpj}"

    @staticmethod
    def _pick_first(data: dict[str, Any], *keys: str) -> str:
        for key in keys:
            value = data.get(key)
            if value:
                return str(value).strip()
        return ""

    def _build_address(self, data: dict[str, Any]) -> str:
        street_type = self._pick_first(data, "descricao_tipo_de_logradouro")
        street_name = self._pick_first(data, "logradouro")
        street = " ".join(part for part in [street_type, street_name] if part).strip()
        number = self._pick_first(data, "numero")
        complement = self._pick_first(data, "complemento")
        neighborhood = self._pick_first(data, "bairro")
        parts = [part for part in [street, number] if part]
        address = ", ".join(parts)
        if complement:
            address = f"{address} {complement}".strip() if address else complement
        if neighborhood:
            address = f"{address} - {neighborhood}" if address else neighborhood
        return address.strip()

    def _format_postal_code(self, value: str) -> str:
        digits = self._normalize_document(value)
        if len(digits) == 8:
            return f"{digits[:5]}-{digits[5:]}"
        return value.strip() if value else ""

    def _build_payload(self, data: dict[str, Any], document: str) -> dict[str, Any]:
        name = self._pick_first(data, "razao_social", "nome", "name")
        trade_name = self._pick_first(data, "nome_fantasia", "fantasia", "trade_name")
        if not name:
            name = trade_name
        email = self._pick_first(data, "email")
        phone = self._pick_first(data, "telefone")
        if not phone:
            phone = " / ".join(
                part
                for part in [
                    self._pick_first(data, "ddd_telefone_1"),
                    self._pick_first(data, "ddd_telefone_2"),
                ]
                if part
            )
        city = self._pick_first(data, "municipio", "cidade", "city")
        state = self._pick_first(data, "uf", "estado", "state")
        postal_code = self._format_postal_code(self._pick_first(data, "cep", "postal_code"))
        return {
            "person_type": SupplierPersonType.PJ,
            "document": document,
            "name": name,
            "trade_name": trade_name,
            "email": email,
            "phone": phone,
            "address_line": self._build_address(data),
            "city": city,
            "state": state,
            "postal_code": postal_code,
            "country": "BR",
        }


class TravelReimbursementCreateView(BaseFormView, FormView):
    template_name = "restricted/form.html"
    form_class = TravelReimbursementForm
    page_title = "Reembolso de viagem"
    submit_label = "Criar titulos"
    cancel_url_name = "cadastros_web:dashboard"
    success_url = reverse_lazy("cadastros_web:dashboard")
    full_width_fields = ("description", "notes", "confirmation_file")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    @staticmethod
    def _resolve_account_plan_item(code: str) -> AccountPlanTemplateItem | None:
        return (
            AccountPlanTemplateItem.objects.filter(
                code=code,
                status=StatusChoices.ACTIVE,
                is_analytic=True,
            )
            .order_by("id")
            .first()
        )

    @staticmethod
    def _resolve_senior_client() -> Client | None:
        client = Client.objects.filter(
            name__iexact="Senior Sistemas",
            status=StatusChoices.ACTIVE,
        ).first()
        if client:
            return client
        return (
            Client.objects.filter(
                name__icontains="Senior Sistemas",
                status=StatusChoices.ACTIVE,
            )
            .order_by("id")
            .first()
        )

    @staticmethod
    def _build_description(value: str) -> str:
        text = (value or "").strip()
        if not text:
            return "Reembolso de viagem"
        if text.lower().startswith("reembolso"):
            return text
        return f"Reembolso de viagem - {text}"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        role = resolve_user_role(self.request.user)
        consultant_field = form.fields.get("consultant")
        if consultant_field and role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(
                user=self.request.user,
                status=StatusChoices.ACTIVE,
            ).first()
            consultant_field.queryset = Consultant.objects.filter(
                pk=getattr(consultant, "pk", None)
            )
            consultant_field.initial = consultant.pk if consultant else None
            consultant_field.required = False
            consultant_field.widget = forms.HiddenInput()
        return form

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        next_url = self.request.GET.get("next") or self.request.POST.get("next")
        if next_url:
            context["cancel_url"] = next_url
        return context

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        role = resolve_user_role(self.request.user)
        if role == UserRole.CONSULTANT:
            return reverse("cadastros_web:consultant_panel")
        if role == UserRole.ADMIN:
            return reverse("cadastros_web:accounts_payable_list")
        return reverse("cadastros_web:dashboard")

    def form_valid(self, form):
        role = resolve_user_role(self.request.user)
        consultant = form.cleaned_data.get("consultant")
        if role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(
                user=self.request.user,
                status=StatusChoices.ACTIVE,
            ).first()
            if not consultant:
                form.add_error(None, "Usuario sem consultor vinculado.")
                return self.form_invalid(form)

        if not consultant:
            form.add_error("consultant", "Selecione o consultor.")
            return self.form_invalid(form)
        if not consultant.supplier_id:
            form.add_error("consultant", "Consultor sem fornecedor vinculado.")
            return self.form_invalid(form)

        payable_account = self._resolve_account_plan_item("3.03.04")
        receivable_account = self._resolve_account_plan_item("1.04.01")
        senior_client = self._resolve_senior_client()

        has_error = False
        if not payable_account:
            form.add_error(None, "Conta 3.03.04 nao encontrada no plano de contas.")
            has_error = True
        if not receivable_account:
            form.add_error(None, "Conta 1.04.01 nao encontrada no plano de contas.")
            has_error = True
        if not senior_client:
            form.add_error(
                None,
                "Cliente Senior Sistemas nao encontrado. Cadastre o cliente com esse nome.",
            )
            has_error = True
        if has_error:
            return self.form_invalid(form)

        document_number = (form.cleaned_data.get("document_number") or "").strip()
        description = self._build_description(form.cleaned_data.get("description"))
        issue_date = form.cleaned_data.get("issue_date")
        due_date = form.cleaned_data.get("due_date")
        amount = form.cleaned_data.get("amount")
        notes = (form.cleaned_data.get("notes") or "").strip()
        confirmation_file = form.cleaned_data.get("confirmation_file")

        if AccountsPayable.objects.filter(
            supplier=consultant.supplier,
            document_number=document_number,
        ).exists():
            form.add_error(
                "document_number",
                "Ja existe um titulo de contas a pagar com esse documento para o fornecedor.",
            )
        if AccountsReceivable.objects.filter(
            client=senior_client,
            document_number=document_number,
        ).exists():
            form.add_error(
                "document_number",
                "Ja existe um titulo de contas a receber com esse documento para a Senior Sistemas.",
            )
        if form.errors:
            return self.form_invalid(form)

        with transaction.atomic():
            payable = AccountsPayable.objects.create(
                supplier=consultant.supplier,
                consultant=consultant,
                account_plan_item=payable_account,
                document_number=document_number,
                description=description,
                issue_date=issue_date,
                due_date=due_date,
                amount=amount,
                discount=Decimal("0.00"),
                interest=Decimal("0.00"),
                penalty=Decimal("0.00"),
                status=FinancialStatus.OPEN,
                settlement_date=None,
                payment_method="",
                notes=notes,
            )
            receivable_notes = notes
            consultant_label = f"Consultor: {consultant.full_name}"
            receivable_notes = (
                f"{receivable_notes}\n{consultant_label}".strip()
                if receivable_notes
                else consultant_label
            )
            receivable = AccountsReceivable.objects.create(
                client=senior_client,
                account_plan_item=receivable_account,
                document_number=document_number,
                description=description,
                issue_date=issue_date,
                due_date=due_date,
                amount=amount,
                discount=Decimal("0.00"),
                interest=Decimal("0.00"),
                penalty=Decimal("0.00"),
                status=FinancialStatus.OPEN,
                settlement_date=None,
                payment_method="",
                notes=receivable_notes,
            )
            AccountsPayableAttachment.objects.create(
                payable=payable,
                description="Confirmacao Senior",
                file=confirmation_file,
            )

        def _notify() -> None:
            notify_admin_payable_created(payable)
            notify_consultant_payable_created(payable)
            notify_admin_receivable_created(receivable)

        transaction.on_commit(_notify)

        messages.success(self.request, "Reembolso de viagem registrado com sucesso.")
        return HttpResponseRedirect(self.get_success_url())


class AccountsCompensationView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/accounts_compensation.html"
    page_title = "Compensacao de titulos"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def _holders_match(self, client: Client, supplier: Supplier) -> bool:
        client_doc = _normalize_doc(
            client.company.tax_id if client.company_id else ""
        )
        supplier_doc = _normalize_doc(supplier.document)
        if client_doc and supplier_doc and client_doc == supplier_doc:
            return True
        client_name = _normalize_name(client.name)
        supplier_name = _normalize_name(supplier.name)
        supplier_trade = _normalize_name(supplier.trade_name)
        if client_name and (client_name == supplier_name or client_name == supplier_trade):
            return True
        company_name = _normalize_name(
            client.company.legal_name if client.company_id else ""
        )
        company_trade = _normalize_name(
            client.company.trade_name if client.company_id else ""
        )
        if company_name and (company_name == supplier_name or company_name == supplier_trade):
            return True
        if company_trade and (company_trade == supplier_name or company_trade == supplier_trade):
            return True
        return False

    def _parse_id(self, value: str | None) -> int | None:
        if not value:
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _open_receivables(self, client_id: int | None):
        qs = (
            AccountsReceivable.objects.select_related("client", "client__company")
            .annotate(paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00"))))
            .order_by("due_date", "document_number")
        )
        if client_id:
            qs = qs.filter(client_id=client_id)
        rows = []
        open_map = {}
        total_open = Decimal("0.00")
        for title in qs:
            if title.status == FinancialStatus.CANCELED:
                continue
            open_amount = _resolve_open_amount(title.total_amount(), title.paid_total)
            if open_amount <= 0:
                continue
            open_map[title.id] = open_amount
            total_open += open_amount
            rows.append(
                {
                    "id": title.id,
                    "document": title.document_number,
                    "due_date": title.due_date,
                    "client": title.client,
                    "description": title.description,
                    "open_amount": open_amount,
                    "open_amount_display": self._format_currency(open_amount),
                    "total_amount_display": self._format_currency(title.total_amount()),
                }
            )
        return rows, open_map, total_open

    def _open_payables(self, supplier_id: int | None):
        qs = (
            AccountsPayable.objects.select_related("supplier")
            .annotate(paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00"))))
            .order_by("due_date", "document_number")
        )
        if supplier_id:
            qs = qs.filter(supplier_id=supplier_id)
        rows = []
        open_map = {}
        total_open = Decimal("0.00")
        for title in qs:
            if title.status == FinancialStatus.CANCELED:
                continue
            open_amount = _resolve_open_amount(title.total_amount(), title.paid_total)
            if open_amount <= 0:
                continue
            open_map[title.id] = open_amount
            total_open += open_amount
            rows.append(
                {
                    "id": title.id,
                    "document": title.document_number,
                    "due_date": title.due_date,
                    "supplier": title.supplier,
                    "description": title.description,
                    "open_amount": open_amount,
                    "open_amount_display": self._format_currency(open_amount),
                    "total_amount_display": self._format_currency(title.total_amount()),
                }
            )
        return rows, open_map, total_open

    @staticmethod
    def _format_currency(value: Decimal) -> str:
        return f"R$ {formats.number_format(value, decimal_pos=2, use_l10n=True, force_grouping=True)}"

    @staticmethod
    def _format_date(value: date | None) -> str:
        return value.strftime("%d/%m/%Y") if value else "-"

    def get(self, request, *args, **kwargs):
        client_id = self._parse_id(request.GET.get("client_id"))
        supplier_id = self._parse_id(request.GET.get("supplier_id"))
        client = Client.objects.select_related("company").filter(pk=client_id).first() if client_id else None
        supplier = Supplier.objects.filter(pk=supplier_id).first() if supplier_id else None
        form = AccountsCompensationForm(
            initial={"payment_date": timezone.localdate()}
        )
        receivable_rows, receivable_map, receivable_total = self._open_receivables(client_id)
        payable_rows, payable_map, payable_total = self._open_payables(supplier_id)
        match_error = False
        if client and supplier and not self._holders_match(client, supplier):
            match_error = True
        context = {
            "page_title": self.page_title,
            "clients": Client.objects.select_related("company").order_by("name"),
            "suppliers": Supplier.objects.order_by("name"),
            "selected_client": client,
            "selected_supplier": supplier,
            "receivable_rows": receivable_rows,
            "payable_rows": payable_rows,
            "receivable_total_display": self._format_currency(receivable_total),
            "payable_total_display": self._format_currency(payable_total),
            "form": form,
            "match_error": match_error,
            "selected_receivable_id": None,
            "selected_payable_id": None,
        }
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        client_id = self._parse_id(request.POST.get("client_id"))
        supplier_id = self._parse_id(request.POST.get("supplier_id"))
        client = Client.objects.select_related("company").filter(pk=client_id).first() if client_id else None
        supplier = Supplier.objects.filter(pk=supplier_id).first() if supplier_id else None
        form = AccountsCompensationForm(request.POST)
        receivable_rows, receivable_map, receivable_total = self._open_receivables(client_id)
        payable_rows, payable_map, payable_total = self._open_payables(supplier_id)
        selected_receivable_id = self._parse_id(request.POST.get("receivable_id"))
        selected_payable_id = self._parse_id(request.POST.get("payable_id"))

        match_error = False
        if not client or not supplier:
            form.add_error(None, "Selecione cliente e fornecedor.")
        elif not self._holders_match(client, supplier):
            form.add_error(None, "Cliente e fornecedor precisam ser o mesmo titular.")
            match_error = True

        if not selected_receivable_id or selected_receivable_id not in receivable_map:
            form.add_error(None, "Selecione um titulo a receber em aberto.")
        if not selected_payable_id or selected_payable_id not in payable_map:
            form.add_error(None, "Selecione um titulo a pagar em aberto.")

        if form.is_valid() and selected_receivable_id in receivable_map and selected_payable_id in payable_map:
            amount = form.cleaned_data.get("amount")
            receivable_open = receivable_map.get(selected_receivable_id, Decimal("0.00"))
            payable_open = payable_map.get(selected_payable_id, Decimal("0.00"))
            if amount and amount > receivable_open:
                form.add_error("amount", "Valor nao pode ser maior que o saldo a receber.")
            if amount and amount > payable_open:
                form.add_error("amount", "Valor nao pode ser maior que o saldo a pagar.")

        if not form.is_valid():
            context = {
                "page_title": self.page_title,
                "clients": Client.objects.select_related("company").order_by("name"),
                "suppliers": Supplier.objects.order_by("name"),
                "selected_client": client,
                "selected_supplier": supplier,
                "receivable_rows": receivable_rows,
                "payable_rows": payable_rows,
                "receivable_total_display": self._format_currency(receivable_total),
                "payable_total_display": self._format_currency(payable_total),
                "form": form,
                "match_error": match_error,
                "selected_receivable_id": selected_receivable_id,
                "selected_payable_id": selected_payable_id,
            }
            return self.render_to_response(context)

        receivable = AccountsReceivable.objects.select_related("client").filter(
            pk=selected_receivable_id
        ).first()
        payable = AccountsPayable.objects.select_related("supplier").filter(
            pk=selected_payable_id
        ).first()
        if not receivable or not payable:
            form.add_error(None, "Titulos selecionados nao encontrados.")
            context = {
                "page_title": self.page_title,
                "clients": Client.objects.select_related("company").order_by("name"),
                "suppliers": Supplier.objects.order_by("name"),
                "selected_client": client,
                "selected_supplier": supplier,
                "receivable_rows": receivable_rows,
                "payable_rows": payable_rows,
                "receivable_total_display": self._format_currency(receivable_total),
                "payable_total_display": self._format_currency(payable_total),
                "form": form,
                "match_error": match_error,
                "selected_receivable_id": selected_receivable_id,
                "selected_payable_id": selected_payable_id,
            }
            return self.render_to_response(context)

        payment_date = form.cleaned_data["payment_date"]
        bank_account = form.cleaned_data["bank_account"]
        notes = _build_compensation_notes(form.cleaned_data.get("notes"))
        amount = form.cleaned_data["amount"]

        was_receivable_paid = receivable.status == FinancialStatus.PAID
        was_payable_paid = payable.status == FinancialStatus.PAID

        if receivable.status == FinancialStatus.CANCELED or payable.status == FinancialStatus.CANCELED:
            form.add_error(None, "Titulos cancelados nao podem ser compensados.")
            context = {
                "page_title": self.page_title,
                "clients": Client.objects.select_related("company").order_by("name"),
                "suppliers": Supplier.objects.order_by("name"),
                "selected_client": client,
                "selected_supplier": supplier,
                "receivable_rows": receivable_rows,
                "payable_rows": payable_rows,
                "receivable_total_display": self._format_currency(receivable_total),
                "payable_total_display": self._format_currency(payable_total),
                "form": form,
                "match_error": match_error,
                "selected_receivable_id": selected_receivable_id,
                "selected_payable_id": selected_payable_id,
            }
            return self.render_to_response(context)

        with transaction.atomic():
            receivable_payment = AccountsReceivablePayment.objects.create(
                receivable=receivable,
                bank_account=bank_account,
                payment_date=payment_date,
                amount=amount,
                payment_method=PaymentMethod.OTHER,
                notes=notes,
            )
            payable_payment = AccountsPayablePayment.objects.create(
                payable=payable,
                bank_account=bank_account,
                payment_date=payment_date,
                amount=amount,
                payment_method=PaymentMethod.OTHER,
                notes=notes,
            )

            receivable_total_paid = (
                receivable.payments.aggregate(total=Sum("amount")).get("total")
                or Decimal("0.00")
            )
            receivable_total_due = receivable.total_amount()
            if receivable_total_paid >= receivable_total_due:
                receivable.settlement_date = receivable_payment.payment_date
                receivable.status = FinancialStatus.PAID
                receivable.payment_method = receivable_payment.payment_method or ""
            else:
                receivable.settlement_date = None
                receivable.status = FinancialStatus.OPEN
            receivable.save()

            payable_total_paid = (
                payable.payments.aggregate(total=Sum("amount")).get("total")
                or Decimal("0.00")
            )
            payable_total_due = payable.total_amount()
            if payable_total_paid >= payable_total_due:
                payable.settlement_date = payable_payment.payment_date
                payable.status = FinancialStatus.PAID
                payable.payment_method = payable_payment.payment_method or ""
            else:
                payable.settlement_date = None
                payable.status = FinancialStatus.OPEN
            payable.save()

        messages.success(request, "Compensacao registrada com sucesso.")
        if not was_receivable_paid and receivable.status == FinancialStatus.PAID:
            transaction.on_commit(lambda: notify_admin_receivable_paid(receivable))
        if not was_payable_paid and payable.status == FinancialStatus.PAID:
            transaction.on_commit(lambda: notify_admin_payable_paid(payable))
            transaction.on_commit(lambda: notify_consultant_payable_paid(payable))

        redirect_url = reverse("cadastros_web:accounts_compensation")
        if client_id:
            redirect_url = f"{redirect_url}?client_id={client_id}&supplier_id={supplier_id or ''}"
        return HttpResponseRedirect(redirect_url)


class AccountsPayableListView(BaseListView):
    model = AccountsPayable
    queryset = AccountsPayable.objects.select_related(
        "supplier",
        "consultant",
        "billing_invoice",
    )
    page_title = "Contas a pagar"
    list_title = "Contas a pagar cadastradas"
    search_placeholder = "Buscar por fornecedor ou documento"
    date_filter_field = "due_date"
    date_filter_label = "Vencimento"
    ordering = ("due_date",)
    table_headers = (
        "Consultor",
        "Fatura",
        "Documento",
        "Vencimento",
        "Valor original",
        "Valor em aberto",
        "Status",
    )
    table_fields = (
        "consultant",
        "billing_invoice.number",
        "document_number",
        "due_date",
        "amount",
        "amount",
        "status",
    )
    search_fields = (
        "document_number",
        "description",
        "supplier__name",
        "supplier__trade_name",
        "supplier__document",
        "consultant__full_name",
        "billing_invoice__number",
    )
    create_url_name = "cadastros_web:accounts_payable_create"
    edit_url_name = "cadastros_web:accounts_payable_update"
    delete_url_name = "cadastros_web:accounts_payable_delete"
    allowed_roles = (UserRole.ADMIN,)

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.annotate(
            paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00")))
        )

    def _get_chart_queryset(self):
        if hasattr(self, "_chart_queryset"):
            return self._chart_queryset
        base_qs = self._apply_filters(AccountsPayable.objects.all())
        self._chart_queryset = base_qs.values(
            "id",
            "status",
            "due_date",
            "amount",
            "discount",
            "interest",
            "penalty",
        ).annotate(
            paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00")))
        ).order_by()
        return self._chart_queryset

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["extra_actions"] = [
            {
                "label": "Compensar titulos",
                "url": reverse("cadastros_web:accounts_compensation"),
            }
        ]
        amount_indices = [
            idx for idx, field in enumerate(self.table_fields) if field == "amount"
        ]
        if len(amount_indices) < 2:
            return context
        original_index, open_index = amount_indices[:2]
        for obj, row in zip(context["object_list"], context["table_rows"]):
            total_amount = obj.total_amount()
            open_amount = _resolve_open_amount(
                total_amount, getattr(obj, "paid_total", None)
            )
            row["values"][original_index] = formats.number_format(
                total_amount,
                decimal_pos=2,
                use_l10n=True,
                force_grouping=True,
            )
            row["values"][open_index] = formats.number_format(
                open_amount,
                decimal_pos=2,
                use_l10n=True,
                force_grouping=True,
            )
            if obj.supplier_id:
                row["lead"] = obj.supplier.name or obj.supplier.trade_name
            elif obj.consultant_id:
                row["lead"] = str(obj.consultant)
        context["list_key"] = "accounts-payable"
        context["due_charts"] = _build_due_charts(
            self._get_chart_queryset().iterator(),
            variant="payable",
            subtitle="Titulos a pagar em aberto",
        )
        return context

    def get_row_actions(self, obj: AccountsPayable) -> list[dict[str, str]]:
        next_url = self.request.get_full_path()
        attachment_url = reverse(
            "cadastros_web:accounts_payable_attachment_create",
            args=[obj.pk],
        )
        if next_url:
            attachment_url = f"{attachment_url}?{urlencode({'next': next_url})}"
        actions = [
            {
                "label": "Baixar",
                "url": reverse(
                    "cadastros_web:accounts_payable_payment_create",
                    args=[obj.pk],
                ),
            },
            {
                "label": "Pagamentos",
                "url": reverse(
                    "cadastros_web:accounts_payable_payment_list",
                    args=[obj.pk],
                ),
            },
            {
                "label": "Documento",
                "url": attachment_url,
            },
        ]
        if obj.billing_invoice_id:
            actions.append(
                {
                    "label": "Apontamentos",
                    "url": reverse(
                        "cadastros_web:accounts_payable_entries", args=[obj.pk]
                    ),
                }
            )
        return actions


class AccountsPayableMissingAttachmentReportView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/accounts_payable_missing_attachments.html"
    page_title = "Titulos sem documento"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _format_currency(value: Decimal) -> str:
        return f"R$ {formats.number_format(value, decimal_pos=2, use_l10n=True, force_grouping=True)}"

    @staticmethod
    def _format_date(value: date | None) -> str:
        return value.strftime("%d/%m/%Y") if value else "-"

    def _resolve_status(
        self,
        status: str,
        due_date: date | None,
        paid_total: Decimal,
        total_amount: Decimal,
    ) -> tuple[str, str]:
        today = timezone.localdate()
        if status == FinancialStatus.CANCELED:
            return "Cancelado", "chip-neutral"
        if total_amount > 0 and paid_total >= total_amount:
            return "Pago", "chip-ok"
        if paid_total > 0:
            return "Pago parcial", "chip-warn"
        if status == FinancialStatus.OVERDUE or (
            status == FinancialStatus.OPEN and due_date and due_date < today
        ):
            return "Atrasado", "chip-danger"
        if status == FinancialStatus.OPEN:
            return "Em aberto", "chip-info"
        if status == FinancialStatus.PAID:
            return "Pago", "chip-ok"
        return "-", "chip-neutral"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        titles = (
            AccountsPayable.objects.select_related(
                "supplier",
                "consultant",
                "billing_invoice",
            )
            .annotate(
                paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00"))),
            )
            .filter(
                consultant__isnull=False,
                billing_invoice__isnull=False,
            )
            .filter(attachments__isnull=True)
            .order_by("due_date", "id")
        )

        rows = []
        next_url = self.request.get_full_path()
        for title in titles:
            total_amount = title.total_amount()
            paid_total = title.paid_total or Decimal("0.00")
            open_amount = _resolve_open_amount(total_amount, paid_total)
            status_label, status_chip = self._resolve_status(
                title.status,
                title.due_date,
                paid_total,
                total_amount,
            )
            attachment_url = reverse(
                "cadastros_web:accounts_payable_attachment_create",
                args=[title.pk],
            )
            if next_url:
                attachment_url = f"{attachment_url}?{urlencode({'next': next_url})}"
            rows.append(
                {
                    "consultant": str(title.consultant),
                    "supplier": str(title.supplier),
                    "invoice_number": title.billing_invoice.number
                    if title.billing_invoice_id
                    else "-",
                    "document_number": title.document_number,
                    "due_date": self._format_date(title.due_date),
                    "original_amount": self._format_currency(total_amount),
                    "open_amount": self._format_currency(open_amount),
                    "status_label": status_label,
                    "status_chip": status_chip,
                    "attachment_url": attachment_url,
                }
            )

        context.update(
            {
                "page_title": self.page_title,
                "rows": rows,
                "title_count": len(rows),
            }
        )
        return context


class AccountsPayableCreateView(BaseCreateView):
    model = AccountsPayable
    form_class = AccountsPayableForm
    page_title = "Nova conta a pagar"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:accounts_payable_list"
    success_url = reverse_lazy("cadastros_web:accounts_payable_list")
    full_width_fields = ("description", "notes")
    allowed_roles = (UserRole.ADMIN,)

    def _build_document_number(
        self,
        base_number: str,
        index: int,
        used_numbers: set[str],
        supplier_id: int,
    ) -> str:
        if index == 1:
            used_numbers.add(base_number)
            return base_number
        suffix_index = index
        max_length = AccountsPayable._meta.get_field("document_number").max_length
        while True:
            suffix = f"-{suffix_index}"
            trimmed_base = base_number
            if len(base_number) + len(suffix) > max_length:
                trimmed_base = base_number[: max_length - len(suffix)]
            candidate = f"{trimmed_base}{suffix}"
            if (
                candidate not in used_numbers
                and not AccountsPayable.objects.filter(
                    supplier_id=supplier_id,
                    document_number=candidate,
                ).exists()
            ):
                used_numbers.add(candidate)
                return candidate
            suffix_index += 1

    def form_valid(self, form):
        interval = form.cleaned_data.get("recurrence_interval_days")
        count = form.cleaned_data.get("recurrence_count") or 1
        if not interval or count <= 1:
            response = super().form_valid(form)
            title = self.object
            if title:
                def _notify_single() -> None:
                    notify_admin_payable_created(title)
                    notify_consultant_payable_created(title)

                transaction.on_commit(_notify_single)
            return response

        base_obj = form.save(commit=False)
        first_due_date = base_obj.due_date
        created = []
        used_numbers: set[str] = set()

        with transaction.atomic():
            for index in range(1, count + 1):
                due_date = first_due_date + timedelta(days=interval * (index - 1))
                document_number = self._build_document_number(
                    base_obj.document_number,
                    index,
                    used_numbers,
                    base_obj.supplier_id,
                )
                obj = AccountsPayable(
                    supplier=base_obj.supplier,
                    consultant=base_obj.consultant,
                    billing_invoice=base_obj.billing_invoice,
                    account_plan_item=base_obj.account_plan_item,
                    document_number=document_number,
                    description=base_obj.description,
                    issue_date=base_obj.issue_date,
                    due_date=due_date,
                    amount=base_obj.amount,
                    discount=base_obj.discount,
                    interest=base_obj.interest,
                    penalty=base_obj.penalty,
                    status=base_obj.status,
                    settlement_date=base_obj.settlement_date,
                    payment_method=base_obj.payment_method,
                    notes=base_obj.notes,
                )
                obj.save()
                created.append(obj)

        self.object = created[0] if created else None
        messages.success(
            self.request,
            f"{len(created)} titulos criados com sucesso.",
        )
        if created:
            def _notify_created() -> None:
                for title in created:
                    notify_admin_payable_created(title)
                    notify_consultant_payable_created(title)

            transaction.on_commit(_notify_created)
        return HttpResponseRedirect(self.get_success_url())


class AccountsPayableUpdateView(BaseUpdateView):
    model = AccountsPayable
    form_class = AccountsPayableForm
    page_title = "Editar conta a pagar"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:accounts_payable_list"
    success_url = reverse_lazy("cadastros_web:accounts_payable_list")
    full_width_fields = ("description", "notes")
    allowed_roles = (UserRole.ADMIN,)

    def form_valid(self, form):
        before = AccountsPayable.objects.get(pk=self.object.pk)
        response = super().form_valid(form)
        if before.status != FinancialStatus.PAID and self.object.status == FinancialStatus.PAID:
            def _notify_paid() -> None:
                notify_admin_payable_paid(self.object)
                notify_consultant_payable_paid(self.object)

            transaction.on_commit(_notify_paid)
        return response


class AccountsPayableDeleteView(BaseDeleteView):
    model = AccountsPayable
    cancel_url_name = "cadastros_web:accounts_payable_list"
    success_url = reverse_lazy("cadastros_web:accounts_payable_list")
    allowed_roles = (UserRole.ADMIN,)


class AccountsPayableEntriesView(BaseListView):
    model = TimeEntry
    page_title = "Apontamentos do titulo"
    list_title = "Apontamentos vinculados"
    search_placeholder = "Buscar por projeto ou atividade"
    ordering = ("start_date", "created_at")
    table_headers = ("Consultor", "Projeto", "Atividade", "Inicio", "Fim", "Horas", "Status")
    table_fields = (
        "consultant",
        "activity.project",
        "activity",
        "start_date",
        "end_date",
        "total_hours",
        "status",
    )
    search_fields = (
        "consultant__full_name",
        "activity__activity",
        "activity__project__description",
        "description",
    )
    create_url_name = "cadastros_web:accounts_payable_list"
    edit_url_name = "cadastros_web:time_entry_update"
    delete_url_name = "cadastros_web:time_entry_update"
    allowed_roles = (UserRole.ADMIN,)
    show_actions = False
    show_create = False

    def get_queryset(self):
        title = get_object_or_404(AccountsPayable, pk=self.kwargs["pk"])
        self.title = title
        if not title.billing_invoice_id or not title.consultant_id:
            return TimeEntry.objects.none()
        return (
            TimeEntry.objects.select_related("consultant", "activity", "activity__project")
            .filter(
                billing_invoice=title.billing_invoice,
                consultant=title.consultant,
            )
            .order_by("start_date", "created_at")
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        title = getattr(self, "title", None)
        invoice_number = title.billing_invoice.number if title and title.billing_invoice_id else "-"
        context["page_title"] = f"Apontamentos do titulo {invoice_number}"
        context["list_title"] = f"Consultor: {title.consultant}" if title else "Apontamentos"
        context["extra_actions"] = [
            {
                "label": "Voltar",
                "url": reverse("cadastros_web:accounts_payable_list"),
            }
        ]
        return context


class AccountsPayableAttachmentCreateView(BaseCreateView):
    model = AccountsPayableAttachment
    form_class = AccountsPayableAttachmentForm
    template_name = "restricted/accounts_payable_attachment.html"
    page_title = "Documento do titulo"
    submit_label = "Enviar documento"
    cancel_url_name = "cadastros_web:accounts_payable_list"
    success_message = "Documento anexado com sucesso."
    full_width_fields = ("description", "file")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def _get_payable(self) -> AccountsPayable:
        if not hasattr(self, "payable"):
            self.payable = get_object_or_404(AccountsPayable, pk=self.kwargs["pk"])
        return self.payable

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
            payable = self._get_payable()
            if role == UserRole.CONSULTANT:
                if not payable.consultant_id or payable.consultant.user_id != request.user.id:
                    raise PermissionDenied("Titulo sem acesso para este consultor.")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["payable"] = self._get_payable()
        return kwargs

    def form_valid(self, form):
        form.instance.payable = self._get_payable()
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse("cadastros_web:accounts_payable_attachment_create", args=[self._get_payable().pk])

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        payable = self._get_payable()
        attachments = payable.attachments.order_by("-created_at")
        next_url = self.request.GET.get("next", "").strip()
        if not next_url:
            next_url = self.request.POST.get("next", "").strip()
        context.update(
            {
                "page_title": f"Documento do titulo {payable.document_number}",
                "payable": payable,
                "attachments": attachments,
                "attachment_count": attachments.count(),
                "next_url": next_url,
                "cancel_url": next_url or context.get("cancel_url"),
            }
        )
        return context


class AccountsPayablePaymentListView(BaseListView):
    model = AccountsPayablePayment
    page_title = "Pagamentos"
    list_title = "Pagamentos do titulo"
    search_placeholder = "Buscar por conta bancaria"
    ordering = ("-payment_date", "-created_at")
    table_headers = ("Data", "Conta bancaria", "Valor", "Forma", "Observacoes")
    table_fields = (
        "payment_date",
        "bank_account",
        "amount",
        "payment_method",
        "notes",
    )
    search_fields = (
        "bank_account__bank_name",
        "bank_account__company__legal_name",
        "bank_account__company__trade_name",
        "notes",
    )
    create_url_name = "cadastros_web:accounts_payable_list"
    edit_url_name = ""
    delete_url_name = ""
    allowed_roles = (UserRole.ADMIN,)
    status_field = None
    show_actions = True
    show_create = False

    def get_queryset(self):
        title = get_object_or_404(AccountsPayable, pk=self.kwargs["pk"])
        self.title = title
        return (
            AccountsPayablePayment.objects.select_related(
                "bank_account",
                "bank_account__company",
            )
            .filter(payable=title)
            .order_by("-payment_date", "-created_at")
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        title = getattr(self, "title", None)
        document = title.document_number if title else "-"
        context["page_title"] = "Pagamentos do titulo"
        context["list_title"] = f"Titulo {document}"
        context["extra_actions"] = [
            {
                "label": "Nova baixa",
                "url": reverse(
                    "cadastros_web:accounts_payable_payment_create",
                    args=[title.pk],
                )
                if title
                else reverse("cadastros_web:accounts_payable_list"),
            },
            {
                "label": "Voltar",
                "url": reverse("cadastros_web:accounts_payable_list"),
            },
        ]
        return context

    def get_row_actions(self, obj: AccountsPayablePayment) -> list[dict[str, str]]:
        next_url = self.request.get_full_path()
        reverse_url = reverse(
            "cadastros_web:accounts_payable_payment_reverse",
            args=[obj.pk],
        )
        if next_url:
            reverse_url = f"{reverse_url}?{urlencode({'next': next_url})}"
        return [{"label": "Estornar", "url": reverse_url}]


def _refresh_payable_after_payment_change(payable: AccountsPayable) -> None:
    if payable.status == FinancialStatus.CANCELED:
        payable.settlement_date = None
        payable.save()
        return
    total_paid = (
        payable.payments.aggregate(total=Sum("amount")).get("total")
        or Decimal("0.00")
    )
    total_due = payable.total_amount()
    last_payment = payable.payments.order_by("-payment_date", "-created_at").first()
    if total_paid >= total_due and total_due > 0:
        payable.settlement_date = last_payment.payment_date if last_payment else None
        payable.status = FinancialStatus.PAID
        if last_payment and last_payment.payment_method:
            payable.payment_method = last_payment.payment_method
    else:
        payable.settlement_date = None
        payable.status = FinancialStatus.OPEN
        if not last_payment:
            payable.payment_method = ""
    payable.save()


def _refresh_receivable_after_payment_change(receivable: AccountsReceivable) -> None:
    if receivable.status == FinancialStatus.CANCELED:
        receivable.settlement_date = None
        receivable.save()
        return
    total_paid = (
        receivable.payments.aggregate(total=Sum("amount")).get("total")
        or Decimal("0.00")
    )
    total_due = receivable.total_amount()
    last_payment = receivable.payments.order_by("-payment_date", "-created_at").first()
    if total_paid >= total_due and total_due > 0:
        receivable.settlement_date = last_payment.payment_date if last_payment else None
        receivable.status = FinancialStatus.PAID
        if last_payment and last_payment.payment_method:
            receivable.payment_method = last_payment.payment_method
    else:
        receivable.settlement_date = None
        receivable.status = FinancialStatus.OPEN
        if not last_payment:
            receivable.payment_method = ""
    receivable.save()


class AccountsPayablePaymentReverseView(LoginRequiredMixin, View):
    template_name = "restricted/payment_reversal_confirm.html"
    page_title = "Estornar pagamento"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def _get_payment(self) -> AccountsPayablePayment:
        if not hasattr(self, "payment"):
            self.payment = get_object_or_404(AccountsPayablePayment, pk=self.kwargs["pk"])
        return self.payment

    def _build_context(self, request) -> dict[str, Any]:
        payment = self._get_payment()
        title = payment.payable
        amount_display = formats.number_format(
            payment.amount,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )
        reconciled = payment.reconciliation_items.exists()
        next_url = request.GET.get("next", "").strip() or request.POST.get("next", "").strip()
        cancel_url = (
            next_url
            or reverse("cadastros_web:accounts_payable_payment_list", args=[title.pk])
        )
        return {
            "page_title": self.page_title,
            "payment": payment,
            "title": title,
            "title_label": "Titulo",
            "party_label": "Fornecedor",
            "party_name": str(title.supplier),
            "payment_type": "Pagamento",
            "amount_display": f"R$ {amount_display}",
            "payment_date": payment.payment_date,
            "can_reverse": not reconciled,
            "reconciled": reconciled,
            "cancel_url": cancel_url,
            "next_url": cancel_url,
        }

    def get(self, request, *args, **kwargs):
        context = self._build_context(request)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        payment = self._get_payment()
        if payment.reconciliation_items.exists():
            messages.error(request, "Movimento conciliado. Remova a conciliacao antes de estornar.")
            context = self._build_context(request)
            return render(request, self.template_name, context)
        payable = payment.payable
        with transaction.atomic():
            payment.delete()
            _refresh_payable_after_payment_change(payable)
        messages.success(request, "Baixa estornada com sucesso.")
        redirect_url = request.POST.get("next") or reverse(
            "cadastros_web:accounts_payable_payment_list",
            args=[payable.pk],
        )
        return HttpResponseRedirect(redirect_url)



class AccountsPayablePaymentCreateView(BaseCreateView):
    model = AccountsPayablePayment
    form_class = AccountsPayablePaymentForm
    page_title = "Baixar conta a pagar"
    submit_label = "Registrar pagamento"
    cancel_url_name = "cadastros_web:accounts_payable_list"
    success_url = reverse_lazy("cadastros_web:accounts_payable_list")
    full_width_fields = ("notes",)
    allowed_roles = (UserRole.ADMIN,)

    def _get_payable(self) -> AccountsPayable:
        if not hasattr(self, "payable"):
            self.payable = get_object_or_404(AccountsPayable, pk=self.kwargs["pk"])
        return self.payable

    def _paid_total(self, payable: AccountsPayable) -> Decimal:
        return (
            payable.payments.aggregate(total=Sum("amount")).get("total")
            or Decimal("0.00")
        )

    def _remaining_amount(self, payable: AccountsPayable) -> Decimal:
        remaining = payable.total_amount() - self._paid_total(payable)
        return remaining.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["payable"] = self._get_payable()
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        remaining = self._remaining_amount(self._get_payable())
        if remaining > 0:
            initial.setdefault("amount", remaining)
        return initial

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse(
            "cadastros_web:accounts_payable_payment_list",
            args=[self._get_payable().pk],
        )

    def form_valid(self, form):
        payable = self._get_payable()
        was_paid = payable.status == FinancialStatus.PAID
        if payable.status == FinancialStatus.CANCELED:
            messages.error(self.request, "Titulo cancelado nao pode receber baixa.")
            return HttpResponseRedirect(self.get_success_url())
        remaining = self._remaining_amount(payable)
        if remaining <= 0:
            messages.error(self.request, "Titulo ja esta liquidado.")
            return HttpResponseRedirect(self.get_success_url())
        with transaction.atomic():
            payment = form.save(commit=False)
            payment.payable = payable
            payment.save()
            total_paid = self._paid_total(payable)
            total_due = payable.total_amount()
            if total_paid >= total_due:
                payable.settlement_date = payment.payment_date
                payable.status = FinancialStatus.PAID
                if payment.payment_method:
                    payable.payment_method = payment.payment_method
            else:
                payable.settlement_date = None
                payable.status = FinancialStatus.OPEN
            payable.save()
        messages.success(self.request, "Pagamento registrado com sucesso.")
        if not was_paid and payable.status == FinancialStatus.PAID:
            def _notify_paid() -> None:
                notify_admin_payable_paid(payable)
                notify_consultant_payable_paid(payable)

            transaction.on_commit(_notify_paid)
        return HttpResponseRedirect(self.get_success_url())


class ConsultantAccountsPayableEntriesView(BaseListView):
    model = TimeEntry
    page_title = "Apontamentos do titulo"
    list_title = "Apontamentos vinculados"
    search_placeholder = "Buscar por projeto ou atividade"
    ordering = ("start_date", "created_at")
    table_headers = ("Projeto", "Atividade", "Inicio", "Fim", "Horas", "Status")
    table_fields = (
        "activity.project",
        "activity",
        "start_date",
        "end_date",
        "total_hours",
        "status",
    )
    search_fields = (
        "activity__activity",
        "activity__project__description",
        "description",
    )
    create_url_name = "cadastros_web:consultant_panel"
    edit_url_name = "cadastros_web:consultant_panel"
    delete_url_name = "cadastros_web:consultant_panel"
    allowed_roles = (UserRole.CONSULTANT,)
    show_actions = False
    show_create = False

    def get_queryset(self):
        consultant = Consultant.objects.filter(user=self.request.user).first()
        if not consultant:
            raise PermissionDenied("Usuario sem consultor vinculado.")
        title = get_object_or_404(
            AccountsPayable,
            pk=self.kwargs["pk"],
            consultant=consultant,
        )
        self.title = title
        if not title.billing_invoice_id:
            return TimeEntry.objects.none()
        return (
            TimeEntry.objects.select_related(
                "consultant",
                "activity",
                "activity__project",
            )
            .filter(
                billing_invoice=title.billing_invoice,
                consultant=consultant,
            )
            .order_by("start_date", "created_at")
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        title = getattr(self, "title", None)
        invoice_number = title.billing_invoice.number if title and title.billing_invoice_id else "-"
        context["page_title"] = f"Apontamentos do titulo {invoice_number}"
        context["list_title"] = f"Fatura {invoice_number}"
        context["extra_actions"] = [
            {
                "label": "Voltar",
                "url": reverse("cadastros_web:consultant_panel"),
            }
        ]
        return context


class AccountsReceivableListView(BaseListView):
    model = AccountsReceivable
    queryset = AccountsReceivable.objects.select_related(
        "client",
        "client__company",
        "billing_invoice",
    )
    page_title = "Contas a receber"
    list_title = "Contas a receber cadastradas"
    search_placeholder = "Buscar por cliente ou documento"
    date_filter_field = "due_date"
    date_filter_label = "Vencimento"
    ordering = ("due_date",)
    table_headers = (
        "Fatura",
        "Documento",
        "Vencimento",
        "Valor original",
        "Valor em aberto",
        "Status",
    )
    table_fields = (
        "billing_invoice.number",
        "document_number",
        "due_date",
        "amount",
        "amount",
        "status",
    )
    search_fields = (
        "document_number",
        "description",
        "client__name",
        "client__company__legal_name",
        "client__company__trade_name",
        "billing_invoice__number",
    )
    create_url_name = "cadastros_web:accounts_receivable_create"
    edit_url_name = "cadastros_web:accounts_receivable_update"
    delete_url_name = "cadastros_web:accounts_receivable_delete"
    allowed_roles = (UserRole.ADMIN,)

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.annotate(
            paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00")))
        )

    def _get_chart_queryset(self):
        if hasattr(self, "_chart_queryset"):
            return self._chart_queryset
        base_qs = self._apply_filters(AccountsReceivable.objects.all())
        self._chart_queryset = base_qs.values(
            "id",
            "status",
            "due_date",
            "amount",
            "discount",
            "interest",
            "penalty",
        ).annotate(
            paid_total=Coalesce(Sum("payments__amount"), Value(Decimal("0.00")))
        ).order_by()
        return self._chart_queryset

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["extra_actions"] = [
            {
                "label": "Compensar titulos",
                "url": reverse("cadastros_web:accounts_compensation"),
            }
        ]
        amount_indices = [
            idx for idx, field in enumerate(self.table_fields) if field == "amount"
        ]
        if len(amount_indices) < 2:
            return context
        original_index, open_index = amount_indices[:2]
        for obj, row in zip(context["object_list"], context["table_rows"]):
            total_amount = obj.total_amount()
            open_amount = _resolve_open_amount(
                total_amount, getattr(obj, "paid_total", None)
            )
            row["values"][original_index] = formats.number_format(
                total_amount,
                decimal_pos=2,
                use_l10n=True,
                force_grouping=True,
            )
            row["values"][open_index] = formats.number_format(
                open_amount,
                decimal_pos=2,
                use_l10n=True,
                force_grouping=True,
            )
            if obj.client_id:
                row["lead"] = obj.client.name
        context["list_key"] = "accounts-receivable"
        context["due_charts"] = _build_due_charts(
            self._get_chart_queryset().iterator(),
            variant="receivable",
            subtitle="Titulos a receber em aberto",
        )
        return context

    def get_row_actions(self, obj: AccountsReceivable) -> list[dict[str, str]]:
        actions = [
            {
                "label": "Baixar",
                "url": reverse(
                    "cadastros_web:accounts_receivable_payment_create",
                    args=[obj.pk],
                ),
            },
            {
                "label": "Pagamentos",
                "url": reverse(
                    "cadastros_web:accounts_receivable_payment_list",
                    args=[obj.pk],
                ),
            }
        ]
        if obj.billing_invoice_id:
            actions.append(
                {
                    "label": "Apontamentos",
                    "url": reverse(
                        "cadastros_web:accounts_receivable_entries", args=[obj.pk]
                    ),
                }
            )
        return actions


class AccountsReceivableCreateView(BaseCreateView):
    model = AccountsReceivable
    form_class = AccountsReceivableForm
    page_title = "Nova conta a receber"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:accounts_receivable_list"
    success_url = reverse_lazy("cadastros_web:accounts_receivable_list")
    full_width_fields = ("description", "notes")
    allowed_roles = (UserRole.ADMIN,)

    def form_valid(self, form):
        response = super().form_valid(form)
        receivable = self.object
        if receivable:
            transaction.on_commit(lambda: notify_admin_receivable_created(receivable))
        return response


class AccountsReceivableUpdateView(BaseUpdateView):
    model = AccountsReceivable
    form_class = AccountsReceivableForm
    page_title = "Editar conta a receber"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:accounts_receivable_list"
    success_url = reverse_lazy("cadastros_web:accounts_receivable_list")
    full_width_fields = ("description", "notes")
    allowed_roles = (UserRole.ADMIN,)

    def form_valid(self, form):
        before = AccountsReceivable.objects.get(pk=self.object.pk)
        response = super().form_valid(form)
        if before.status != FinancialStatus.PAID and self.object.status == FinancialStatus.PAID:
            transaction.on_commit(lambda: notify_admin_receivable_paid(self.object))
        return response


class AccountsReceivableDeleteView(BaseDeleteView):
    model = AccountsReceivable
    cancel_url_name = "cadastros_web:accounts_receivable_list"
    success_url = reverse_lazy("cadastros_web:accounts_receivable_list")
    allowed_roles = (UserRole.ADMIN,)


class AccountsReceivableEntriesView(BaseListView):
    model = TimeEntry
    page_title = "Apontamentos do titulo"
    list_title = "Apontamentos vinculados"
    search_placeholder = "Buscar por consultor, projeto ou atividade"
    ordering = ("start_date", "created_at")
    table_headers = ("Consultor", "Projeto", "Atividade", "Inicio", "Fim", "Horas", "Status")
    table_fields = (
        "consultant",
        "activity.project",
        "activity",
        "start_date",
        "end_date",
        "total_hours",
        "status",
    )
    search_fields = (
        "consultant__full_name",
        "activity__activity",
        "activity__project__description",
        "description",
    )
    create_url_name = "cadastros_web:accounts_receivable_list"
    edit_url_name = "cadastros_web:time_entry_update"
    delete_url_name = "cadastros_web:time_entry_update"
    allowed_roles = (UserRole.ADMIN,)
    show_actions = False
    show_create = False

    def get_queryset(self):
        title = get_object_or_404(AccountsReceivable, pk=self.kwargs["pk"])
        self.title = title
        if not title.billing_invoice_id:
            return TimeEntry.objects.none()
        return (
            TimeEntry.objects.select_related("consultant", "activity", "activity__project")
            .filter(billing_invoice=title.billing_invoice)
            .order_by("start_date", "created_at")
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        title = getattr(self, "title", None)
        invoice_number = title.billing_invoice.number if title and title.billing_invoice_id else "-"
        context["page_title"] = f"Apontamentos do titulo {invoice_number}"
        context["list_title"] = f"Cliente: {title.client}" if title else "Apontamentos"
        context["extra_actions"] = [
            {
                "label": "Voltar",
                "url": reverse("cadastros_web:accounts_receivable_list"),
            }
        ]
        return context


class AccountsReceivablePaymentListView(BaseListView):
    model = AccountsReceivablePayment
    page_title = "Recebimentos"
    list_title = "Recebimentos do titulo"
    search_placeholder = "Buscar por conta bancaria"
    ordering = ("-payment_date", "-created_at")
    table_headers = ("Data", "Conta bancaria", "Valor", "Forma", "Observacoes")
    table_fields = (
        "payment_date",
        "bank_account",
        "amount",
        "payment_method",
        "notes",
    )
    search_fields = (
        "bank_account__bank_name",
        "bank_account__company__legal_name",
        "bank_account__company__trade_name",
        "notes",
    )
    create_url_name = "cadastros_web:accounts_receivable_list"
    edit_url_name = ""
    delete_url_name = ""
    allowed_roles = (UserRole.ADMIN,)
    status_field = None
    show_actions = True
    show_create = False

    def get_queryset(self):
        title = get_object_or_404(AccountsReceivable, pk=self.kwargs["pk"])
        self.title = title
        return (
            AccountsReceivablePayment.objects.select_related(
                "bank_account",
                "bank_account__company",
            )
            .filter(receivable=title)
            .order_by("-payment_date", "-created_at")
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        title = getattr(self, "title", None)
        document = title.document_number if title else "-"
        context["page_title"] = "Recebimentos do titulo"
        context["list_title"] = f"Titulo {document}"
        context["extra_actions"] = [
            {
                "label": "Nova baixa",
                "url": reverse(
                    "cadastros_web:accounts_receivable_payment_create",
                    args=[title.pk],
                )
                if title
                else reverse("cadastros_web:accounts_receivable_list"),
            },
            {
                "label": "Voltar",
                "url": reverse("cadastros_web:accounts_receivable_list"),
            },
        ]
        return context

    def get_row_actions(self, obj: AccountsReceivablePayment) -> list[dict[str, str]]:
        next_url = self.request.get_full_path()
        reverse_url = reverse(
            "cadastros_web:accounts_receivable_payment_reverse",
            args=[obj.pk],
        )
        if next_url:
            reverse_url = f"{reverse_url}?{urlencode({'next': next_url})}"
        return [{"label": "Estornar", "url": reverse_url}]


class AccountsReceivablePaymentReverseView(LoginRequiredMixin, View):
    template_name = "restricted/payment_reversal_confirm.html"
    page_title = "Estornar recebimento"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def _get_payment(self) -> AccountsReceivablePayment:
        if not hasattr(self, "payment"):
            self.payment = get_object_or_404(AccountsReceivablePayment, pk=self.kwargs["pk"])
        return self.payment

    def _build_context(self, request) -> dict[str, Any]:
        payment = self._get_payment()
        title = payment.receivable
        amount_display = formats.number_format(
            payment.amount,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )
        reconciled = payment.reconciliation_items.exists()
        next_url = request.GET.get("next", "").strip() or request.POST.get("next", "").strip()
        cancel_url = (
            next_url
            or reverse("cadastros_web:accounts_receivable_payment_list", args=[title.pk])
        )
        return {
            "page_title": self.page_title,
            "payment": payment,
            "title": title,
            "title_label": "Titulo",
            "party_label": "Cliente",
            "party_name": str(title.client),
            "payment_type": "Recebimento",
            "amount_display": f"R$ {amount_display}",
            "payment_date": payment.payment_date,
            "can_reverse": not reconciled,
            "reconciled": reconciled,
            "cancel_url": cancel_url,
            "next_url": cancel_url,
        }

    def get(self, request, *args, **kwargs):
        context = self._build_context(request)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        payment = self._get_payment()
        if payment.reconciliation_items.exists():
            messages.error(request, "Movimento conciliado. Remova a conciliacao antes de estornar.")
            context = self._build_context(request)
            return render(request, self.template_name, context)
        receivable = payment.receivable
        with transaction.atomic():
            payment.delete()
            _refresh_receivable_after_payment_change(receivable)
        messages.success(request, "Baixa estornada com sucesso.")
        redirect_url = request.POST.get("next") or reverse(
            "cadastros_web:accounts_receivable_payment_list",
            args=[receivable.pk],
        )
        return HttpResponseRedirect(redirect_url)


class AccountsReceivablePaymentCreateView(BaseCreateView):
    model = AccountsReceivablePayment
    form_class = AccountsReceivablePaymentForm
    page_title = "Baixar conta a receber"
    submit_label = "Registrar recebimento"
    cancel_url_name = "cadastros_web:accounts_receivable_list"
    success_url = reverse_lazy("cadastros_web:accounts_receivable_list")
    full_width_fields = ("notes",)
    allowed_roles = (UserRole.ADMIN,)

    def _get_receivable(self) -> AccountsReceivable:
        if not hasattr(self, "receivable"):
            self.receivable = get_object_or_404(AccountsReceivable, pk=self.kwargs["pk"])
        return self.receivable

    def _paid_total(self, receivable: AccountsReceivable) -> Decimal:
        return (
            receivable.payments.aggregate(total=Sum("amount")).get("total")
            or Decimal("0.00")
        )

    def _remaining_amount(self, receivable: AccountsReceivable) -> Decimal:
        remaining = receivable.total_amount() - self._paid_total(receivable)
        return remaining.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["receivable"] = self._get_receivable()
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        remaining = self._remaining_amount(self._get_receivable())
        if remaining > 0:
            initial.setdefault("amount", remaining)
        return initial

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return reverse(
            "cadastros_web:accounts_receivable_payment_list",
            args=[self._get_receivable().pk],
        )

    def form_valid(self, form):
        receivable = self._get_receivable()
        was_paid = receivable.status == FinancialStatus.PAID
        if receivable.status == FinancialStatus.CANCELED:
            messages.error(self.request, "Titulo cancelado nao pode receber baixa.")
            return HttpResponseRedirect(self.get_success_url())
        remaining = self._remaining_amount(receivable)
        if remaining <= 0:
            messages.error(self.request, "Titulo ja esta liquidado.")
            return HttpResponseRedirect(self.get_success_url())
        with transaction.atomic():
            payment = form.save(commit=False)
            payment.receivable = receivable
            payment.save()
            total_paid = self._paid_total(receivable)
            total_due = receivable.total_amount()
            if total_paid >= total_due:
                receivable.settlement_date = payment.payment_date
                receivable.status = FinancialStatus.PAID
                if payment.payment_method:
                    receivable.payment_method = payment.payment_method
            else:
                receivable.settlement_date = None
                receivable.status = FinancialStatus.OPEN
            receivable.save()
        messages.success(self.request, "Recebimento registrado com sucesso.")
        if not was_paid and receivable.status == FinancialStatus.PAID:
            transaction.on_commit(lambda: notify_admin_receivable_paid(receivable))
        return HttpResponseRedirect(self.get_success_url())


class ClientListView(BaseListView):
    model = Client
    queryset = Client.objects.select_related("company")
    page_title = "Clientes"
    list_title = "Clientes cadastrados"
    search_placeholder = "Buscar por cliente"
    ordering = ("name",)
    table_headers = ("Cliente", "Empresa", "Ciclo", "Prazo (dias)", "Status")
    table_fields = ("name", "company", "billing_cycle", "payment_terms_days", "status")
    search_fields = ("name", "company__legal_name", "company__trade_name")
    create_url_name = "cadastros_web:client_create"
    edit_url_name = "cadastros_web:client_update"
    delete_url_name = "cadastros_web:client_delete"


class ClientCreateView(BaseCreateView):
    model = Client
    form_class = ClientForm
    page_title = "Novo cliente"
    submit_label = "Salvar cliente"
    cancel_url_name = "cadastros_web:client_list"
    success_url = reverse_lazy("cadastros_web:client_list")
    full_width_fields = ("address_line", "commercial_notes")


class ClientUpdateView(BaseUpdateView):
    model = Client
    form_class = ClientForm
    page_title = "Editar cliente"
    submit_label = "Salvar cliente"
    cancel_url_name = "cadastros_web:client_list"
    success_url = reverse_lazy("cadastros_web:client_list")
    full_width_fields = ("address_line", "commercial_notes")


class ClientDeleteView(BaseDeleteView):
    model = Client
    cancel_url_name = "cadastros_web:client_list"
    success_url = reverse_lazy("cadastros_web:client_list")


class ClientContactListView(BaseListView):
    model = ClientContact
    queryset = ClientContact.objects.select_related("client")
    page_title = "Contatos"
    list_title = "Contatos de clientes"
    search_placeholder = "Buscar por nome ou email"
    ordering = ("client__name", "name")
    table_headers = ("Contato", "Cliente", "Email", "Principal", "Status")
    table_fields = ("name", "client", "email", "is_primary", "status")
    search_fields = ("name", "email", "client__name")
    create_url_name = "cadastros_web:contact_create"
    edit_url_name = "cadastros_web:contact_update"
    delete_url_name = "cadastros_web:contact_delete"
    filter_params = {"client_id": "client_id"}


class ClientContactCreateView(BaseCreateView):
    model = ClientContact
    form_class = ClientContactForm
    page_title = "Novo contato"
    submit_label = "Salvar contato"
    cancel_url_name = "cadastros_web:contact_list"
    success_url = reverse_lazy("cadastros_web:contact_list")


class ClientContactUpdateView(BaseUpdateView):
    model = ClientContact
    form_class = ClientContactForm
    page_title = "Editar contato"
    submit_label = "Salvar contato"
    cancel_url_name = "cadastros_web:contact_list"
    success_url = reverse_lazy("cadastros_web:contact_list")


class ClientContactDeleteView(BaseDeleteView):
    model = ClientContact
    cancel_url_name = "cadastros_web:contact_list"
    success_url = reverse_lazy("cadastros_web:contact_list")


class ConsultantListView(BaseListView):
    model = Consultant
    queryset = Consultant.objects.select_related("company", "supplier")
    page_title = "Consultores"
    list_title = "Consultores cadastrados"
    search_placeholder = "Buscar por consultor"
    ordering = ("full_name",)
    table_headers = ("Consultor", "Tipo", "Empresa", "Fornecedor", "Status")
    table_fields = ("full_name", "contract_type", "company", "supplier", "status")
    search_fields = (
        "full_name",
        "email",
        "document",
        "supplier__name",
        "supplier__trade_name",
        "supplier__document",
    )
    create_url_name = "cadastros_web:consultant_create"
    edit_url_name = "cadastros_web:consultant_update"
    delete_url_name = "cadastros_web:consultant_delete"


class ConsultantCreateView(BaseCreateView):
    model = Consultant
    form_class = ConsultantForm
    template_name = "restricted/consultant_form.html"
    page_title = "Novo consultor"
    submit_label = "Salvar consultor"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("notes", "competencies", "certifications")

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "consultant": None,
                "attachments": [],
                "rates": [],
                "bank_accounts": [],
                "bank_account_form": None,
                "attachment_form": None,
                "rate_form": None,
                "bank_account_create_url": "",
                "attachment_create_url": "",
                "rate_create_url": "",
                "bank_account_next_url": "",
                "attachment_next_url": "",
                "rate_next_url": "",
            }
        )
        return context


class ConsultantUpdateView(BaseUpdateView):
    model = Consultant
    form_class = ConsultantForm
    template_name = "restricted/consultant_form.html"
    page_title = "Editar consultor"
    submit_label = "Salvar consultor"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("notes", "competencies", "certifications")

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        attachments = self.object.attachments.order_by("-created_at")
        rates = self.object.rates.order_by("-start_date", "-created_at")
        bank_accounts = self.object.bank_accounts.order_by("bank_name", "account_number")
        attachment_form = ConsultantAttachmentForm(initial={"consultant": self.object.pk})
        attachment_form.fields["consultant"].widget = forms.HiddenInput()
        rate_form = ConsultantRateForm(initial={"consultant": self.object.pk})
        rate_form.fields["consultant"].widget = forms.HiddenInput()
        bank_account_form = ConsultantBankAccountForm(
            initial={"consultant": self.object.pk}
        )
        bank_account_form.fields["consultant"].widget = forms.HiddenInput()
        context.update(
            {
                "consultant": self.object,
                "attachments": attachments,
                "rates": rates,
                "bank_accounts": bank_accounts,
                "attachment_form": attachment_form,
                "rate_form": rate_form,
                "bank_account_form": bank_account_form,
                "attachment_create_url": reverse(
                    "cadastros_web:consultant_attachment_create"
                ),
                "rate_create_url": reverse("cadastros_web:consultant_rate_create"),
                "bank_account_create_url": reverse(
                    "cadastros_web:consultant_bank_account_create"
                ),
                "attachment_next_url": reverse(
                    "cadastros_web:consultant_update",
                    args=[self.object.pk],
                ),
                "rate_next_url": reverse(
                    "cadastros_web:consultant_update",
                    args=[self.object.pk],
                ),
                "bank_account_next_url": reverse(
                    "cadastros_web:consultant_update",
                    args=[self.object.pk],
                ),
            }
        )
        return context


class ConsultantDeleteView(BaseDeleteView):
    model = Consultant
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")


class ConsultantAttachmentCreateView(BaseCreateView):
    model = ConsultantAttachment
    form_class = ConsultantAttachmentForm
    page_title = "Novo arquivo do consultor"
    submit_label = "Salvar arquivo"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("description", "file")

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantAttachmentUpdateView(BaseUpdateView):
    model = ConsultantAttachment
    form_class = ConsultantAttachmentForm
    page_title = "Editar arquivo do consultor"
    submit_label = "Salvar arquivo"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("description", "file")

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantAttachmentDeleteView(BaseDeleteView):
    model = ConsultantAttachment
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantRateCreateView(BaseCreateView):
    model = ConsultantRate
    form_class = ConsultantRateForm
    page_title = "Nova tarifa do consultor"
    submit_label = "Salvar tarifa"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("notes",)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantRateUpdateView(BaseUpdateView):
    model = ConsultantRate
    form_class = ConsultantRateForm
    page_title = "Editar tarifa do consultor"
    submit_label = "Salvar tarifa"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("notes",)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantRateDeleteView(BaseDeleteView):
    model = ConsultantRate
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantBankAccountCreateView(BaseCreateView):
    model = ConsultantBankAccount
    form_class = ConsultantBankAccountForm
    page_title = "Nova conta bancaria"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("pix_keys",)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantBankAccountUpdateView(BaseUpdateView):
    model = ConsultantBankAccount
    form_class = ConsultantBankAccountForm
    page_title = "Editar conta bancaria"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")
    full_width_fields = ("pix_keys",)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ConsultantBankAccountDeleteView(BaseDeleteView):
    model = ConsultantBankAccount
    cancel_url_name = "cadastros_web:consultant_list"
    success_url = reverse_lazy("cadastros_web:consultant_list")

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class CompetencyListView(BaseListView):
    model = Competency
    page_title = "Competencias"
    list_title = "Competencias cadastradas"
    search_placeholder = "Buscar competencia"
    ordering = ("name",)
    table_headers = ("Competencia", "Status")
    table_fields = ("name", "status")
    search_fields = ("name",)
    create_url_name = "cadastros_web:competency_create"
    edit_url_name = "cadastros_web:competency_update"
    delete_url_name = "cadastros_web:competency_delete"


class CompetencyCreateView(BaseCreateView):
    model = Competency
    form_class = CompetencyForm
    page_title = "Nova competencia"
    submit_label = "Salvar competencia"
    cancel_url_name = "cadastros_web:competency_list"
    success_url = reverse_lazy("cadastros_web:competency_list")
    full_width_fields = ("description",)


class CompetencyUpdateView(BaseUpdateView):
    model = Competency
    form_class = CompetencyForm
    page_title = "Editar competencia"
    submit_label = "Salvar competencia"
    cancel_url_name = "cadastros_web:competency_list"
    success_url = reverse_lazy("cadastros_web:competency_list")
    full_width_fields = ("description",)


class CompetencyDeleteView(BaseDeleteView):
    model = Competency
    cancel_url_name = "cadastros_web:competency_list"
    success_url = reverse_lazy("cadastros_web:competency_list")


class CertificationListView(BaseListView):
    model = Certification
    page_title = "Certificacoes"
    list_title = "Certificacoes cadastradas"
    search_placeholder = "Buscar certificacao"
    ordering = ("name",)
    table_headers = ("Certificacao", "Emissor", "Status")
    table_fields = ("name", "issuer", "status")
    search_fields = ("name", "issuer")
    create_url_name = "cadastros_web:certification_create"
    edit_url_name = "cadastros_web:certification_update"
    delete_url_name = "cadastros_web:certification_delete"


class CertificationCreateView(BaseCreateView):
    model = Certification
    form_class = CertificationForm
    page_title = "Nova certificacao"
    submit_label = "Salvar certificacao"
    cancel_url_name = "cadastros_web:certification_list"
    success_url = reverse_lazy("cadastros_web:certification_list")
    full_width_fields = ("description",)


class CertificationUpdateView(BaseUpdateView):
    model = Certification
    form_class = CertificationForm
    page_title = "Editar certificacao"
    submit_label = "Salvar certificacao"
    cancel_url_name = "cadastros_web:certification_list"
    success_url = reverse_lazy("cadastros_web:certification_list")
    full_width_fields = ("description",)


class CertificationDeleteView(BaseDeleteView):
    model = Certification
    cancel_url_name = "cadastros_web:certification_list"
    success_url = reverse_lazy("cadastros_web:certification_list")


class PhaseListView(BaseListView):
    model = Phase
    page_title = "Fases"
    list_title = "Fases cadastradas"
    search_placeholder = "Buscar fase"
    ordering = ("description",)
    table_headers = ("Descricao", "Situacao")
    table_fields = ("description", "status")
    search_fields = ("description",)
    create_url_name = "cadastros_web:phase_create"
    edit_url_name = "cadastros_web:phase_update"
    delete_url_name = "cadastros_web:phase_delete"


class PhaseCreateView(BaseCreateView):
    model = Phase
    form_class = PhaseForm
    page_title = "Nova fase"
    submit_label = "Salvar fase"
    cancel_url_name = "cadastros_web:phase_list"
    success_url = reverse_lazy("cadastros_web:phase_list")


class PhaseUpdateView(BaseUpdateView):
    model = Phase
    form_class = PhaseForm
    page_title = "Editar fase"
    submit_label = "Salvar fase"
    cancel_url_name = "cadastros_web:phase_list"
    success_url = reverse_lazy("cadastros_web:phase_list")


class PhaseDeleteView(BaseDeleteView):
    model = Phase
    cancel_url_name = "cadastros_web:phase_list"
    success_url = reverse_lazy("cadastros_web:phase_list")


class ProjectRoleListView(BaseListView):
    model = ProjectRole
    page_title = "Papeis de projeto"
    list_title = "Papeis de projeto cadastrados"
    search_placeholder = "Buscar papel"
    ordering = ("name",)
    table_headers = ("Papel", "Situacao")
    table_fields = ("name", "status")
    search_fields = ("name",)
    create_url_name = "cadastros_web:project_role_create"
    edit_url_name = "cadastros_web:project_role_update"
    delete_url_name = "cadastros_web:project_role_delete"


class ProjectRoleCreateView(BaseCreateView):
    model = ProjectRole
    form_class = ProjectRoleForm
    page_title = "Novo papel de projeto"
    submit_label = "Salvar papel"
    cancel_url_name = "cadastros_web:project_role_list"
    success_url = reverse_lazy("cadastros_web:project_role_list")


class ProjectRoleUpdateView(BaseUpdateView):
    model = ProjectRole
    form_class = ProjectRoleForm
    page_title = "Editar papel de projeto"
    submit_label = "Salvar papel"
    cancel_url_name = "cadastros_web:project_role_list"
    success_url = reverse_lazy("cadastros_web:project_role_list")


class ProjectRoleDeleteView(BaseDeleteView):
    model = ProjectRole
    cancel_url_name = "cadastros_web:project_role_list"
    success_url = reverse_lazy("cadastros_web:project_role_list")


class ProductListView(BaseListView):
    model = Product
    page_title = "Produtos"
    list_title = "Produtos cadastrados"
    search_placeholder = "Buscar produto"
    ordering = ("description",)
    table_headers = ("Descricao", "Situacao")
    table_fields = ("description", "status")
    search_fields = ("description",)
    create_url_name = "cadastros_web:product_create"
    edit_url_name = "cadastros_web:product_update"
    delete_url_name = "cadastros_web:product_delete"


class ProductCreateView(BaseCreateView):
    model = Product
    form_class = ProductForm
    page_title = "Novo produto"
    submit_label = "Salvar produto"
    cancel_url_name = "cadastros_web:product_list"
    success_url = reverse_lazy("cadastros_web:product_list")


class ProductUpdateView(BaseUpdateView):
    model = Product
    form_class = ProductForm
    page_title = "Editar produto"
    submit_label = "Salvar produto"
    cancel_url_name = "cadastros_web:product_list"
    success_url = reverse_lazy("cadastros_web:product_list")


class ProductDeleteView(BaseDeleteView):
    model = Product
    cancel_url_name = "cadastros_web:product_list"
    success_url = reverse_lazy("cadastros_web:product_list")


class ModuleListView(BaseListView):
    model = Module
    queryset = Module.objects.select_related("product")
    page_title = "Modulos"
    list_title = "Modulos cadastrados"
    search_placeholder = "Buscar modulo"
    ordering = ("description",)
    table_headers = ("Descricao", "Produto", "Situacao")
    table_fields = ("description", "product", "status")
    search_fields = ("description", "product__description")
    create_url_name = "cadastros_web:module_create"
    edit_url_name = "cadastros_web:module_update"
    delete_url_name = "cadastros_web:module_delete"


class ModuleCreateView(BaseCreateView):
    model = Module
    form_class = ModuleForm
    page_title = "Novo modulo"
    submit_label = "Salvar modulo"
    cancel_url_name = "cadastros_web:module_list"
    success_url = reverse_lazy("cadastros_web:module_list")


class ModuleUpdateView(BaseUpdateView):
    model = Module
    form_class = ModuleForm
    page_title = "Editar modulo"
    submit_label = "Salvar modulo"
    cancel_url_name = "cadastros_web:module_list"
    success_url = reverse_lazy("cadastros_web:module_list")


class ModuleDeleteView(BaseDeleteView):
    model = Module
    cancel_url_name = "cadastros_web:module_list"
    success_url = reverse_lazy("cadastros_web:module_list")


class SubmoduleListView(BaseListView):
    model = Submodule
    queryset = Submodule.objects.select_related("product", "module")
    page_title = "Submodulos"
    list_title = "Submodulos cadastrados"
    search_placeholder = "Buscar submodulo"
    ordering = ("description",)
    table_headers = ("Descricao", "Produto", "Modulo", "Situacao")
    table_fields = ("description", "product", "module", "status")
    search_fields = ("description", "product__description", "module__description")
    create_url_name = "cadastros_web:submodule_create"
    edit_url_name = "cadastros_web:submodule_update"
    delete_url_name = "cadastros_web:submodule_delete"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["extra_actions"] = [
            {
                "label": "Cadastro em massa",
                "url": reverse("cadastros_web:submodule_bulk_create"),
            }
        ]
        return context


class SubmoduleCreateView(BaseCreateView):
    model = Submodule
    form_class = SubmoduleForm
    page_title = "Novo submodulo"
    submit_label = "Salvar submodulo"
    cancel_url_name = "cadastros_web:submodule_list"
    success_url = reverse_lazy("cadastros_web:submodule_list")


class SubmoduleUpdateView(BaseUpdateView):
    model = Submodule
    form_class = SubmoduleForm
    page_title = "Editar submodulo"
    submit_label = "Salvar submodulo"
    cancel_url_name = "cadastros_web:submodule_list"
    success_url = reverse_lazy("cadastros_web:submodule_list")


class SubmoduleDeleteView(BaseDeleteView):
    model = Submodule
    cancel_url_name = "cadastros_web:submodule_list"
    success_url = reverse_lazy("cadastros_web:submodule_list")


class SubmoduleBulkCreateView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/submodule_bulk_create.html"

    def _get_product_id(self, request) -> str:
        return (request.POST.get("product_id") or request.GET.get("product_id") or "").strip()

    def _get_modules(self, product_id: str):
        queryset = Module.objects.select_related("product").order_by(
            "product__description",
            "description",
        )
        if product_id:
            try:
                queryset = queryset.filter(product_id=int(product_id))
            except (TypeError, ValueError):
                queryset = queryset.none()
        return list(queryset)

    def _build_context(
        self,
        request,
        form,
        product_id: str,
        modules: list[Module],
        selected_module_ids: list[int],
    ) -> dict[str, Any]:
        products = Product.objects.order_by("description")
        selected_set = set(selected_module_ids)
        return {
            "page_title": "Cadastro de submodulos em massa",
            "form": form,
            "products": products,
            "modules": modules,
            "selected_product_id": product_id,
            "selected_module_ids": selected_set,
        }

    def get(self, request, *args, **kwargs):
        product_id = self._get_product_id(request)
        modules = self._get_modules(product_id)
        form = SubmoduleBulkCreateForm(
            initial={"status": StatusChoices.ACTIVE}
        )
        context = self._build_context(request, form, product_id, modules, [])
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        product_id = self._get_product_id(request)
        modules = self._get_modules(product_id)
        form = SubmoduleBulkCreateForm(request.POST)
        module_ids_raw = request.POST.getlist("module_ids")
        selected_module_ids: list[int] = []
        for value in module_ids_raw:
            try:
                selected_module_ids.append(int(value))
            except (TypeError, ValueError):
                continue

        if not selected_module_ids:
            form.add_error(None, "Selecione ao menos um modulo.")
            context = self._build_context(
                request, form, product_id, modules, selected_module_ids
            )
            return self.render_to_response(context)

        if not form.is_valid():
            context = self._build_context(
                request, form, product_id, modules, selected_module_ids
            )
            return self.render_to_response(context)

        description = form.cleaned_data["description"].strip()
        status = form.cleaned_data["status"]
        module_queryset = Module.objects.select_related("product").filter(
            id__in=selected_module_ids
        )
        found_ids = set(module_queryset.values_list("id", flat=True))
        missing_count = len(set(selected_module_ids) - found_ids)

        existing_ids = set(
            Submodule.objects.filter(
                module_id__in=found_ids,
                description__iexact=description,
            ).values_list("module_id", flat=True)
        )

        to_create = []
        skipped = 0
        for module in module_queryset:
            if module.id in existing_ids:
                skipped += 1
                continue
            to_create.append(
                Submodule(
                    product=module.product,
                    module=module,
                    description=description,
                    status=status,
                )
            )

        if to_create:
            Submodule.objects.bulk_create(to_create)
        created_count = len(to_create)

        message = f"{created_count} submodulos criados."
        if skipped:
            message += f" {skipped} ja existiam."
        if missing_count:
            message += f" {missing_count} modulos nao encontrados."
        messages.success(request, message)

        redirect_url = reverse("cadastros_web:submodule_bulk_create")
        if product_id:
            redirect_url = f"{redirect_url}?product_id={product_id}"
        return HttpResponseRedirect(redirect_url)


class DeploymentTemplateListView(BaseListView):
    model = DeploymentTemplateHeader
    page_title = "Templates de implantacao"
    list_title = "Templates de implantacao"
    search_placeholder = "Buscar template"
    ordering = ("name",)
    table_headers = ("Nome do template", "Criado em")
    table_fields = ("name", "created_at")
    search_fields = ("name",)
    status_field = None
    create_url_name = "cadastros_web:deployment_template_create"
    edit_url_name = "cadastros_web:deployment_template_update"
    delete_url_name = "cadastros_web:deployment_template_delete"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["extra_actions"] = [
            {
                "label": "Importar Excel/MS Project",
                "url": reverse("cadastros_web:deployment_template_import"),
            }
        ]
        return context

    def get_row_actions(self, obj: models.Model) -> list[dict[str, str]]:
        return [
            {
                "label": "Manutencao em massa",
                "url": reverse(
                    "cadastros_web:deployment_template_maintenance",
                    args=[obj.pk],
                ),
            }
        ]


class DeploymentTemplateHierarchyView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/deployment_template_hierarchy.html"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        headers = DeploymentTemplateHeader.objects.prefetch_related(
            models.Prefetch(
                "items",
                queryset=DeploymentTemplate.objects.select_related(
                    "phase",
                    "product",
                    "module",
                    "submodule",
                ).order_by("seq"),
            )
        ).order_by("name")
        templates = []
        for header in headers:
            items = list(header.items.all())
            phases = []
            phase_index = {}
            for item in items:
                phase_name = str(item.phase)
                phase_entry = phase_index.get(phase_name)
                if phase_entry is None:
                    phase_entry = {"name": phase_name, "activities": []}
                    phase_index[phase_name] = phase_entry
                    phases.append(phase_entry)
                activity_index = phase_entry.setdefault("_activity_index", {})
                activity_entry = activity_index.get(item.activity)
                if activity_entry is None:
                    activity_entry = {"name": item.activity, "subactivities": []}
                    activity_index[item.activity] = activity_entry
                    phase_entry["activities"].append(activity_entry)
                subactivity = item.subactivity.strip() if item.subactivity else "Sem subatividade"
                activity_entry["subactivities"].append(subactivity)
            for phase_entry in phases:
                phase_entry.pop("_activity_index", None)
            templates.append(
                {
                    "id": header.id,
                    "name": header.name,
                    "phases": phases,
                    "phase_count": len(phases),
                    "item_count": len(items),
                }
            )
        context["page_title"] = "Hierarquia de templates"
        context["templates"] = templates
        return context


class DeploymentTemplateCreateView(BaseCreateView):
    model = DeploymentTemplateHeader
    form_class = DeploymentTemplateHeaderForm
    template_name = "restricted/deployment_template_form.html"
    page_title = "Novo template de implantacao"
    submit_label = "Salvar template"
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")
    form_columns = 1
    full_width_fields = ("name",)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "template": None,
                "items": [],
                "item_form": None,
                "item_create_url": "",
                "item_next_url": "",
            }
        )
        return context


class DeploymentTemplateUpdateView(BaseUpdateView):
    model = DeploymentTemplateHeader
    form_class = DeploymentTemplateHeaderForm
    template_name = "restricted/deployment_template_form.html"
    page_title = "Editar template de implantacao"
    submit_label = "Salvar template"
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")
    form_columns = 1
    full_width_fields = ("name",)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        items = (
            self.object.items.select_related(
                "phase",
                "product",
                "module",
                "submodule",
            )
            .order_by("seq")
        )
        item_form = DeploymentTemplateItemForm(initial={"template": self.object.pk})
        item_form.fields["template"].widget = forms.HiddenInput()
        context.update(
            {
                "template": self.object,
                "items": items,
                "item_form": item_form,
                "item_create_url": reverse(
                    "cadastros_web:deployment_template_item_create"
                ),
                "item_next_url": reverse(
                    "cadastros_web:deployment_template_update",
                    args=[self.object.pk],
                ),
            }
        )
        return context


class DeploymentTemplateMaintenanceView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/deployment_template_maintenance.html"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["page_classes"] = "content-wide"
        return context

    def _missing_fields(self, item: DeploymentTemplate) -> list[str]:
        missing = []
        if item.days is None or item.days <= 0:
            missing.append("Dias")
        if item.hours is None or item.hours <= 0:
            missing.append("Horas")
        subactivity = (item.subactivity or "").strip().lower()
        if not subactivity or subactivity in {"a definir", "sem subatividade"}:
            missing.append("Subatividade")
        return missing

    def _build_formset(self, data, queryset):
        FormSet = modelformset_factory(
            DeploymentTemplate,
            form=DeploymentTemplateMaintenanceForm,
            extra=0,
        )
        return FormSet(data=data, queryset=queryset)

    def _get_context(self, request, template: DeploymentTemplateHeader, formset, items, missing_map):
        missing_ids = [item_id for item_id, fields in missing_map.items() if fields]
        rows = []
        for form in formset:
            row_missing = missing_map.get(form.instance.id, [])
            rows.append(
                {
                    "form": form,
                    "missing_fields": row_missing,
                    "is_missing": bool(row_missing),
                }
            )
        return {
            "template": template,
            "formset": formset,
            "rows": rows,
            "total_items": len(items),
            "missing_count": len(missing_ids),
            "missing_only": request.GET.get("missing") == "1",
        }

    def get(self, request, *args, **kwargs):
        template = get_object_or_404(DeploymentTemplateHeader, pk=kwargs.get("pk"))
        missing_only = request.GET.get("missing") == "1"
        base_qs = template.items.select_related(
            "phase",
            "product",
            "module",
            "submodule",
        ).order_by("seq")
        items = list(base_qs)
        missing_map = {item.id: self._missing_fields(item) for item in items}
        if missing_only:
            missing_ids = [
                item_id for item_id, fields in missing_map.items() if fields
            ]
            queryset = base_qs.filter(pk__in=missing_ids) if missing_ids else base_qs.none()
        else:
            queryset = base_qs
        formset = self._build_formset(None, queryset)
        context = self._get_context(request, template, formset, items, missing_map)
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        template = get_object_or_404(DeploymentTemplateHeader, pk=kwargs.get("pk"))
        missing_only = request.GET.get("missing") == "1"
        base_qs = template.items.select_related(
            "phase",
            "product",
            "module",
            "submodule",
        ).order_by("seq")
        items = list(base_qs)
        missing_map = {item.id: self._missing_fields(item) for item in items}
        if missing_only:
            missing_ids = [
                item_id for item_id, fields in missing_map.items() if fields
            ]
            queryset = base_qs.filter(pk__in=missing_ids) if missing_ids else base_qs.none()
        else:
            queryset = base_qs
        formset = self._build_formset(request.POST, queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Itens atualizados com sucesso.")
            redirect_url = reverse(
                "cadastros_web:deployment_template_maintenance",
                args=[template.pk],
            )
            if missing_only:
                redirect_url = f"{redirect_url}?missing=1"
            return HttpResponseRedirect(redirect_url)
        context = self._get_context(request, template, formset, items, missing_map)
        return self.render_to_response(context)


class DeploymentTemplateDeleteView(BaseDeleteView):
    model = DeploymentTemplateHeader
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")


class DeploymentTemplateItemCreateView(BaseCreateView):
    model = DeploymentTemplate
    form_class = DeploymentTemplateItemForm
    page_title = "Novo item do template"
    submit_label = "Salvar item"
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")
    full_width_fields = ("activity", "subactivity")

    def get_initial(self):
        initial = super().get_initial()
        template_id = self.request.GET.get("template")
        if template_id:
            initial["template"] = template_id
        return initial

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class DeploymentTemplateItemUpdateView(BaseUpdateView):
    model = DeploymentTemplate
    form_class = DeploymentTemplateItemForm
    page_title = "Editar item do template"
    submit_label = "Salvar item"
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")
    full_width_fields = ("activity", "subactivity")

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class DeploymentTemplateItemDeleteView(BaseDeleteView):
    model = DeploymentTemplate
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class DeploymentTemplateImportView(BaseFormView, FormView):
    template_name = "restricted/form.html"
    form_class = DeploymentTemplateImportForm
    page_title = "Importar template de implantacao"
    submit_label = "Importar"
    cancel_url_name = "cadastros_web:deployment_template_list"
    success_url = reverse_lazy("cadastros_web:deployment_template_list")
    full_width_fields = ("file",)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["download_template_url"] = static(
            "modelos/modelo_importacao_templates.xlsx"
        )
        context["download_template_label"] = "Baixar modelo"
        return context

    def form_valid(self, form):
        uploaded_file = form.cleaned_data["file"]
        try:
            count, errors = import_deployment_templates(uploaded_file)
        except ValueError as exc:
            form.add_error("file", str(exc))
            messages.error(self.request, str(exc))
            return self.form_invalid(form)
        if errors:
            form.add_error("file", "Arquivo com erros. Corrija e tente novamente.")
            for error in errors[:20]:
                messages.error(self.request, error)
            if len(errors) > 20:
                messages.error(
                    self.request,
                    f"Mais {len(errors) - 20} erros nao exibidos.",
                )
            return self.form_invalid(form)

        messages.success(
            self.request, f"{count} templates importados com sucesso."
        )
        return HttpResponseRedirect(self.get_success_url())


class AccountPlanTemplateListView(BaseListView):
    model = AccountPlanTemplateHeader
    page_title = "Modelos de plano de contas"
    list_title = "Modelos de plano de contas"
    search_placeholder = "Buscar modelo"
    ordering = ("name",)
    table_headers = ("Nome do modelo", "Criado em")
    table_fields = ("name", "created_at")
    search_fields = ("name",)
    status_field = None
    create_url_name = "cadastros_web:account_plan_template_create"
    edit_url_name = "cadastros_web:account_plan_template_update"
    delete_url_name = "cadastros_web:account_plan_template_delete"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["extra_actions"] = [
            {
                "label": "Importar Excel",
                "url": reverse("cadastros_web:account_plan_template_import"),
            }
        ]
        return context


class AccountPlanTemplateHierarchyView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/account_plan_template_hierarchy.html"

    def _build_node(self, item) -> dict[str, Any]:
        level = item.level or 1
        level_class = f"level-{level}" if level < 3 else "level-3"
        return {
            "id": item.id,
            "code": item.code,
            "description": item.description,
            "level": level,
            "level_class": level_class,
            "indent": max(level - 1, 0) * 18,
            "children": [],
        }

    def _sort_children(self, nodes: list[dict[str, Any]]) -> None:
        nodes.sort(key=lambda node: node.get("code", ""))
        for node in nodes:
            self._sort_children(node["children"])

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        headers = AccountPlanTemplateHeader.objects.prefetch_related(
            models.Prefetch(
                "items",
                queryset=AccountPlanTemplateItem.objects.select_related("parent").order_by("code"),
            )
        ).order_by("name")
        templates = []
        for header in headers:
            items = list(header.items.all())
            node_map = {item.id: self._build_node(item) for item in items}
            roots = []
            max_level = 0
            for item in items:
                node = node_map[item.id]
                max_level = max(max_level, node["level"])
                if item.parent_id and item.parent_id in node_map:
                    node_map[item.parent_id]["children"].append(node)
                else:
                    roots.append(node)
            self._sort_children(roots)
            templates.append(
                {
                    "id": header.id,
                    "name": header.name,
                    "description": header.description,
                    "roots": roots,
                    "item_count": len(items),
                    "max_level": max_level,
                }
            )
        context["page_title"] = "Hierarquia do plano de contas"
        context["templates"] = templates
        return context


class AccountPlanTemplateCreateView(BaseCreateView):
    model = AccountPlanTemplateHeader
    form_class = AccountPlanTemplateHeaderForm
    template_name = "restricted/account_plan_template_form.html"
    page_title = "Novo modelo de plano de contas"
    submit_label = "Salvar modelo"
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")
    form_columns = 1
    full_width_fields = ("name", "description")

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "template": None,
                "items": [],
                "item_form": None,
                "item_create_url": "",
                "item_next_url": "",
            }
        )
        return context


class AccountPlanTemplateUpdateView(BaseUpdateView):
    model = AccountPlanTemplateHeader
    form_class = AccountPlanTemplateHeaderForm
    template_name = "restricted/account_plan_template_form.html"
    page_title = "Editar modelo de plano de contas"
    submit_label = "Salvar modelo"
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")
    form_columns = 1
    full_width_fields = ("name", "description")

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        items = self.object.items.select_related("parent").order_by("code")
        item_form = AccountPlanTemplateItemForm(initial={"template": self.object.pk})
        item_form.fields["template"].widget = forms.HiddenInput()
        context.update(
            {
                "template": self.object,
                "items": items,
                "item_form": item_form,
                "item_create_url": reverse(
                    "cadastros_web:account_plan_template_item_create"
                ),
                "item_next_url": reverse(
                    "cadastros_web:account_plan_template_update",
                    args=[self.object.pk],
                ),
            }
        )
        return context


class AccountPlanTemplateDeleteView(BaseDeleteView):
    model = AccountPlanTemplateHeader
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")


class AccountPlanTemplateItemCreateView(BaseCreateView):
    model = AccountPlanTemplateItem
    form_class = AccountPlanTemplateItemForm
    page_title = "Nova conta do modelo"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")
    full_width_fields = ("description", "dre_group", "dre_subgroup")

    def get_initial(self):
        initial = super().get_initial()
        template_id = self.request.GET.get("template")
        if template_id:
            initial["template"] = template_id
        return initial

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class AccountPlanTemplateItemUpdateView(BaseUpdateView):
    model = AccountPlanTemplateItem
    form_class = AccountPlanTemplateItemForm
    page_title = "Editar conta do modelo"
    submit_label = "Salvar conta"
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")
    full_width_fields = ("description", "dre_group", "dre_subgroup")

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class AccountPlanTemplateItemDeleteView(BaseDeleteView):
    model = AccountPlanTemplateItem
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class AccountPlanTemplateImportView(BaseFormView, FormView):
    template_name = "restricted/form.html"
    form_class = AccountPlanTemplateImportForm
    page_title = "Importar plano de contas"
    submit_label = "Importar"
    cancel_url_name = "cadastros_web:account_plan_template_list"
    success_url = reverse_lazy("cadastros_web:account_plan_template_list")
    full_width_fields = ("file",)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["download_template_url"] = static(
            "modelos/modelo_importacao_plano_contas.xlsx"
        )
        context["download_template_label"] = "Baixar modelo"
        return context

    def form_valid(self, form):
        uploaded_file = form.cleaned_data["file"]
        try:
            count, errors = import_account_plan_templates(uploaded_file)
        except ValueError as exc:
            form.add_error("file", str(exc))
            messages.error(self.request, str(exc))
            return self.form_invalid(form)
        if errors:
            form.add_error("file", "Arquivo com erros. Corrija e tente novamente.")
            for error in errors[:20]:
                messages.error(self.request, error)
            if len(errors) > 20:
                messages.error(
                    self.request,
                    f"Mais {len(errors) - 20} erros nao exibidos.",
                )
            return self.form_invalid(form)

        messages.success(self.request, f"{count} contas importadas com sucesso.")
        return super().form_valid(form)


class ProjectListView(BaseListView):
    model = Project
    queryset = Project.objects.select_related(
        "billing_client",
        "project_client",
        "internal_manager",
        "external_manager",
        "client_user",
    )
    page_title = "Projetos"
    list_title = "Projetos cadastrados"
    search_placeholder = "Buscar projeto"
    ordering = ("description",)
    table_headers = (
        "Projeto",
        "Recebimento",
        "Go live planejado",
        "Cliente faturamento",
        "Cliente projeto",
        "Status",
        "Valor total",
        "Valor hora",
        "Horas contratadas",
        "Contingencia (%)",
        "Horas disponiveis",
        "Valor disponivel",
    )
    table_fields = (
        "description",
        "received_date",
        "planned_go_live_date",
        "billing_client",
        "project_client",
        "status",
        "total_value",
        "hourly_rate",
        "contracted_hours",
        "contingency_percent",
        "available_hours",
        "available_value",
    )
    search_fields = ("description",)
    create_url_name = "cadastros_web:project_create"
    edit_url_name = "cadastros_web:project_update"
    delete_url_name = "cadastros_web:project_delete"

    def get_queryset(self):
        queryset = super().get_queryset()
        return filter_projects_for_user(queryset, self.request.user)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        role = resolve_user_role(self.request.user)
        if can_view_financial(self.request.user):
            return context

        if role == UserRole.CLIENT:
            headers = (
                "Projeto",
                "Recebimento",
                "Go live planejado",
                "Cliente projeto",
                "Status",
            )
        elif role in {UserRole.CONSULTANT, UserRole.GP_EXTERNAL}:
            headers = (
                "Projeto",
                "Recebimento",
                "Go live planejado",
                "Cliente projeto",
                "Status",
                "Horas disponiveis",
            )
        else:
            headers = (
                "Projeto",
                "Recebimento",
                "Go live planejado",
                "Cliente projeto",
                "Status",
                "Horas contratadas",
                "Contingencia (%)",
                "Horas disponiveis",
            )
        can_edit_delete = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        can_create = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        show_actions = True
        rows = []
        for project in context["object_list"]:
            values = [
                project.description,
                _format_value(project, "received_date"),
                _format_value(project, "planned_go_live_date"),
                str(project.project_client),
                project.get_status_display(),
            ]
            if role == UserRole.CLIENT:
                pass
            elif role in {UserRole.CONSULTANT, UserRole.GP_EXTERNAL}:
                values.append(_format_value(project, "available_hours"))
            else:
                values.extend(
                    [
                        _format_value(project, "contracted_hours"),
                        _format_value(project, "contingency_percent"),
                        _format_value(project, "available_hours"),
                    ]
                )
            rows.append(
                {
                    "values": values,
                    "edit_url": reverse(self.edit_url_name, args=[project.pk])
                    if can_edit_delete
                    else None,
                    "delete_url": reverse(self.delete_url_name, args=[project.pk])
                    if can_edit_delete
                    else None,
                    "extra_actions": [
                        {
                            "label": "Ocorrencias",
                            "url": f"{reverse('cadastros_web:project_occurrence_list')}?project_id={project.pk}",
                        }
                    ],
                }
            )
        context.update(
            {
                "table_headers": list(headers),
                "table_rows": rows,
                "column_count": len(headers) + (1 if show_actions else 0),
                "show_actions": show_actions,
                "show_create": can_create,
            }
        )
        return context


class ProjectCreateView(BaseCreateView):
    model = Project
    form_class = ProjectForm
    template_name = "restricted/project_form.html"
    page_title = "Novo projeto"
    submit_label = "Salvar projeto"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    full_width_fields = ("description", "hml_url", "prd_url", "explanation")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
    )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if not can_view_financial(self.request.user):
            for field_name in ("total_value", "hourly_rate", "available_value"):
                form.fields.pop(field_name, None)
        return form

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "project": None,
                "project_contacts": [],
                "project_contact_form": None,
                "project_contact_create_url": "",
                "project_contact_next_url": "",
                "attachments": [],
                "attachment_form": None,
                "attachment_create_url": "",
                "attachment_next_url": "",
                "observations": [],
                "observation_form": None,
                "observation_create_url": "",
                "observation_next_url": "",
                "go_no_go_items": [],
                "go_no_go_form": None,
                "go_no_go_create_url": "",
                "go_no_go_next_url": "",
                "go_no_go_summary": [],
                "occurrences": [],
                "occurrence_form": None,
                "occurrence_create_url": "",
                "occurrence_next_url": "",
                "occurrence_attachment_create_url": "",
                "project_history_url": "",
            }
        )
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        create_project_receipt_observation(self.object, self.request.user)
        return response


class ProjectUpdateView(BaseUpdateView):
    model = Project
    form_class = ProjectForm
    template_name = "restricted/project_form.html"
    page_title = "Editar projeto"
    submit_label = "Salvar projeto"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    full_width_fields = ("description", "hml_url", "prd_url", "explanation")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
    )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if not can_view_financial(self.request.user):
            for field_name in ("total_value", "hourly_rate", "available_value"):
                form.fields.pop(field_name, None)
        return form

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        project_contacts = self.object.contacts.select_related("role").order_by("name")
        attachments = self.object.attachments.order_by("-created_at")
        observations = (
            ProjectObservation.objects.select_related("created_by")
            .filter(project=self.object)
            .order_by("-created_at")
        )
        attachment_form = ProjectAttachmentForm()
        contact_form = ProjectContactForm(initial={"project": self.object.pk})
        contact_form.fields["project"].widget = forms.HiddenInput()
        observation_form = None
        go_no_go_form = None
        occurrence_form = None
        role = resolve_user_role(self.request.user)
        can_manage_sensitive = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        if can_manage_sensitive:
            observation_form = ProjectObservationForm(initial={"project": self.object.pk})
            observation_form.fields["project"].widget = forms.HiddenInput()
            go_no_go_form = ProjectGoNoGoChecklistItemForm(
                initial={"project": self.object.pk}
            )
            go_no_go_form.fields["project"].widget = forms.HiddenInput()
            _apply_visibility_choices(go_no_go_form, role)
            occurrence_form = ProjectOccurrenceForm(initial={"project": self.object.pk})
            occurrence_form.fields["project"].widget = forms.HiddenInput()
            _apply_visibility_choices(occurrence_form, role)
        go_no_go_items = list(
            filter_by_visibility(
                self.object.go_no_go_items.order_by("id"),
                role,
            )
        )
        go_no_go_result_map = {
            GoNoGoResult.OK: "chip-ok",
            GoNoGoResult.NO: "chip-danger",
            GoNoGoResult.PENDING: "chip-warn",
        }
        go_no_go_counts = {key: 0 for key, _ in GoNoGoResult.choices}
        for item in go_no_go_items:
            item.result_chip = go_no_go_result_map.get(item.result, "chip-neutral")
            if item.result in go_no_go_counts:
                go_no_go_counts[item.result] += 1
        go_no_go_summary = [
            {
                "label": label,
                "count": go_no_go_counts.get(value, 0),
                "chip": go_no_go_result_map.get(value, "chip-neutral"),
            }
            for value, label in GoNoGoResult.choices
        ]
        occurrences = list(
            filter_by_visibility(
                self.object.occurrences.select_related("created_by").prefetch_related(
                    "attachments"
                ),
                role,
            )
        )
        context.update(
            {
                "project": self.object,
                "project_contacts": project_contacts,
                "project_contact_form": contact_form,
                "attachments": attachments,
                "observations": observations,
                "attachment_form": attachment_form,
                "observation_form": observation_form,
                "go_no_go_items": go_no_go_items,
                "go_no_go_form": go_no_go_form,
                "go_no_go_summary": go_no_go_summary,
                "go_no_go_result_map": go_no_go_result_map,
                "occurrences": occurrences,
                "occurrence_form": occurrence_form,
                "project_contact_create_url": reverse(
                    "cadastros_web:project_contact_create"
                ),
                "project_contact_next_url": reverse(
                    "cadastros_web:project_update",
                    args=[self.object.pk],
                ),
                "attachment_create_url": reverse(
                    "cadastros_web:project_attachment_create"
                ),
                "attachment_next_url": reverse(
                    "cadastros_web:project_update",
                    args=[self.object.pk],
                ),
                "observation_create_url": reverse(
                    "cadastros_web:project_observation_create"
                )
                if observation_form
                else "",
                "observation_next_url": reverse(
                    "cadastros_web:project_update",
                    args=[self.object.pk],
                )
                if observation_form
                else "",
                "go_no_go_create_url": reverse(
                    "cadastros_web:project_go_no_go_create"
                )
                if go_no_go_form
                else "",
                "go_no_go_next_url": reverse(
                    "cadastros_web:project_update",
                    args=[self.object.pk],
                )
                if go_no_go_form
                else "",
                "occurrence_create_url": reverse(
                    "cadastros_web:project_occurrence_create"
                )
                if occurrence_form
                else "",
                "occurrence_next_url": reverse(
                    "cadastros_web:project_update",
                    args=[self.object.pk],
                )
                if occurrence_form
                else "",
                "occurrence_attachment_create_url": reverse(
                    "cadastros_web:project_occurrence_attachment_create"
                ),
                "project_history_url": reverse(
                    "cadastros_web:project_history",
                    args=[self.object.pk],
                ),
            }
        )
        return context

    def form_valid(self, form):
        before = Project.objects.get(pk=self.object.pk)
        response = super().form_valid(form)
        create_project_change_observation(before, self.object, self.request.user)
        if before.received_date != self.object.received_date:
            create_project_receipt_observation(
                self.object,
                self.request.user,
                previous_date=before.received_date,
            )
        return response


class ProjectObservationCreateView(BaseCreateView):
    model = ProjectObservation
    form_class = ProjectObservationForm
    page_title = "Nova observacao do projeto"
    submit_label = "Salvar observacao"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    full_width_fields = ("note",)
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def form_valid(self, form):
        project = form.cleaned_data.get("project")
        if not project:
            raise PermissionDenied("Projeto nao informado.")
        allowed_project = filter_projects_for_user(
            Project.objects.filter(pk=project.pk),
            self.request.user,
        ).exists()
        if not allowed_project:
            raise PermissionDenied("Perfil sem permissao para alterar projetos.")
        form.instance.created_by = self.request.user
        form.instance.observation_type = ProjectObservationType.MANUAL
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectGoNoGoChecklistItemCreateView(BaseCreateView):
    model = ProjectGoNoGoChecklistItem
    form_class = ProjectGoNoGoChecklistItemForm
    page_title = "Novo item Go/No-Go"
    submit_label = "Salvar item"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    full_width_fields = ("required_evidence", "observation")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        project_id = (self.request.GET.get("project_id") or "").strip()
        if project_id and project_id.isdigit() and int(project_id) in project_ids:
            form.initial.setdefault("project", int(project_id))
        _apply_visibility_choices(form, resolve_user_role(self.request.user))
        return form

    def form_valid(self, form):
        project = form.cleaned_data.get("project")
        if not project:
            raise PermissionDenied("Projeto nao informado.")
        allowed_project = filter_projects_for_user(
            Project.objects.filter(pk=project.pk),
            self.request.user,
        ).exists()
        if not allowed_project:
            raise PermissionDenied("Perfil sem permissao para alterar projetos.")
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectGoNoGoChecklistItemUpdateView(BaseUpdateView):
    model = ProjectGoNoGoChecklistItem
    form_class = ProjectGoNoGoChecklistItemForm
    page_title = "Editar item Go/No-Go"
    submit_label = "Salvar item"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    full_width_fields = ("required_evidence", "observation")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        _apply_visibility_choices(form, resolve_user_role(self.request.user))
        return form

    def get_queryset(self):
        queryset = super().get_queryset().select_related("project")
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        queryset = queryset.filter(project_id__in=project_ids)
        return filter_by_visibility(queryset, resolve_user_role(self.request.user))

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectGoNoGoChecklistItemDeleteView(BaseDeleteView):
    model = ProjectGoNoGoChecklistItem
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_queryset(self):
        queryset = super().get_queryset().select_related("project")
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        queryset = queryset.filter(project_id__in=project_ids)
        return filter_by_visibility(queryset, resolve_user_role(self.request.user))

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectOccurrenceListView(BaseListView):
    model = ProjectOccurrence
    queryset = ProjectOccurrence.objects.select_related("project", "created_by")
    page_title = "Ocorrencias"
    list_title = "Ocorrencias do projeto"
    search_placeholder = "Buscar ocorrencia"
    ordering = ("-created_at",)
    status_field = None
    table_headers = (
        "Projeto",
        "Ocorrencia",
        "Visibilidade",
        "Registrado por",
        "Criado em",
    )
    table_fields = ("project", "title", "visibility", "created_by", "created_at")
    search_fields = ("title", "description", "project__description")
    create_url_name = "cadastros_web:project_occurrence_create"
    edit_url_name = "cadastros_web:project_occurrence_update"
    delete_url_name = "cadastros_web:project_occurrence_delete"
    filter_params = {"project_id": "project_id"}
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
        UserRole.CONSULTANT,
        UserRole.CLIENT,
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        queryset = queryset.filter(project_id__in=project_ids)
        return filter_by_visibility(queryset, resolve_user_role(self.request.user))

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        role = resolve_user_role(self.request.user)
        can_manage = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        self.show_actions = can_manage
        self.show_create = can_manage
        return super().get_context_data(**kwargs)


class ProjectOccurrenceCreateView(BaseCreateView):
    model = ProjectOccurrence
    form_class = ProjectOccurrenceForm
    page_title = "Nova ocorrencia do projeto"
    submit_label = "Salvar ocorrencia"
    cancel_url_name = "cadastros_web:project_occurrence_list"
    success_url = reverse_lazy("cadastros_web:project_occurrence_list")
    full_width_fields = ("description",)
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        _apply_visibility_choices(form, resolve_user_role(self.request.user))
        return form

    def form_valid(self, form):
        project = form.cleaned_data.get("project")
        if not project:
            raise PermissionDenied("Projeto nao informado.")
        allowed_project = filter_projects_for_user(
            Project.objects.filter(pk=project.pk),
            self.request.user,
        ).exists()
        if not allowed_project:
            raise PermissionDenied("Perfil sem permissao para alterar projetos.")
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectOccurrenceUpdateView(BaseUpdateView):
    model = ProjectOccurrence
    form_class = ProjectOccurrenceForm
    page_title = "Editar ocorrencia do projeto"
    submit_label = "Salvar ocorrencia"
    cancel_url_name = "cadastros_web:project_occurrence_list"
    success_url = reverse_lazy("cadastros_web:project_occurrence_list")
    full_width_fields = ("description",)
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        _apply_visibility_choices(form, resolve_user_role(self.request.user))
        return form

    def get_queryset(self):
        queryset = super().get_queryset().select_related("project")
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        queryset = queryset.filter(project_id__in=project_ids)
        return filter_by_visibility(queryset, resolve_user_role(self.request.user))

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectOccurrenceDeleteView(BaseDeleteView):
    model = ProjectOccurrence
    cancel_url_name = "cadastros_web:project_occurrence_list"
    success_url = reverse_lazy("cadastros_web:project_occurrence_list")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_queryset(self):
        queryset = super().get_queryset().select_related("project")
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        queryset = queryset.filter(project_id__in=project_ids)
        return filter_by_visibility(queryset, resolve_user_role(self.request.user))

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectOccurrenceAttachmentCreateView(BaseCreateView):
    model = ProjectOccurrenceAttachment
    form_class = ProjectOccurrenceAttachmentForm
    page_title = "Novo anexo da ocorrencia"
    submit_label = "Salvar anexo"
    cancel_url_name = "cadastros_web:project_occurrence_list"
    success_url = reverse_lazy("cadastros_web:project_occurrence_list")
    full_width_fields = ("file",)
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        role = resolve_user_role(self.request.user)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        occurrences = ProjectOccurrence.objects.select_related("project").filter(
            project_id__in=project_ids
        )
        occurrences = filter_by_visibility(occurrences, role)
        form.fields["occurrence"].queryset = occurrences
        occurrence_id = (self.request.GET.get("occurrence_id") or "").strip()
        if occurrence_id and occurrence_id.isdigit():
            occurrence_id = int(occurrence_id)
            if occurrences.filter(pk=occurrence_id).exists():
                form.initial.setdefault("occurrence", occurrence_id)
        return form

    def form_valid(self, form):
        occurrence = form.cleaned_data.get("occurrence")
        if not occurrence:
            raise PermissionDenied("Ocorrencia nao informada.")
        allowed_project = filter_projects_for_user(
            Project.objects.filter(pk=occurrence.project_id),
            self.request.user,
        ).exists()
        if not allowed_project:
            raise PermissionDenied("Perfil sem permissao para alterar projetos.")
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectOccurrenceAttachmentDeleteView(BaseDeleteView):
    model = ProjectOccurrenceAttachment
    cancel_url_name = "cadastros_web:project_occurrence_list"
    success_url = reverse_lazy("cadastros_web:project_occurrence_list")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_queryset(self):
        queryset = super().get_queryset().select_related("occurrence", "occurrence__project")
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        queryset = queryset.filter(occurrence__project_id__in=project_ids)
        return filter_by_visibility(queryset, resolve_user_role(self.request.user), "occurrence__visibility")

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectHistoryView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/project_history.html"

    def _build_timeline(self, project: Project) -> dict[str, object]:
        activities = (
            ProjectActivity.objects.select_related("phase")
            .filter(project=project)
            .order_by("seq")
        )
        transition_activity = activities.filter(
            phase__description__icontains="transicao"
        ).first()
        planned_end = activities.aggregate(max_value=models.Max("planned_end")).get(
            "max_value"
        )
        actual_end = activities.aggregate(max_value=models.Max("actual_end")).get(
            "max_value"
        )
        transition_label = None
        if transition_activity:
            transition_label = transition_activity.activity
            subactivities_label = _format_activity_subactivities(
                transition_activity, ""
            )
            if subactivities_label:
                transition_label = f"{transition_label} / {subactivities_label}"
        return {
            "received_date": project.received_date,
            "planned_go_live_date": project.planned_go_live_date,
            "transition_activity": transition_activity,
            "transition_label": transition_label,
            "planned_end": planned_end,
            "actual_end": actual_end,
        }

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get("pk")
        project = (
            filter_projects_for_user(
                Project.objects.select_related(
                    "billing_client",
                    "project_client",
                    "internal_manager",
                    "external_manager",
                    "client_user",
                ),
                self.request.user,
            )
            .filter(pk=project_id)
            .first()
        )
        if not project:
            raise PermissionDenied("Perfil sem acesso ao projeto.")
        observations = (
            ProjectObservation.objects.select_related("created_by")
            .filter(project=project)
            .order_by("-created_at")
        )
        role = resolve_user_role(self.request.user)
        project_edit_url = ""
        if role in {UserRole.ADMIN, UserRole.GP_INTERNAL}:
            project_edit_url = reverse(
                "cadastros_web:project_update",
                args=[project.pk],
            )
        context.update(
            {
                "page_title": "Historico do projeto",
                "project": project,
                "observations": observations,
                "timeline": self._build_timeline(project),
                "project_edit_url": project_edit_url,
            }
        )
        return context


class ProjectDeleteView(BaseDeleteView):
    model = Project
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
    )


class ProjectAttachmentListView(BaseListView):
    model = ProjectAttachment
    queryset = ProjectAttachment.objects.select_related("project")
    page_title = "Arquivos de projeto"
    list_title = "Arquivos de projeto"
    search_placeholder = "Buscar arquivo"
    ordering = ("-created_at",)
    table_headers = ("Projeto", "Tipo", "Descricao", "Arquivo", "Enviado em")
    table_fields = ("project", "attachment_type", "description", "file", "created_at")
    search_fields = ("description", "project__description")
    status_field = None
    create_url_name = "cadastros_web:project_attachment_create"
    edit_url_name = "cadastros_web:project_attachment_update"
    delete_url_name = "cadastros_web:project_attachment_delete"
    filter_params = {"project_id": "project_id"}
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def get_queryset(self):
        queryset = super().get_queryset()
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        return queryset.filter(project_id__in=project_ids)


class ProjectAttachmentCreateView(BaseCreateView):
    model = ProjectAttachment
    form_class = ProjectAttachmentForm
    page_title = "Novo arquivo do projeto"
    submit_label = "Salvar arquivo"
    cancel_url_name = "cadastros_web:project_attachment_list"
    success_url = reverse_lazy("cadastros_web:project_attachment_list")
    full_width_fields = ("description", "file")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        return form

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectAttachmentUpdateView(BaseUpdateView):
    model = ProjectAttachment
    form_class = ProjectAttachmentForm
    page_title = "Editar arquivo do projeto"
    submit_label = "Salvar arquivo"
    cancel_url_name = "cadastros_web:project_attachment_list"
    success_url = reverse_lazy("cadastros_web:project_attachment_list")
    full_width_fields = ("description", "file")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )


class ProjectAttachmentDeleteView(BaseDeleteView):
    model = ProjectAttachment
    cancel_url_name = "cadastros_web:project_attachment_list"
    success_url = reverse_lazy("cadastros_web:project_attachment_list")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )


class ProjectContactCreateView(BaseCreateView):
    model = ProjectContact
    form_class = ProjectContactForm
    page_title = "Nova pessoa do projeto"
    submit_label = "Salvar pessoa"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        return form

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectContactUpdateView(BaseUpdateView):
    model = ProjectContact
    form_class = ProjectContactForm
    page_title = "Editar pessoa do projeto"
    submit_label = "Salvar pessoa"
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        return form

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectContactDeleteView(BaseDeleteView):
    model = ProjectContact
    cancel_url_name = "cadastros_web:project_list"
    success_url = reverse_lazy("cadastros_web:project_list")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def get_success_url(self):
        next_url = self.request.POST.get("next") or self.request.GET.get("next")
        if next_url:
            return next_url
        return super().get_success_url()


class ProjectActivityListView(LoginRequiredMixin, ListView):
    model = ProjectActivity
    template_name = "restricted/project_activity_list.html"
    paginate_by = 15
    page_title = "Atividades do projeto"

    def _format_decimal(self, value) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _build_querystring(
        self,
        prefix: str = "&",
        exclude: Iterable[str] = (),
    ) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        for key in exclude:
            params.pop(key, None)
        if not params:
            return ""
        return f"{prefix}{params.urlencode()}"

    def _can_export(self) -> bool:
        role = resolve_user_role(self.request.user)
        return role in {
            UserRole.ADMIN,
            UserRole.GP_INTERNAL,
            UserRole.GP_EXTERNAL,
        }

    def get_queryset(self):
        queryset = (
            ProjectActivity.objects.select_related(
                "project",
                "template_item",
                "phase",
                "product",
                "module",
                "submodule",
                "account_plan_item",
                "client_feedback_by",
            )
            .prefetch_related("consultants", "predecessors", "subactivity_items")
            .order_by("project_id", "seq")
        )
        queryset = filter_activities_for_user(queryset, self.request.user)
        query = self.request.GET.get("q", "").strip()
        project_id = self.request.GET.get("project_id")
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        if query:
            queryset = queryset.filter(
                Q(project__description__icontains=query)
                | Q(activity__icontains=query)
                | Q(subactivity__icontains=query)
                | Q(subactivity_items__description__icontains=query)
            ).distinct()
        return queryset

    def get(self, request, *args, **kwargs):
        export_kind = request.GET.get("export", "").strip().lower()
        if export_kind in {"excel", "msproject", "ms-project", "ms_project"}:
            if not self._can_export():
                raise PermissionDenied("Perfil sem permissao para exportar atividades.")
            queryset = self.get_queryset()
            if export_kind == "excel":
                return self._export_excel(queryset)
            return self._export_msproject(queryset)
        return super().get(request, *args, **kwargs)

    def _export_excel(self, queryset) -> HttpResponse:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Atividades"

        headers = [
            "Projeto",
            "Item do template",
            "Seq",
            "Seq predecessoras",
            "Fase",
            "Produto",
            "Modulo",
            "Submodulo",
            "Atividade",
            "Subatividades",
            "Criticidade",
            "Dias",
            "Horas",
            "Classificacao de horas",
            "Valor hora consultor (R$)",
            "Conta plano de contas",
            "Inicio previsto",
            "Fim previsto",
            "Inicio real",
            "Fim real",
            "Status",
            "Consultores",
            "Visivel ao cliente",
            "Concluida pelo cliente",
            "Comentario do cliente",
            "Feedback por",
            "Feedback em",
            "Horas contingencia",
            "Horas liberadas",
            "Criado em",
            "Atualizado em",
        ]
        sheet.append(headers)

        for activity in queryset:
            consultants = ", ".join(
                consultant.full_name for consultant in activity.consultants.all()
            )
            predecessor_list = sorted(
                activity.predecessors.all(),
                key=lambda item: item.seq,
            )
            predecessor_seqs = ", ".join(str(pred.seq) for pred in predecessor_list)
            if not predecessor_seqs and activity.seq_predecessor:
                predecessor_seqs = str(activity.seq_predecessor)
            row = [
                str(activity.project),
                _format_value(activity, "template_item"),
                activity.seq,
                predecessor_seqs,
                str(activity.phase),
                str(activity.product),
                str(activity.module),
                str(activity.submodule),
                activity.activity,
                _format_activity_subactivities(activity),
                activity.get_criticality_display(),
                _format_value(activity, "days"),
                _format_value(activity, "hours"),
                activity.billing_type_label(),
                _format_value(activity, "consultant_hourly_rate"),
                _format_value(activity, "account_plan_item"),
                _format_value(activity, "planned_start"),
                _format_value(activity, "planned_end"),
                _format_value(activity, "actual_start"),
                _format_value(activity, "actual_end"),
                activity.get_status_display(),
                consultants or "-",
                "Sim" if activity.client_visible else "Nao",
                "-"
                if activity.client_completed is None
                else ("Sim" if activity.client_completed else "Nao"),
                activity.client_comment or "-",
                _format_value(activity, "client_feedback_by"),
                _format_value(activity, "client_feedback_at"),
                self._format_decimal(activity.hours_contingency()),
                self._format_decimal(activity.hours_available()),
                _format_value(activity, "created_at"),
                _format_value(activity, "updated_at"),
            ]
            sheet.append(row)

        for idx, header in enumerate(headers, start=1):
            column_letter = get_column_letter(idx)
            width = max(14, len(str(header)) + 4)
            sheet.column_dimensions[column_letter].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="atividades_projeto.xlsx"'
        )
        return response

    def _format_msproject_datetime(self, value, fallback_time: str) -> str | None:
        if not value:
            return None
        if hasattr(value, "strftime"):
            return value.strftime(f"%Y-%m-%dT{fallback_time}")
        return None

    def _format_msproject_work(self, hours: Decimal | None) -> str | None:
        if hours is None:
            return None
        total_minutes = int(
            (hours * Decimal("60")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        )
        hour_part = total_minutes // 60
        minute_part = total_minutes % 60
        return f"PT{hour_part}H{minute_part}M0S"

    def _format_msproject_duration(
        self, days: Decimal | int | None, hours: Decimal | None
    ) -> str | None:
        if days is not None and days > 0:
            try:
                amount = Decimal(days)
            except Exception:
                amount = None
            if amount is not None:
                if amount == amount.to_integral_value():
                    return f"P{int(amount)}D"
                work = self._format_msproject_work(hours)
                if work:
                    return work
                total_hours = (amount * Decimal("8")).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                return self._format_msproject_work(total_hours) or "PT0H0M0S"
        work = self._format_msproject_work(hours)
        return work or "PT0H0M0S"

    def _export_msproject(self, queryset) -> HttpResponse:
        activities = list(queryset)
        activity_keys_with_sub: set[tuple[int, int, int, int, int, str]] = set()
        for activity in activities:
            if _get_activity_subactivities(activity):
                activity_keys_with_sub.add(
                    (
                        activity.project_id,
                        activity.phase_id,
                        activity.product_id,
                        activity.module_id,
                        activity.submodule_id,
                        activity.activity,
                    )
                )

        root = ET.Element("Project", xmlns="http://schemas.microsoft.com/project")
        name_el = ET.SubElement(root, "Name")
        name_el.text = "Exportacao de atividades"
        tasks_el = ET.SubElement(root, "Tasks")

        uid_counter = 1
        seq_uid_map: dict[tuple[int, int], int] = {}
        pending_predecessors: list[tuple[ET.Element, tuple[int, int]]] = []

        current_project_id = None
        current_phase_id = None
        current_product_id = None
        current_module_id = None
        current_submodule_id = None
        current_activity_name = None

        def add_task(
            name: str,
            outline_level: int,
            summary: bool,
            start=None,
            finish=None,
            duration: str | None = None,
            work: str | None = None,
        ) -> ET.Element:
            nonlocal uid_counter
            task_el = ET.SubElement(tasks_el, "Task")
            uid_el = ET.SubElement(task_el, "UID")
            uid_el.text = str(uid_counter)
            id_el = ET.SubElement(task_el, "ID")
            id_el.text = str(uid_counter)
            name_el = ET.SubElement(task_el, "Name")
            name_el.text = name
            outline_el = ET.SubElement(task_el, "OutlineLevel")
            outline_el.text = str(outline_level)
            summary_el = ET.SubElement(task_el, "Summary")
            summary_el.text = "1" if summary else "0"
            if start:
                start_el = ET.SubElement(task_el, "Start")
                start_el.text = start
            if finish:
                finish_el = ET.SubElement(task_el, "Finish")
                finish_el.text = finish
            if duration:
                duration_el = ET.SubElement(task_el, "Duration")
                duration_el.text = duration
            if work:
                work_el = ET.SubElement(task_el, "Work")
                work_el.text = work
            uid_counter += 1
            return task_el

        for activity in activities:
            if activity.project_id != current_project_id:
                current_project_id = activity.project_id
                current_phase_id = None
                current_product_id = None
                current_module_id = None
                current_submodule_id = None
                current_activity_name = None
                add_task(str(activity.project), 1, True)

            if activity.phase_id != current_phase_id:
                current_phase_id = activity.phase_id
                current_product_id = None
                current_module_id = None
                current_submodule_id = None
                current_activity_name = None
                add_task(str(activity.phase), 2, True)

            if activity.product_id != current_product_id:
                current_product_id = activity.product_id
                current_module_id = None
                current_submodule_id = None
                current_activity_name = None
                add_task(str(activity.product), 3, True)

            if activity.module_id != current_module_id:
                current_module_id = activity.module_id
                current_submodule_id = None
                current_activity_name = None
                add_task(str(activity.module), 4, True)

            if activity.submodule_id != current_submodule_id:
                current_submodule_id = activity.submodule_id
                current_activity_name = None
                add_task(str(activity.submodule), 5, True)

            start = self._format_msproject_datetime(
                activity.actual_start or activity.planned_start,
                "08:00:00",
            )
            finish = self._format_msproject_datetime(
                activity.actual_end or activity.planned_end,
                "17:00:00",
            )
            work = self._format_msproject_work(activity.hours)
            duration = self._format_msproject_duration(activity.days, activity.hours)

            activity_key = (
                activity.project_id,
                activity.phase_id,
                activity.product_id,
                activity.module_id,
                activity.submodule_id,
                activity.activity,
            )
            if activity_key in activity_keys_with_sub:
                if activity.activity != current_activity_name:
                    current_activity_name = activity.activity
                    add_task(activity.activity, 6, True)
                subactivities_label = _format_activity_subactivities(activity, "")
                detail_name = subactivities_label or activity.activity
                task_el = add_task(detail_name, 7, False, start, finish, duration, work)
            else:
                current_activity_name = activity.activity
                task_el = add_task(activity.activity, 6, False, start, finish, duration, work)

            if activity.seq:
                seq_uid_map[(activity.project_id, activity.seq)] = int(
                    task_el.findtext("UID")
                )
            predecessors = list(activity.predecessors.all())
            if predecessors:
                for predecessor in predecessors:
                    pending_predecessors.append(
                        (task_el, (activity.project_id, predecessor.seq))
                    )
            elif activity.seq_predecessor:
                pending_predecessors.append(
                    (task_el, (activity.project_id, activity.seq_predecessor))
                )

        for task_el, key in pending_predecessors:
            predecessor_uid = seq_uid_map.get(key)
            if predecessor_uid:
                link_el = ET.SubElement(task_el, "PredecessorLink")
                pred_el = ET.SubElement(link_el, "PredecessorUID")
                pred_el.text = str(predecessor_uid)

        xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        response = HttpResponse(xml_bytes, content_type="application/xml")
        response["Content-Disposition"] = (
            'attachment; filename="atividades_projeto.xml"'
        )
        return response

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        role = resolve_user_role(self.request.user)
        can_manage = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        can_edit_delete = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        can_create = role in {UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.GP_EXTERNAL}
        can_feedback = role == UserRole.CLIENT
        show_consultants = role in {
            UserRole.ADMIN,
            UserRole.GP_INTERNAL,
            UserRole.GP_EXTERNAL,
        }
        show_hours_total = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        show_hours_contingency = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        show_hours_available = role in {
            UserRole.ADMIN,
            UserRole.GP_EXTERNAL,
            UserRole.CONSULTANT,
        }

        headers = [
            "Projeto",
            "Fase",
            "Atividade",
            "Subatividades",
            "Status",
            "Criticidade",
            "Classificacao de horas",
            "Prev. inicio",
            "Prev. fim",
            "Real inicio",
            "Real fim",
            "Prazo",
        ]
        if show_consultants:
            headers.append("Consultores")
        if show_hours_total:
            headers.append("Horas")
        if show_hours_contingency:
            headers.append("Horas contingencia")
        if show_hours_available:
            headers.append("Horas liberadas")
        headers.extend(["Cliente concluiu", "Comentario cliente"])

        rows = []
        today = timezone.localdate()
        schedule_chip_map = {
            "late": "chip-danger",
            "on_time": "chip-ok",
            "not_started": "chip-info",
        }
        for activity in context["object_list"]:
            schedule_state = activity.schedule_state(today)
            values = [
                str(activity.project),
                str(activity.phase),
                activity.activity,
                _format_activity_subactivities(activity),
                activity.get_status_display(),
                activity.get_criticality_display(),
                activity.billing_type_label(),
                _format_value(activity, "planned_start"),
                _format_value(activity, "planned_end"),
                _format_value(activity, "actual_start"),
                _format_value(activity, "actual_end"),
                {
                    "text": activity.schedule_label(today),
                    "chip": schedule_chip_map.get(schedule_state, "chip-neutral"),
                },
            ]
            if show_consultants:
                consultants = ", ".join(
                    consultant.full_name for consultant in activity.consultants.all()
                )
                values.append(consultants or "-")
            if show_hours_total:
                values.append(self._format_decimal(activity.hours))
            if show_hours_contingency:
                values.append(self._format_decimal(activity.hours_contingency()))
            if show_hours_available:
                values.append(self._format_decimal(activity.hours_available()))
            values.append(
                "-"
                if activity.client_completed is None
                else ("Sim" if activity.client_completed else "Nao")
            )
            values.append(activity.client_comment or "-")

            actions = []
            if can_edit_delete:
                actions.append(
                    {
                        "label": "Editar",
                        "url": reverse(
                            "cadastros_web:project_activity_update", args=[activity.pk]
                        ),
                        "style": "ghost",
                    }
                )
                actions.append(
                    {
                        "label": "Excluir",
                        "url": reverse(
                            "cadastros_web:project_activity_delete", args=[activity.pk]
                        ),
                        "style": "outline",
                    }
                )
            if can_feedback and activity.client_visible:
                actions.append(
                    {
                        "label": "Atualizar status",
                        "url": reverse(
                            "cadastros_web:project_activity_feedback", args=[activity.pk]
                        ),
                        "style": "primary",
                    }
                )

            rows.append({"values": values, "actions": actions})

        projects = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).order_by("description")
        current_project = self.request.GET.get("project_id", "")
        create_url = reverse("cadastros_web:project_activity_create")
        generate_url = reverse("cadastros_web:project_activity_generate")
        if current_project:
            create_url = f"{create_url}?project_id={current_project}"
            generate_url = f"{generate_url}?project_id={current_project}"

        export_params = self.request.GET.copy()
        export_params.pop("page", None)
        export_params["export"] = "excel"
        export_excel_url = (
            f"{reverse('cadastros_web:project_activity_list')}?{export_params.urlencode()}"
        )
        export_params["export"] = "msproject"
        export_msproject_url = (
            f"{reverse('cadastros_web:project_activity_list')}?{export_params.urlencode()}"
        )

        extra_actions = []
        if self._can_export():
            extra_actions.extend(
                [
                    {"label": "Exportar Excel", "url": export_excel_url},
                    {"label": "Exportar MS Project", "url": export_msproject_url},
                ]
            )
        if can_create:
            extra_actions.append({"label": "Gerar do template", "url": generate_url})

        context.update(
            {
                "page_title": self.page_title,
                "list_title": "Atividades cadastradas",
                "table_headers": headers,
                "table_rows": rows,
                "show_actions": any(row["actions"] for row in rows),
                "show_create": can_create,
                "create_url": create_url,
                "extra_actions": extra_actions,
                "projects": projects,
                "current_project": current_project,
                "query": self.request.GET.get("q", ""),
                "querystring": self._build_querystring(),
                "column_count": len(headers)
                + (1 if any(row["actions"] for row in rows) else 0),
                "can_manage": can_manage,
            }
        )
        return context


class ProjectActivityCreateView(BaseCreateView):
    model = ProjectActivity
    form_class = ProjectActivityForm
    template_name = "restricted/project_activity_form.html"
    page_title = "Nova atividade"
    submit_label = "Salvar atividade"
    cancel_url_name = "cadastros_web:project_activity_list"
    success_url = reverse_lazy("cadastros_web:project_activity_list")
    full_width_fields = ("activity", "subactivities")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def _get_next_seq(self, project_id: int) -> int:
        max_seq = (
            ProjectActivity.objects.filter(project_id=project_id)
            .aggregate(max_seq=models.Max("seq"))
            .get("max_seq")
            or 0
        )
        return int(max_seq) + 1

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project_id")
        if project_id:
            initial["project"] = project_id
            if not initial.get("seq"):
                initial["seq"] = self._get_next_seq(int(project_id))
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        project_id = self.request.GET.get("project_id")
        if project_id:
            form.fields["project"].widget = forms.HiddenInput()
        return form

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        form = context.get("form")
        consultants_field = getattr(form, "fields", {}).get("consultants") if form else None
        if consultants_field:
            consultants = list(consultants_field.queryset)
            context["consultant_rate_map"] = _build_consultant_rate_map(consultants)
        context.setdefault("predecessor_options", [])
        context.setdefault("selected_predecessor_ids", [])
        if form and "predecessors" in getattr(form, "fields", {}):
            raw_selected = form["predecessors"].value() if hasattr(form, "__getitem__") else []
            selected_ids = [int(item) for item in (raw_selected or [])]
            context.update(
                {
                    "predecessor_options": list(form.fields["predecessors"].queryset),
                    "selected_predecessor_ids": selected_ids,
                }
            )
        context.setdefault("subactivity_items", [])
        if form and "subactivities" in getattr(form, "fields", {}):
            raw_items = form["subactivities"].value() if hasattr(form, "__getitem__") else []
            if isinstance(raw_items, str):
                raw_items = [raw_items]
            items = [str(item).strip() for item in (raw_items or []) if str(item).strip()]
            context["subactivity_items"] = items
        return context

    def form_valid(self, form):
        cleaned = form.cleaned_data
        project = cleaned["project"]
        seq = cleaned.get("seq")
        if seq:
            if ProjectActivity.objects.filter(project=project, seq=seq).exists():
                form.add_error("seq", "Sequencia ja utilizada no projeto.")
                return self.form_invalid(form)
        else:
            form.instance.seq = self._get_next_seq(project.id)

        predecessors = list(cleaned.get("predecessors") or [])
        form.instance.seq_predecessor = (
            min((item.seq for item in predecessors), default=None) if predecessors else None
        )
        response = super().form_valid(form)
        _sync_subactivity_items(self.object, cleaned.get("subactivities_list") or [])
        notify_consultant_activity_assigned(
            self.object,
            list(self.object.consultants.all()),
        )
        _warn_project_activity_overage(self.object.project, self.request.user, self.request)
        return response


class ProjectActivityUpdateView(BaseUpdateView):
    model = ProjectActivity
    form_class = ProjectActivityForm
    template_name = "restricted/project_activity_form.html"
    page_title = "Editar atividade"
    submit_label = "Salvar atividade"
    cancel_url_name = "cadastros_web:project_activity_list"
    success_url = reverse_lazy("cadastros_web:project_activity_list")
    full_width_fields = ("activity", "subactivities")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
    )

    def get_queryset(self):
        queryset = ProjectActivity.objects.select_related("project")
        return filter_activities_for_user(queryset, self.request.user)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        form = context.get("form")
        consultants_field = getattr(form, "fields", {}).get("consultants") if form else None
        if consultants_field:
            consultants = list(consultants_field.queryset)
            context["consultant_rate_map"] = _build_consultant_rate_map(consultants)
        context.setdefault("predecessor_options", [])
        context.setdefault("selected_predecessor_ids", [])
        if form and "predecessors" in getattr(form, "fields", {}):
            raw_selected = form["predecessors"].value() if hasattr(form, "__getitem__") else []
            selected_ids = [int(item) for item in (raw_selected or [])]
            context.update(
                {
                    "predecessor_options": list(form.fields["predecessors"].queryset),
                    "selected_predecessor_ids": selected_ids,
                }
            )
        context.setdefault("subactivity_items", [])
        if form and "subactivities" in getattr(form, "fields", {}):
            raw_items = form["subactivities"].value() if hasattr(form, "__getitem__") else []
            if isinstance(raw_items, str):
                raw_items = [raw_items]
            items = [str(item).strip() for item in (raw_items or []) if str(item).strip()]
            context["subactivity_items"] = items
        return context

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        return form

    def form_valid(self, form):
        previous_ids = set(self.object.consultants.values_list("id", flat=True))
        predecessors = list(form.cleaned_data.get("predecessors") or [])
        form.instance.seq_predecessor = (
            min((item.seq for item in predecessors), default=None) if predecessors else None
        )
        response = super().form_valid(form)
        _sync_subactivity_items(self.object, form.cleaned_data.get("subactivities_list") or [])
        current_ids = set(self.object.consultants.values_list("id", flat=True))
        new_ids = current_ids - previous_ids
        if new_ids:
            consultants = list(Consultant.objects.filter(id__in=new_ids))
            notify_consultant_activity_assigned(self.object, consultants)
        return response


class ProjectActivityDeleteView(BaseDeleteView):
    model = ProjectActivity
    cancel_url_name = "cadastros_web:project_activity_list"
    success_url = reverse_lazy("cadastros_web:project_activity_list")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
    )

    def get_queryset(self):
        queryset = ProjectActivity.objects.select_related("project")
        return filter_activities_for_user(queryset, self.request.user)


class ProjectActivityGenerateView(BaseFormView, FormView):
    template_name = "restricted/form.html"
    form_class = ProjectActivityGenerateForm
    page_title = "Gerar atividades do template"
    submit_label = "Gerar atividades"
    cancel_url_name = "cadastros_web:project_activity_list"
    success_url = reverse_lazy("cadastros_web:project_activity_list")
    full_width_fields = ("template", "replace_existing")
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
    )

    def get_initial(self):
        initial = super().get_initial()
        project_id = self.request.GET.get("project_id")
        if project_id:
            initial["project"] = project_id
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        project_ids = filter_projects_for_user(
            Project.objects.all(), self.request.user
        ).values_list("id", flat=True)
        form.fields["project"].queryset = Project.objects.filter(id__in=project_ids)
        project_id = self.request.GET.get("project_id")
        if project_id:
            form.fields["project"].widget = forms.HiddenInput()
        return form

    def form_valid(self, form):
        project = form.cleaned_data["project"]
        template = form.cleaned_data["template"]
        replace_existing = form.cleaned_data.get("replace_existing")
        if replace_existing:
            ProjectActivity.objects.filter(project=project).delete()
        elif ProjectActivity.objects.filter(project=project).exists():
            form.add_error(
                None,
                "O projeto ja possui atividades. Use a opcao de substituir.",
            )
            return self.form_invalid(form)

        items = list(
            template.items.select_related(
                "phase",
                "product",
                "module",
                "submodule",
            ).order_by("seq")
        )
        if not items:
            form.add_error("template", "Template selecionado nao possui itens.")
            return self.form_invalid(form)

        groups: dict[
            tuple[int, int, int, int, str],
            dict[str, Any],
        ] = {}
        for item in items:
            key = (
                item.phase_id,
                item.product_id,
                item.module_id,
                item.submodule_id,
                item.activity,
            )
            entry = groups.get(key)
            if entry is None:
                entry = {
                    "items": [],
                    "min_seq": item.seq or 0,
                    "days": Decimal("0.00"),
                    "hours": Decimal("0.00"),
                    "pred_seqs": set(),
                }
                groups[key] = entry
            entry["items"].append(item)
            if item.seq is not None and item.seq < entry["min_seq"]:
                entry["min_seq"] = item.seq
            entry["days"] += Decimal(item.days or 0)
            entry["hours"] += item.hours or Decimal("0.00")
            if item.seq_predecessor:
                entry["pred_seqs"].add(item.seq_predecessor)

        sorted_groups = sorted(groups.values(), key=lambda entry: entry["min_seq"])
        original_seq_to_new: dict[int, int] = {}
        for idx, entry in enumerate(sorted_groups, start=1):
            entry["new_seq"] = idx
            for item in entry["items"]:
                if item.seq is not None:
                    original_seq_to_new[item.seq] = idx

        for entry in sorted_groups:
            mapped_pred = {
                original_seq_to_new.get(seq)
                for seq in entry["pred_seqs"]
                if seq in original_seq_to_new
            }
            mapped_pred.discard(None)
            mapped_pred.discard(entry["new_seq"])
            entry["mapped_pred_seqs"] = sorted(mapped_pred)

        activities = []
        for entry in sorted_groups:
            primary_item = min(entry["items"], key=lambda item: item.seq or 0)
            mapped_pred_seqs = entry.get("mapped_pred_seqs") or []
            activities.append(
                ProjectActivity(
                    project=project,
                    template_item=primary_item,
                    seq=entry["new_seq"],
                    seq_predecessor=min(mapped_pred_seqs) if mapped_pred_seqs else None,
                    phase=primary_item.phase,
                    product=primary_item.product,
                    module=primary_item.module,
                    submodule=primary_item.submodule,
                    activity=primary_item.activity,
                    days=entry["days"],
                    hours=entry["hours"],
                    status=ActivityStatus.PLANNED,
                )
            )
        ProjectActivity.objects.bulk_create(activities)
        seq_map = {
            activity.seq: activity
            for activity in ProjectActivity.objects.filter(
                project=project,
                seq__in=[entry["new_seq"] for entry in sorted_groups],
            )
        }
        subactivity_records = []
        for entry in sorted_groups:
            activity = seq_map.get(entry["new_seq"])
            if not activity:
                continue
            subactivities = []
            seen = set()
            for item in entry["items"]:
                subactivity = (item.subactivity or "").strip()
                if not subactivity:
                    continue
                key = subactivity.lower()
                if key in seen:
                    continue
                seen.add(key)
                subactivities.append(subactivity)
            if subactivities:
                activity.subactivity = subactivities[0]
                activity.save(update_fields=["subactivity"])
                for index, subactivity in enumerate(subactivities, start=1):
                    subactivity_records.append(
                        ProjectActivitySubactivity(
                            activity=activity,
                            description=subactivity,
                            order=index,
                        )
                    )
            mapped_pred_seqs = entry.get("mapped_pred_seqs") or []
            if mapped_pred_seqs:
                predecessors = [
                    seq_map[pred_seq]
                    for pred_seq in mapped_pred_seqs
                    if pred_seq in seq_map
                ]
                if predecessors:
                    activity.predecessors.set(predecessors)
        if subactivity_records:
            ProjectActivitySubactivity.objects.bulk_create(subactivity_records)
        _warn_project_activity_overage(project, self.request.user, self.request)
        messages.success(
            self.request,
            f"{len(activities)} atividades geradas. Ajuste consultores e visibilidade do cliente.",
        )
        return super().form_valid(form)


class ProjectActivityFeedbackView(BaseUpdateView):
    model = ProjectActivity
    form_class = ProjectActivityFeedbackForm
    page_title = "Atualizar status do cliente"
    submit_label = "Salvar feedback"
    cancel_url_name = "cadastros_web:project_activity_list"
    success_url = reverse_lazy("cadastros_web:project_activity_list")
    full_width_fields = ("client_comment",)
    allowed_roles = (UserRole.CLIENT, UserRole.ADMIN)

    def get_queryset(self):
        queryset = ProjectActivity.objects.select_related("project")
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset
        return queryset.filter(
            project__client_user=self.request.user, client_visible=True
        )

    def form_valid(self, form):
        activity = form.instance
        activity.client_feedback_by = self.request.user
        activity.client_feedback_at = timezone.now()
        return super().form_valid(form)


class KnowledgeCategoryListView(BaseListView):
    model = KnowledgeCategory
    page_title = "Categorias de conhecimento"
    list_title = "Categorias cadastradas"
    search_placeholder = "Buscar por categoria"
    ordering = ("name",)
    table_headers = ("Categoria", "Status")
    table_fields = ("name", "status")
    search_fields = ("name",)
    create_url_name = "cadastros_web:knowledge_category_create"
    edit_url_name = "cadastros_web:knowledge_category_update"
    delete_url_name = "cadastros_web:knowledge_category_delete"
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)


class KnowledgeCategoryCreateView(BaseCreateView):
    model = KnowledgeCategory
    form_class = KnowledgeCategoryForm
    page_title = "Nova categoria"
    submit_label = "Salvar categoria"
    cancel_url_name = "cadastros_web:knowledge_category_list"
    success_url = reverse_lazy("cadastros_web:knowledge_category_list")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)


class KnowledgeCategoryUpdateView(BaseUpdateView):
    model = KnowledgeCategory
    form_class = KnowledgeCategoryForm
    page_title = "Editar categoria"
    submit_label = "Salvar categoria"
    cancel_url_name = "cadastros_web:knowledge_category_list"
    success_url = reverse_lazy("cadastros_web:knowledge_category_list")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)


class KnowledgeCategoryDeleteView(BaseDeleteView):
    model = KnowledgeCategory
    page_title = "Excluir categoria"
    cancel_url_name = "cadastros_web:knowledge_category_list"
    success_url = reverse_lazy("cadastros_web:knowledge_category_list")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)


class KnowledgePostListView(BaseListView):
    model = KnowledgePost
    queryset = KnowledgePost.objects.select_related("category", "author")
    page_title = "Conhecimento"
    list_title = "Base de conhecimento"
    search_placeholder = "Buscar por titulo, categoria ou conteudo"
    ordering = ("-updated_at", "-created_at")
    table_headers = ("Titulo", "Categoria", "Autor", "Atualizado", "Status")
    table_fields = ("title", "category", "author", "updated_at", "status")
    search_fields = (
        "title",
        "content",
        "category__name",
        "author__username",
        "author__email",
    )
    create_url_name = "cadastros_web:knowledge_post_create"
    edit_url_name = "cadastros_web:knowledge_post_update"
    delete_url_name = "cadastros_web:knowledge_post_delete"
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def get_row_actions(self, obj: KnowledgePost) -> list[dict[str, str]]:
        return [
            {
                "label": "Ver",
                "url": reverse("cadastros_web:knowledge_post_detail", args=[obj.pk]),
            }
        ]

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        role = resolve_user_role(self.request.user)
        for obj, row in zip(context.get("object_list", []), context.get("table_rows", [])):
            can_edit = role == UserRole.ADMIN or obj.author_id == self.request.user.id
            if not can_edit:
                row["edit_url"] = None
                row["delete_url"] = None
        return context


class KnowledgePostDetailView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/knowledge_post_detail.html"
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if resolve_user_role(request.user) not in self.allowed_roles:
            raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        post = get_object_or_404(
            KnowledgePost.objects.select_related("category", "author").prefetch_related(
                "attachments"
            ),
            pk=self.kwargs.get("pk"),
        )
        role = resolve_user_role(self.request.user)
        can_edit = role == UserRole.ADMIN or post.author_id == self.request.user.id
        context.update(
            {
                "page_title": post.title,
                "post": post,
                "attachments": post.attachments.all(),
                "can_edit": can_edit,
                "edit_url": reverse("cadastros_web:knowledge_post_update", args=[post.pk]),
                "back_url": reverse("cadastros_web:knowledge_post_list"),
            }
        )
        return context


class KnowledgePostCreateView(BaseCreateView):
    model = KnowledgePost
    form_class = KnowledgePostForm
    template_name = "restricted/knowledge_post_form.html"
    page_title = "Novo post"
    submit_label = "Publicar"
    cancel_url_name = "cadastros_web:knowledge_post_list"
    success_url = reverse_lazy("cadastros_web:knowledge_post_list")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def form_valid(self, form):
        form.instance.author = self.request.user
        response = super().form_valid(form)
        _save_knowledge_attachments(
            self.object,
            self.request.FILES.getlist("attachments"),
            self.request.user,
        )
        return response


class KnowledgePostUpdateView(BaseUpdateView):
    model = KnowledgePost
    form_class = KnowledgePostForm
    template_name = "restricted/knowledge_post_form.html"
    page_title = "Editar post"
    submit_label = "Salvar"
    cancel_url_name = "cadastros_web:knowledge_post_list"
    success_url = reverse_lazy("cadastros_web:knowledge_post_list")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def get_queryset(self):
        queryset = KnowledgePost.objects.select_related("category", "author")
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset
        return queryset.filter(author=self.request.user)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["attachments"] = self.object.attachments.all()
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        _save_knowledge_attachments(
            self.object,
            self.request.FILES.getlist("attachments"),
            self.request.user,
        )
        return response


class KnowledgePostDeleteView(BaseDeleteView):
    model = KnowledgePost
    page_title = "Excluir post"
    cancel_url_name = "cadastros_web:knowledge_post_list"
    success_url = reverse_lazy("cadastros_web:knowledge_post_list")
    allowed_roles = (UserRole.ADMIN, UserRole.CONSULTANT)

    def get_queryset(self):
        queryset = KnowledgePost.objects.all()
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset
        return queryset.filter(author=self.request.user)


class TicketListView(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = "restricted/ticket_list.html"
    paginate_by = 15
    page_title = "Chamados"

    def _format_datetime(self, value) -> str:
        if not value:
            return "-"
        local = timezone.localtime(value) if timezone.is_aware(value) else value
        return local.strftime("%d/%m/%Y %H:%M")

    def _build_querystring(self) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        if not params:
            return ""
        return "&" + params.urlencode()

    def get_queryset(self):
        queryset = (
            Ticket.objects.select_related(
                "project",
                "project__project_client",
                "assigned_to",
                "created_by",
                "consultant_responsible",
            )
            .order_by("-updated_at", "-created_at")
        )
        queryset = _filter_tickets_for_user(queryset, self.request.user)

        params = self.request.GET
        query = params.get("q", "").strip()
        if query:
            queryset = queryset.filter(
                Q(title__icontains=query) | Q(description__icontains=query)
            )
        if params.get("status"):
            queryset = queryset.filter(status=params["status"])
        if params.get("project_id"):
            queryset = queryset.filter(project_id=params["project_id"])
        if params.get("client_id"):
            queryset = queryset.filter(project__project_client_id=params["client_id"])
        if params.get("consultant_id"):
            queryset = queryset.filter(consultant_responsible_id=params["consultant_id"])
        if params.get("assigned_to_id"):
            queryset = queryset.filter(assigned_to_id=params["assigned_to_id"])
        return queryset

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        rows = []
        for ticket in context["object_list"]:
            rows.append(
                {
                    "id": ticket.id,
                    "title": ticket.title,
                    "ticket_type": ticket.ticket_type,
                    "ticket_type_label": ticket.get_ticket_type_display(),
                    "criticality": ticket.criticality,
                    "criticality_label": ticket.get_criticality_display(),
                    "status": ticket.get_status_display(),
                    "status_key": ticket.status,
                    "created_at": self._format_datetime(ticket.created_at),
                    "updated_at": self._format_datetime(ticket.updated_at),
                    "detail_url": reverse("cadastros_web:ticket_detail", args=[ticket.pk]),
                }
            )

        projects = filter_projects_for_user(
            Project.objects.order_by("description"),
            self.request.user,
        )
        clients = (
            Client.objects.filter(projects__in=projects)
            .distinct()
            .order_by("name")
        )
        consultants = (
            Consultant.objects.filter(project_activities__project__in=projects)
            .distinct()
            .order_by("full_name")
        )
        assignees = _get_assignable_users()

        context.update(
            {
                "page_title": self.page_title,
                "list_title": "Chamados registrados",
                "table_rows": rows,
                "create_url": reverse("cadastros_web:ticket_create"),
                "dashboard_url": reverse("cadastros_web:ticket_dashboard"),
                "querystring": self._build_querystring(),
                "projects": projects,
                "clients": clients,
                "consultants": consultants,
                "assignees": assignees,
                "current_filters": {
                    "query": self.request.GET.get("q", ""),
                    "status": self.request.GET.get("status", ""),
                    "project_id": self.request.GET.get("project_id", ""),
                    "client_id": self.request.GET.get("client_id", ""),
                    "consultant_id": self.request.GET.get("consultant_id", ""),
                    "assigned_to_id": self.request.GET.get("assigned_to_id", ""),
                },
            }
        )
        return context


class TicketCreateView(BaseCreateView):
    model = Ticket
    form_class = TicketForm
    template_name = "restricted/ticket_form.html"
    page_title = "Novo chamado"
    submit_label = "Abrir chamado"
    cancel_url_name = "cadastros_web:ticket_list"
    success_url = reverse_lazy("cadastros_web:ticket_list")
    success_message = "Chamado aberto com sucesso."
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
        UserRole.CONSULTANT,
        UserRole.CLIENT,
    )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        projects = filter_projects_for_user(
            Project.objects.order_by("description"),
            self.request.user,
        )
        form.fields["project"].queryset = projects

        project_id = self.request.POST.get("project") or self.request.GET.get("project")
        project = projects.filter(id=project_id).first() if project_id else None

        activities = filter_activities_for_user(
            ProjectActivity.objects.select_related("project"),
            self.request.user,
        )
        if project:
            activities = activities.filter(project=project)
        form.fields["activity"].queryset = activities.order_by(
            "project__description",
            "seq",
        )

        consultants = Consultant.objects.order_by("full_name")
        if project:
            consultants = consultants.filter(
                project_activities__project=project
            ).distinct()
        form.fields["consultant_responsible"].queryset = consultants
        form.fields["assigned_to"].queryset = _get_assignable_users(project)
        return form

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.status = TicketStatus.OPEN
        form.instance.closed_at = None
        form.instance.closed_by = None
        if form.instance.assigned_to and not form.instance.consultant_responsible:
            consultant = Consultant.objects.filter(
                user=form.instance.assigned_to
            ).first()
            if consultant:
                form.instance.consultant_responsible = consultant
        response = super().form_valid(form)
        _save_ticket_attachments(
            self.object,
            self.request.FILES.getlist("attachments"),
        )
        notify_ticket_created(self.object)
        return response


class TicketDetailView(LoginRequiredMixin, FormView):
    form_class = TicketReplyForm
    template_name = "restricted/ticket_detail.html"

    def dispatch(self, request, *args, **kwargs):
        ticket_qs = Ticket.objects.select_related(
            "project",
            "project__project_client",
            "assigned_to",
            "created_by",
            "consultant_responsible",
        ).prefetch_related(
            "attachments",
            "replies__author",
            "replies__attachments",
        )
        ticket_qs = _filter_tickets_for_user(ticket_qs, request.user)
        self.ticket = get_object_or_404(ticket_qs, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def _status_chip(self) -> str:
        return "chip-ok" if self.ticket.status == TicketStatus.CLOSED else "chip-warn"

    def _can_close(self) -> bool:
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return True
        return self.ticket.created_by_id == self.request.user.id or (
            self.ticket.assigned_to_id == self.request.user.id
        )

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "page_title": f"Chamado #{self.ticket.id}",
                "ticket": self.ticket,
                "status_chip": self._status_chip(),
                "attachments": self.ticket.attachments.order_by("-created_at"),
                "replies": self.ticket.replies.all(),
                "can_reply": self.ticket.status == TicketStatus.OPEN,
                "can_close": self._can_close(),
                "close_url": reverse(
                    "cadastros_web:ticket_close",
                    args=[self.ticket.pk],
                ),
            }
        )
        return context

    def form_valid(self, form):
        if self.ticket.status == TicketStatus.CLOSED:
            messages.error(self.request, "Chamado encerrado nao aceita novas respostas.")
            return self.form_invalid(form)
        reply = form.save(commit=False)
        reply.ticket = self.ticket
        reply.author = self.request.user
        reply.save()
        _save_ticket_reply_attachments(
            reply,
            self.request.FILES.getlist("attachments"),
        )
        self.ticket.save(update_fields=["updated_at"])
        notify_ticket_reply(self.ticket, reply)
        messages.success(self.request, "Resposta enviada com sucesso.")
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse("cadastros_web:ticket_detail", args=[self.ticket.pk])


class TicketCloseView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        ticket_qs = _filter_tickets_for_user(Ticket.objects.all(), request.user)
        ticket = get_object_or_404(ticket_qs, pk=kwargs["pk"])
        role = resolve_user_role(request.user)
        can_close = role == UserRole.ADMIN or ticket.created_by_id == request.user.id
        if not can_close and ticket.assigned_to_id != request.user.id:
            raise PermissionDenied("Perfil sem acesso a este chamado.")
        if ticket.status == TicketStatus.CLOSED:
            messages.info(request, "Chamado ja estava encerrado.")
            return HttpResponseRedirect(
                reverse("cadastros_web:ticket_detail", args=[ticket.pk])
            )
        ticket.status = TicketStatus.CLOSED
        ticket.closed_at = timezone.now()
        ticket.closed_by = request.user
        ticket.save(update_fields=["status", "closed_at", "closed_by", "updated_at"])
        notify_ticket_closed(ticket)
        messages.success(request, "Chamado encerrado com sucesso.")
        return HttpResponseRedirect(
            reverse("cadastros_web:ticket_detail", args=[ticket.pk])
        )


class TicketDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/ticket_dashboard.html"
    page_title = "Dashboard de chamados"
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
        UserRole.CONSULTANT,
        UserRole.CLIENT,
    )

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def _average_timedelta(self, values: list[timedelta]) -> timedelta | None:
        if not values:
            return None
        total_seconds = sum((delta.total_seconds() for delta in values), 0.0)
        return timedelta(seconds=(total_seconds / len(values)))

    def _parse_date(self, value: str | None) -> date | None:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    def _format_duration(self, delta: timedelta | None) -> str:
        if not delta:
            return "-"
        hours = Decimal(delta.total_seconds()) / Decimal("3600")
        return f"{formats.number_format(hours, decimal_pos=1, use_l10n=True)}h"

    def _build_counts(self, items, label_func, empty_label: str) -> list[dict[str, Any]]:
        counts: dict[str, int] = defaultdict(int)
        for item in items:
            label = label_func(item) or empty_label
            counts[label] += 1
        ordered = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        max_count = max((count for _, count in ordered), default=0)
        rows = []
        for label, count in ordered:
            percent = int(round((count / max_count) * 100)) if max_count else 0
            rows.append({"label": label, "count": count, "percent": percent})
        return rows

    def _build_open_by_day(self, tickets: list[Ticket], days: int = 14):
        today = timezone.localdate()
        days_list = [today - timedelta(days=offset) for offset in range(days - 1, -1, -1)]
        counts = {day: 0 for day in days_list}
        for ticket in tickets:
            created = timezone.localtime(ticket.created_at).date()
            if created in counts:
                counts[created] += 1
        max_count = max(counts.values()) if counts else 0
        return [
            {
                "label": day.strftime("%d/%m"),
                "count": counts[day],
                "percent": int(round((counts[day] / max_count) * 100)) if max_count else 0,
            }
            for day in days_list
        ]

    def _first_response_deltas(self, tickets: list[Ticket]) -> list[timedelta]:
        deltas: list[timedelta] = []
        for ticket in tickets:
            replies = [
                reply
                for reply in ticket.replies.all()
                if reply.author_id and reply.author_id != ticket.created_by_id
            ]
            if not replies:
                continue
            replies.sort(key=lambda reply: reply.created_at)
            delta = replies[0].created_at - ticket.created_at
            if delta.total_seconds() >= 0:
                deltas.append(delta)
        return deltas

    def _solution_deltas(self, tickets: list[Ticket]) -> list[timedelta]:
        deltas: list[timedelta] = []
        for ticket in tickets:
            if ticket.status != TicketStatus.CLOSED or not ticket.closed_at:
                continue
            delta = ticket.closed_at - ticket.created_at
            if delta.total_seconds() >= 0:
                deltas.append(delta)
        return deltas

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        tickets_qs = Ticket.objects.select_related(
            "project",
            "project__project_client",
            "consultant_responsible",
        ).prefetch_related("replies")
        tickets_qs = _filter_tickets_for_user(tickets_qs, self.request.user)
        params = self.request.GET
        project_id = params.get("project_id")
        client_id = params.get("client_id")
        consultant_id = params.get("consultant_id")
        period_start = self._parse_date(params.get("period_start"))
        period_end = self._parse_date(params.get("period_end"))
        if period_start and period_end and period_end < period_start:
            period_start, period_end = period_end, period_start

        if project_id:
            tickets_qs = tickets_qs.filter(project_id=project_id)
        if client_id:
            tickets_qs = tickets_qs.filter(project__project_client_id=client_id)
        if consultant_id:
            tickets_qs = tickets_qs.filter(consultant_responsible_id=consultant_id)
        if period_start:
            tickets_qs = tickets_qs.filter(created_at__date__gte=period_start)
        if period_end:
            tickets_qs = tickets_qs.filter(created_at__date__lte=period_end)

        tickets = list(tickets_qs)
        open_tickets = [ticket for ticket in tickets if ticket.status == TicketStatus.OPEN]
        closed_tickets = [
            ticket
            for ticket in tickets
            if ticket.status == TicketStatus.CLOSED and ticket.closed_at
        ]

        response_deltas = self._first_response_deltas(tickets)
        solution_deltas = self._solution_deltas(tickets)

        projects = filter_projects_for_user(
            Project.objects.order_by("description"),
            self.request.user,
        )
        clients = (
            Client.objects.filter(projects__in=projects)
            .distinct()
            .order_by("name")
        )
        consultants = (
            Consultant.objects.filter(project_activities__project__in=projects)
            .distinct()
            .order_by("full_name")
        )

        context.update(
            {
                "page_title": self.page_title,
                "ticket_total": len(tickets),
                "ticket_open": len(open_tickets),
                "ticket_closed": len(closed_tickets),
                "first_response_avg": self._format_duration(
                    self._average_timedelta(response_deltas)
                ),
                "first_response_count": len(response_deltas),
                "solution_avg": self._format_duration(
                    self._average_timedelta(solution_deltas)
                ),
                "solution_count": len(solution_deltas),
                "open_by_day": self._build_open_by_day(tickets),
                "open_by_client": self._build_counts(
                    open_tickets,
                    lambda ticket: (
                        ticket.project.project_client.name
                        if ticket.project and ticket.project.project_client
                        else None
                    ),
                    "Sem cliente",
                ),
                "open_by_project": self._build_counts(
                    open_tickets,
                    lambda ticket: ticket.project.description if ticket.project else None,
                    "Sem projeto",
                ),
                "open_by_consultant": self._build_counts(
                    open_tickets,
                    lambda ticket: (
                        ticket.consultant_responsible.full_name
                        if ticket.consultant_responsible
                        else None
                    ),
                    "Sem consultor",
                ),
                "projects": projects,
                "clients": clients,
                "consultants": consultants,
                "current_filters": {
                    "project_id": project_id or "",
                    "client_id": client_id or "",
                    "consultant_id": consultant_id or "",
                    "period_start": params.get("period_start", ""),
                    "period_end": params.get("period_end", ""),
                },
            }
        )
        return context


class BillingClosureView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/billing_closure.html"
    page_title = "Fechamento de faturamento"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _parse_date(value: str | None) -> date | None:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    @staticmethod
    def _format_decimal(value: Decimal) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value: Decimal) -> str:
        return f"R$ {self._format_decimal(value)}"

    @staticmethod
    def _resolve_account_plan_item(code: str) -> AccountPlanTemplateItem | None:
        return (
            AccountPlanTemplateItem.objects.filter(
                code=code,
                status=StatusChoices.ACTIVE,
                is_analytic=True,
            )
            .order_by("id")
            .first()
        )

    def _parse_filters(self, params) -> dict[str, Any]:
        period_start = self._parse_date(params.get("period_start", "").strip())
        period_end = self._parse_date(params.get("period_end", "").strip())
        billing_client_id = params.get("billing_client_id", "").strip()
        project_id = params.get("project_id", "").strip()
        consultant_ids = [value for value in params.getlist("consultant_ids") if value]
        invoice_number = params.get("invoice_number", "").strip()
        return {
            "period_start": period_start,
            "period_end": period_end,
            "billing_client_id": billing_client_id,
            "project_id": project_id,
            "consultant_ids": consultant_ids,
            "invoice_number": invoice_number,
        }

    @staticmethod
    def _has_filter_params(params) -> bool:
        if params.get("period_start", "").strip():
            return True
        if params.get("period_end", "").strip():
            return True
        if params.get("billing_client_id", "").strip():
            return True
        if params.get("project_id", "").strip():
            return True
        if params.get("invoice_number", "").strip():
            return True
        return any(value for value in params.getlist("consultant_ids") if value)

    def _filters_ready(self, filters: dict[str, Any]) -> bool:
        return bool(
            filters["period_start"]
            and filters["period_end"]
            and filters["billing_client_id"]
        )

    def _get_eligible_entries(self, filters: dict[str, Any]):
        queryset = (
            TimeEntry.objects.select_related(
                "consultant",
                "activity",
                "activity__project",
                "activity__project__billing_client",
            )
            .filter(status=TimeEntryStatus.APPROVED, billing_invoice__isnull=True)
            .order_by("consultant__full_name", "start_date", "created_at")
        )
        if filters["billing_client_id"]:
            queryset = queryset.filter(
                activity__project__billing_client_id=filters["billing_client_id"]
            )
        if filters["project_id"]:
            queryset = queryset.filter(activity__project_id=filters["project_id"])
        if filters["consultant_ids"]:
            queryset = queryset.filter(consultant_id__in=filters["consultant_ids"])
        if filters["period_start"]:
            queryset = queryset.filter(end_date__gte=filters["period_start"])
        if filters["period_end"]:
            queryset = queryset.filter(start_date__lte=filters["period_end"])
        return list(queryset)

    def _group_entries(
        self, entries: list[TimeEntry]
    ) -> tuple[list[dict[str, Any]], dict[str, Decimal]]:
        grouped: dict[int, dict[str, Any]] = {}
        totals = {"hours": Decimal("0.00"), "value": Decimal("0.00")}
        for entry in entries:
            hours = entry.total_hours or Decimal("0.00")
            project_rate = entry.activity.project.hourly_rate or Decimal("0.00")
            entry_value = (hours * project_rate).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            totals["hours"] += hours
            totals["value"] += entry_value
            group = grouped.get(entry.consultant_id)
            if not group:
                group = {
                    "consultant": entry.consultant,
                    "hours": Decimal("0.00"),
                    "total": Decimal("0.00"),
                    "rate": Decimal("0.00"),
                }
                grouped[entry.consultant_id] = group
            group["hours"] += hours
            group["total"] += entry_value

        groups = list(grouped.values())
        for group in groups:
            if group["hours"] > 0:
                group["rate"] = (group["total"] / group["hours"]).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
        groups.sort(key=lambda item: (item["consultant"].full_name or "").lower())
        return groups, totals

    def _build_preview_items(self, groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
        items = []
        for group in groups:
            items.append(
                {
                    "consultant": group["consultant"].full_name,
                    "hours": group["hours"],
                    "hours_display": self._format_decimal(group["hours"]),
                    "rate": group["rate"],
                    "rate_display": self._format_currency(group["rate"]),
                    "total": group["total"],
                    "total_display": self._format_currency(group["total"]),
                }
            )
        return items

    def _build_report_url(self, filters: dict[str, Any]) -> str:
        params = {}
        if filters["period_start"]:
            params["period_start"] = filters["period_start"].isoformat()
        if filters["period_end"]:
            params["period_end"] = filters["period_end"].isoformat()
        if filters["project_id"]:
            params["project_id"] = filters["project_id"]
        if len(filters["consultant_ids"]) == 1:
            params["consultant_id"] = filters["consultant_ids"][0]
        base = reverse("cadastros_web:time_entry_report")
        if not params:
            return base
        return f"{base}?{urlencode(params)}"

    def _build_context(self, params) -> dict[str, Any]:
        snapshot_invoice = None
        snapshot_id = params.get("invoice_id", "").strip()
        if snapshot_id and not self._has_filter_params(params):
            try:
                snapshot_invoice = BillingInvoice.objects.select_related(
                    "billing_client", "project"
                ).prefetch_related("items__consultant").get(pk=int(snapshot_id))
            except (BillingInvoice.DoesNotExist, ValueError):
                snapshot_invoice = None

        if snapshot_invoice:
            items = list(
                snapshot_invoice.items.select_related("consultant").order_by(
                    "consultant__full_name", "id"
                )
            )
            preview_groups = [
                {
                    "consultant": item.consultant,
                    "hours": item.hours or Decimal("0.00"),
                    "total": item.total or Decimal("0.00"),
                    "rate": item.rate or Decimal("0.00"),
                }
                for item in items
            ]
            totals = {
                "hours": snapshot_invoice.total_hours or Decimal("0.00"),
                "value": snapshot_invoice.total_value or Decimal("0.00"),
            }
            filters = {
                "period_start": snapshot_invoice.period_start,
                "period_end": snapshot_invoice.period_end,
                "billing_client_id": str(snapshot_invoice.billing_client_id),
                "project_id": str(snapshot_invoice.project_id or ""),
                "consultant_ids": [str(item.consultant_id) for item in items],
                "invoice_number": snapshot_invoice.number,
            }
            preview_items = self._build_preview_items(preview_groups)
            preview_available = bool(preview_items)
            preview_message = ""
            snapshot_mode = True
        else:
            filters = self._parse_filters(params)
            preview_items = []
            preview_available = False
            preview_message = ""
            totals = {"hours": Decimal("0.00"), "value": Decimal("0.00")}
            snapshot_mode = False
            if self._filters_ready(filters):
                entries = self._get_eligible_entries(filters)
                preview_groups, totals = self._group_entries(entries)
                preview_items = self._build_preview_items(preview_groups)
                preview_available = bool(preview_items)
                if not preview_available:
                    preview_message = "Nenhum apontamento aprovado elegivel."
            else:
                preview_message = "Informe periodo e cliente de faturamento para ver o preview."

        current_filters = {
            "period_start": filters["period_start"].isoformat()
            if filters["period_start"]
            else "",
            "period_end": filters["period_end"].isoformat()
            if filters["period_end"]
            else "",
            "billing_client_id": filters["billing_client_id"],
            "project_id": filters["project_id"],
            "consultant_ids": filters["consultant_ids"],
            "invoice_number": filters["invoice_number"],
        }

        billing_clients = Client.objects.order_by("name")
        projects = Project.objects.select_related("billing_client").order_by("description")
        if filters["billing_client_id"]:
            projects = projects.filter(billing_client_id=filters["billing_client_id"])

        preview_badge = "Snapshot" if snapshot_mode else "Preview"
        preview_badge_class = "chip-neutral" if snapshot_mode else "chip-info"

        return {
            "page_title": self.page_title,
            "billing_clients": billing_clients,
            "projects": projects,
            "consultants": Consultant.objects.order_by("full_name"),
            "preview_items": preview_items,
            "preview_available": preview_available,
            "preview_message": preview_message,
            "preview_total_hours": self._format_decimal(totals["hours"]),
            "preview_total_value": self._format_currency(totals["value"]),
            "current_filters": current_filters,
            "preview_badge": preview_badge,
            "preview_badge_class": preview_badge_class,
            "report_url": self._build_report_url(filters),
        }

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(self._build_context(self.request.GET))
        return context

    def post(self, request, *args, **kwargs):
        params = request.POST
        filters = self._parse_filters(params)
        errors = []

        if not filters["invoice_number"]:
            errors.append("Informe o numero da fatura.")
        elif BillingInvoice.objects.filter(number=filters["invoice_number"]).exists():
            errors.append("Numero de fatura ja utilizado.")
        if not filters["billing_client_id"]:
            errors.append("Informe o cliente de faturamento.")
        if not filters["period_start"] or not filters["period_end"]:
            errors.append("Informe o periodo completo para faturamento.")
        if (
            filters["period_start"]
            and filters["period_end"]
            and filters["period_end"] < filters["period_start"]
        ):
            errors.append("Periodo final deve ser maior ou igual ao inicial.")

        billing_client = None
        if filters["billing_client_id"]:
            billing_client = Client.objects.filter(
                pk=filters["billing_client_id"]
            ).first()
            if not billing_client:
                errors.append("Cliente de faturamento invalido.")

        project = None
        if filters["project_id"]:
            project = Project.objects.filter(pk=filters["project_id"]).first()
            if not project:
                errors.append("Projeto invalido.")
            elif billing_client and project.billing_client_id != billing_client.id:
                errors.append("Projeto nao pertence ao cliente de faturamento selecionado.")

        if errors:
            for error in errors:
                messages.error(request, error)
            context = self._build_context(params)
            return self.render_to_response(context)

        entries = self._get_eligible_entries(filters)
        if not entries:
            messages.error(request, "Nenhum apontamento aprovado elegivel para o faturamento.")
            context = self._build_context(params)
            return self.render_to_response(context)

        groups, totals = self._group_entries(entries)
        total_hours = totals["hours"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total_value = totals["value"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        missing_suppliers = [
            group["consultant"]
            for group in groups
            if not getattr(group["consultant"], "supplier_id", None)
        ]
        if missing_suppliers:
            names = ", ".join(
                sorted({consultant.full_name for consultant in missing_suppliers})
            )
            messages.error(
                request,
                f"Consultores sem fornecedor vinculado: {names}.",
            )
            context = self._build_context(params)
            return self.render_to_response(context)

        payable_totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0.00"))
        missing_rates = set()
        for entry in entries:
            consultant = entry.consultant
            activity_rate = entry.activity.consultant_hourly_rate
            if activity_rate is None or activity_rate <= 0:
                missing_rates.add(consultant.full_name)
                continue
            hours = entry.total_hours or Decimal("0.00")
            payable_totals[consultant.id] += hours * activity_rate
        if missing_rates:
            names = ", ".join(sorted(missing_rates))
            messages.error(
                request,
                f"Consultores com atividades sem taxa hora definida: {names}.",
            )
            context = self._build_context(params)
            return self.render_to_response(context)

        with transaction.atomic():
            issue_date = timezone.localdate()
            due_date = issue_date + timedelta(days=7)
            receivable_plan_item = self._resolve_account_plan_item("1.01.01")
            payable_plan_item = self._resolve_account_plan_item("3.01.01")
            invoice = BillingInvoice.objects.create(
                number=filters["invoice_number"],
                billing_client=billing_client,
                project=project,
                period_start=filters["period_start"],
                period_end=filters["period_end"],
                total_hours=total_hours,
                total_value=total_value,
                payment_status=BillingPaymentStatus.UNPAID,
                created_by=request.user,
            )
            items = [
                BillingInvoiceItem(
                    invoice=invoice,
                    consultant=group["consultant"],
                    hours=group["hours"].quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    ),
                    rate=group["rate"],
                    total=group["total"].quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    ),
                )
                for group in groups
            ]
            BillingInvoiceItem.objects.bulk_create(items)
            TimeEntry.objects.filter(pk__in=[entry.pk for entry in entries]).update(
                billing_invoice=invoice,
                billing_invoice_number=invoice.number,
            )
            description = f"TITULO GERADO PELA FATURA {invoice.number}"
            payables = [
                AccountsPayable(
                    supplier=group["consultant"].supplier,
                    consultant=group["consultant"],
                    billing_invoice=invoice,
                    document_number=f"{invoice.number}-{group['consultant'].id}",
                    description=description,
                    issue_date=issue_date,
                    due_date=due_date,
                    amount=payable_totals.get(
                        group["consultant"].id, Decimal("0.00")
                    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                    account_plan_item=payable_plan_item,
                )
                for group in groups
            ]
            AccountsPayable.objects.bulk_create(payables)
            receivable_amount = total_value
            if project and project.hourly_rate is not None:
                receivable_amount = (project.hourly_rate * total_hours).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
            receivable = AccountsReceivable.objects.create(
                client=billing_client,
                billing_invoice=invoice,
                document_number=invoice.number,
                description=description,
                issue_date=issue_date,
                due_date=due_date,
                amount=receivable_amount,
                account_plan_item=receivable_plan_item,
            )

        def _notify_closure() -> None:
            notify_admin_receivable_created(receivable)
            for payable in payables:
                notify_admin_payable_created(payable)
                notify_consultant_payable_created(payable)
            for group in groups:
                notify_consultant_billing_closure(
                    group["consultant"],
                    filters["period_start"],
                    filters["period_end"],
                    group["hours"],
                    group["total"],
                    due_date,
                )

        transaction.on_commit(_notify_closure)

        messages.success(request, "Fatura gerada com sucesso.")
        return HttpResponseRedirect(
            f"{reverse('cadastros_web:billing_closure')}?invoice_id={invoice.id}"
        )


class BillingInvoiceReportView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/billing_invoice_report.html"
    page_title = "Consulta de faturas"
    allowed_roles = (UserRole.ADMIN,)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _parse_date(value: str | None) -> date | None:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    @staticmethod
    def _format_decimal(value: Decimal) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value: Decimal) -> str:
        return f"R$ {self._format_decimal(value)}"

    @staticmethod
    def _format_period(start_date: date | None, end_date: date | None) -> str:
        if not start_date and not end_date:
            return "-"
        start_label = start_date.strftime("%d/%m/%Y") if start_date else "-"
        end_label = end_date.strftime("%d/%m/%Y") if end_date else "-"
        if start_label == end_label:
            return start_label
        return f"{start_label} a {end_label}"

    def _parse_filters(self, params) -> dict[str, Any]:
        period_start = self._parse_date(params.get("period_start", "").strip())
        period_end = self._parse_date(params.get("period_end", "").strip())
        if period_start and period_end and period_end < period_start:
            period_start, period_end = period_end, period_start
        return {
            "period_start": period_start,
            "period_end": period_end,
            "billing_client_id": params.get("billing_client_id", "").strip(),
            "project_id": params.get("project_id", "").strip(),
            "consultant_id": params.get("consultant_id", "").strip(),
            "payment_status": params.get("payment_status", "").strip(),
        }

    def _get_invoice_queryset(self, filters: dict[str, Any]):
        queryset = BillingInvoice.objects.select_related(
            "billing_client", "project"
        ).order_by("-period_end", "-created_at")
        if filters["billing_client_id"]:
            queryset = queryset.filter(billing_client_id=filters["billing_client_id"])
        if filters["payment_status"]:
            queryset = queryset.filter(payment_status=filters["payment_status"])
        if filters["period_start"]:
            queryset = queryset.filter(period_end__gte=filters["period_start"])
        if filters["period_end"]:
            queryset = queryset.filter(period_start__lte=filters["period_end"])
        if filters["project_id"]:
            queryset = queryset.filter(
                time_entries__activity__project_id=filters["project_id"]
            )
        if filters["consultant_id"]:
            queryset = queryset.filter(time_entries__consultant_id=filters["consultant_id"])
        return queryset.distinct()

    def _build_invoice_payload(self, filters: dict[str, Any]) -> dict[str, Any]:
        invoices = list(self._get_invoice_queryset(filters))
        invoice_ids = [invoice.id for invoice in invoices]
        invoice_map = {invoice.id: invoice for invoice in invoices}

        items = BillingInvoiceItem.objects.select_related("consultant", "invoice").filter(
            invoice_id__in=invoice_ids
        )
        items_map = {
            (item.invoice_id, item.consultant_id): item for item in items
        }

        entries = (
            TimeEntry.objects.select_related(
                "consultant",
                "activity",
                "activity__project",
                "billing_invoice",
            )
            .filter(billing_invoice_id__in=invoice_ids)
            .order_by(
                "billing_invoice_id",
                "consultant__full_name",
                "activity__project__description",
                "start_date",
                "created_at",
            )
        )
        if filters["consultant_id"]:
            entries = entries.filter(consultant_id=filters["consultant_id"])
        if filters["project_id"]:
            entries = entries.filter(activity__project_id=filters["project_id"])
        entries = list(entries)

        overall_hours = Decimal("0.00")
        overall_value = Decimal("0.00")
        invoice_data_map: dict[int, dict[str, Any]] = {}

        for invoice in invoices:
            payment_chip = "chip-ok" if invoice.payment_status == BillingPaymentStatus.PAID else "chip-warn"
            invoice_data_map[invoice.id] = {
                "invoice": invoice,
                "period_label": self._format_period(invoice.period_start, invoice.period_end),
                "payment_label": invoice.get_payment_status_display(),
                "payment_chip": payment_chip,
                "consultants_map": {},
                "consultants": [],
                "entry_count": 0,
                "total_hours": Decimal("0.00"),
                "total_value": Decimal("0.00"),
            }

        for entry in entries:
            invoice_data = invoice_data_map.get(entry.billing_invoice_id)
            if not invoice_data:
                continue
            invoice = invoice_map.get(entry.billing_invoice_id)
            if not invoice:
                continue
            invoice_data["entry_count"] += 1

            consultant_group = invoice_data["consultants_map"].get(entry.consultant_id)
            if consultant_group is None:
                item = items_map.get((invoice.id, entry.consultant_id))
                rate = item.rate if item else None
                consultant_group = {
                    "consultant": entry.consultant,
                    "rate": rate,
                    "hours": Decimal("0.00"),
                    "total": Decimal("0.00"),
                    "projects_map": {},
                    "projects": [],
                }
                invoice_data["consultants_map"][entry.consultant_id] = consultant_group

            project_group = consultant_group["projects_map"].get(entry.activity.project_id)
            if project_group is None:
                project_group = {
                    "project": entry.activity.project,
                    "hours": Decimal("0.00"),
                    "total": Decimal("0.00"),
                    "entries": [],
                }
                consultant_group["projects_map"][entry.activity.project_id] = project_group

            hours = entry.total_hours or Decimal("0.00")
            rate = consultant_group["rate"]
            if rate is None:
                rate = entry.activity.project.hourly_rate or Decimal("0.00")
            entry_total = (hours * rate).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            project_group["hours"] += hours
            project_group["total"] += entry_total
            consultant_group["hours"] += hours
            consultant_group["total"] += entry_total
            project_group["entries"].append(
                {
                    "activity": entry.activity.activity,
                    "subactivity": _format_activity_subactivities(entry.activity),
                    "period": self._format_period(entry.start_date, entry.end_date),
                    "hours_display": self._format_decimal(hours),
                }
            )

        invoices_data = []
        for invoice_id, invoice_data in invoice_data_map.items():
            consultants = list(invoice_data["consultants_map"].values())
            invoice_total_hours = Decimal("0.00")
            invoice_total_value = Decimal("0.00")
            for consultant_group in consultants:
                projects = list(consultant_group["projects_map"].values())
                for project_group in projects:
                    project_group["hours_display"] = self._format_decimal(
                        project_group["hours"]
                    )
                    project_group["total_display"] = self._format_currency(
                        project_group["total"]
                    )
                    project_group["entry_count"] = len(project_group["entries"])
                projects.sort(
                    key=lambda item: (item["project"].description or "").lower()
                )
                consultant_group["projects"] = projects
                consultant_group["project_count"] = len(projects)
                consultant_group["entry_count"] = sum(
                    project["entry_count"] for project in projects
                )
                consultant_group["hours_display"] = self._format_decimal(
                    consultant_group["hours"]
                )
                consultant_group["total_display"] = self._format_currency(
                    consultant_group["total"]
                )
                invoice_total_hours += consultant_group["hours"]
                invoice_total_value += consultant_group["total"]
            consultants.sort(
                key=lambda item: (item["consultant"].full_name or "").lower()
            )
            invoice_data["consultants"] = consultants
            invoice_data["consultant_count"] = len(consultants)
            invoice_data["total_hours"] = invoice_total_hours
            invoice_data["total_value"] = invoice_total_value
            invoice_data["total_hours_display"] = self._format_decimal(invoice_total_hours)
            invoice_data["total_value_display"] = self._format_currency(invoice_total_value)
            overall_hours += invoice_total_hours
            overall_value += invoice_total_value
            if invoice_data["entry_count"] <= 0:
                continue
            invoices_data.append(invoice_data)

        invoices_data.sort(key=lambda item: item["invoice"].created_at, reverse=True)

        return {
            "invoices": invoices_data,
            "invoice_count": len(invoices_data),
            "entry_count": len(entries),
            "overall_hours": self._format_decimal(overall_hours),
            "overall_value": self._format_currency(overall_value),
        }

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        filters = self._parse_filters(self.request.GET)
        payload = self._build_invoice_payload(filters)
        context.update(payload)
        context.update(
            {
                "page_title": self.page_title,
                "billing_clients": Client.objects.order_by("name"),
                "projects": Project.objects.order_by("description"),
                "consultants": Consultant.objects.order_by("full_name"),
                "payment_status_options": [
                    {"value": BillingPaymentStatus.UNPAID, "label": "Nao pago"},
                    {"value": BillingPaymentStatus.PAID, "label": "Pago"},
                ],
                "current_filters": {
                    "period_start": filters["period_start"].isoformat()
                    if filters["period_start"]
                    else "",
                    "period_end": filters["period_end"].isoformat()
                    if filters["period_end"]
                    else "",
                    "billing_client_id": filters["billing_client_id"],
                    "project_id": filters["project_id"],
                    "consultant_id": filters["consultant_id"],
                    "payment_status": filters["payment_status"],
                },
            }
        )
        return context


class TimeEntryListView(LoginRequiredMixin, ListView):
    model = TimeEntry
    template_name = "restricted/time_entry_list.html"
    paginate_by = 15
    page_title = "Apontamentos"

    def _format_decimal(self, value) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_period(self, start_date, end_date) -> str:
        if not start_date:
            return "-"
        start_label = start_date.strftime("%d/%m/%Y")
        end_label = end_date.strftime("%d/%m/%Y") if end_date else "-"
        if start_label == end_label:
            return start_label
        return f"{start_label} a {end_label}"

    def _filter_for_user(self, queryset):
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset
        if role == UserRole.GP_INTERNAL:
            return queryset.filter(activity__project__internal_manager=self.request.user)
        if role == UserRole.GP_EXTERNAL:
            return queryset.filter(activity__project__external_manager=self.request.user)
        if role == UserRole.CONSULTANT:
            return queryset.filter(consultant__user=self.request.user)
        return queryset.none()

    def _apply_activity_status_filter(self, queryset, status: str):
        if status == "pending":
            return queryset.filter(activity__status=ActivityStatus.IN_PROGRESS)
        if status == "planned":
            return queryset.filter(activity__status=ActivityStatus.PLANNED)
        if status == "released":
            return queryset.filter(activity__status=ActivityStatus.RELEASED)
        if status == "done":
            return queryset.filter(activity__status=ActivityStatus.DONE)
        if status == "paused":
            return queryset.filter(activity__status=ActivityStatus.BLOCKED)
        if status == "canceled":
            return queryset.filter(activity__status=ActivityStatus.CANCELED)
        return queryset

    def _build_querystring(self) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        if not params:
            return ""
        return "&" + params.urlencode()

    def get_queryset(self):
        queryset = (
            TimeEntry.objects.select_related(
                "consultant",
                "activity",
                "activity__project",
                "activity__product",
                "activity__module",
                "activity__submodule",
                "billing_invoice",
            )
            .prefetch_related("activity__subactivity_items")
            .order_by("-created_at")
        )
        queryset = self._filter_for_user(queryset)

        params = self.request.GET
        if params.get("project_id"):
            queryset = queryset.filter(activity__project_id=params["project_id"])
        if params.get("consultant_id"):
            queryset = queryset.filter(consultant_id=params["consultant_id"])
        if params.get("internal_manager_id"):
            queryset = queryset.filter(
                activity__project__internal_manager_id=params["internal_manager_id"]
            )
        if params.get("product_id"):
            queryset = queryset.filter(activity__product_id=params["product_id"])
        if params.get("module_id"):
            queryset = queryset.filter(activity__module_id=params["module_id"])
        if params.get("submodule_id"):
            queryset = queryset.filter(activity__submodule_id=params["submodule_id"])
        if params.get("activity"):
            queryset = queryset.filter(activity__activity__icontains=params["activity"])
        if params.get("subactivity"):
            queryset = _filter_by_subactivity(
                queryset,
                params["subactivity"],
                "activity",
            )
        if params.get("status"):
            queryset = self._apply_activity_status_filter(queryset, params["status"])
        return queryset

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        role = resolve_user_role(self.request.user)
        can_review = role in {UserRole.ADMIN, UserRole.GP_INTERNAL}
        can_edit = role == UserRole.CONSULTANT

        rows = []
        for entry in context["object_list"]:
            activity = entry.activity
            project = activity.project
            if entry.status == TimeEntryStatus.APPROVED:
                status_chip = "chip-ok"
            elif entry.status == TimeEntryStatus.REJECTED:
                status_chip = "chip-danger"
            else:
                status_chip = "chip-warn"
            status_chips = [
                {
                    "label": entry.get_status_display(),
                    "class": status_chip,
                    "title": "",
                }
            ]
            if entry.status == TimeEntryStatus.APPROVED and (
                entry.billing_invoice_id or entry.billing_invoice_number
            ):
                invoice_number = entry.billing_invoice_number
                if not invoice_number and entry.billing_invoice:
                    invoice_number = entry.billing_invoice.number
                status_chips.append(
                    {
                        "label": "Faturado",
                        "class": "chip-info",
                        "title": f"Fatura {invoice_number}" if invoice_number else "",
                    }
                )
                payment_status = (
                    entry.billing_invoice.payment_status
                    if entry.billing_invoice
                    else BillingPaymentStatus.UNPAID
                )
                if payment_status == BillingPaymentStatus.PAID:
                    status_chips.append({"label": "Pago", "class": "chip-ok", "title": ""})
                else:
                    status_chips.append(
                        {"label": "Nao pago", "class": "chip-warn", "title": ""}
                    )
            rows.append(
                {
                    "project": project.description,
                    "consultant": str(entry.consultant),
                    "product": str(activity.product),
                    "module": str(activity.module),
                    "submodule": str(activity.submodule),
                    "activity": activity.activity,
                    "subactivity": _format_activity_subactivities(activity),
                    "period": self._format_period(entry.start_date, entry.end_date),
                    "entry_type": entry.get_entry_type_display(),
                    "hours": self._format_decimal(entry.total_hours),
                    "status": entry.get_status_display(),
                    "status_chip": status_chip,
                    "status_chips": status_chips,
                    "edit_url": reverse("cadastros_web:time_entry_update", args=[entry.pk])
                    if can_edit
                    and entry.consultant.user_id == self.request.user.id
                    and entry.status in {TimeEntryStatus.PENDING, TimeEntryStatus.REJECTED}
                    else None,
                    "review_url": reverse("cadastros_web:time_entry_review", args=[entry.pk])
                    if can_review and entry.status == TimeEntryStatus.PENDING
                    else None,
                }
            )

        projects = filter_projects_for_user(
            Project.objects.select_related("internal_manager"),
            self.request.user,
        ).order_by("description")

        internal_manager_map = {}
        for project in projects:
            manager = project.internal_manager
            if manager and manager.pk not in internal_manager_map:
                internal_manager_map[manager.pk] = manager
        internal_managers = sorted(
            internal_manager_map.values(),
            key=lambda user: user.get_short_name()
            or user.get_full_name()
            or user.username,
        )

        consultants = Consultant.objects.order_by("full_name")
        if role == UserRole.CONSULTANT:
            consultants = consultants.filter(user=self.request.user)

        context.update(
            {
                "page_title": self.page_title,
                "list_title": "Apontamentos registrados",
                "table_rows": rows,
                "column_count": 6 + (1 if can_edit or can_review else 0),
                "projects": projects,
                "consultants": consultants,
                "internal_managers": internal_managers,
                "products": Product.objects.order_by("description"),
                "modules": Module.objects.order_by("description"),
                "submodules": Submodule.objects.order_by("description"),
                "status_options": [
                    {"value": "pending", "label": "Pendente"},
                    {"value": "planned", "label": "Planejada"},
                    {"value": "released", "label": "Liberada"},
                    {"value": "done", "label": "Concluida"},
                    {"value": "paused", "label": "Paralizada"},
                    {"value": "canceled", "label": "Cancelada"},
                ],
                "current_filters": {
                    "project_id": self.request.GET.get("project_id", ""),
                    "consultant_id": self.request.GET.get("consultant_id", ""),
                    "internal_manager_id": self.request.GET.get("internal_manager_id", ""),
                    "product_id": self.request.GET.get("product_id", ""),
                    "module_id": self.request.GET.get("module_id", ""),
                    "submodule_id": self.request.GET.get("submodule_id", ""),
                    "activity": self.request.GET.get("activity", ""),
                    "subactivity": self.request.GET.get("subactivity", ""),
                    "status": self.request.GET.get("status", ""),
                },
                "show_actions": can_edit or can_review,
                "show_create": role in {
                    UserRole.ADMIN,
                    UserRole.GP_INTERNAL,
                    UserRole.CONSULTANT,
                },
                "create_url": reverse("cadastros_web:time_entry_create"),
                "querystring": self._build_querystring(),
            }
        )
        return context


class TimeEntryReportView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/time_entry_report.html"
    page_title = "Relatorio de apontamentos"
    allowed_roles = (
        UserRole.ADMIN,
        UserRole.GP_INTERNAL,
        UserRole.GP_EXTERNAL,
        UserRole.CONSULTANT,
    )

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        export_kind = request.GET.get("export", "").strip().lower()
        if export_kind in {"excel", "pdf"}:
            payload = self._get_report_payload()
            if export_kind == "excel":
                return self._export_excel(payload)
            return self._export_pdf(payload)
        return super().get(request, *args, **kwargs)

    def _format_decimal(self, value, places: int = 2) -> str:
        return formats.number_format(
            value,
            decimal_pos=places,
            use_l10n=True,
            force_grouping=True,
        )

    def _format_currency(self, value) -> str:
        return f"R$ {self._format_decimal(value, 2)}"

    def _format_period(self, start_date, end_date) -> str:
        if not start_date:
            return "-"
        start_label = start_date.strftime("%d/%m/%Y")
        end_label = end_date.strftime("%d/%m/%Y") if end_date else "-"
        if start_label == end_label:
            return start_label
        return f"{start_label} a {end_label}"

    def _parse_date(self, value: str | None):
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    def _filter_for_user(self, queryset):
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset
        if role == UserRole.GP_INTERNAL:
            return queryset.filter(activity__project__internal_manager=self.request.user)
        if role == UserRole.GP_EXTERNAL:
            return queryset.filter(activity__project__external_manager=self.request.user)
        if role == UserRole.CONSULTANT:
            return queryset.filter(consultant__user=self.request.user)
        return queryset.none()

    def _build_querystring(self, exclude: Iterable[str] = ()) -> str:
        params = self.request.GET.copy()
        for key in exclude:
            params.pop(key, None)
        if not params:
            return ""
        return params.urlencode()

    def _resolve_rate_value(self, rates, target_date) -> Decimal:
        for rate in rates:
            if rate.start_date <= target_date and (
                rate.end_date is None or rate.end_date >= target_date
            ):
                return rate.rate or Decimal("0.00")
        return Decimal("0.00")

    def _get_report_payload(self) -> dict[str, Any]:
        params = self.request.GET
        role = resolve_user_role(self.request.user)
        show_consultant_value = role in {UserRole.ADMIN, UserRole.CONSULTANT}
        show_consultancy_value = role == UserRole.ADMIN

        filters = {
            "project_id": params.get("project_id", "").strip(),
            "consultant_id": params.get("consultant_id", "").strip(),
            "internal_manager_id": params.get("internal_manager_id", "").strip(),
            "payment_status": params.get("payment_status", "").strip(),
        }

        entries = (
            TimeEntry.objects.select_related(
                "consultant",
                "activity",
                "activity__project",
                "activity__product",
                "activity__module",
                "activity__submodule",
                "billing_invoice",
            )
            .prefetch_related("activity__subactivity_items")
            .order_by("consultant__full_name", "-start_date", "-created_at")
        )
        entries = self._filter_for_user(entries)

        if filters["project_id"]:
            entries = entries.filter(activity__project_id=filters["project_id"])
        if filters["consultant_id"]:
            entries = entries.filter(consultant_id=filters["consultant_id"])
        if filters["internal_manager_id"]:
            entries = entries.filter(
                activity__project__internal_manager_id=filters["internal_manager_id"]
            )
        if filters["payment_status"]:
            entries = entries.filter(
                billing_invoice__payment_status=filters["payment_status"]
            )

        period_start = self._parse_date(params.get("period_start", "").strip())
        period_end = self._parse_date(params.get("period_end", "").strip())
        if period_start:
            entries = entries.filter(end_date__gte=period_start)
        if period_end:
            entries = entries.filter(start_date__lte=period_end)

        entries = list(entries)

        status_meta = {
            TimeEntryStatus.APPROVED: {"label": "Aprovada", "chip": "chip-ok"},
            TimeEntryStatus.PENDING: {"label": "Pendente", "chip": "chip-warn"},
            TimeEntryStatus.REJECTED: {"label": "Reprovada", "chip": "chip-danger"},
        }
        status_order = [
            TimeEntryStatus.APPROVED,
            TimeEntryStatus.PENDING,
            TimeEntryStatus.REJECTED,
        ]

        rates_map: dict[int, list[ConsultantRate]] = defaultdict(list)
        consultant_ids = {entry.consultant_id for entry in entries}
        if consultant_ids:
            rates = ConsultantRate.objects.filter(
                consultant_id__in=consultant_ids
            ).order_by("consultant_id", "-start_date")
            for rate in rates:
                rates_map[rate.consultant_id].append(rate)

        rows = []
        status_totals = {
            status_key: {
                "hours": Decimal("0.00"),
                "consultant_value": Decimal("0.00"),
                "consultancy_value": Decimal("0.00"),
            }
            for status_key in status_order
        }

        for entry in entries:
            status_key = entry.status
            hours = entry.total_hours or Decimal("0.00")
            activity_rate = entry.activity.consultant_hourly_rate
            if activity_rate is not None:
                rate_value = activity_rate
            else:
                rate_value = self._resolve_rate_value(
                    rates_map.get(entry.consultant_id, []),
                    entry.start_date,
                )
            consultant_value = (hours * rate_value).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            project_rate = entry.activity.project.hourly_rate or Decimal("0.00")
            consultancy_value = (hours * project_rate).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            status_chips = [
                {"label": entry.get_status_display(), "class": status_meta[status_key]["chip"], "title": ""}
            ]
            if entry.status == TimeEntryStatus.APPROVED and (
                entry.billing_invoice_id or entry.billing_invoice_number
            ):
                invoice_number = entry.billing_invoice_number
                if not invoice_number and entry.billing_invoice:
                    invoice_number = entry.billing_invoice.number
                status_chips.append(
                    {
                        "label": "Faturado",
                        "class": "chip-info",
                        "title": f"Fatura {invoice_number}" if invoice_number else "",
                    }
                )
                payment_status = (
                    entry.billing_invoice.payment_status
                    if entry.billing_invoice
                    else BillingPaymentStatus.UNPAID
                )
                if payment_status == BillingPaymentStatus.PAID:
                    status_chips.append({"label": "Pago", "class": "chip-ok", "title": ""})
                else:
                    status_chips.append({"label": "Nao pago", "class": "chip-warn", "title": ""})

            rows.append(
                {
                    "consultant_id": entry.consultant_id,
                    "consultant": str(entry.consultant),
                    "consultant_obj": entry.consultant,
                    "status_key": status_key,
                    "status_label": entry.get_status_display(),
                    "status_chip": status_meta[status_key]["chip"],
                    "project": entry.activity.project.description,
                    "product": str(entry.activity.product),
                    "module": str(entry.activity.module),
                    "submodule": str(entry.activity.submodule),
                    "activity": entry.activity.activity,
                    "subactivity": _format_activity_subactivities(entry.activity),
                    "period": self._format_period(entry.start_date, entry.end_date),
                    "hours": hours,
                    "hours_display": self._format_decimal(hours),
                    "consultant_value": consultant_value,
                    "consultant_value_display": self._format_currency(
                        consultant_value
                    ),
                    "consultancy_value": consultancy_value,
                    "consultancy_value_display": self._format_currency(
                        consultancy_value
                    ),
                    "status_chips": status_chips,
                }
            )

            totals = status_totals[status_key]
            totals["hours"] += hours
            totals["consultant_value"] += consultant_value
            totals["consultancy_value"] += consultancy_value

        groups_map: dict[int, dict[str, Any]] = {}
        for row in rows:
            group = groups_map.get(row["consultant_id"])
            if not group:
                group = {
                    "consultant": row["consultant_obj"],
                    "totals": {
                        status_key: {
                            "hours": Decimal("0.00"),
                            "consultant_value": Decimal("0.00"),
                            "consultancy_value": Decimal("0.00"),
                        }
                        for status_key in status_order
                    },
                    "statuses": {
                        status_key: {
                            "label": status_meta[status_key]["label"],
                            "chip": status_meta[status_key]["chip"],
                            "entries": [],
                            "total_hours": Decimal("0.00"),
                            "consultant_value": Decimal("0.00"),
                            "consultancy_value": Decimal("0.00"),
                        }
                        for status_key in status_order
                    },
                }
                groups_map[row["consultant_id"]] = group

            status_bucket = group["statuses"][row["status_key"]]
            status_bucket["entries"].append(row)
            status_bucket["total_hours"] += row["hours"]
            status_bucket["consultant_value"] += row["consultant_value"]
            status_bucket["consultancy_value"] += row["consultancy_value"]

            group_totals = group["totals"][row["status_key"]]
            group_totals["hours"] += row["hours"]
            group_totals["consultant_value"] += row["consultant_value"]
            group_totals["consultancy_value"] += row["consultancy_value"]

        groups = []
        for _, group in sorted(
            groups_map.items(),
            key=lambda item: (item[1]["consultant"].full_name or "").lower(),
        ):
            statuses = []
            for status_key in status_order:
                status_data = group["statuses"][status_key]
                if not status_data["entries"]:
                    continue
                statuses.append(
                    {
                        "label": status_data["label"],
                        "chip": status_data["chip"],
                        "entries": status_data["entries"],
                        "total_hours": self._format_decimal(
                            status_data["total_hours"]
                        ),
                        "total_consultant_value": self._format_currency(
                            status_data["consultant_value"]
                        ),
                        "total_consultancy_value": self._format_currency(
                            status_data["consultancy_value"]
                        ),
                    }
                )
            groups.append(
                {
                    "consultant": group["consultant"],
                    "totals": {
                        "approved_hours": self._format_decimal(
                            group["totals"][TimeEntryStatus.APPROVED]["hours"]
                        ),
                        "pending_hours": self._format_decimal(
                            group["totals"][TimeEntryStatus.PENDING]["hours"]
                        ),
                        "approved_consultant_value": self._format_currency(
                            group["totals"][TimeEntryStatus.APPROVED]["consultant_value"]
                        ),
                        "pending_consultant_value": self._format_currency(
                            group["totals"][TimeEntryStatus.PENDING]["consultant_value"]
                        ),
                        "approved_consultancy_value": self._format_currency(
                            group["totals"][TimeEntryStatus.APPROVED]["consultancy_value"]
                        ),
                        "pending_consultancy_value": self._format_currency(
                            group["totals"][TimeEntryStatus.PENDING]["consultancy_value"]
                        ),
                    },
                    "statuses": statuses,
                }
            )

        consultants = Consultant.objects.order_by("full_name")
        if role == UserRole.CONSULTANT:
            consultants = consultants.filter(user=self.request.user)

        projects = filter_projects_for_user(
            Project.objects.select_related("internal_manager"),
            self.request.user,
        ).order_by("description")
        internal_manager_map = {}
        for project in projects:
            manager = project.internal_manager
            if manager and manager.pk not in internal_manager_map:
                internal_manager_map[manager.pk] = manager
        internal_managers = sorted(
            internal_manager_map.values(),
            key=lambda user: user.get_short_name()
            or user.get_full_name()
            or user.username,
        )

        current_filters = {
            "project_id": filters["project_id"],
            "consultant_id": filters["consultant_id"],
            "internal_manager_id": filters["internal_manager_id"],
            "period_start": period_start.isoformat() if period_start else "",
            "period_end": period_end.isoformat() if period_end else "",
            "payment_status": filters["payment_status"],
        }
        filters_active = any(value for value in current_filters.values())
        export_querystring = self._build_querystring(exclude=("export",))
        if export_querystring:
            export_querystring = f"{export_querystring}&"

        column_count = 4
        if show_consultant_value:
            column_count += 1
        if show_consultancy_value:
            column_count += 1

        report_totals = {
            "approved_hours": self._format_decimal(
                status_totals[TimeEntryStatus.APPROVED]["hours"]
            ),
            "pending_hours": self._format_decimal(
                status_totals[TimeEntryStatus.PENDING]["hours"]
            ),
            "approved_consultant_value": self._format_currency(
                status_totals[TimeEntryStatus.APPROVED]["consultant_value"]
            ),
            "pending_consultant_value": self._format_currency(
                status_totals[TimeEntryStatus.PENDING]["consultant_value"]
            ),
            "approved_consultancy_value": self._format_currency(
                status_totals[TimeEntryStatus.APPROVED]["consultancy_value"]
            ),
            "pending_consultancy_value": self._format_currency(
                status_totals[TimeEntryStatus.PENDING]["consultancy_value"]
            ),
        }

        return {
            "page_title": self.page_title,
            "report_title": "Relatorio de apontamentos",
            "rows": rows,
            "groups": groups,
            "group_count": len(groups),
            "entry_count": len(rows),
            "report_totals": report_totals,
            "status_totals": status_totals,
            "status_order": status_order,
            "status_meta": status_meta,
            "consultants": consultants,
            "projects": projects,
            "internal_managers": internal_managers,
            "payment_status_options": [
                {"value": BillingPaymentStatus.UNPAID, "label": "Nao pago"},
                {"value": BillingPaymentStatus.PAID, "label": "Pago"},
            ],
            "current_filters": current_filters,
            "filters_active": filters_active,
            "export_querystring": export_querystring,
            "show_consultant_value": show_consultant_value,
            "show_consultancy_value": show_consultancy_value,
            "column_count": column_count,
        }

    def _export_excel(self, payload: dict[str, Any]) -> HttpResponse:
        rows = payload["rows"]
        show_consultant_value = payload["show_consultant_value"]
        show_consultancy_value = payload["show_consultancy_value"]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Apontamentos"

        headers = [
            "Consultor",
            "Situacao",
            "Projeto",
            "Produto",
            "Modulo",
            "Submodulo",
            "Atividade",
            "Subatividades",
            "Periodo",
            "Horas",
        ]
        if show_consultant_value:
            headers.append("Valor consultor")
        if show_consultancy_value:
            headers.append("Valor consultoria")
        sheet.append(headers)

        for row in rows:
            data = [
                row["consultant"],
                row["status_label"],
                row["project"],
                row["product"],
                row["module"],
                row["submodule"],
                row["activity"],
                row["subactivity"],
                row["period"],
                float(row["hours"]),
            ]
            if show_consultant_value:
                data.append(float(row["consultant_value"]))
            if show_consultancy_value:
                data.append(float(row["consultancy_value"]))
            sheet.append(data)

        hours_col = headers.index("Horas") + 1
        currency_format = "R$ #,##0.00"
        for cell in sheet[get_column_letter(hours_col)][1:]:
            cell.number_format = "0.00"
        if show_consultant_value:
            col = headers.index("Valor consultor") + 1
            for cell in sheet[get_column_letter(col)][1:]:
                cell.number_format = currency_format
        if show_consultancy_value:
            col = headers.index("Valor consultoria") + 1
            for cell in sheet[get_column_letter(col)][1:]:
                cell.number_format = currency_format

        summary = workbook.create_sheet("Resumo")
        summary_headers = ["Situacao", "Horas"]
        if show_consultant_value:
            summary_headers.append("Valor consultor")
        if show_consultancy_value:
            summary_headers.append("Valor consultoria")
        summary.append(summary_headers)

        status_totals = payload["status_totals"]
        status_order = payload["status_order"]
        status_meta = payload["status_meta"]

        for status_key in status_order:
            totals = status_totals[status_key]
            summary_row = [
                status_meta[status_key]["label"],
                float(totals["hours"]),
            ]
            if show_consultant_value:
                summary_row.append(float(totals["consultant_value"]))
            if show_consultancy_value:
                summary_row.append(float(totals["consultancy_value"]))
            summary.append(summary_row)

        total_hours = sum(
            status_totals[status_key]["hours"] for status_key in status_order
        )
        total_consultant = sum(
            status_totals[status_key]["consultant_value"] for status_key in status_order
        )
        total_consultancy = sum(
            status_totals[status_key]["consultancy_value"] for status_key in status_order
        )
        total_row = ["Total", float(total_hours)]
        if show_consultant_value:
            total_row.append(float(total_consultant))
        if show_consultancy_value:
            total_row.append(float(total_consultancy))
        summary.append(total_row)

        hours_col = summary_headers.index("Horas") + 1
        for cell in summary[get_column_letter(hours_col)][1:]:
            cell.number_format = "0.00"
        if show_consultant_value:
            col = summary_headers.index("Valor consultor") + 1
            for cell in summary[get_column_letter(col)][1:]:
                cell.number_format = currency_format
        if show_consultancy_value:
            col = summary_headers.index("Valor consultoria") + 1
            for cell in summary[get_column_letter(col)][1:]:
                cell.number_format = currency_format

        for sheet_obj in (sheet, summary):
            for idx, header in enumerate(sheet_obj[1], start=1):
                column_letter = get_column_letter(idx)
                width = max(14, len(str(header.value)) + 4)
                sheet_obj.column_dimensions[column_letter].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="relatorio_apontamentos.xlsx"'
        )
        return response

    def _export_pdf(self, payload: dict[str, Any]) -> HttpResponse:
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
        except ImportError:
            return HttpResponse("Exportacao PDF indisponivel.", status=500)

        rows = payload["rows"]
        show_consultant_value = payload["show_consultant_value"]
        show_consultancy_value = payload["show_consultancy_value"]

        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)
        margin = 14 * mm
        y = height - margin

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(margin, y, payload["report_title"])
        y -= 8 * mm

        summary_parts = [
            f"Horas aprovadas: {payload['report_totals']['approved_hours']}h",
            f"Horas pendentes: {payload['report_totals']['pending_hours']}h",
        ]
        if show_consultant_value:
            summary_parts.append(
                f"Valor consultor aprovado: {payload['report_totals']['approved_consultant_value']}"
            )
            summary_parts.append(
                f"Valor consultor pendente: {payload['report_totals']['pending_consultant_value']}"
            )
        if show_consultancy_value:
            summary_parts.append(
                f"Valor consultoria aprovado: {payload['report_totals']['approved_consultancy_value']}"
            )
            summary_parts.append(
                f"Valor consultoria pendente: {payload['report_totals']['pending_consultancy_value']}"
            )

        pdf.setFont("Helvetica", 8.5)
        pdf.drawString(margin, y, " | ".join(summary_parts))
        y -= 6 * mm

        columns = [
            {"label": "Consultor", "key": "consultant", "width": 40 * mm},
            {"label": "Situacao", "key": "status_label", "width": 22 * mm},
            {"label": "Projeto / Atividade", "key": "project_activity", "width": 76 * mm},
            {"label": "Periodo", "key": "period", "width": 34 * mm},
            {"label": "Horas", "key": "hours_display", "width": 18 * mm},
        ]
        if show_consultant_value:
            columns.append(
                {
                    "label": "Valor consultor",
                    "key": "consultant_value_display",
                    "width": 28 * mm,
                }
            )
        if show_consultancy_value:
            columns.append(
                {
                    "label": "Valor consultoria",
                    "key": "consultancy_value_display",
                    "width": 28 * mm,
                }
            )

        def fit_text(text, max_width):
            if pdf.stringWidth(text, "Helvetica", 8) <= max_width:
                return text
            clipped = text
            while clipped and pdf.stringWidth(f"{clipped}...", "Helvetica", 8) > max_width:
                clipped = clipped[:-1]
            return f"{clipped}..." if clipped else ""

        def draw_header(current_y):
            pdf.setFont("Helvetica-Bold", 8.5)
            x = margin
            for column in columns:
                pdf.drawString(x, current_y, column["label"])
                x += column["width"]
            current_y -= 2 * mm
            pdf.setLineWidth(0.4)
            pdf.line(margin, current_y, margin + sum(col["width"] for col in columns), current_y)
            return current_y - 4 * mm

        y = draw_header(y)
        pdf.setFont("Helvetica", 8)

        for row in rows:
            if y < margin + 12 * mm:
                pdf.showPage()
                y = height - margin
                y = draw_header(y)
                pdf.setFont("Helvetica", 8)

            project_activity = row["project"]
            if row["activity"]:
                project_activity = f"{project_activity} - {row['activity']}"
            if row["subactivity"]:
                project_activity = f"{project_activity} / {row['subactivity']}"

            data = {
                "consultant": row["consultant"],
                "status_label": row["status_label"],
                "project_activity": project_activity,
                "period": row["period"],
                "hours_display": row["hours_display"],
                "consultant_value_display": row["consultant_value_display"],
                "consultancy_value_display": row["consultancy_value_display"],
            }

            x = margin
            for column in columns:
                text = str(data.get(column["key"], ""))
                pdf.drawString(x, y, fit_text(text, column["width"] - 2))
                x += column["width"]
            y -= 5 * mm
        pdf.save()
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="relatorio_apontamentos.pdf"'
        )
        return response

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.update(self._get_report_payload())
        return context


class NotificationsView(LoginRequiredMixin, TemplateView):
    template_name = "restricted/notifications.html"
    page_title = "Pendencias"

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["page_title"] = self.page_title
        return context


class TimeEntryActivityInfoView(LoginRequiredMixin, View):
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    def dispatch(self, request, *args, **kwargs):
        if self.allowed_roles is not None:
            role = resolve_user_role(request.user)
            if role not in self.allowed_roles:
                raise PermissionDenied("Perfil sem acesso a esta area.")
        return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def _format_decimal(value: Decimal) -> str:
        return formats.number_format(
            value,
            decimal_pos=2,
            use_l10n=True,
            force_grouping=True,
        )

    def get(self, request, *args, **kwargs):
        activity_id_raw = request.GET.get("activity_id", "").strip()
        if not activity_id_raw:
            return JsonResponse(
                {"ok": False, "error": "Atividade nao informada."},
                status=400,
            )
        try:
            activity_id = int(activity_id_raw)
        except ValueError:
            return JsonResponse(
                {"ok": False, "error": "Atividade invalida."},
                status=400,
            )

        entry_id = None
        entry_id_raw = request.GET.get("entry_id", "").strip()
        if entry_id_raw:
            try:
                entry_id = int(entry_id_raw)
            except ValueError:
                return JsonResponse(
                    {"ok": False, "error": "Apontamento invalido."},
                    status=400,
                )

        activity_qs = filter_activities_for_user(
            ProjectActivity.objects.filter(pk=activity_id).prefetch_related(
                "subactivity_items"
            ),
            request.user,
        )
        activity = activity_qs.first()
        if not activity:
            return JsonResponse(
                {"ok": False, "error": "Atividade nao encontrada."},
                status=404,
            )

        entries = TimeEntry.objects.filter(activity=activity)
        if entry_id:
            entries = entries.exclude(pk=entry_id)
        approved = (
            entries.filter(status=TimeEntryStatus.APPROVED)
            .aggregate(total=Sum("total_hours"))
            .get("total")
            or Decimal("0.00")
        )
        pending = (
            entries.filter(status=TimeEntryStatus.PENDING)
            .aggregate(total=Sum("total_hours"))
            .get("total")
            or Decimal("0.00")
        )
        total_hours = activity.hours or Decimal("0.00")
        balance = total_hours - approved - pending

        return JsonResponse(
            {
                "ok": True,
                "data": {
                    "subactivities": _get_activity_subactivities(activity),
                    "activity_hours": self._format_decimal(total_hours),
                    "approved_hours": self._format_decimal(approved),
                    "pending_hours": self._format_decimal(pending),
                    "balance_hours": self._format_decimal(balance),
                },
            }
        )


class TimeEntryCreateView(BaseCreateView):
    model = TimeEntry
    form_class = TimeEntryForm
    template_name = "restricted/time_entry_form.html"
    page_title = "Novo apontamento"
    submit_label = "Enviar para aprovacao"
    cancel_url_name = "cadastros_web:time_entry_list"
    success_url = reverse_lazy("cadastros_web:time_entry_list")
    full_width_fields = ("description",)
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["activity_info_url"] = reverse("cadastros_web:time_entry_activity_info")
        context["time_entry_id"] = ""
        consultant = Consultant.objects.filter(user=self.request.user).first()
        context["consultant_name"] = consultant.full_name if consultant else ""
        context["attachments"] = []
        return context

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        role = resolve_user_role(self.request.user)
        form.fields["activity"].queryset = filter_activities_for_user(
            ProjectActivity.objects.filter(status=ActivityStatus.RELEASED),
            self.request.user,
        ).order_by("project__description", "seq")
        consultant_field = form.fields.get("consultant")
        if consultant_field and role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(user=self.request.user).first()
            consultant_field.queryset = Consultant.objects.filter(
                pk=getattr(consultant, "pk", None)
            )
            consultant_field.initial = consultant.pk if consultant else None
            consultant_field.required = False
            consultant_field.widget = forms.HiddenInput()
        return form

    def form_valid(self, form):
        role = resolve_user_role(self.request.user)
        if role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(user=self.request.user).first()
            if not consultant:
                form.add_error(None, "Usuario sem consultor vinculado.")
                return self.form_invalid(form)
            form.instance.consultant = consultant
        form.instance.status = TimeEntryStatus.PENDING
        form.instance.reviewed_by = None
        form.instance.reviewed_at = None
        response = super().form_valid(form)
        _save_time_entry_attachments(
            self.object,
            self.request.FILES.getlist("attachments"),
        )
        if self.object and self.object.status == TimeEntryStatus.PENDING:
            transaction.on_commit(lambda: notify_time_entry_pending(self.object))
        return response


class TimeEntryUpdateView(BaseUpdateView):
    model = TimeEntry
    form_class = TimeEntryForm
    template_name = "restricted/time_entry_form.html"
    page_title = "Editar apontamento"
    submit_label = "Reenviar para aprovacao"
    cancel_url_name = "cadastros_web:time_entry_list"
    success_url = reverse_lazy("cadastros_web:time_entry_list")
    full_width_fields = ("description",)
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL, UserRole.CONSULTANT)

    def get_context_data(self, **kwargs: Any) -> dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["activity_info_url"] = reverse("cadastros_web:time_entry_activity_info")
        context["time_entry_id"] = self.object.pk if self.object else ""
        consultant = Consultant.objects.filter(user=self.request.user).first()
        context["consultant_name"] = consultant.full_name if consultant else ""
        context["attachments"] = (
            self.object.attachments.order_by("-created_at") if self.object else []
        )
        return context

    def get_queryset(self):
        queryset = TimeEntry.objects.select_related("consultant", "activity", "activity__project")
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset
        if role == UserRole.GP_INTERNAL:
            return queryset.filter(activity__project__internal_manager=self.request.user)
        if role == UserRole.GP_EXTERNAL:
            return queryset.filter(activity__project__external_manager=self.request.user)
        if role == UserRole.CONSULTANT:
            return queryset.filter(consultant__user=self.request.user)
        return queryset.none()

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        role = resolve_user_role(self.request.user)
        allowed_activities = filter_activities_for_user(
            ProjectActivity.objects.filter(status=ActivityStatus.RELEASED),
            self.request.user,
        )
        if self.object:
            allowed_activities = ProjectActivity.objects.filter(
                Q(pk=self.object.activity_id) | Q(pk__in=allowed_activities)
            )
        form.fields["activity"].queryset = allowed_activities.order_by(
            "project__description",
            "seq",
        )
        consultant_field = form.fields.get("consultant")
        if consultant_field and role == UserRole.CONSULTANT:
            consultant_field.queryset = Consultant.objects.filter(user=self.request.user)
            consultant_field.required = False
            consultant_field.widget = forms.HiddenInput()
        return form

    def form_valid(self, form):
        role = resolve_user_role(self.request.user)
        if role == UserRole.CONSULTANT:
            consultant = Consultant.objects.filter(user=self.request.user).first()
            if not consultant:
                form.add_error(None, "Usuario sem consultor vinculado.")
                return self.form_invalid(form)
            form.instance.consultant = consultant
            if form.instance.status == TimeEntryStatus.APPROVED:
                form.add_error(None, "Apontamento aprovado nao pode ser alterado.")
                return self.form_invalid(form)
            form.instance.status = TimeEntryStatus.PENDING
            form.instance.reviewed_by = None
            form.instance.reviewed_at = None
        response = super().form_valid(form)
        _save_time_entry_attachments(
            self.object,
            self.request.FILES.getlist("attachments"),
        )
        if self.object and self.object.status == TimeEntryStatus.PENDING:
            transaction.on_commit(lambda: notify_time_entry_pending(self.object))
        return response


class TimeEntryReviewView(BaseUpdateView):
    model = TimeEntry
    form_class = TimeEntryReviewForm
    template_name = "restricted/form.html"
    page_title = "Aprovar apontamento"
    submit_label = "Salvar decisao"
    cancel_url_name = "cadastros_web:time_entry_list"
    success_url = reverse_lazy("cadastros_web:time_entry_list")
    allowed_roles = (UserRole.ADMIN, UserRole.GP_INTERNAL)

    def get_queryset(self):
        queryset = TimeEntry.objects.select_related("activity", "activity__project", "consultant")
        role = resolve_user_role(self.request.user)
        if role == UserRole.ADMIN:
            return queryset.filter(status=TimeEntryStatus.PENDING)
        return queryset.filter(
            activity__project__internal_manager=self.request.user,
            status=TimeEntryStatus.PENDING,
        )

    def form_valid(self, form):
        form.instance.reviewed_by = self.request.user
        form.instance.reviewed_at = timezone.now()
        if form.instance.status == TimeEntryStatus.APPROVED:
            form.instance.rejection_reason = ""
        response = super().form_valid(form)
        entry = self.object
        if entry:
            transaction.on_commit(lambda: notify_time_entry_reviewed(entry))
        return response
