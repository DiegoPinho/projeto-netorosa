from __future__ import annotations

import csv
import io
import os
import re
import unicodedata
import xml.etree.ElementTree as ET
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import transaction

from .models import (
    AccountNature,
    AccountPlanTemplateHeader,
    AccountPlanTemplateItem,
    AccountType,
    DreSign,
    DeploymentTemplate,
    DeploymentTemplateHeader,
    Module,
    Phase,
    Product,
    StatusChoices,
    Submodule,
)

_COLUMN_MAP = {
    "descricao": "template",
    "template": "template",
    "nome do template": "template",
    "seq": "seq",
    "seq predecessora": "seq_predecessor",
    "seq_predecessora": "seq_predecessor",
    "seq predecessor": "seq_predecessor",
    "seq_predecessor": "seq_predecessor",
    "projeto": "template",
    "fase": "phase",
    "produto": "product",
    "modulo": "module",
    "submodulo": "submodule",
    "atividade": "activity",
    "subatividade": "subactivity",
    "dias": "days",
    "horas": "hours",
}

_REQUIRED_FIELDS = {
    "template",
    "phase",
    "product",
    "module",
    "submodule",
    "activity",
}

_DEFAULT_SUBACTIVITY = "A definir"

_ACCOUNT_PLAN_COLUMN_MAP = {
    "modelo": "template",
    "nome do modelo": "template",
    "template": "template",
    "plano de contas": "template",
    "codigo": "code",
    "código": "code",
    "conta": "description",
    "descricao": "description",
    "descricao da conta": "description",
    "nivel": "level",
    "nível": "level",
    "conta pai": "parent_code",
    "codigo conta pai": "parent_code",
    "codigo pai": "parent_code",
    "tipo": "account_type",
    "tipo de conta": "account_type",
    "natureza": "nature",
    "analitica": "is_analytic",
    "analítica": "is_analytic",
    "status": "status",
    "grupo dre": "dre_group",
    "linha dre": "dre_subgroup",
    "ordem dre": "dre_order",
    "sinal dre": "dre_sign",
}

_ACCOUNT_PLAN_REQUIRED_FIELDS = {
    "template",
    "code",
    "description",
    "level",
    "account_type",
    "nature",
    "is_analytic",
    "dre_group",
    "dre_order",
    "dre_sign",
}

_DURATION_RE = re.compile(
    r"P(?:(?P<days>\d+)D)?"
    r"(?:T(?:(?P<hours>\d+)H)?(?:(?P<minutes>\d+)M)?(?:(?P<seconds>\d+)S)?)?"
)


def import_deployment_templates(uploaded_file) -> tuple[int, list[str]]:
    extension = os.path.splitext(uploaded_file.name or "")[1].lower()
    if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        rows = _read_xlsx(uploaded_file)
    elif extension == ".csv":
        rows = _read_csv(uploaded_file)
    elif extension == ".xml":
        rows = _read_msproject_xml(uploaded_file)
    else:
        return 0, ["Formato de arquivo nao suportado. Use .xlsx, .csv ou .xml."]

    return _import_rows(rows)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text)


def _build_header_map(headers: list[object]) -> dict[int, str]:
    header_map: dict[int, str] = {}
    for idx, header in enumerate(headers):
        normalized = _normalize_header(header)
        if normalized in _COLUMN_MAP:
            header_map[idx] = _COLUMN_MAP[normalized]
    missing = _REQUIRED_FIELDS.difference(header_map.values())
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Colunas obrigatorias ausentes: {missing_list}.")
    return header_map


def _read_xlsx(uploaded_file) -> list[tuple[int, dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Dependencia openpyxl nao instalada.") from exc

    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise ValueError("Planilha vazia.")
    header_map = _build_header_map(list(headers))
    rows: list[tuple[int, dict[str, object]]] = []
    for row_index, row in enumerate(rows_iter, start=2):
        if not row or all(_is_empty(cell) for cell in row):
            continue
        row_data = {
            field_name: row[col_index] if col_index < len(row) else None
            for col_index, field_name in header_map.items()
        }
        rows.append((row_index, row_data))
    return rows


def _read_csv(uploaded_file) -> list[tuple[int, dict[str, object]]]:
    wrapper = io.TextIOWrapper(uploaded_file, encoding="utf-8-sig")
    try:
        sample = wrapper.read(2048)
        wrapper.seek(0)
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(wrapper, dialect)
    headers = next(reader, None)
    if not headers:
        raise ValueError("Planilha vazia.")
    header_map = _build_header_map(headers)
    rows: list[tuple[int, dict[str, object]]] = []
    for row_index, row in enumerate(reader, start=2):
        if not row or all(_is_empty(cell) for cell in row):
            continue
        row_data = {
            field_name: row[col_index] if col_index < len(row) else None
            for col_index, field_name in header_map.items()
        }
        rows.append((row_index, row_data))
    return rows


def _read_msproject_xml(uploaded_file) -> list[tuple[int, dict[str, object]]]:
    tree = ET.parse(uploaded_file)
    root = tree.getroot()
    project_name = _text_or_none(root.find(".//{*}Name")) or "Template importado"

    tasks = []
    for task in root.findall(".//{*}Task"):
        task_id = _to_int(_text_or_none(task.find("{*}ID")))
        if task_id == 0:
            continue
        tasks.append(
            {
                "uid": _to_int(_text_or_none(task.find("{*}UID"))),
                "id": task_id,
                "name": _text_or_none(task.find("{*}Name")),
                "outline_level": _to_int(_text_or_none(task.find("{*}OutlineLevel"))) or 0,
                "summary": _to_int(_text_or_none(task.find("{*}Summary"))) or 0,
                "duration": _text_or_none(task.find("{*}Duration")),
                "work": _text_or_none(task.find("{*}Work")),
                "predecessor": _to_int(
                    _text_or_none(task.find("{*}PredecessorLink/{*}PredecessorUID"))
                ),
            }
        )

    tasks.sort(key=lambda item: item.get("id") or 0)
    uid_to_id = {task["uid"]: task["id"] for task in tasks if task.get("uid")}

    outline_stack: dict[int, str] = {}
    rows: list[tuple[int, dict[str, object]]] = []
    for idx, task in enumerate(tasks, start=1):
        level = task["outline_level"]
        name = task.get("name") or ""
        if level > 0:
            outline_stack[level] = name
            for key in list(outline_stack.keys()):
                if key > level:
                    outline_stack.pop(key, None)

        if task["summary"] == 1:
            continue

        activity = outline_stack.get(5) or name
        subactivity = outline_stack.get(6) or ""
        hours_value = _parse_msp_duration(task.get("work")) or _parse_msp_duration(
            task.get("duration")
        )
        days_value = _duration_hours_to_days(hours_value)

        rows.append(
            (
                idx,
                {
                    "template": project_name,
                    "seq": task.get("id") or idx,
                    "seq_predecessor": _resolve_predecessor(task.get("predecessor"), uid_to_id),
                    "phase": outline_stack.get(1),
                    "product": outline_stack.get(2),
                    "module": outline_stack.get(3),
                    "submodule": outline_stack.get(4),
                    "activity": activity,
                    "subactivity": subactivity,
                    "days": days_value,
                    "hours": hours_value or Decimal("0"),
                },
            )
        )
    return rows


def _resolve_predecessor(predecessor_uid: int | None, uid_to_id: dict[int, int]) -> int | None:
    if predecessor_uid is None:
        return None
    return uid_to_id.get(predecessor_uid, predecessor_uid)


def _parse_msp_duration(value: str | None) -> Decimal | None:
    if not value:
        return None
    match = _DURATION_RE.match(value)
    if not match:
        return None
    days = int(match.group("days") or 0)
    hours = int(match.group("hours") or 0)
    minutes = int(match.group("minutes") or 0)
    seconds = int(match.group("seconds") or 0)
    total_hours = days * 24 + hours + (minutes / 60) + (seconds / 3600)
    return Decimal(str(round(total_hours, 2)))


def _duration_hours_to_days(hours: Decimal | None) -> int:
    if hours is None:
        return 0
    if hours <= 0:
        return 0
    return max(1, int(round(float(hours) / 8)))


def _text_or_none(element: ET.Element | None) -> str | None:
    if element is None or element.text is None:
        return None
    return element.text.strip() or None


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _import_rows(rows: list[tuple[int, dict[str, object]]]) -> tuple[int, list[str]]:
    errors: list[str] = []
    instances: list[DeploymentTemplate] = []

    for row_index, row in rows:
        try:
            instance = _build_instance(row, row_index)
            instances.append(instance)
        except ValueError as exc:
            errors.append(f"Linha {row_index}: {exc}")

    if errors:
        return 0, errors

    with transaction.atomic():
        for instance in instances:
            instance.full_clean()
            instance.save()

    return len(instances), []


def _build_instance(row: dict[str, object], row_index: int) -> DeploymentTemplate:
    missing = [
        field for field in _REQUIRED_FIELDS if _is_empty(row.get(field, None))
    ]
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Campos obrigatorios vazios: {missing_list}.")

    template_name = _to_text(row.get("template"), "Nome do template")
    seq = _to_int(row.get("seq"), "Seq", required=False)
    if seq is None:
        seq = row_index
    seq_predecessor = _to_int(row.get("seq_predecessor"), "Seq predecessora", required=False)
    activity = _to_text(row.get("activity"), "Atividade")
    subactivity = _to_text(row.get("subactivity"), "Subatividade", allow_empty=True)
    if not subactivity:
        subactivity = _DEFAULT_SUBACTIVITY
    days = _to_int(row.get("days"), "Dias", required=False)
    if days is None:
        days = 0
    hours = _to_decimal_optional(row.get("hours"), "Horas")

    template = _resolve_template(template_name)
    phase = _resolve_related(Phase, row.get("phase"), "Fase")
    product = _resolve_related(Product, row.get("product"), "Produto")
    module = _resolve_related(Module, row.get("module"), "Modulo", product=product)
    submodule = _resolve_related(
        Submodule,
        row.get("submodule"),
        "Submodulo",
        product=product,
        module=module,
    )

    if days < 0:
        raise ValueError("Dias deve ser maior ou igual a zero.")
    if hours < 0:
        raise ValueError("Horas deve ser maior ou igual a zero.")

    return DeploymentTemplate(
        template=template,
        seq=seq,
        seq_predecessor=seq_predecessor,
        phase=phase,
        product=product,
        module=module,
        submodule=submodule,
        activity=activity,
        subactivity=subactivity,
        days=days,
        hours=hours,
    )


def _resolve_template(name: str) -> DeploymentTemplateHeader:
    template, _ = DeploymentTemplateHeader.objects.get_or_create(name=name)
    return template


def _resolve_related(model, value: object, label: str, **filters):
    if _is_empty(value):
        raise ValueError(f"{label} obrigatorio.")
    obj = None
    text = _to_text(value, label, allow_empty=True)
    if text:
        obj = model.objects.filter(description__iexact=text, **filters).first()
    if obj is None:
        raise ValueError(_build_missing_message(label, value, filters))
    return obj


def _to_text(value: object, label: str, allow_empty: bool = False) -> str:
    if value is None:
        if allow_empty:
            return ""
        raise ValueError(f"{label} obrigatorio.")
    text = str(value).strip()
    if not text and not allow_empty:
        raise ValueError(f"{label} obrigatorio.")
    return text


def _build_missing_message(label: str, value: object, filters: dict) -> str:
    details = []
    product = filters.get("product")
    module = filters.get("module")
    if product:
        details.append(f"produto {product}")
    if module:
        details.append(f"modulo {module}")
    suffix = f" para {', '.join(details)}" if details else ""
    return f"{label} nao encontrado: {value}{suffix}. Cadastre antes de importar."


def _to_int(value: object, label: str, required: bool = False) -> int | None:
    if _is_empty(value):
        if required:
            raise ValueError(f"{label} obrigatorio.")
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        if required:
            raise ValueError(f"{label} obrigatorio.")
        return None
    try:
        return int(float(text.replace(",", ".")))
    except ValueError as exc:
        raise ValueError(f"{label} deve ser numerico.") from exc


def _to_decimal(value: object, label: str) -> Decimal:
    if _is_empty(value):
        raise ValueError(f"{label} obrigatorio.")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        raise ValueError(f"{label} obrigatorio.")
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{label} deve ser numerico.") from exc


def _to_decimal_optional(value: object, label: str) -> Decimal:
    if _is_empty(value):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return Decimal("0")
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{label} deve ser numerico.") from exc


def import_account_plan_templates(uploaded_file) -> tuple[int, list[str]]:
    extension = os.path.splitext(uploaded_file.name or "")[1].lower()
    if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        rows = _read_account_plan_xlsx(uploaded_file)
    elif extension == ".csv":
        rows = _read_account_plan_csv(uploaded_file)
    else:
        return 0, ["Formato de arquivo nao suportado. Use .xlsx ou .csv."]

    return _import_account_plan_rows(rows)


def _build_account_plan_header_map(headers: list[object]) -> dict[int, str]:
    header_map: dict[int, str] = {}
    for idx, header in enumerate(headers):
        normalized = _normalize_header(header)
        if normalized in _ACCOUNT_PLAN_COLUMN_MAP:
            header_map[idx] = _ACCOUNT_PLAN_COLUMN_MAP[normalized]
    missing = _ACCOUNT_PLAN_REQUIRED_FIELDS.difference(header_map.values())
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Colunas obrigatorias ausentes: {missing_list}.")
    return header_map


def _read_account_plan_xlsx(uploaded_file) -> list[tuple[int, dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Dependencia openpyxl nao instalada.") from exc

    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise ValueError("Planilha vazia.")
    header_map = _build_account_plan_header_map(list(headers))
    rows: list[tuple[int, dict[str, object]]] = []
    for row_index, row in enumerate(rows_iter, start=2):
        if not row or all(_is_empty(cell) for cell in row):
            continue
        row_data = {
            field_name: row[col_index] if col_index < len(row) else None
            for col_index, field_name in header_map.items()
        }
        rows.append((row_index, row_data))
    return rows


def _read_account_plan_csv(uploaded_file) -> list[tuple[int, dict[str, object]]]:
    wrapper = io.TextIOWrapper(uploaded_file, encoding="utf-8-sig")
    try:
        sample = wrapper.read(2048)
        wrapper.seek(0)
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(wrapper, dialect)
    headers = next(reader, None)
    if not headers:
        raise ValueError("Planilha vazia.")
    header_map = _build_account_plan_header_map(headers)
    rows: list[tuple[int, dict[str, object]]] = []
    for row_index, row in enumerate(reader, start=2):
        if not row or all(_is_empty(cell) for cell in row):
            continue
        row_data = {
            field_name: row[col_index] if col_index < len(row) else None
            for col_index, field_name in header_map.items()
        }
        rows.append((row_index, row_data))
    return rows


def _import_account_plan_rows(
    rows: list[tuple[int, dict[str, object]]],
) -> tuple[int, list[str]]:
    errors: list[str] = []
    instances: list[tuple[int, AccountPlanTemplateItem]] = []
    seen_codes: set[tuple[int, str]] = set()

    for row_index, row in rows:
        try:
            instance = _build_account_plan_instance(row)
            code_key = (instance.template_id or 0, instance.code.strip().lower())
            if code_key in seen_codes:
                raise ValueError("Codigo duplicado dentro do mesmo modelo.")
            seen_codes.add(code_key)
            instances.append((row_index, instance))
        except ValueError as exc:
            errors.append(f"Linha {row_index}: {exc}")

    if errors:
        return 0, errors

    for row_index, instance in instances:
        try:
            instance.full_clean()
        except ValidationError as exc:
            errors.append(f"Linha {row_index}: {'; '.join(exc.messages)}")

    if errors:
        return 0, errors

    with transaction.atomic():
        for _, instance in instances:
            instance.save()

    return len(instances), []


def _build_account_plan_instance(row: dict[str, object]) -> AccountPlanTemplateItem:
    missing = [
        field for field in _ACCOUNT_PLAN_REQUIRED_FIELDS if _is_empty(row.get(field, None))
    ]
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Campos obrigatorios vazios: {missing_list}.")

    template_name = _to_text(row.get("template"), "Nome do modelo")
    code = _to_text(row.get("code"), "Codigo")
    description = _to_text(row.get("description"), "Descricao")
    level = _to_int(row.get("level"), "Nivel", required=True)
    parent_code = _to_text(row.get("parent_code"), "Conta pai", allow_empty=True)
    account_type = _parse_account_type(row.get("account_type"))
    nature = _parse_account_nature(row.get("nature"))
    is_analytic = _parse_bool_value(row.get("is_analytic"), "Analitica")
    status = _parse_status(row.get("status"))
    dre_group = _to_text(row.get("dre_group"), "Grupo DRE")
    dre_subgroup = _to_text(row.get("dre_subgroup"), "Linha DRE", allow_empty=True)
    dre_order = _to_int(row.get("dre_order"), "Ordem DRE", required=True)
    dre_sign = _parse_dre_sign(row.get("dre_sign"))

    if parent_code and parent_code.strip().lower() == code.strip().lower():
        raise ValueError("Conta pai nao pode ser a propria conta.")
    if level is not None and level < 1:
        raise ValueError("Nivel deve ser maior ou igual a 1.")
    if parent_code and level == 1:
        raise ValueError("Conta pai nao deve ser informada para nivel 1.")
    if not parent_code and level and level > 1:
        raise ValueError("Conta pai obrigatoria para niveis acima de 1.")

    template = _resolve_account_plan_template(template_name)
    parent = _resolve_account_plan_parent(template, parent_code) if parent_code else None

    return AccountPlanTemplateItem(
        template=template,
        code=code,
        description=description,
        level=level,
        parent=parent,
        account_type=account_type,
        nature=nature,
        is_analytic=is_analytic,
        status=status,
        dre_group=dre_group,
        dre_subgroup=dre_subgroup,
        dre_order=dre_order,
        dre_sign=dre_sign,
    )


def _resolve_account_plan_template(name: str) -> AccountPlanTemplateHeader:
    template, _ = AccountPlanTemplateHeader.objects.get_or_create(name=name)
    return template


def _resolve_account_plan_parent(
    template: AccountPlanTemplateHeader, parent_code: str
) -> AccountPlanTemplateItem:
    parent = AccountPlanTemplateItem.objects.filter(
        template=template, code__iexact=parent_code
    ).first()
    if parent is None:
        raise ValueError(f"Conta pai nao encontrada: {parent_code}. Cadastre antes de importar.")
    return parent


def _normalize_value(value: object) -> str:
    return _normalize_header(value)


def _parse_account_type(value: object) -> str:
    text = _normalize_value(value)
    if not text:
        raise ValueError("Tipo de conta obrigatorio.")
    if text in {choice.value for choice in AccountType}:
        return text
    mapping = {
        "ativo": AccountType.ASSET,
        "passivo": AccountType.LIABILITY,
        "patrimonio": AccountType.EQUITY,
        "patrimonio liquido": AccountType.EQUITY,
        "receita": AccountType.REVENUE,
        "custo": AccountType.COST,
        "despesa": AccountType.EXPENSE,
        "outro": AccountType.OTHER,
        "outros": AccountType.OTHER,
    }
    if text in mapping:
        return mapping[text]
    raise ValueError(f"Tipo de conta invalido: {value}.")


def _parse_account_nature(value: object) -> str:
    text = _normalize_value(value)
    if not text:
        raise ValueError("Natureza obrigatoria.")
    if text in {choice.value for choice in AccountNature}:
        return text
    mapping = {
        "debito": AccountNature.DEBIT,
        "credito": AccountNature.CREDIT,
        "d": AccountNature.DEBIT,
        "c": AccountNature.CREDIT,
        "debit": AccountNature.DEBIT,
        "credit": AccountNature.CREDIT,
    }
    if text in mapping:
        return mapping[text]
    raise ValueError(f"Natureza invalida: {value}.")


def _parse_status(value: object) -> str:
    text = _normalize_value(value)
    if not text:
        return StatusChoices.ACTIVE
    mapping = {
        "ativo": StatusChoices.ACTIVE,
        "active": StatusChoices.ACTIVE,
        "inativo": StatusChoices.INACTIVE,
        "inactive": StatusChoices.INACTIVE,
        "pendente": StatusChoices.PENDING,
        "pending": StatusChoices.PENDING,
    }
    if text in mapping:
        return mapping[text]
    raise ValueError(f"Status invalido: {value}.")


def _parse_bool_value(value: object, label: str) -> bool:
    text = _normalize_value(value)
    if not text:
        raise ValueError(f"{label} obrigatorio.")
    if text in {"true", "1", "yes", "y", "sim", "s", "x"}:
        return True
    if text in {"false", "0", "no", "n", "nao", "não"}:
        return False
    raise ValueError(f"{label} deve ser Sim/Nao.")


def _parse_dre_sign(value: object) -> str:
    text = _normalize_value(value)
    if not text:
        raise ValueError("Sinal DRE obrigatorio.")
    if text in {choice.value for choice in DreSign}:
        return text
    mapping = {
        "somar": DreSign.ADD,
        "soma": DreSign.ADD,
        "adicionar": DreSign.ADD,
        "positivo": DreSign.ADD,
        "+": DreSign.ADD,
        "subtrair": DreSign.SUBTRACT,
        "subtracao": DreSign.SUBTRACT,
        "negativo": DreSign.SUBTRACT,
        "-": DreSign.SUBTRACT,
    }
    if text in mapping:
        return mapping[text]
    raise ValueError(f"Sinal DRE invalido: {value}.")
